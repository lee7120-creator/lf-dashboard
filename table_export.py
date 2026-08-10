"""표 → 엑셀(xlsx) 내보내기 — 검은 헤더 스타일, 화면과 같은 서식.

발송성과·주간보고 두 대시보드가 같이 쓴다. Streamlit 기본 툴바의 'Download as CSV'는
서식이 다 날아가고 한글 인코딩도 흔들려서, 보고서에 바로 붙일 수 있는 표를 따로 만든다.

핵심은 **값을 문자열이 아니라 '숫자 + 엑셀 표시형식'으로 넣는 것** — 그래야 받는 쪽에서
정렬·합계가 된다. 화면 표시 문자열(천단위·%·부호)에서 표시형식을 역으로 추론한다.
"""
import io
import re

import numpy as np
import pandas as pd


XL_HEAD_BG = "111827"      # 거의 검정 (헤더)
XL_HEAD_FG = "FFFFFF"
XL_ZEBRA = "F6F7F9"
XL_BORDER = "D9DDE3"
_XL_PCT_RE = re.compile(r"^([+-]?)[\d,]*\.?(\d*)\s*%$")
_XL_NUM_RE = re.compile(r"^([+-]?)[\d,]+\.?(\d*)$")


def _xl_numfmt(vals):
    """화면 표시 문자열들 → 엑셀 숫자 서식. 숫자로 볼 수 없으면 None(=문자로 씀)."""
    vals = [str(v).strip() for v in vals if str(v).strip() not in ("", "–", "-", "nan", "None")]
    if not vals:
        return None
    pcts = [_XL_PCT_RE.match(v) for v in vals]
    if all(pcts):
        dec = max(len(m.group(2)) for m in pcts)
        base = "0." + "0" * dec + "%" if dec else "0%"
        return f"+{base};-{base}" if any(m.group(1) == "+" for m in pcts) else base
    nums = [_XL_NUM_RE.match(v) for v in vals]
    if all(nums):
        dec = max(len(m.group(2)) for m in nums)
        base = ("#,##0." + "0" * dec) if dec else "#,##0"
        return f"+{base};-{base}" if any(m.group(1) == "+" for m in nums) else base
    return None


_CSS_NAMED = {
    "red": "FF0000", "green": "008000", "blue": "0000FF", "black": "000000",
    "white": "FFFFFF", "gray": "808080", "grey": "808080", "orange": "FFA500",
    "crimson": "DC143C", "teal": "008080", "purple": "800080",
}
_RGB_RE = re.compile(r"rgba?\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)")


def _css_color(v):
    """CSS 색 → openpyxl용 6자리 RGB. 못 읽으면 None."""
    if not v:
        return None
    s = str(v).strip().lower()
    if s.startswith("#"):
        h = s[1:]
        if len(h) == 3:
            h = "".join(c * 2 for c in h)
        return h.upper()[:6] if len(h) >= 6 else None
    m = _RGB_RE.match(s)
    if m:
        return "".join(f"{min(255, int(g)):02X}" for g in m.groups())
    return _CSS_NAMED.get(s)


def _xl_display(obj):
    """Styler든 DataFrame이든 (원본 df, 표시 문자열 df, 셀 CSS)로 돌려준다.

    셀 CSS는 화면의 빨강(감소)·초록(증가) 글자색을 엑셀에도 그대로 입히려고 같이 뽑는다.
    """
    df = getattr(obj, "data", obj)
    if not isinstance(df, pd.DataFrame):
        df = pd.DataFrame(df)
    disp = df.astype(object).map(lambda v: "" if pd.isna(v) else str(v))
    css = {}
    if hasattr(obj, "_translate"):                        # Styler — 화면과 같은 서식으로
        # 순서가 중요하다: _translate는 끝나면서 ctx를 비운다. 색을 먼저 떠 놓고
        # 그다음에 표시값을 읽어야 한다. (반대로 하면 색이 통째로 날아간다)
        try:
            obj._compute()
            # ctx: (행, 열) → [('color', '#c0392b'), ('font-weight', '600'), ...]
            for pos, props in dict(getattr(obj, "ctx", {})).items():
                css[pos] = {str(k).strip().lower(): str(v).strip() for k, v in props}
        except Exception:                                 # noqa: BLE001
            css = {}
        try:
            body = obj._translate(False, False)["body"]
            rows = [[c.get("display_value") for c in r if c.get("type") == "td"] for r in body]
            rows = [r for r in rows if r]
            if len(rows) == len(df) and all(len(r) == df.shape[1] for r in rows):
                disp = pd.DataFrame(rows, index=df.index, columns=df.columns)
        except Exception:                                 # noqa: BLE001
            pass                                          # 실패해도 원본 문자열로 내보낸다
    return df, disp, css


def xlsx_bytes(obj, sheet_name="데이터", title=None, index=False):
    """표 → 검은 헤더 스타일 xlsx 바이트. Styler·DataFrame 모두 받는다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    df, disp, css = _xl_display(obj)
    ncol0 = df.shape[1]
    if index:
        df = df.reset_index()
        disp = disp.reset_index()
        # 인덱스를 앞에 끼우면 CSS의 열 번호가 한 칸씩 밀린다
        _shift = df.shape[1] - ncol0
        css = {(i, j + _shift): v for (i, j), v in css.items()}
    cols = ["" if c is None else str(c) for c in df.columns]

    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r"[\[\]:*?/\\]", " ", str(sheet_name))[:31] or "데이터"
    r0 = 1
    if title:
        ws.cell(row=1, column=1, value=str(title)).font = Font(bold=True, size=13, color="111827")
        r0 = 3
    thin = Side(style="thin", color=XL_BORDER)
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=r0, column=j, value=c)
        cell.fill = PatternFill("solid", fgColor=XL_HEAD_BG)
        cell.font = Font(bold=True, color=XL_HEAD_FG, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bd
    fmts = [_xl_numfmt(disp.iloc[:, j].tolist()) for j in range(df.shape[1])]
    zebra = PatternFill("solid", fgColor=XL_ZEBRA)
    for i in range(len(df)):
        for j in range(df.shape[1]):
            raw, txt = df.iat[i, j], disp.iat[i, j]
            cell = ws.cell(row=r0 + 1 + i, column=j + 1)
            if fmts[j] is not None and isinstance(raw, (int, float, np.integer, np.floating)) \
                    and not pd.isna(raw):
                cell.value = float(raw)
                cell.number_format = fmts[j]
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = "" if txt in ("nan", "None") else txt
                cell.alignment = Alignment(horizontal="left", vertical="center")
            # 화면의 글자색(빨강=감소·초록=증가)과 배경을 그대로 옮긴다
            _st = css.get((i, j), {})
            _fg = _css_color(_st.get("color")) or "1F2937"
            _bold = str(_st.get("font-weight", "")).strip() in ("bold", "600", "700", "800", "900")
            cell.font = Font(size=10, color=_fg, bold=_bold)
            _bgc = _css_color(_st.get("background-color"))
            if _bgc:                                      # 화면에서 칠한 배경이 우선
                cell.fill = PatternFill("solid", fgColor=_bgc)
            elif i % 2 == 1:
                cell.fill = zebra
            cell.border = bd
    for j, c in enumerate(cols, start=1):
        width = max([len(str(c))] + [len(str(x)) for x in disp.iloc[:, j - 1].tolist()[:200]])
        ws.column_dimensions[get_column_letter(j)].width = min(max(width * 1.6 + 2, 9), 42)
    ws.row_dimensions[r0].height = 24
    ws.freeze_panes = ws.cell(row=r0 + 1, column=1)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
