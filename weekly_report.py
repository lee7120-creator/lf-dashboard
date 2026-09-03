"""주간보고 — 첫구매 통합 실적 대시보드
여러 원천 엑셀(전체관점 마스터 + 지표별 파일)을 업로드하면 하나의 통합 뷰로 종합하고,
전년(YoY)·전주(WoW) 증감현황과 보고란을 갖춘 주간보고 화면을 만든다.
통합 결과는 (월/주/첫구매_요약 + 차트) 엑셀 워크북으로 다운로드할 수 있다.
"""

import datetime
import io, itertools, json, os, re
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st

# table_export는 같은 폴더의 모듈이다. 테스트 하네스가 앱만 임시 폴더로 복사해 돌리는
# 경우가 있어 경로를 직접 얹는다 (없으면 엑셀 버튼만 빠지고 앱은 정상 동작).
import sys as _sys
_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in _sys.path:
    _sys.path.insert(0, _HERE)
try:
    from table_export import xlsx_bytes
except Exception:                                         # noqa: BLE001
    xlsx_bytes = None

try:
    from streamlit_quill import st_quill
    HAS_QUILL = True
except Exception:
    HAS_QUILL = False

# ══════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════
st.set_page_config(page_title="주간보고 — 첫구매 통합 실적", page_icon="📋",
                   layout="wide", initial_sidebar_state="expanded")

INSIGHT_FILE = "wr_insights.json"

KST = datetime.timezone(datetime.timedelta(hours=9))
def today_kst():
    """서버가 UTC라 date.today()는 KST 새벽(00~09시)에 하루 밀린다 — '오늘'은 반드시 이걸 쓸 것"""
    return datetime.datetime.now(KST).date()

st.markdown("""
<style>
[data-testid="stAppViewContainer"]{background:#f8f9fc}
[data-testid="stSidebar"]{background:#ffffff;border-right:1px solid #e2e8f0}
[data-testid="stMetric"]{background:#ffffff;border-radius:8px;padding:12px 16px;border:1px solid #e2e8f0;box-shadow:0 1px 3px rgba(0,0,0,0.06)}
[data-testid="stMetricLabel"]{color:#64748b!important;font-size:12px!important}
[data-testid="stMetricValue"]{color:#1e293b!important;font-size:20px!important}
.sdiv{border-top:1px solid #e2e8f0;margin:22px 0}
.report-box{border:1px solid #e2e8f0;border-radius:8px;padding:12px 16px;margin:8px 0;line-height:1.7;background:#ffffff}
.report-box p{margin:0 0 4px}
.kpi-card{background:#ffffff;border:1px solid #e2e8f0;border-radius:8px;padding:12px 16px;box-shadow:0 1px 3px rgba(0,0,0,0.06)}
.kpi-label{color:#64748b;font-size:12px}
.kpi-value{color:#1e293b;font-size:21px;font-weight:600;margin:2px 0 8px}
.kpi-delta{display:block;width:fit-content;font-size:12px;border-radius:6px;padding:2px 8px;margin:4px 0 0;font-weight:500;white-space:nowrap}
.kpi-delta.up{background:#ecfdf5;color:#15803d}
.kpi-delta.down{background:#fef2f2;color:#dc2626}
.kpi-delta.na{background:#f1f5f9;color:#94a3b8}
@media print {
  @page { margin: 12mm; }
  [data-testid="stSidebar"], [data-testid="stHeader"], [data-testid="stToolbar"],
  header, .stButton, .no-print, iframe,
  [data-testid="stExpander"] { display:none !important; }
  [data-testid="stAppViewContainer"], .main, .block-container { background:#fff !important; }
  .block-container { max-width:100% !important; padding-top:0 !important; }

  /* 겹침 방지: 레이아웃 블록을 정적 배치하고 넘침을 그대로 노출 */
  [data-testid="stVerticalBlock"], [data-testid="stHorizontalBlock"],
  [data-testid="stVerticalBlockBorderWrapper"], [data-testid="column"],
  .element-container { position:static !important; transform:none !important;
    overflow:visible !important; }

  /* 인쇄 시 Plotly 차트 높이 붕괴 → 아래 요소가 제목 위로 밀려 겹치는 문제 차단 */
  .stPlotlyChart, .js-plotly-plot, [data-testid="stPlotlyChart"] {
    min-height:240px !important; break-inside:avoid; page-break-inside:avoid; }

  /* 제목이 투명 배경으로 다른 요소와 겹쳐 보이지 않도록 */
  h1, h2, h3, h4 { background:#fff !important; position:relative; z-index:1;
    page-break-after:avoid; break-after:avoid; }

  .stPlotlyChart, .report-box, table,
  [data-testid="stMetric"], [data-testid="column"] {
    break-inside:avoid; page-break-inside:avoid; }

  /* 증감 색상(빨강/초록) 인쇄에 유지 */
  * { -webkit-print-color-adjust:exact !important; print-color-adjust:exact !important; }
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
# 색상 팔레트 (기존 first_purchase 와 동일 계열)
# ══════════════════════════════════════════════════════
PALETTE = {
    "blue":   ("rgba(79,143,255,1)",  "rgba(79,143,255,0.15)"),
    "red":    ("rgba(245,101,101,1)", "rgba(245,101,101,0.15)"),
    "green":  ("rgba(72,187,120,1)",  "rgba(72,187,120,0.15)"),
    "amber":  ("rgba(237,137,54,1)",  "rgba(237,137,54,0.15)"),
    "purple": ("rgba(159,122,234,1)", "rgba(159,122,234,0.15)"),
    "teal":   ("rgba(56,178,172,1)",  "rgba(56,178,172,0.15)"),
    "orange": ("rgba(249,115,22,1)",  "rgba(249,115,22,0.15)"),
    "slate":  ("rgba(100,116,139,1)", "rgba(100,116,139,0.15)"),
}
def clr(n): return PALETTE.get(n, PALETTE["blue"])[0]
def cbg(n): return PALETTE.get(n, PALETTE["blue"])[1]

CHANNEL_PAL = {
    "직접": "blue", "광고": "amber", "EP": "green", "PUSH": "purple",
    "제휴": "red", "브랜드광고": "teal", "미디어커머스": "orange", "*TOTAL": "slate",
}
CHANNELS = ["직접", "광고", "EP", "PUSH", "제휴", "브랜드광고", "미디어커머스"]
YEAR_PAL = ["slate", "blue", "red", "green", "purple", "amber", "teal"]

# ══════════════════════════════════════════════════════
# 지표 정의
# ══════════════════════════════════════════════════════
METRICS7 = ["첫구매 거래액", "첫구매 고객수", "첫구매 객단가",
            "비회원트래픽", "가입자수", "가입율", "당일가입CR"]
PCT_METRICS = {"가입율", "당일가입CR", "진입률", "CR", "거래액비중", "고객비중", "동의율", "유입율",
               "상품CR"}
# 마스터 파일 지표 → 보고서 지표 매핑
MASTER_MAP = {"일평균거래액": "첫구매 거래액", "일평균고객수": "첫구매 고객수",
              "일평균객단가": "첫구매 객단가"}
# 지표별 파일명 → 보고서 지표 매핑 (공백 제거 후 매칭)
METRIC_FILE_MAP = {
    "가입율": "가입율", "가입률": "가입율",
    "가입자수": "가입자수",
    "당일가입첫구매율": "당일가입CR", "당일가입CR": "당일가입CR",
    "비회원트래픽": "비회원트래픽",
}

METRIC_UNIT = {
    "첫구매 거래액": ("백만원", 1e6), "첫구매 고객수": ("명", 1),
    "첫구매 객단가": ("원", 1), "비회원트래픽": ("명", 1),
    "가입자수": ("명", 1), "가입율": ("%", 1), "당일가입CR": ("%", 1),
    "앱푸시수신동의": ("명", 1), "앱푸시_동의자수": ("명", 1),
    "앱푸시_신규추가": ("명", 1), "앱푸시_이탈": ("명", 1),
    "앱푸시_유효회원": ("명", 1), "앱푸시_수신동의전체": ("명", 1),
    "상품UV": ("명", 1), "상품CR": ("%", 1), "거래액비중": ("%", 1), "고객비중": ("%", 1),
}

# 숫자·영문으로 끝나는 이름(e-영업1·SPACE-R)도 읽는 소리로 받침을 판단한다
_JOSA_TAIL = {**{d: t for d, t in zip("0123456789", [1, 1, 0, 1, 0, 0, 1, 1, 1, 0])},
              **{c: 1 for c in "LMNR"}, **{c.lower(): 1 for c in "LMNR"}}


def josa(word, pair="은는"):
    """받침 유무로 조사를 고른다 — '객단가이 가장'·'조직는' 같은 문장을 막는다."""
    w = str(word).strip()
    if not w:
        return pair[1]
    ch = w[-1]
    if "가" <= ch <= "힣":
        return pair[0] if (ord(ch) - 0xAC00) % 28 else pair[1]
    return pair[0] if _JOSA_TAIL.get(ch) else pair[1]


def esc(v):
    """업로드 파일에서 온 문자열을 unsafe_allow_html에 넣기 전에 이스케이프.
    조직·카테고리 이름에 &·< 가 섞이면 태그로 오해석된다."""
    return (str("" if v is None else v)
            .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;").replace("'", "&#39;"))


def fmt_value(metric, v):
    if v is None or (isinstance(v, float) and np.isnan(v)): return "–"
    if metric in PCT_METRICS: return f"{v*100:.2f}%"
    if metric == "첫구매 거래액": return f"{v/1e6:,.1f}백만원"
    if metric == "첫구매 객단가": return f"{v:,.0f}원"
    return f"{int(v):,}명"  # 명 단위(고객수 등)는 소수점 버림

def fmt_delta(metric, cur, prev):
    """전년비/전주비 문자열: 비율 지표는 %p 차이, 그 외 증감율"""
    if cur is None or prev is None: return None
    if isinstance(cur, float) and np.isnan(cur): return None
    if isinstance(prev, float) and np.isnan(prev): return None
    if metric in PCT_METRICS:
        d = (cur - prev) * 100
        return f"△{abs(d):.2f}%p" if d < 0 else f"+{d:.2f}%p"
    if prev == 0: return None
    d = (cur - prev) / prev * 100
    return f"△{abs(d):.1f}%" if d < 0 else f"+{d:.1f}%"

def style_delta_cols(tbl):
    """증감 컬럼(△/+)에 빨강/초록 색상 적용한 Styler 반환"""
    delta_cols = [c for c in tbl.columns
                  if any(k in str(c) for k in ("전년비", "전주비", "전월비", "증감"))]
    def _color(v):
        s = str(v)
        if s.startswith("△"): return "color:#dc2626;font-weight:600"
        if s.startswith("+"): return "color:#16a34a;font-weight:600"
        return ""
    try:
        return tbl.style.map(_color, subset=delta_cols)
    except Exception:
        return tbl

# ══════════════════════════════════════════════════════
# 파싱 — 원천 파일 → 통합 long DataFrame
# ══════════════════════════════════════════════════════
YEAR_RE   = re.compile(r"^(20\d{2})(\.0)?$")
PERIOD_RE = re.compile(r"^\s*(\d{1,2}\s*월(\s*\d\s*주차)?|\d{1,2}/\d{1,2})\s*$")

def detect_file(name):
    """파일명에서 (kind, granularity 힌트, metric) 감지.
    단위(일/주/월)와 마스터 여부는 parse_file에서 내용으로 최종 판별하므로 여기선 힌트만."""
    base = os.path.basename(name)
    base = re.sub(r"\.(xlsx|xls|csv)$", "", base, flags=re.I)
    key = base.replace(" ", "")
    # 단위 힌트: 일자별/주별/월별 키워드 또는 일_/주_/월_ 접두사
    gran = None
    if "일자별" in key or "데일리" in key: gran = "일"
    elif "주별" in key or "주간" in key:   gran = "주"
    elif "월별" in key or "월간" in key:   gran = "월"
    else:
        m = re.match(r"^(일|주|월)[_\s]", base)
        if m: gran = m.group(1)
    if "전체관점" in key or "마스터" in key:
        return "master", gran, None
    # 지표 키워드는 파일명 어느 위치에 있어도 인식
    for k, v in METRIC_FILE_MAP.items():
        if k.replace(" ", "") in key:
            return "metric", gran, v
    # 접두사형(월_xxx)인데 모르는 지표면 정리된 이름 그대로 사용
    m = re.match(r"^(일|주|월)[_\s]+(.+)$", base)
    if m:
        rest = re.sub(r"\(.*?\)", "", m.group(2)).strip()
        return "metric", m.group(1), rest or "기타"
    return None, gran, None

def _decode_text(data: bytes) -> str:
    for enc in ("utf-16", "utf-8-sig", "cp949", "utf-8"):
        try: return data.decode(enc)
        except (UnicodeDecodeError, UnicodeError): pass
    return data.decode("utf-8", "replace")

def read_grid(name, data: bytes):
    """csv(UTF-16/탭 포함)·xlsx 모두 2차원 셀 그리드로 읽는다"""
    if name.lower().endswith((".csv", ".txt", ".tsv")):
        text = _decode_text(data)
        return [line.split("\t") for line in text.splitlines()]
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    return [list(row) for row in ws.iter_rows(values_only=True)]

def _cell(v):
    return "" if v is None else str(v).strip()

def _num(v):
    """셀 값 → float (콤마·% 처리, %는 비율로 변환)"""
    if v is None: return np.nan
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(",", "")
    if s in ("", "-", "–"): return np.nan
    pct = s.endswith("%")
    if pct: s = s[:-1]
    try: f = float(s)
    except ValueError: return np.nan
    return f / 100 if pct else f

def period_parts(gran, year, plabel):
    """기간 라벨 → (표준라벨, 정렬키)"""
    p = plabel.replace(" ", "")
    if gran == "월":
        m = re.match(r"(\d{1,2})월", p)
        if not m: return None
        mo = int(m.group(1))
        return f"{mo}월", year * 10000 + mo * 100
    if gran == "주":
        m = re.match(r"(\d{1,2})월(\d)주차", p)
        if not m: return None
        mo, wk = int(m.group(1)), int(m.group(2))
        return f"{mo:02d}월 {wk}주차", year * 10000 + mo * 100 + wk
    m = re.match(r"(\d{1,2})/(\d{1,2})", p)
    if not m: return None
    mo, dd = int(m.group(1)), int(m.group(2))
    return f"{mo}/{dd}", year * 10000 + mo * 100 + dd

def parse_file(name, data: bytes) -> pd.DataFrame:
    # 누적 데이터 백업 CSV(long 포맷) 재업로드 시 그대로 복원
    if name.lower().endswith(".csv"):
        first = data[:400].decode("utf-8", "ignore").splitlines()[0] if data else ""
        if "gran" in first and "metric" in first and "sortkey" in first:
            try:
                d = pd.read_csv(io.BytesIO(data), encoding="utf-8-sig")
                if set(STORE_COLS) <= set(d.columns):
                    return d[STORE_COLS]
            except Exception:
                pass
    kind, gran, file_metric = detect_file(name)
    try:
        rows = read_grid(name, data)
    except Exception:
        return pd.DataFrame()
    if not rows: return pd.DataFrame()

    # 헤더 행 탐색 (앞 6행): 연도 / 기간라벨 / 마감구분
    year_row = period_row = close_row = None
    for ri in range(min(6, len(rows))):
        cells = [_cell(c) for c in rows[ri]]
        nY = sum(1 for c in cells if YEAR_RE.match(c))
        nP = sum(1 for c in cells if PERIOD_RE.match(c))
        nC = sum(1 for c in cells if "마감" in c)
        if year_row is None and nY >= 1 and nP == 0: year_row = ri
        if period_row is None and nP >= 2: period_row = ri
        if close_row is None and nC >= 2: close_row = ri
    if period_row is None: return pd.DataFrame()
    data_start = max(r for r in (year_row, period_row, close_row) if r is not None) + 1

    # ── 내용 기반 최종 판별 (파일명은 힌트일 뿐)
    # 단위: 기간 라벨 형태가 결정 (N월 N주차 → 주, N/N → 일, N월 → 월)
    plabels = [_cell(c) for c in rows[period_row] if PERIOD_RE.match(_cell(c))]
    if any("주차" in p for p in plabels):  gran = "주"
    elif any("/" in p for p in plabels):   gran = "일"
    else:                                   gran = "월"
    # 마스터 여부: 헤더에 '구분01'이 있으면 마스터(전체관점) 구조
    if any("구분01" in _cell(c) for ri in range(data_start) for c in rows[ri]):
        kind = "master"
    if kind is None or (kind == "metric" and not file_metric):
        return pd.DataFrame()

    ncols = max(len(r) for r in rows)
    def cell(ri, ci):
        return _cell(rows[ri][ci]) if ri is not None and ci < len(rows[ri]) else ""

    # 컬럼별 연도(좌측 ffill)·기간·마감
    col_year, cur_y = {}, None
    for ci in range(ncols):
        m = YEAR_RE.match(cell(year_row, ci)) if year_row is not None else None
        if m: cur_y = int(m.group(1))
        col_year[ci] = cur_y
    # 기간 라벨이 빈 셀(병합)인 일마감/MTD 컬럼은 직전 라벨을 이어받는다
    data_cols, col_label, last_lbl = [], {}, None
    for ci in range(ncols):
        lbl = cell(period_row, ci)
        if PERIOD_RE.match(lbl) and col_year[ci]:
            last_lbl = lbl
            col_label[ci] = lbl
            data_cols.append(ci)
        elif (close_row is not None and "마감" in cell(close_row, ci)
              and last_lbl and col_year[ci]):
            col_label[ci] = last_lbl
            data_cols.append(ci)

    seg_col = 1 if kind == "master" else 0
    records, cur_metric = [], None
    for ri in range(data_start, len(rows)):
        if kind == "master":
            m0 = cell(ri, 0)
            if m0 and m0 not in ("-", "–"): cur_metric = m0
            metric = MASTER_MAP.get(cur_metric, cur_metric)
        else:
            metric = file_metric
        seg = cell(ri, seg_col)
        # '채널'은 세그먼트 컬럼의 헤더 텍스트 (해당 행 값은 요일) — 데이터 행이 아니므로 스킵
        if not seg or seg in ("-", "–", "채널") or metric is None: continue
        for ci in data_cols:
            pp = period_parts(gran, col_year[ci], col_label[ci])
            if pp is None: continue
            label, sortkey = pp
            close = cell(close_row, ci) if close_row is not None else ""
            records.append({
                "gran": gran, "metric": metric, "segment": seg,
                "year": col_year[ci], "label": label, "sortkey": sortkey,
                "close": "mtd" if "일마감" in close and gran != "일" else "final",
                "value": _num(rows[ri][ci] if ci < len(rows[ri]) else None),
            })
    return pd.DataFrame(records)

# ══════════════════════════════════════════════════════
# 조직 × 카테고리 실적 (MICRO 대시보드 export)
# ══════════════════════════════════════════════════════
# 마스터(전체관점)와 달리 세그먼트 축이 **둘**이다 — 구분06=조직(BPU), 구분07=카테고리.
# 기존 store는 `segment` 한 칸뿐이라 여기에 밀어 넣으면 두 축이 뭉개진다. 그래서
# 별도 store(`wr_orgcat_store.csv`)에 org/cat 두 칸으로 쌓는다.
#
# LFMS 포함여부(Y/N)는 헤더 2행에 실려 오는 **다른 모집단**이다. 키에 넣지 않으면
# 같은 기간·같은 조직 값이 서로를 조용히 덮어쓴다 — 화면에서 필터로 고르게 둔다.
ORGCAT_STORE = "wr_orgcat_store.csv"
ORGCAT_KEY = ["gran", "metric", "org", "cat", "lfms", "year", "label", "close"]
ORGCAT_COLS = ORGCAT_KEY + ["sortkey", "value"]
# 마스터 파일과 같은 지표는 같은 이름으로 — fmt_value·PCT_METRICS를 그대로 태운다
ORGCAT_MAP = {"일평균거래액": "첫구매 거래액", "일평균고객수": "첫구매 고객수",
              "일평균객단가": "첫구매 객단가"}
ORGCAT_TOTAL = "*TOTAL"


def is_orgcat_grid(rows):
    """구분06·구분07 헤더가 있으면 조직×카테고리 export."""
    for r in rows[:8]:
        cells = [_cell(c) for c in r]
        if "구분06" in cells and "구분07" in cells:
            return True
    return False


def parse_orgcat_grid(rows):
    """조직×카테고리 그리드 → long DataFrame.

    헤더 구성이 단위마다 다르다. 월·주는 `연도 / LFMS / 기간 / (구분+마감)` 4행이고,
    일은 마감 행이 없어 `연도 / LFMS / (구분+기간)` 3행이다. 그래서 행 번호를 박지 않고
    내용으로 찾는다 — 구분06이 있는 행, 연도가 있는 행, Y/N만 있는 행, 기간 라벨 행.
    """
    hdr = c_org = c_cat = None
    for ri, r in enumerate(rows[:8]):
        cells = [_cell(c) for c in r]
        if "구분06" in cells and "구분07" in cells:
            hdr, c_org, c_cat = ri, cells.index("구분06"), cells.index("구분07")
            break
    if hdr is None:
        return pd.DataFrame()
    ncols = max(len(r) for r in rows)

    def cell(ri, ci):
        return _cell(rows[ri][ci]) if ri is not None and ci < len(rows[ri]) else ""

    year_row = lfms_row = period_row = close_row = None
    for ri in range(min(hdr + 1, len(rows))):
        cells = [_cell(c) for c in rows[ri]]
        nz = [c for c in cells if c]
        nY = sum(1 for c in cells if YEAR_RE.match(c))
        nP = sum(1 for c in cells if PERIOD_RE.match(c))
        nC = sum(1 for c in cells if "마감" in c)
        if year_row is None and nY >= 1 and nP == 0: year_row = ri
        if lfms_row is None and nz and all(c in ("Y", "N") for c in nz): lfms_row = ri
        if period_row is None and nP >= 2: period_row = ri
        if close_row is None and nC >= 2: close_row = ri
    if period_row is None:
        return pd.DataFrame()
    data_start = max(r for r in (hdr, year_row, period_row, close_row)
                     if r is not None) + 1

    plabels = [_cell(c) for c in rows[period_row] if PERIOD_RE.match(_cell(c))]
    if any("주차" in p for p in plabels):  gran = "주"
    elif any("/" in p for p in plabels):   gran = "일"
    else:                                   gran = "월"

    # 연도·LFMS는 병합셀이라 왼쪽 값을 오른쪽으로 이어받는다 (한 파일에 2개년이 온다)
    col_year, col_lfms, cur_y, cur_l = {}, {}, None, ""
    for ci in range(ncols):
        m = YEAR_RE.match(cell(year_row, ci)) if year_row is not None else None
        if m: cur_y = int(m.group(1))
        v = cell(lfms_row, ci) if lfms_row is not None else ""
        if v in ("Y", "N"): cur_l = v
        col_year[ci], col_lfms[ci] = cur_y, cur_l

    # 기간 라벨이 빈 '일마감'(MTD) 칼럼은 직전 라벨을 이어받는다 (마스터 파서와 동일)
    data_cols, col_label, last_lbl = [], {}, None
    for ci in range(ncols):
        lbl = cell(period_row, ci)
        if PERIOD_RE.match(lbl) and col_year[ci]:
            last_lbl = lbl; col_label[ci] = lbl; data_cols.append(ci)
        elif (close_row is not None and "마감" in cell(close_row, ci)
              and last_lbl and col_year[ci]):
            col_label[ci] = last_lbl; data_cols.append(ci)

    records, metric, org = [], None, None
    for ri in range(data_start, len(rows)):
        m0 = cell(ri, 0)
        if m0 and m0 not in ("-", "–"):
            metric = ORGCAT_MAP.get(m0, m0)
        o = cell(ri, c_org)
        if o and o not in ("-", "–"):
            org = o                                   # 조직은 병합셀 — 아래로 이어받는다
        cat = cell(ri, c_cat)
        # 카테고리 '-'는 카테고리 구분이 없는 조직(SPACE-R 등)의 자리표시라 *TOTAL과 겹친다
        if not metric or not org or not cat or cat in ("-", "–"):
            continue
        for ci in data_cols:
            pp = period_parts(gran, col_year[ci], col_label[ci])
            if pp is None: continue
            label, sortkey = pp
            close = cell(close_row, ci) if close_row is not None else ""
            records.append({
                "gran": gran, "metric": metric, "org": org, "cat": cat,
                "lfms": col_lfms[ci] or "N", "year": col_year[ci],
                "label": label, "sortkey": sortkey,
                "close": "mtd" if "일마감" in close and gran != "일" else "final",
                "value": _num(rows[ri][ci] if ci < len(rows[ri]) else None),
            })
    return pd.DataFrame(records)


@st.cache_data(show_spinner=False)
def parse_orgcat_file(name, data: bytes) -> pd.DataFrame:
    """업로드 1건 → 조직×카테고리 long DF. 이 형식이 아니면 빈 DF.

    라우팅·인식목록·누적 병합 세 군데서 같은 파일을 물어보므로 캐시해 둔다
    (일별 파일이 250칼럼짜리라 매번 다시 읽으면 업로드가 눈에 띄게 느려진다)."""
    # 백업 CSV 재업로드는 그대로 복원
    if name.lower().endswith(".csv"):
        head = data[:400].decode("utf-8", "ignore").splitlines()
        if head and all(k in head[0] for k in ("org", "cat", "lfms")):
            try:
                d = pd.read_csv(io.BytesIO(data), encoding="utf-8-sig")
                if set(ORGCAT_COLS) <= set(d.columns):
                    return d[ORGCAT_COLS]
            except Exception:
                pass
        return pd.DataFrame()
    if not name.lower().endswith((".xlsx", ".xls")):
        return pd.DataFrame()
    try:
        rows = read_grid(name, data)
    except Exception:
        return pd.DataFrame()
    if not rows or not is_orgcat_grid(rows):
        return pd.DataFrame()
    return parse_orgcat_grid(rows)


def _zip_entry_name(info):
    """zip 내 한글 파일명 복원 (UTF-8 플래그 없으면 cp437→cp949 재해석)"""
    if info.flag_bits & 0x800:
        return info.filename
    for enc in ("cp949", "utf-8", "euc-kr"):
        try: return info.filename.encode("cp437").decode(enc)
        except (UnicodeDecodeError, UnicodeEncodeError): pass
    return info.filename

def expand_uploads(uploads):
    """업로드 목록 → (이름, bytes) 목록. zip은 풀어서 내부 엑셀/CSV를 꺼낸다"""
    out = []
    for f in uploads:
        data = f.getvalue()
        if f.name.lower().endswith(".zip"):
            import zipfile
            try:
                zf = zipfile.ZipFile(io.BytesIO(data))
            except zipfile.BadZipFile:
                continue
            with zf:
                for info in zf.infolist():
                    if info.is_dir(): continue
                    name = _zip_entry_name(info)
                    if "__MACOSX" in name: continue
                    base = os.path.basename(name)
                    if base.startswith((".", "~")): continue
                    if not base.lower().endswith((".xlsx", ".xls", ".csv")): continue
                    out.append((base, zf.read(info)))
        else:
            out.append((f.name, data))
    return out

# (지표·세그먼트·연도)별 일 |중앙값| 대비 이 배수 초과 시 일회성 스파이크로 간주.
# 실측: 정상 피크 최대 2.9배(빼빼로데이 등 실제 이벤트 2.4~2.9배) vs
#       인공 스파이크 최소 5.4배(7/28~30 이관성 이탈 5.4~8.7배, 재동의·재분류 40~711배)
#       → 4로 설정하면 실제 이벤트는 보존되고 인공 스파이크만 제거된다
PUSH_SPIKE_RATIO = 4
# 흐름(flow) 지표만 마스킹 대상 — 잔고(앱푸시_동의자수)는 수준값이라 별도 글리치 규칙 사용
PUSH_FLOW_METRICS = {"앱푸시수신동의", "앱푸시_신규추가", "앱푸시_이탈"}
# 잔고(수준) 지표 — 하루짜리 급변 글리치 규칙 대상
PUSH_LEVEL_METRICS = {"앱푸시_동의자수", "앱푸시_유효회원", "앱푸시_수신동의전체"}

def mask_push_spikes(d: pd.DataFrame) -> pd.DataFrame:
    """앱푸시 흐름 지표의 대량 이관/재동의/재분류 스파이크 마스킹 (NaN 처리).
    (지표,세그먼트,연도)별 일 |중앙값|의 PUSH_SPIKE_RATIO배 초과 값을 이상치로 본다.
    (실데이터: 2025-04-19 재동의 16,496건=중앙값 76배, 2026-04-24 15,441건=84배,
     2025-01-01 세그먼트 재분류 6만건대. 연도마다 날짜가 달라 하드코딩 대신 통계 규칙 사용)"""
    m = d["metric"].isin(PUSH_FLOW_METRICS)
    if not m.any():
        return d
    med = (d.loc[m].groupby(["metric", "segment", "year"])["value"]
           .transform(lambda s: s.abs().median()))
    spike = d.loc[m, "value"].abs() > med * PUSH_SPIKE_RATIO
    d.loc[spike[spike].index, "value"] = np.nan
    return d

def mask_level_glitches(d: pd.DataFrame, rel=0.3) -> pd.DataFrame:
    """잔고(수준) 지표의 하루짜리 급변 글리치 마스킹.
    양옆 이웃과 rel(30%) 이상 괴리하면서 이웃끼리는 서로 비슷(rel/2 이내)하면 그 날만 NaN.
    지속되는 계단(연초 신규→기존 재분류, 4/19 재이관 등)은 한쪽 이웃과 일치하므로 보존.
    (실데이터: 2025-04-18 이관일에 잔고가 0으로 기록된 글리치 — 3개 세그먼트 공통)
    d에는 'dt'(datetime) 컬럼이 있어야 한다."""
    m = d["metric"].isin(PUSH_LEVEL_METRICS)
    if not m.any():
        return d
    for _seg, g in d[m].groupby("segment"):
        g = g.sort_values("dt")
        v = g["value"].to_numpy(float)
        if len(v) < 3:
            continue
        prev, nxt = np.roll(v, 1), np.roll(v, -1)
        dev_p = np.abs(v - prev) / np.maximum(np.abs(prev), 1)
        dev_n = np.abs(v - nxt) / np.maximum(np.abs(nxt), 1)
        agree = np.abs(nxt - prev) / np.maximum(np.abs(prev), 1) < rel / 2
        bad = (dev_p > rel) & (dev_n > rel) & agree
        bad[0] = bad[-1] = False
        d.loc[g.index[bad], "value"] = np.nan
    return d

def parse_push_file(name, data: bytes) -> pd.DataFrame:
    """PUSH 원천: 기존/신규/Total 3섹션 × (수신동의 누적·신규추가·이탈) 전부 파싱.
    순증감은 저장하지 않고 화면에서 신규추가-이탈로 파생 (마스킹 일관성 유지)."""
    try:
        rows = read_grid(name, data)
    except Exception:
        return pd.DataFrame()
    if not rows or len(rows) < 6: return pd.DataFrame()

    # 1) 날짜 헤더 행 탐색 — 파일마다 'Date' 라벨 유무·행 위치가 달라(0행 또는 1행)
    #    M/D 형태 셀이 가장 많은 행을 날짜 행으로 본다 (앞 6행 중)
    def _date_cols_of(row):
        out = []
        for ci in range(2, len(row)):
            m = re.match(r"^(\d{1,2})/\d{1,2}$", _cell(row[ci]))
            if m: out.append((ci, _cell(row[ci]), int(m.group(1))))
        return out
    date_cols, date_ri = [], None
    for ri in range(min(6, len(rows))):
        cand = _date_cols_of(rows[ri])
        if len(cand) > len(date_cols):
            date_cols, date_ri = cand, ri
    if len(date_cols) < 3:
        return pd.DataFrame()

    # 2) 뒤에서부터 역순으로 읽으면서 해(year)가 바뀌는 지점 계산
    # 맨 마지막 데이터를 0으로 두고, (1월 <- 12월)로 넘어갈 때마다 연도를 -1
    rel_years = {}
    current_rel_year = 0
    last_month = date_cols[-1][2]

    for ci, d, month in reversed(date_cols):
        # 역순으로 읽을 때 월이 1에서 12로 커지면 (실제로는 12월 -> 1월) 해가 바뀐 것
        if last_month < 6 and month > 6:
            current_rel_year -= 1
        elif last_month > 6 and month < 6:
            current_rel_year += 1

        rel_years[ci] = current_rel_year
        last_month = month

    # 3) 섹션(A열) × 행유형(B열) → (지표, 세그먼트) 매핑으로 대상 행 수집
    SECTION_SEG = {"기존": "기존", "신규": "신규", "Total": "*TOTAL"}
    # 행 라벨은 공백 유무가 파일마다 달라(수신동의/수신 동의) 공백 제거 후 매칭
    _rk = lambda s: _cell(s).replace(" ", "")
    labels = {_rk(r[1]) for r in rows if len(r) > 1}
    # 3단 표(전체 유효 회원/수신 동의/타겟팅 가능)인지 판별.
    # 이 표에서 '수신 동의'는 앱 수신동의 전체(수백만)이고, 실제 발송 가능 모수는
    # '타겟팅 가능'이다. 구 2단 표의 '수신동의'는 이 '타겟팅 가능'과 동일한 수치이므로
    # (실측 대조 확인) 두 레이아웃 모두 앱푸시_동의자수로 이어붙여 시계열 연속성을 유지한다.
    three_tier = "타겟팅가능" in labels
    ROW_METRIC = {"신규추가(+)": "앱푸시_신규추가", "기존이탈(-)": "앱푸시_이탈",
                  "전체유효회원": "앱푸시_유효회원", "타겟팅가능": "앱푸시_동의자수"}
    ROW_METRIC["수신동의"] = "앱푸시_수신동의전체" if three_tier else "앱푸시_동의자수"
    targets = []  # (metric, segment, row)
    cur_seg = None
    for row in rows:
        c0 = _cell(row[0] if len(row) > 0 else "")
        c1 = _rk(row[1] if len(row) > 1 else "")
        if c0 in SECTION_SEG: cur_seg = SECTION_SEG[c0]
        if cur_seg and c1 in ROW_METRIC:
            targets.append((ROW_METRIC[c1], cur_seg, row))
            # 신규 섹션 신규추가는 기존 지표명으로도 저장 (구버전 누적 데이터·동의율 로직과 연속성)
            if cur_seg == "신규" and c1 == "신규추가(+)":
                targets.append(("앱푸시수신동의", "*TOTAL", row))

    if not targets: return pd.DataFrame()

    records = []
    for metric, seg, trow in targets:
        for ci, d, month in date_cols:
            val = _num(trow[ci] if ci < len(trow) else None)
            if pd.isna(val): continue
            records.append({
                "gran": "일", "metric": metric, "segment": seg,
                "year": rel_years[ci], "label": d, "sortkey": 0, "close": "final", "value": val
            })
    return pd.DataFrame(records)

def looks_like_push_name(name: str) -> bool:
    """파일명이 앱푸시 원천으로 보이는가 (한글명 포함)"""
    up = name.upper()
    return "PUSH" in up or any(h in name for h in ("앱푸시", "푸시", "수신동의"))

def route_push(n, b):
    """엑셀 1건 → (push_df 또는 None, 일반_df 또는 None).
    이름 힌트가 있으면 PUSH 우선 시도, 없으면 일반 파싱 후 빈 결과면 내용 기반으로 PUSH 재시도.
    → 파일명이 'PUSH'가 아니어도(예: 앱푸시수신동의현황.xlsx) 인식된다.
    조직×카테고리 export는 combine_orgcat이 따로 받으므로 여기선 둘 다 None."""
    if not parse_orgcat_file(n, b).empty:
        return None, None
    is_xlsx = n.lower().endswith((".xlsx", ".xls"))
    if is_xlsx and looks_like_push_name(n):
        pf = parse_push_file(n, b)
        if not pf.empty: return pf, None
    d = parse_file(n, b)
    if not d.empty: return None, d
    if is_xlsx:  # 이름 힌트 없이도 헤더가 PUSH 형식이면 인식 (내용 기반 폴백)
        pf = parse_push_file(n, b)
        if not pf.empty: return pf, None
    return None, None

@st.cache_data(show_spinner=False)
def combine_files(file_tuples) -> pd.DataFrame:
    """업로드 파일들 → 통합 long DF. 동일 키는 마지막 파일 우선"""
    frames = []
    push_frames = []
    for n, b in file_tuples:
        pf, df = route_push(n, b)
        if pf is not None: push_frames.append(pf)
        elif df is not None: frames.append(df)

    inferred_years = set()
    for f in frames:
        if "year" in f.columns:
            inferred_years.update(f["year"].dropna().unique())
            
    default_year = int(max(inferred_years)) if inferred_years else today_kst().year

    # PUSH 파일엔 연도가 없어 상대연도(rel)만 있다 — 함께 올린 일자 데이터와 (연도,일자)
    # 겹침이 가장 큰 기준연도를 고른다 (예: 2025 가입자수만 곁들여 올려도 2024/2025 오배정 방지)
    if push_frames and frames:
        daily_keys = set()
        for f in frames:
            d = f[f["gran"] == "일"][["year", "label"]].dropna()
            daily_keys.update(map(tuple, d.values.tolist()))
        if daily_keys:
            def _overlap(base):
                return sum(1 for pf in push_frames
                           for ry, lb in pf[["year", "label"]].values.tolist()
                           if (base + ry, lb) in daily_keys)
            cands = sorted({default_year, default_year + 1, today_kst().year})
            best = max(cands, key=_overlap)
            if _overlap(best) > _overlap(default_year):
                default_year = int(best)

    for pf in push_frames:
        pf["year"] = default_year + pf["year"]
        def calc_sortkey(row):
            m = re.match(r"(\d+)/(\d+)", str(row["label"]))
            if m: return int(row["year"]) * 10000 + int(m.group(1))*100 + int(m.group(2))
            return 0
        pf["sortkey"] = pf.apply(calc_sortkey, axis=1)
        frames.append(pf)

    if not frames: return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    df = df.drop_duplicates(subset=KEY_COLS, keep="last")
    return df

# ══════════════════════════════════════════════════════
# 데이터 누적 저장소 — 업로드할 때마다 병합·저장
# ══════════════════════════════════════════════════════
DATA_STORE = "wr_data_store.csv"
KEY_COLS = ["gran", "metric", "segment", "year", "label", "close"]
STORE_COLS = KEY_COLS + ["sortkey", "value"]

def load_store() -> pd.DataFrame:
    if os.path.exists(DATA_STORE):
        try:
            d = pd.read_csv(DATA_STORE, encoding="utf-8-sig")
            if set(STORE_COLS) <= set(d.columns):
                # 과거 버전이 요일 헤더 행을 '채널' 세그먼트(전부 NaN)로 오파싱해 저장한 쓰레기 제거
                d = d[~((d["segment"] == "채널") & d["value"].isna())]
                return d[STORE_COLS]
        except Exception:
            pass
    return pd.DataFrame()

def save_store(df: pd.DataFrame):
    df[STORE_COLS].to_csv(DATA_STORE, index=False, encoding="utf-8-sig")

def merge_store(old: pd.DataFrame, new: pd.DataFrame) -> pd.DataFrame:
    """기존 누적 + 신규 업로드 병합 — 같은 (단위·지표·채널·기간) 키는 신규 우선"""
    if old is None or old.empty: return new
    if new is None or new.empty: return old
    return (pd.concat([old[STORE_COLS], new[STORE_COLS]], ignore_index=True)
            .drop_duplicates(subset=KEY_COLS, keep="last"))

def load_orgcat_store() -> pd.DataFrame:
    if os.path.exists(ORGCAT_STORE):
        try:
            d = pd.read_csv(ORGCAT_STORE, encoding="utf-8-sig")
            if set(ORGCAT_COLS) <= set(d.columns):
                return d[ORGCAT_COLS]
        except Exception:
            pass
    return pd.DataFrame(columns=ORGCAT_COLS)


def save_orgcat_store(df: pd.DataFrame):
    df[ORGCAT_COLS].to_csv(ORGCAT_STORE, index=False, encoding="utf-8-sig")


def merge_orgcat(old: pd.DataFrame, new: pd.DataFrame) -> pd.DataFrame:
    """조직×카테고리 누적 병합 — 같은 (단위·지표·조직·카테고리·LFMS·기간) 키는 신규 우선"""
    if old is None or old.empty: return new if new is not None else pd.DataFrame(columns=ORGCAT_COLS)
    if new is None or new.empty: return old
    return (pd.concat([old[ORGCAT_COLS], new[ORGCAT_COLS]], ignore_index=True)
            .drop_duplicates(subset=ORGCAT_KEY, keep="last"))


@st.cache_data(show_spinner=False)
def combine_orgcat(file_tuples) -> pd.DataFrame:
    """업로드 파일들 중 조직×카테고리 형식만 모아 하나로 — 같은 키는 마지막 파일 우선"""
    frames = []
    for n, b in file_tuples:
        d = parse_orgcat_file(n, b)
        if not d.empty:
            frames.append(d)
    if not frames:
        return pd.DataFrame(columns=ORGCAT_COLS)
    return (pd.concat(frames, ignore_index=True)
            .drop_duplicates(subset=ORGCAT_KEY, keep="last"))


def _period_set(d, cols):
    if d is None or d.empty: return set()
    return set(map(tuple, d[cols].drop_duplicates().values.tolist()))

def upload_diff(stored, df_new):
    """업로드 데이터가 기존 누적 대비 추가/갱신하는 기간 수 (저장 전 미리보기용)"""
    cols = ["gran", "year", "label"]
    nc, oc = _period_set(df_new, cols), _period_set(stored, cols)
    return len(nc - oc), len(nc & oc)

# ── 백업/업로드 고도화 헬퍼 ────────────────────────────────
@st.cache_data(show_spinner=False)
def classify_uploads(file_tuples):
    """업로드된 각 파일이 무엇으로 인식됐는지 (파일명, 인식결과, 행수) 목록.
    미인식 파일을 눈에 보이게 해 조용한 누락을 방지한다."""
    out = []
    for n, b in file_tuples:
        is_backup = False
        if n.lower().endswith(".csv"):
            head = b[:400].decode("utf-8", "ignore").splitlines()
            if head and all(k in head[0] for k in ("gran", "metric", "sortkey")):
                is_backup = True
        oc = parse_orgcat_file(n, b)
        if not oc.empty:
            _g = "".join(g for g in ("일", "주", "월") if g in set(oc["gran"]))
            _lf = "/".join(sorted(set(oc["lfms"].astype(str))))
            out.append((n, f"✅ 조직×카테고리 · {_g} · LFMS {_lf} · "
                           f"조직 {oc['org'].nunique()}", len(oc)))
            continue
        pf, d = route_push(n, b)  # combine_files와 동일한 라우팅 (한글명 PUSH 포함)
        if pf is not None:
            nseries = pf.groupby(["metric", "segment"]).ngroups
            out.append((n, f"✅ 앱푸시 원천 · {nseries}시리즈", len(pf)))
        elif d is None:
            hint = ("❌ PUSH 인식 실패 (헤더 확인)" if looks_like_push_name(n)
                    else "❌ 미인식 (파일명·형식 확인)")
            out.append((n, hint, 0))
        elif is_backup:
            out.append((n, "♻ 누적 백업 복원", len(d)))
        else:
            grans = "".join(g for g in ("일", "주", "월") if g in set(d["gran"]))
            nm = d["metric"].nunique()
            kind = (f"{grans} · {nm}지표" if nm > 1
                    else f"{grans} · {d['metric'].iloc[0]}")
            out.append((n, f"✅ {kind}", len(d)))
    return out

def make_backup_zip(df, texts, odf=None) -> bytes:
    """누적 데이터 CSV + 조직×카테고리 CSV + 보고란·메모 JSON + manifest를 한 ZIP으로."""
    import zipfile
    buf = io.BytesIO()
    ts = datetime.datetime.now(KST).strftime("%Y-%m-%d %H:%M KST")
    yrs = (f"{int(df['year'].min())}–{int(df['year'].max())}년"
           if not df.empty else "-")
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("wr_data_store.csv",
                    df[STORE_COLS].to_csv(index=False).encode("utf-8-sig"))
        if odf is not None and not odf.empty:
            zf.writestr("wr_orgcat_store.csv",
                        odf[ORGCAT_COLS].to_csv(index=False).encode("utf-8-sig"))
        zf.writestr("wr_insights.json",
                    json.dumps(texts, ensure_ascii=False, indent=2).encode("utf-8"))
        zf.writestr("manifest.txt",
                    (f"LF Mall 주간보고 통합 백업\n생성: {ts}\n"
                     f"누적 데이터: {len(df):,}행 · {df['metric'].nunique() if not df.empty else 0}지표 · {yrs}\n"
                     f"조직×카테고리: {0 if odf is None else len(odf):,}행\n"
                     f"보고란·메모: {len(texts)}개 항목\n"
                     "복원: 사이드바 '백업 복원'에 이 ZIP을 그대로 올리세요.\n").encode("utf-8"))
    return buf.getvalue()

def _apply_memos(memos: dict):
    """메모 dict를 세션·파일에 병합 (업로드 값 우선)"""
    merged = {**st.session_state.wr_texts, **memos}
    st.session_state.wr_texts = merged
    all_d = load_insights(); all_d.update(merged); save_insights(all_d)

def restore_from_upload(upload):
    """백업 파일(zip/csv/json) 자동 인식 복원 → (요약 메시지, 데이터스토어_갱신여부).
    형식 불일치는 ValueError로 던진다 (호출측에서 표시)."""
    name = upload.name.lower()
    raw = upload.getvalue()
    data_restored, msgs = False, []

    def _try_csv(b):
        nonlocal data_restored
        d = pd.read_csv(io.BytesIO(b), encoding="utf-8-sig")
        # 조직×카테고리 백업이 먼저 — segment가 없어 STORE_COLS와 헷갈릴 일은 없다
        if set(ORGCAT_COLS) <= set(d.columns):
            save_orgcat_store(d[ORGCAT_COLS]); data_restored = True
            msgs.append(f"조직×카테고리 {len(d):,}행")
            return True
        if set(STORE_COLS) <= set(d.columns):
            save_store(d[STORE_COLS]); data_restored = True
            msgs.append(f"누적 데이터 {len(d):,}행")
            return True
        return False

    def _try_json(b):
        memos = json.loads(b.decode("utf-8"))
        if isinstance(memos, dict):
            _apply_memos(memos); msgs.append(f"메모 {len(memos)}개")
            return True
        return False

    if name.endswith(".zip"):
        import zipfile
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            for info in zf.infolist():
                if info.is_dir(): continue
                base = os.path.basename(info.filename).lower()
                if "__macosx" in info.filename.lower(): continue
                b = zf.read(info)
                if base.endswith(".csv"): _try_csv(b)
                elif base.endswith(".json"): _try_json(b)
    elif name.endswith(".csv"):
        if not _try_csv(raw):
            raise ValueError("누적 데이터 CSV 형식이 아니에요. gran·metric·sortkey 같은 필수 컬럼이 없어요.")
    elif name.endswith(".json"):
        if not _try_json(raw):
            raise ValueError("메모 JSON 형식이 아니에요. 최상위가 객체여야 해요.")
    else:
        raise ValueError("지원 형식: .zip / .csv / .json")

    if not msgs:
        raise ValueError("복원할 내용을 찾지 못했어요. ZIP 안에 백업 파일이 없어요.")
    return " · ".join(msgs), data_restored

def restore_widget(key, label="백업 복원 (ZIP / CSV / JSON)"):
    """백업 복원 업로더 + 처리. 빈 상태(메인)·사이드바 공용. 성공 시 rerun."""
    up = st.file_uploader(label, type=["zip", "csv", "json"], key=key)
    if up is not None and st.session_state.get("wr_restored") != up.name:
        try:
            summary, data_changed = restore_from_upload(up)
            st.session_state["wr_restored"] = up.name
            if data_changed:
                st.session_state.pop("wr_saved_sig", None)
                st.cache_data.clear()
            st.success(f"복원 완료 ✓ — {summary}"); st.rerun()
        except Exception as e:
            st.error(f"복원 실패: {e}")

def source_upload_widget(key):
    """빈 상태 메인용 원천 업로더 — 인식되면 즉시 저장 후 rerun (첫 업로드는 누적할 게 없음)."""
    mfiles = st.file_uploader("엑셀 / CSV / ZIP (복수 선택)",
                              type=["xlsx", "xls", "csv", "zip"],
                              accept_multiple_files=True, key=key)
    if mfiles:
        exp = expand_uploads(mfiles)
        dnew = combine_files(tuple(exp)) if exp else pd.DataFrame()
        if not dnew.empty:
            save_store(dnew)
            st.session_state.pop("wr_saved_sig", None)
            st.cache_data.clear()
            st.success(f"{len(dnew):,}행 인식 — 저장됨 ✓"); st.rerun()
        else:
            st.error("인식된 데이터가 없어요. 파일명과 형식을 확인해 주세요.")


# ── 표 렌더 + 엑셀 내려받기 ────────────────────────────────────────────
# Streamlit 기본 툴바의 'Download as CSV'는 서식이 다 날아간다. 발송성과 대시보드와
# 같은 검은 헤더 스타일 xlsx를 같은 자리에서 받을 수 있게 감싼다.
# 바이트는 콜러블로 넘겨 '누를 때만' 만든다.
_WTBL_SEQ = itertools.count(1)


def wtable(data, *args, dl=True, dl_name=None, **kw):
    """st.dataframe 과 같게 쓰되, 아래에 엑셀 다운로드 버튼을 붙인다."""
    ev = st.dataframe(data, *args, **kw)
    if dl and xlsx_bytes is not None:
        _i = next(_WTBL_SEQ)
        _nm = dl_name or "표"
        _fn = re.sub(r"[^\w가-힣.\- ]", "", str(_nm)).strip() or "표"
        # set_index로 축을 세운 표가 많아 인덱스도 같이 내보낸다(기본 RangeIndex면 제외)
        _df = getattr(data, "data", data)
        _idx = (not kw.get("hide_index", False)) and not isinstance(
            getattr(_df, "index", None), pd.RangeIndex)
        try:
            st.download_button(
                "⬇️ 엑셀", key=f"_wxl{_i}",
                data=lambda d=data, n=_nm, x=_idx: xlsx_bytes(d, sheet_name=n, title=n, index=x),
                file_name=f"{_fn}_{today_kst():%Y%m%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="화면에 보이는 서식 그대로 받아요. 숫자는 엑셀 숫자로 들어가서 정렬·합계가 돼요.")
        except Exception:                                 # noqa: BLE001
            pass                                          # 다운로드가 안 되더라도 표는 보여야 한다
    return ev


# ══════════════════════════════════════════════════════
# 조회 헬퍼 — mtd(당월/당주 일마감) vs final 선택
# ══════════════════════════════════════════════════════
def pick(df, gran, metric, seg, year, label, prefer="final"):
    sub = df[(df["gran"] == gran) & (df["metric"] == metric) &
             (df["segment"] == seg) & (df["year"] == year) & (df["label"] == label)]
    if sub.empty: return np.nan
    order = ["final", "mtd"] if prefer == "final" else ["mtd", "final"]
    for c in order:
        s = sub[sub["close"] == c]["value"].dropna()
        if len(s): return s.iloc[-1]
    return np.nan

def series_by_label(df, gran, metric, seg, year, prefer="final"):
    """한 연도의 기간라벨 → 값 Series (sortkey 순)"""
    sub = df[(df["gran"] == gran) & (df["metric"] == metric) &
             (df["segment"] == seg) & (df["year"] == year)].copy()
    if sub.empty: return pd.Series(dtype=float)
    pref = {"final": 0, "mtd": 1} if prefer == "final" else {"mtd": 0, "final": 1}
    sub["_p"] = sub["close"].map(pref)
    sub = sub.sort_values(["sortkey", "_p"]).drop_duplicates("label", keep="first")
    return sub.set_index("label")["value"]

def labels_sorted(df, gran, years=None):
    sub = df[df["gran"] == gran]
    if years is not None: sub = sub[sub["year"].isin(years)]
    return (sub[["label", "sortkey"]].assign(k=lambda d: d["sortkey"] % 10000)
            .drop_duplicates("label").sort_values("k")["label"].tolist())

def latest_period(df, gran):
    """가장 최근 (year, label)"""
    sub = df[(df["gran"] == gran) & df["value"].notna()]
    if sub.empty: return None, None
    row = sub.loc[sub["sortkey"].idxmax()]
    return int(row["year"]), row["label"]

def prev_label(df, gran, year, label):
    """직전 기간 (연도 경계 포함)"""
    sub = (df[(df["gran"] == gran) & df["value"].notna()]
           [["year", "label", "sortkey"]].drop_duplicates().sort_values("sortkey"))
    keys = sub[["year", "label"]].apply(tuple, axis=1).tolist()
    try: i = keys.index((year, label))
    except ValueError: return None, None
    return keys[i - 1] if i > 0 else (None, None)

def week_like(df, year, month, wk):
    """(year, 'MM월 N주차') 라벨 찾기 — 같은 주차 번호가 없으면 그 달의 '마지막 주차'로 대체.

    주차는 달마다 4~5개로 들쭉날쭉해서(전체 342주 중 31주가 5주차) 07월 5주차의
    전월인 06월엔 5주차가 아예 없다. 그대로 두면 전월비가 '–'로 비어버리므로
    가장 가까운 대응 주차(= 그 달 마지막 주)로 떨어뜨린다.
    반환: (label, exact) — exact=False면 대체된 것이라 화면에 기준을 같이 표기해야 한다.
          label이 None이면 그 달 데이터 자체가 없음.
    """
    labs = set(df[(df["gran"] == "주") & (df["year"] == year)]["label"].dropna().unique())
    exact = f"{month:02d}월 {wk}주차"
    if exact in labs:
        return exact, True

    def _wknum(s):
        """라벨에서 주차 번호. 'N주차'가 없으면 -1 (후보에서 제외).

        파서는 주간 라벨을 'MM월 N주차'로 강제하지만, 백업 CSV 복원은 컬럼만 보고
        내용은 검증하지 않는다. 형식이 어긋난 행이 하나 섞이면 여기서 None.group()으로
        터져 앱 전체가 초기 렌더부터 죽으므로 방어한다.
        """
        m = re.search(r"(\d)주차", s)
        return int(m.group(1)) if m else -1

    cand = [l for l in labs if l.startswith(f"{month:02d}월") and _wknum(l) > 0]
    if not cand:
        return None, False
    return max(cand, key=_wknum), False

def growth_pace_note(s_cur, s_prev=None):
    """잔고 시계열의 증가속도 분석 문구(HTML). s_cur/s_prev: dt 인덱스 Series(당해/전년).
    연초 대비 증감·일평균 속도, 전년 동기 속도 대비, 최근 4주 가속/감속, 연말 단순추정."""
    s_cur = s_cur.dropna().sort_index()
    if len(s_cur) < 8:
        return None
    as_of = s_cur.index[-1]
    span = max((as_of - s_cur.index[0]).days, 1)
    ytd = float(s_cur.iloc[-1] - s_cur.iloc[0])
    pace = ytd / span

    def _c(v, txt):
        return f'<span style="color:{"#dc2626" if v < 0 else "#16a34a"};font-weight:700">{txt}</span>'
    bits = [f"연초 대비 {_c(ytd, f'{ytd:+,.0f}명')} · 일평균 {_c(pace, f'{pace:+,.1f}명/일')}"]

    # 전년 동기(같은 연중 위치)까지의 속도와 비교
    if s_prev is not None:
        sp = s_prev.dropna().sort_index()
        if len(sp) >= 8:
            py = int(sp.index[0].year)
            try:
                cutoff = as_of.replace(year=py)
            except ValueError:          # 2/29 → 평년
                cutoff = as_of.replace(year=py, day=28)
            spc = sp[sp.index <= cutoff]
            if len(spc) >= 8:
                p_pace = (float(spc.iloc[-1] - spc.iloc[0])
                          / max((spc.index[-1] - spc.index[0]).days, 1))
                diff = pace - p_pace
                bits.append(f"전년 동기 {p_pace:+,.1f}명/일 대비 "
                            f"{_c(diff, f'{diff:+,.1f}명/일')} {'빠름' if diff > 0 else '느림'}")

    # 최근 4주 속도 → 추세 판정. 감소 추세에서는 '가속/감속'이 뒤집혀 읽히므로
    # 진행 방향(pace 부호)에 맞춰 문구를 고른다.
    recent = s_cur[s_cur.index >= as_of - pd.Timedelta(days=28)]
    if len(recent) >= 5:
        r_pace = (float(recent.iloc[-1] - recent.iloc[0])
                  / max((recent.index[-1] - recent.index[0]).days, 1))
        tol = max(abs(pace) * 0.05, 0.5)      # 이 폭 안이면 속도 변화 없음으로 본다
        if abs(r_pace - pace) <= tol:
            tag = "속도 유지"
        elif pace >= 0:
            tag = "감소 전환" if r_pace < 0 else ("증가 가속" if r_pace > pace else "증가 둔화")
        else:
            tag = "증가 전환" if r_pace > 0 else ("감소 둔화" if r_pace > pace else "감소 심화")
        bits.append(f"최근 4주 {_c(r_pace, f'{r_pace:+,.1f}명/일')} → <b>{tag}</b>")

    # 연말 단순 선형 추정 (현 속도 유지 가정)
    rest = (datetime.date(as_of.year, 12, 31) - as_of.date()).days
    if rest > 0:
        bits.append(f"현 속도 유지 시 연말 ≈ {float(s_cur.iloc[-1]) + pace * rest:,.0f}명")
    return " · ".join(bits)

def week_ref(df, ref_year, ref_week):
    """기준 주차 (year, label): 사이드바 선택 우선, 없으면 데이터상 최신 주차"""
    return (ref_year, ref_week) if ref_week else latest_period(df, "주")

# ══════════════════════════════════════════════════════
# 보고란 영속화
# ══════════════════════════════════════════════════════
def load_insights():
    if os.path.exists(INSIGHT_FILE):
        with open(INSIGHT_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_insights(d):
    with open(INSIGHT_FILE, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=2)

def report_text_block(key, title, default="", regen=None, ai_fn=None):
    """편집 가능한 보고란 텍스트 박스 (JSON 저장).
    regen 텍스트를 주면 '자동 생성'(템플릿) 버튼, ai_fn을 주면 'AI 생성'(Claude) 버튼이 뜬다.
    좁은 컬럼 안에서도 안 깨지도록 버튼은 박스 위/아래에 배치한다."""
    store = st.session_state.wr_texts
    if not store.get(key): store[key] = default
    ekey = f"__wr_edit_{key}__"
    if ekey not in st.session_state: st.session_state[ekey] = False

    # 제목 + 액션 버튼 (제목 한 줄, 버튼은 그 아래 정상 너비 컬럼)
    st.markdown(f"**{title}**")

    if st.session_state[ekey]:
        if HAS_QUILL:
            # Word 수준 리치 에디터 (글자 크기·색·굵게·기울임·밑줄·목록·정렬 등)
            toolbar = [
                [{"size": ["small", False, "large", "huge"]}],
                ["bold", "italic", "underline", "strike"],
                [{"color": []}, {"background": []}],
                [{"list": "ordered"}, {"list": "bullet"}],
                [{"align": []}], ["clean"],
            ]
            # value(초기값)는 최초 진입 또는 store[key]가 바뀐 경우에만 주입한다.
            # 매 rerun마다 value를 다시 넣으면 타이핑 중 저장값으로 되돌아가
            # '계속 리프레시'되는 버그가 발생한다. (st_quill은 입력마다 rerun 유발)
            qkey = f"wr_quill_{key}"
            skey = f"{qkey}__seed"
            if st.session_state.get(skey) != store[key]:
                st.session_state[skey] = store[key]
                new = st_quill(value=store[key], html=True, toolbar=toolbar, key=qkey)
            else:
                new = st_quill(html=True, toolbar=toolbar, key=qkey)
        else:
            new = st.text_area("", store[key], key=f"wr_ta_{key}", height=180,
                               label_visibility="collapsed")
        if st.button("저장", key=f"wr_save_{key}", type="primary",
                     width="stretch"):
            store[key] = new if new is not None else store[key]
            all_d = load_insights(); all_d[key] = store[key]; save_insights(all_d)
            st.session_state[ekey] = False; st.rerun()
    else:
        st.markdown(f"<div class='report-box'>{store[key] or '내용을 입력해 주세요.'}</div>",
                    unsafe_allow_html=True)
                    
    # AI 참고 메모: 데이터에 안 나오는 배경(프로모션·이벤트·이슈)을 적으면
    # AI 생성 시 [배경 메모]로 분리 주입돼 원인·맥락 해석에 활용된다.
    memo_val = ""
    if ai_fn is not None:
        mkey = f"{key}__memo"
        if mkey not in store: store[mkey] = ""
        with st.expander("🧠 AI 참고 메모 (프로모션·이벤트·운영 이슈 등 배경)",
                         expanded=bool(store[mkey])):
            memo_val = st.text_area(
                "데이터에 안 나오는 배경을 적으면 AI가 원인과 맥락을 풀 때 참고해요. "
                "수치는 데이터에서만 가져와요.",
                store[mkey], key=f"wr_memo_{key}", height=120)
            if st.button("메모 저장", key=f"wr_memosave_{key}",
                         width="stretch"):
                store[mkey] = memo_val
                all_d = load_insights(); all_d[mkey] = memo_val; save_insights(all_d)
                st.rerun()
                    
    n = 2 + (1 if regen is not None else 0) + (1 if ai_fn is not None else 0)
    bcols = st.columns(n)
    bi = 0
    edit_on = st.session_state[ekey]
    if bcols[bi].button("편집" if not edit_on else "보기",
                        key=f"wr_edit_{key}", width="stretch"):
        st.session_state[ekey] = not edit_on; st.rerun()
    bi += 1
    if regen is not None:
        if bcols[bi].button("자동 생성", key=f"wr_regen_{key}", width="stretch",
                            help="기준 주차 실적으로 템플릿 문구를 채워요. 기존 내용은 지워져요."):
            store[key] = regen
            all_d = load_insights(); all_d[key] = regen; save_insights(all_d)
            st.session_state[ekey] = False; st.rerun()
        bi += 1
    if ai_fn is not None:
        if bcols[bi].button("AI 생성", key=f"wr_ai_{key}", width="stretch",
                            help="Claude가 데이터와 참고 메모를 보고 인사이트 문구를 써요. 기존 내용은 지워져요."):
            store[mkey] = memo_val
            all_d = load_insights(); all_d[mkey] = memo_val; save_insights(all_d)
            with st.spinner("AI가 인사이트를 작성 중…"):
                text, err = ai_fn(memo_val)
            if err:
                st.error(err)
            else:
                store[key] = text
                all_d = load_insights(); all_d[key] = text; save_insights(all_d)
                st.session_state[ekey] = False; st.rerun()
    return store[key]

# ══════════════════════════════════════════════════════
# YoY 요약표 / 추이표 빌더
# ══════════════════════════════════════════════════════
def month_label(n): return f"{n}월"

def week_disp(year, label):
    """주차 표시: 2026년 06월 2주차"""
    return f"{year}년 {label}" if label else "-"

def week_back(df, wy, wlabel, years):
    """wlabel의 `years`년 전 대응 주차 (label, exact). 5주차 등은 그 달 마지막 주로 대체."""
    wm = re.match(r"(\d{1,2})월 (\d)주차", wlabel or "")
    if not wm:
        return wlabel, True
    return week_like(df, wy - years, int(wm.group(1)), int(wm.group(2)))

def wow_summary_table(df, wy, wlabel, metrics):
    """전주비 요약표 (실적 요약과 동일 구조): 전주·기준주·전주비 + 전년동주·전년비"""
    py, plb = prev_label(df, "주", wy, wlabel)
    ylb, _ = week_back(df, wy, wlabel, 1)
    cols = [week_disp(py, plb), week_disp(wy, wlabel), "전주비",
            week_disp(wy - 1, ylb), "전년비"]
    rows = []
    for met in metrics:
        cur = pick(df, "주", met, "*TOTAL", wy, wlabel, "mtd")
        prv = pick(df, "주", met, "*TOTAL", py, plb, "final") if plb else np.nan
        yoy = pick(df, "주", met, "*TOTAL", wy - 1, ylb, "final") if ylb else np.nan
        rows.append({
            "구분": met,
            cols[0]: fmt_value(met, prv), cols[1]: fmt_value(met, cur),
            cols[2]: fmt_delta(met, cur, prv) or "–",
            cols[3]: fmt_value(met, yoy), cols[4]: fmt_delta(met, cur, yoy) or "–",
        })
    return pd.DataFrame(rows).set_index("구분")

def yoy2_summary_table(df, wy, wlabel, metrics):
    """전년비 · 전전년비 비교표 — 올해 흐름이 전전년에도 있었는지 확인용.

    같은 주차의 3개 연도 값과 세 가지 증감을 나란히 둔다.
      · 전년비        = 금년 vs 전년   (올해 무슨 일이 있었나)
      · 전전년비      = 금년 vs 전전년 (2년 새 얼마나 움직였나)
      · 전년의 전년비 = 전년 vs 전전년 (작년에도 같은 방향이었나)
    '추세' 칸은 전년의 전년비 → 전년비 순서로 방향을 읽어 '2년 연속'인지 판정한다.
    반환: (표, 전년라벨, 전전년라벨, 전년정확여부, 전전년정확여부)
    """
    l1, e1 = week_back(df, wy, wlabel, 1)
    l2, e2 = week_back(df, wy, wlabel, 2)

    def _neg(d):
        return None if not d or d == "–" else d.startswith("△")

    def _trend(d_now, d_prev):
        a, b = _neg(d_now), _neg(d_prev)
        if a is None or b is None: return "–"
        if a and b:       return "2년 연속 감소"
        if not a and not b: return "2년 연속 증가"
        return "증가 → 감소" if a else "감소 → 증가"

    cols = [week_disp(wy, wlabel), week_disp(wy - 1, l1), week_disp(wy - 2, l2),
            "전년비", "전전년비", "전년의 전년비", "추세"]
    rows = []
    for met in metrics:
        cur = pick(df, "주", met, "*TOTAL", wy, wlabel, "mtd")
        p1 = pick(df, "주", met, "*TOTAL", wy - 1, l1, "final") if l1 else np.nan
        p2 = pick(df, "주", met, "*TOTAL", wy - 2, l2, "final") if l2 else np.nan
        d1 = fmt_delta(met, cur, p1)   # 금년 vs 전년
        d2 = fmt_delta(met, cur, p2)   # 금년 vs 전전년
        dp = fmt_delta(met, p1, p2)    # 전년 vs 전전년
        rows.append({
            "구분": met,
            cols[0]: fmt_value(met, cur), cols[1]: fmt_value(met, p1),
            cols[2]: fmt_value(met, p2),
            cols[3]: d1 or "–", cols[4]: d2 or "–", cols[5]: dp or "–",
            cols[6]: _trend(d1, dp),
        })
    return pd.DataFrame(rows).set_index("구분"), l1, l2, e1, e2

def yoy_summary_table(df, ref_year, ref_month, metrics):
    """참고본 '실적 요약' 표: 전월·당월 × (전년, 당년, 전년비) + 당월의 전월비.

    맨 끝 '전월비(당월)'은 같은 해 당월(MTD) ↔ 전월(월마감) 비교다. 당월은 아직
    진행 중이라 누계로는 못 맞대지만, 이 표의 값이 전부 **일평균**이라 기간 길이가
    달라도 그대로 견줄 수 있다.
    """
    rows = []
    pm_y, pm_m = (ref_year, ref_month - 1) if ref_month > 1 else (ref_year - 1, 12)
    cols = [f"{pm_y-1}년 {pm_m}월", f"{pm_y}년 {pm_m}월", "전년비(전월)",
            f"{ref_year-1}년 {ref_month}월", f"{ref_year}년 {ref_month}월", "전년비(당월)",
            "전월비(당월)"]
    for met in metrics:
        pm_prev = pick(df, "월", met, "*TOTAL", pm_y - 1, month_label(pm_m), "final")
        pm_cur  = pick(df, "월", met, "*TOTAL", pm_y,     month_label(pm_m), "final")
        cm_prev = pick(df, "월", met, "*TOTAL", ref_year - 1, month_label(ref_month), "mtd")
        cm_cur  = pick(df, "월", met, "*TOTAL", ref_year,     month_label(ref_month), "mtd")
        rows.append({
            "구분": met,
            cols[0]: fmt_value(met, pm_prev), cols[1]: fmt_value(met, pm_cur),
            cols[2]: fmt_delta(met, pm_cur, pm_prev) or "–",
            cols[3]: fmt_value(met, cm_prev), cols[4]: fmt_value(met, cm_cur),
            cols[5]: fmt_delta(met, cm_cur, cm_prev) or "–",
            cols[6]: fmt_delta(met, cm_cur, pm_cur) or "–",
        })
    return pd.DataFrame(rows).set_index("구분"), (pm_y, pm_m)

def trend_table(df, gran, metrics, years, seg="*TOTAL"):
    """추이표: 행=지표, 열=(연도, 기간)"""
    out, columns = {}, []
    for y in years:
        for lb in labels_sorted(df, gran, [y]):
            sub = df[(df["gran"] == gran) & (df["year"] == y) & (df["label"] == lb) &
                     df["value"].notna()]
            if sub.empty: continue
            columns.append((y, lb))
    if not columns:
        return pd.DataFrame()
    for met in metrics:
        vals = []
        for y, lb in columns:
            vals.append(pick(df, gran, met, seg, y, lb, "final"))
        out[met] = vals
    tbl = pd.DataFrame(out, index=pd.MultiIndex.from_tuples(columns, names=["연도", "기간"])).T
    return tbl

def style_trend(tbl, metrics):
    # 최신 pandas는 float 컬럼에 문자열 대입을 금지하므로 object로 변환 후 포맷
    disp = tbl.astype(object).copy()
    for met in disp.index:
        disp.loc[met] = [fmt_value(met, v) for v in tbl.loc[met]]
    return disp

# ══════════════════════════════════════════════════════
# YoY 라인차트 (Plotly)
# ══════════════════════════════════════════════════════
def base_layout(h=300, ysuffix="", title=""):
    # 제목은 최상단(container 기준), 범례는 그 아래 별도 줄(plot 상단 바로 위)로 분리해
    # 제목·범례가 같은 높이에서 겹치지 않게 한다. 상단 여백(t)을 넉넉히 확보.
    return dict(
        paper_bgcolor="rgba(248,249,252,0)", plot_bgcolor="rgba(248,249,252,0)",
        font=dict(color="#475569", size=11), margin=dict(l=10, r=10, t=64, b=10),
        height=h, showlegend=True,
        legend=dict(orientation="h", yref="paper", yanchor="bottom", y=1.0,
                    xanchor="left", x=0, bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#64748b", size=10)),
        title=dict(text=title, font=dict(color="#94a3b8", size=13),
                   x=0, xanchor="left", yref="container", y=0.98, yanchor="top"),
        xaxis=dict(gridcolor="rgba(0,0,0,0)", linecolor="#e2e8f0",
                   tickfont=dict(color="#64748b", size=10)),
        yaxis=dict(gridcolor="#f1f5f9", linecolor="#e2e8f0",
                   tickfont=dict(color="#64748b", size=10), ticksuffix=ysuffix),
    )

def yoy_chart(df, gran, metric, years, seg="*TOTAL", h=300):
    unit, div = METRIC_UNIT.get(metric, ("", 1))
    if metric in PCT_METRICS: div, unit = 0.01, "%"
    if gran == "월":
        x_all = [month_label(i) for i in range(1, 13)]
    else:
        x_all = labels_sorted(df, gran, years)
    fig = go.Figure()
    for i, y in enumerate(sorted(years)):
        # 연도마다 없는 주차(5주차 등)는 건너뛰고 선을 잇는다
        s = series_by_label(df, gran, metric, seg, y, prefer="final").reindex(x_all).dropna()
        fig.add_trace(go.Scatter(
            x=s.index.tolist(), y=(s / div).tolist(), mode="lines+markers", name=str(y),
            line=dict(color=clr(YEAR_PAL[i % len(YEAR_PAL)]), width=2),
            marker=dict(size=5),
        ))
    gname = "월별" if gran == "월" else "주차별"
    ly = base_layout(h, ysuffix=unit if unit == "%" else "",
                     title=f"{metric} {gname} 추이 ({unit})")
    ly["xaxis"]["categoryorder"] = "array"
    ly["xaxis"]["categoryarray"] = x_all
    if gran == "주": ly["xaxis"]["tickangle"] = -45; ly["xaxis"]["nticks"] = 20
    fig.update_layout(**ly)
    return fig

# ══════════════════════════════════════════════════════
# 엑셀 워크북 내보내기 (월/주/첫구매_요약 + 차트)
# ══════════════════════════════════════════════════════
def build_workbook(df, texts, ref_year, ref_month, chart_years):
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_font = Font(bold=True, size=11)
    head_fill = PatternFill("solid", fgColor="EEF3FA")
    title_font = Font(bold=True, size=13)

    # ── 데이터 시트 (월/주): 지표 블록 스택
    extra_metrics = [m for m in df["metric"].unique() if m not in METRICS7]
    for gran, sheet in (("월", "월"), ("주", "주")):
        ws = wb.create_sheet(sheet)
        sub_g = df[df["gran"] == gran]
        if sub_g.empty: continue
        cols = (sub_g[["year", "label", "sortkey"]].drop_duplicates()
                .sort_values("sortkey")[["year", "label"]].apply(tuple, axis=1).tolist())
        r = 1
        for met in METRICS7 + sorted(extra_metrics):
            sub_m = sub_g[sub_g["metric"] == met]
            if sub_m.empty: continue
            ws.cell(r, 1, met).font = title_font
            for j, (y, lb) in enumerate(cols):
                ws.cell(r, 2 + j, y).font = head_font
                ws.cell(r + 1, 2 + j, lb).font = head_font
                ws.cell(r + 1, 2 + j).fill = head_fill
            segs = ["*TOTAL"] + [s for s in CHANNELS if s in set(sub_m["segment"])]
            for i, seg in enumerate(segs):
                ws.cell(r + 2 + i, 1, seg).font = head_font
                for j, (y, lb) in enumerate(cols):
                    v = pick(df, gran, met, seg, y, lb, "final")
                    c = ws.cell(r + 2 + i, 2 + j)
                    if not (isinstance(v, float) and np.isnan(v)):
                        c.value = v
                        c.number_format = "0.00%" if met in PCT_METRICS else "#,##0"
            r += 2 + len(segs) + 2

    # ── 요약 시트
    ws = wb.create_sheet("첫구매_요약", 0)
    wb.remove(wb["Sheet"])
    r = 1
    ws.cell(r, 1, f"첫구매 주간보고 — {ref_year}년 {ref_month}월 기준").font = Font(bold=True, size=15)
    r += 2

    # 실적 요약 YoY 표
    ws.cell(r, 1, "실적 요약 (일평균)").font = title_font; r += 1
    tbl, _ = yoy_summary_table(df, ref_year, ref_month, METRICS7)
    ws.cell(r, 1, "구분").font = head_font
    for j, cname in enumerate(tbl.columns):
        c = ws.cell(r, 2 + j, cname); c.font = head_font; c.fill = head_fill
    red_font = Font(color="DC2626")
    green_font = Font(color="16A34A")
    for i, met in enumerate(tbl.index):
        ws.cell(r + 1 + i, 1, met).font = head_font
        for j, cname in enumerate(tbl.columns):
            c = ws.cell(r + 1 + i, 2 + j, tbl.loc[met, cname])
            s = str(c.value)
            if "전년비" in str(cname) or "전월비" in str(cname):
                if s.startswith("△"): c.font = red_font
                elif s.startswith("+"): c.font = green_font
    r += len(tbl) + 3

    # 보고란
    ws.cell(r, 1, "전주 주요 지표 현황").font = title_font
    ws.cell(r, 8, "금주 집행 내용 요약").font = title_font
    c1 = ws.cell(r + 1, 1, texts.get("wr_metrics_summary", ""))
    c2 = ws.cell(r + 1, 8, texts.get("wr_exec_summary", ""))
    for c in (c1, c2): c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 8, end_column=6)
    ws.merge_cells(start_row=r + 1, start_column=8, end_row=r + 8, end_column=13)
    r += 11

    # 차트 데이터 블록 + 라인차트 (월별/주차별 × 거래액·고객수·객단가)
    chart_metrics = ["첫구매 거래액", "첫구매 고객수", "첫구매 객단가"]
    for gran, gname in (("월", "월별"), ("주", "주차별")):
        x_all = ([month_label(i) for i in range(1, 13)] if gran == "월"
                 else labels_sorted(df, gran, chart_years))
        if not x_all: continue
        anchor_row = r
        for k, met in enumerate(chart_metrics):
            ws.cell(r, 1, f"{met} {gname} (차트 데이터)").font = head_font
            for j, lb in enumerate(x_all):
                ws.cell(r, 2 + j, lb).fill = head_fill
            nrow = 0
            for y in sorted(chart_years):
                s = series_by_label(df, gran, met, "*TOTAL", y).reindex(x_all)
                if s.dropna().empty: continue
                nrow += 1
                ws.cell(r + nrow, 1, y)
                for j, v in enumerate(s.tolist()):
                    if not (isinstance(v, float) and np.isnan(v)):
                        c = ws.cell(r + nrow, 2 + j, v)
                        c.number_format = "#,##0"
            if nrow:
                ch = LineChart()
                ch.title = f"{met} {gname} 추이"
                ch.height, ch.width = 7.5, 13
                data = Reference(ws, min_col=1, min_row=r + 1,
                                 max_col=1 + len(x_all), max_row=r + nrow)
                cats = Reference(ws, min_col=2, min_row=r, max_col=1 + len(x_all))
                ch.add_data(data, titles_from_data=True, from_rows=True)
                ch.set_categories(cats)
                ws.add_chart(ch, f"{get_column_letter(2 + len(x_all) + 1 + (k % 3) * 8)}{anchor_row}")
            r += nrow + 2
        r += 14

    # 채널별 실적 (당월 YoY)
    ws.cell(r, 1, f"채널별 실적 — {ref_year}년 {ref_month}월 (전년비)").font = title_font; r += 1
    for met in chart_metrics:
        ws.cell(r, 1, met).font = head_font
        heads = [f"{ref_year-1}년 {ref_month}월", f"{ref_year}년 {ref_month}월", "전년비"]
        for j, hd in enumerate(heads):
            c = ws.cell(r, 2 + j, hd); c.font = head_font; c.fill = head_fill
        segs = ["*TOTAL"] + CHANNELS
        for i, seg in enumerate(segs):
            pv = pick(df, "월", met, seg, ref_year - 1, month_label(ref_month), "mtd")
            cv = pick(df, "월", met, seg, ref_year, month_label(ref_month), "mtd")
            ws.cell(r + 1 + i, 1, seg)
            ws.cell(r + 1 + i, 2, None if np.isnan(pv) else pv).number_format = "#,##0"
            ws.cell(r + 1 + i, 3, None if np.isnan(cv) else cv).number_format = "#,##0"
            dcell = ws.cell(r + 1 + i, 4, fmt_delta(met, cv, pv) or "–")
            ds = str(dcell.value)
            if ds.startswith("△"): dcell.font = red_font
            elif ds.startswith("+"): dcell.font = green_font
        r += len(segs) + 3

    ws.column_dimensions["A"].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ══════════════════════════════════════════════════════
# 자동 보고 초안
# ══════════════════════════════════════════════════════
def _delta_html(d):
    """증감 문자열 → 색상 span (역신장 △ 빨강 / 신장 + 초록)"""
    if not d: return ""
    if d.startswith("△"): return f'<span style="color:#dc2626;font-weight:700">{d}</span>'
    return f'<span style="color:#16a34a;font-weight:700">{d}</span>'

def auto_draft(df, ref_year, ref_month, ref_week=None):
    """기준 주차(없으면 기준 월) 실적으로 보고 문구 자동 생성 (HTML, 역신장 빨강)"""
    lines = []
    for met in METRICS7:
        if ref_week:
            cur = pick(df, "주", met, "*TOTAL", ref_year, ref_week, "mtd")
            if isinstance(cur, float) and np.isnan(cur): continue
            py, plb = prev_label(df, "주", ref_year, ref_week)
            prv = pick(df, "주", met, "*TOTAL", py, plb, "final") if plb else np.nan
            yoy = pick(df, "주", met, "*TOTAL", ref_year - 1, ref_week, "final")
            parts = [f" - {met} — {fmt_value(met, cur)}"]
            d_w = fmt_delta(met, cur, prv)
            d_y = fmt_delta(met, cur, yoy)
            if d_w: parts.append(f"전주비 {_delta_html(d_w)}")
            if d_y: parts.append(f"전년비 {_delta_html(d_y)}")
            lines.append(", ".join(parts))
        else:
            cur = pick(df, "월", met, "*TOTAL", ref_year, month_label(ref_month), "mtd")
            prv = pick(df, "월", met, "*TOTAL", ref_year - 1, month_label(ref_month), "mtd")
            if isinstance(cur, float) and np.isnan(cur): continue
            d = fmt_delta(met, cur, prv)
            tail = f", 전년비 {_delta_html(d)}" if d else ""
            lines.append(f" - {met} — {fmt_value(met, cur)}" + tail)
    return "<br>".join(lines)

# ══════════════════════════════════════════════════════
# AI 인사이트 생성 (Claude API) — 모델 교체 가능
# ══════════════════════════════════════════════════════
AI_MODELS = {
    "Claude Sonnet 5 (균형·기본)": "claude-sonnet-5",
    "Claude Opus 4.8 (최고 품질)": "claude-opus-4-8",
    "Claude Haiku 4.5 (빠름·저렴)": "claude-haiku-4-5",
}
DEFAULT_AI_MODEL = "claude-sonnet-5"

def _anthropic_key():
    """Streamlit secrets 또는 환경변수에서 API 키 조회"""
    try:
        if "ANTHROPIC_API_KEY" in st.secrets:
            return st.secrets["ANTHROPIC_API_KEY"]
    except Exception:
        pass
    return os.environ.get("ANTHROPIC_API_KEY")

def _ai_metric_facts(df, ref_year, ref_month, ref_week=None):
    """모델에 넘길 지표·증감 요약 텍스트"""
    rows = []
    for met in METRICS7:
        if ref_week:
            cur = pick(df, "주", met, "*TOTAL", ref_year, ref_week, "mtd")
            if isinstance(cur, float) and np.isnan(cur): continue
            py, plb = prev_label(df, "주", ref_year, ref_week)
            prv = pick(df, "주", met, "*TOTAL", py, plb, "final") if plb else np.nan
            wm = re.match(r"(\d{1,2})월 (\d)주차", ref_week)
            mom = np.nan
            ylb = ref_week
            if wm:
                # 5주차처럼 대응 주차가 없으면 그 달 마지막 주차로 대체 (KPI 카드와 동일 규칙)
                mo, wk = int(wm.group(1)), int(wm.group(2))
                my, mm = (ref_year, mo - 1) if mo > 1 else (ref_year - 1, 12)
                mlb, _ = week_like(df, my, mm, wk)
                if mlb:
                    mom = pick(df, "주", met, "*TOTAL", my, mlb, "final")
                ylb, _ = week_like(df, ref_year - 1, mo, wk)
            yoy = pick(df, "주", met, "*TOTAL", ref_year - 1, ylb, "final") if ylb else np.nan
            rows.append(f"- {met}: {fmt_value(met, cur)} "
                        f"(전주비 {fmt_delta(met, cur, prv) or '–'}, "
                        f"전월비 {fmt_delta(met, cur, mom) or '–'}, "
                        f"전년비 {fmt_delta(met, cur, yoy) or '–'})")
        else:
            cur = pick(df, "월", met, "*TOTAL", ref_year, month_label(ref_month), "mtd")
            prv = pick(df, "월", met, "*TOTAL", ref_year - 1, month_label(ref_month), "mtd")
            if isinstance(cur, float) and np.isnan(cur): continue
            rows.append(f"- {met}: {fmt_value(met, cur)} (전년비 {fmt_delta(met, cur, prv) or '–'})")
    return "\n".join(rows)

def ai_generate_insight(df, ref_year, ref_month, ref_week, model,
                        focus="전주 주요 지표 현황", memo=""):
    """Claude API로 보고 인사이트 생성 → HTML(역신장 빨강) 반환. (텍스트, 에러) 튜플.
    memo: 사용자가 적은 정성 배경(프로모션·이벤트 등). 수치와 분리해 [배경 메모]로 주입."""
    key = _anthropic_key()
    if not key:
        return None, ("ANTHROPIC_API_KEY가 없어요. "
                      "Streamlit Cloud → Settings → Secrets에 추가해 주세요.")
    try:
        import anthropic
    except ImportError:
        return None, "anthropic 패키지가 없어요. requirements.txt에 넣고 다시 배포해 주세요."

    period = f"{ref_year}년 {ref_week}" if ref_week else f"{ref_year}년 {ref_month}월"
    facts = _ai_metric_facts(df, ref_year, ref_month, ref_week)
    system = (
        "당신은 LF몰 CRM 첫구매 보고서를 작성하는 데이터 분석가입니다. "
        "한국어 실무 보고 문구를 작성할 때, 긴 문장형 불릿(• ...은 ...으로 ...)을 피하고 "
        "반드시 '- '(메인 요약)와 'ㄴ '(세부 수치/해석)를 사용하는 계층형 불릿 구조로 출력하세요.\n\n"
        "예시:\n"
        "- 첫구매 거래액 및 고객수 동반 감소\n"
        "ㄴ 거래액 88.9백만원 (전주비 △1.5% · 전년비 △24.9%)\n"
        "ㄴ 고객수 639명 (전주비 △3.7% · 전년비 △34.1%)\n"
        "ㄴ 프로모션 종료 여파로 유입 대비 전환 효율이 저조한 것으로 보임\n\n"
        "[확정 수치]에 있는 숫자만 인용하고 수치를 지어내지 마세요. "
        "[배경 메모]는 원인 추정/해석에만 활용하며(단정짓지 말고 '~로 보임' 등 사용), 비어있으면 수치 팩트만 기재하세요. "
        "출력은 HTML로만 (<br> 로 줄바꿈). "
        "증감 수치 중 역신장(감소)은 <span style=\"color:#dc2626;font-weight:700\">…</span>, "
        "신장(증가)은 <span style=\"color:#16a34a;font-weight:700\">…</span>로 감싸세요. "
        "서론·맺음말 없이 바로 계층형 불릿만 출력하세요."
    )
    memo_block = (memo or "").strip() or "(없음)"
    user = (f"[기준: {period}]\n\n"
            f"[확정 수치]\n{facts}\n\n"
            f"[배경 메모]\n{memo_block}\n\n"
            f"위 자료로 '{focus}' 문구를 작성하세요.")
    try:
        client = anthropic.Anthropic(api_key=key)
        # Sonnet 5는 thinking 미지정 시 adaptive가 기본 켜져 max_tokens를 나눠 쓰므로 여유 있게
        resp = client.messages.create(
            model=model, max_tokens=4000, system=system,
            messages=[{"role": "user", "content": user}],
        )
        text = "".join(b.text for b in resp.content if b.type == "text").strip()
        return (text or None), (None if text else "응답이 비어 있어요.")
    except anthropic.AuthenticationError:
        return None, "API 키 인증에 실패했어요. 키를 확인해 주세요."
    except anthropic.RateLimitError:
        return None, "요청이 많아 잠시 제한됐어요. 조금 뒤에 다시 시도해 주세요."
    except Exception as e:
        return None, f"생성 중 오류: {e}"

# ══════════════════════════════════════════════════════
# 06. 앱푸시 동의 현황 페이지
# ══════════════════════════════════════════════════════
def render_push_page(df, ref_year, chart_years):
    st.markdown("## 앱푸시 동의 현황")
    gran_opt = st.radio("차트 보기 기준", ["일자별", "주차별", "월별"], horizontal=True, key="push_gran_opt")
    gran_map = {"일자별": "일", "주차별": "주", "월별": "월"}
    sel_gran = gran_map[gran_opt]

    cyrs = sorted(chart_years)  # multiselect 클릭 순서에 의존하지 않도록 정렬
    prev_year = cyrs[-2] if len(cyrs) >= 2 else None

    res_df = pd.DataFrame()  # 데이터 없으면 빈 채로 남음 — 하단 '상세 데이터' 가드용
    df_gran = df[df["gran"] == sel_gran]

    if df_gran.empty:
        st.info(f"{sel_gran} 단위 데이터가 없어요.")
    else:
        st.subheader(f"{ref_year}년 {gran_opt} 신규가입 및 앱푸시 수신동의율 추이")
        dates = labels_sorted(df, sel_gran, [ref_year])

        if not dates:
            st.info(f"{ref_year}년 {gran_opt} 데이터가 없어요.")
        else:
            rows = []
            # PUSH 데이터 집계용 (일자별 데이터를 주/월로 변환)
            df_push_daily = df[(df["metric"] == "앱푸시수신동의") & (df["gran"] == "일") & (df["segment"] == "*TOTAL") & (df["close"] == "final")].copy()
            if not df_push_daily.empty:
                df_push_daily["month"] = df_push_daily["label"].apply(lambda x: int(str(x).split('/')[0]) if '/' in str(x) else 0)
                df_push_daily["day"] = df_push_daily["label"].apply(lambda x: int(str(x).split('/')[1]) if '/' in str(x) else 0)
                # 대량 이관/재동의 스파이크 마스킹 (연도별 중앙값 규칙)
                df_push_daily = mask_push_spikes(df_push_daily)

                df_push_daily["주"] = df_push_daily.apply(lambda r: f"{r['month']:02d}월 {(int(r['day'])-1)//7 + 1}주차" if r['month']>0 else "", axis=1)
                df_push_daily["월"] = df_push_daily.apply(lambda r: f"{r['month']}월" if r['month']>0 else "", axis=1)
                df_push_daily["일"] = df_push_daily["label"]

                # 집계
                push_agg = df_push_daily.groupby(["year", sel_gran])["value"].mean().reset_index()
            else:
                push_agg = pd.DataFrame()

            # 월 필터링 로직 (업로드한 해당 월까지만)
            if not df_push_daily.empty:
                last_month = int(df_push_daily["month"].max())
                if sel_gran == "일":
                    dates = [d for d in dates if (int(d.split('/')[0]) if '/' in d else 0) <= last_month]
                elif sel_gran == "주":
                    dates = [d for d in dates if (int(d.split('월')[0]) if '월' in d else 0) <= last_month]
                elif sel_gran == "월":
                    dates = [d for d in dates if (int(d.replace('월', '')) if '월' in d else 0) <= last_month]

            for d in dates:
                join_val = pick(df, sel_gran, "가입자수", "*TOTAL", ref_year, d, "final")

                push_val = float('nan')
                if not push_agg.empty:
                    v = push_agg[(push_agg["year"] == ref_year) & (push_agg[sel_gran] == d)]["value"]
                    if not v.empty:
                        push_val = v.values[0]
                        if push_val == 0: push_val = float('nan')

                rate = float('nan')
                if pd.notna(push_val) and pd.notna(join_val) and join_val > 0:
                    rate = push_val / join_val
                    if rate > 1.0:
                        rate = float('nan')

                prev_rate = float('nan')
                if prev_year:
                    prev_join_val = pick(df, sel_gran, "가입자수", "*TOTAL", prev_year, d, "final")
                    prev_push_val = float('nan')

                    if not push_agg.empty:
                        v = push_agg[(push_agg["year"] == prev_year) & (push_agg[sel_gran] == d)]["value"]
                        if not v.empty:
                            prev_push_val = v.values[0]
                            if prev_push_val == 0: prev_push_val = float('nan')

                    if pd.notna(prev_push_val) and pd.notna(prev_join_val) and prev_join_val > 0:
                        prev_rate = prev_push_val / prev_join_val
                        if prev_rate > 1.0:
                            prev_rate = float('nan')

                rows.append({
                    "날짜": d,
                    "가입자수": join_val,
                    "앱푸시수신동의": push_val,
                    "동의율": rate,
                    "전년동의율": prev_rate
                })

            res_df = pd.DataFrame(rows)

            if not res_df.empty:
                fig = go.Figure()
                fig.add_trace(go.Bar(x=res_df["날짜"], y=res_df["가입자수"], name="가입자수", marker_color=clr("slate"), yaxis="y1"))
                fig.add_trace(go.Bar(x=res_df["날짜"], y=res_df["앱푸시수신동의"], name="앱푸시수신동의", marker_color=clr("blue"), yaxis="y1"))

                fig.add_trace(go.Scatter(x=res_df["날짜"], y=res_df["동의율"]*100, name=f"{ref_year} 동의율(%)", mode="lines+markers", line=dict(color=clr("red"), width=2), connectgaps=True, yaxis="y2"))

                if prev_year:
                    fig.add_trace(go.Scatter(x=res_df["날짜"], y=res_df["전년동의율"]*100, name=f"{prev_year} 동의율(%)", mode="lines", line=dict(color=clr("slate"), width=2, dash="dot"), connectgaps=True, yaxis="y2"))

                max_rate = max(res_df["동의율"].dropna()*100, default=10)
                if prev_year:
                    max_prev = max(res_df["전년동의율"].dropna()*100, default=10)
                    max_rate = max(max_rate, max_prev)

                if pd.isna(max_rate) or max_rate == 0:
                    max_rate = 10

                fig.update_layout(
                    barmode="group",
                    yaxis=dict(title="명", side="left", showgrid=False),
                    yaxis2=dict(title="%", side="right", overlaying="y", range=[0, max_rate * 1.2], showgrid=False),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                    margin=dict(l=0, r=0, t=30, b=0),
                    xaxis=dict(type="category")
                )
                st.plotly_chart(fig, width="stretch")

    # ── 확장 뷰: 동의자 잔고·이탈·순증·요일 패턴 (앱푸시_* 지표)
    ext = df[(df["gran"] == "일") & df["metric"].isin(
        ["앱푸시_동의자수", "앱푸시_신규추가", "앱푸시_이탈"]) & (df["close"] == "final")].copy()
    if ext.empty:
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.info("동의자 잔고·이탈·순증 뷰는 PUSH 원천 엑셀을 다시 올리면 보여요. "
                "기존·신규·Total 섹션을 새로 인식해 누적에 저장해요.")
    else:
        ext[["_m", "_d"]] = ext["label"].str.extract(r"^(\d{1,2})/(\d{1,2})$")
        ext = ext.dropna(subset=["_m"]).copy()
        ext["dt"] = pd.to_datetime(dict(year=ext["year"].astype(int),
                                        month=ext["_m"].astype(int),
                                        day=ext["_d"].astype(int)), errors="coerce")
        ext = ext.dropna(subset=["dt"])
        ext = mask_push_spikes(ext)       # 흐름 지표 스파이크 (이관·재동의)
        ext = mask_level_glitches(ext)    # 잔고 하루짜리 글리치 (0 기록 등)

        bal = ext[ext["metric"] == "앱푸시_동의자수"]
        add = ext[ext["metric"] == "앱푸시_신규추가"]
        out = ext[ext["metric"] == "앱푸시_이탈"]

        # KPI — 최신 잔고(Total) + 당월 흐름
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        bt = bal[bal["segment"] == "*TOTAL"].sort_values("dt")
        if not bt.empty:
            last_dt = bt["dt"].iloc[-1]
            cur_bal = bt["value"].iloc[-1]
            pm_end = bt[bt["dt"] < last_dt.replace(day=1)]  # 전월말 잔고
            mom = (f"{cur_bal - pm_end['value'].iloc[-1]:+,.0f}명 (전월말比)"
                   if not pm_end.empty else None)
            in_mo = lambda s: s[(s["segment"] == "*TOTAL") &
                                (s["dt"].dt.year == last_dt.year) &
                                (s["dt"].dt.month == last_dt.month)]["value"].sum()
            mo_add, mo_out = in_mo(add), in_mo(out)
            k1, k2, k3, k4 = st.columns(4)
            k1.metric(f"총 동의자 ({last_dt:%m/%d} 기준)", f"{cur_bal:,.0f}명", mom)
            k2.metric(f"{last_dt.month}월 신규추가", f"{mo_add:,.0f}명")
            k3.metric(f"{last_dt.month}월 이탈", f"{mo_out:,.0f}명")
            k4.metric(f"{last_dt.month}월 순증", f"{mo_add - mo_out:+,.0f}명")

        # 동의자 잔고 추이 — 기존/신규/Total (선택 단위 기말값)
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.subheader("동의자 잔고 추이 — 기존/신규/Total")
        rule = {"일": "D", "주": "W-SUN", "월": "MS"}[sel_gran]
        fig_b = go.Figure()
        seg_pal = {"*TOTAL": "slate", "기존": "blue", "신규": "teal"}
        for seg in ["*TOTAL", "기존", "신규"]:
            s = (bal[bal["segment"] == seg].set_index("dt").sort_index()["value"]
                 .resample(rule).last().dropna())
            if s.empty: continue
            fig_b.add_trace(go.Scatter(x=s.index, y=s.values, mode="lines",
                                       name="Total" if seg == "*TOTAL" else seg,
                                       line=dict(color=clr(seg_pal[seg]), width=2)))
        fig_b.update_layout(**base_layout(300, title="수신동의 누적 잔고 (명)"))
        st.plotly_chart(fig_b, width="stretch")

        # 타겟팅 가능 모수 — 연중 추이 (전년 비교). 세그먼트별로 각각.
        # 잔고(앱푸시_동의자수)는 실측 대조 결과 원천의 '타겟팅 가능' 행과 동일한 모수다.
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.subheader("타겟팅 가능 모수 — 연중 추이 (전년 비교)")
        st.caption("실제로 발송할 수 있는 수신동의 모수예요. 연도별 라인을 같은 연중 위치(월·일)에 "
                   "겹쳐 작년과 비교해요.")
        byr = bal.copy()
        # 각 연도를 같은 X축(연중 위치)에 겹치려면 공통 연도로 정규화 (2000=윤년이라 2/29 안전)
        byr["mdt"] = byr["dt"].apply(lambda d: d.replace(year=2000))
        yrs_b = sorted(byr["year"].dropna().unique().astype(int))
        SEG_LABEL = {"*TOTAL": "Total", "기존": "기존", "신규": "신규"}
        for seg in ["*TOTAL", "기존", "신규"]:
            sb = byr[byr["segment"] == seg]
            if sb.empty: continue
            figy = go.Figure()
            for i, yr in enumerate(yrs_b):
                s = (sb[sb["year"] == yr].sort_values("mdt")
                     .set_index("mdt")["value"].dropna())
                if s.empty: continue
                figy.add_trace(go.Scatter(
                    x=s.index, y=s.values, mode="lines", name=str(yr),
                    line=dict(color=clr(YEAR_PAL[i % len(YEAR_PAL)]), width=2),
                    hovertemplate="%{x|%m/%d}<br>" + str(yr) + " %{y:,.0f}명<extra></extra>"))
            lyy = base_layout(260, title=f"{SEG_LABEL[seg]} — 타겟팅 가능 (명)")
            lyy["xaxis"]["tickformat"] = "%m월"
            lyy["xaxis"]["dtick"] = "M1"
            figy.update_layout(**lyy)
            st.plotly_chart(figy, width="stretch")

            # 증가속도 분석 코멘트 (실제 날짜 인덱스로 계산 — 정규화된 mdt 아님)
            _ser = lambda y: (sb[sb["year"] == y].sort_values("dt")
                              .set_index("dt")["value"]) if y in yrs_b else None
            note = growth_pace_note(_ser(yrs_b[-1]),
                                    _ser(yrs_b[-2]) if len(yrs_b) >= 2 else None)
            if note:
                st.markdown(
                    "<div style='font-size:12px;color:#64748b;margin:-6px 0 16px 2px'>"
                    f"📈 <b>{SEG_LABEL[seg]} 증감 속도</b> — {note}</div>",
                    unsafe_allow_html=True)

        # 순증감 분해 — 신규추가(+) vs 이탈(−) + 순증 라인 (Total)
        st.subheader("동의 순증감 분해 — 신규추가 vs 이탈 (Total)")
        a = (add[add["segment"] == "*TOTAL"].set_index("dt").sort_index()["value"]
             .resample(rule).sum(min_count=1))
        o = (out[out["segment"] == "*TOTAL"].set_index("dt").sort_index()["value"]
             .resample(rule).sum(min_count=1))
        flow = pd.DataFrame({"신규추가": a, "이탈": o}).dropna(how="all")
        if not flow.empty:
            fig_f = go.Figure()
            fig_f.add_trace(go.Bar(x=flow.index, y=flow["신규추가"], name="신규추가(+)",
                                   marker_color=clr("green")))
            fig_f.add_trace(go.Bar(x=flow.index, y=-flow["이탈"], name="이탈(−)",
                                   marker_color=clr("red")))
            net = flow["신규추가"].fillna(0) - flow["이탈"].fillna(0)
            fig_f.add_trace(go.Scatter(x=flow.index, y=net, name="순증",
                                       mode="lines+markers",
                                       line=dict(color=clr("slate"), width=2),
                                       marker=dict(size=4)))
            lyf = base_layout(320, title=f"{gran_opt} 신규추가·이탈·순증 (명)")
            lyf["barmode"] = "relative"
            fig_f.update_layout(**lyf)
            st.plotly_chart(fig_f, width="stretch")
            st.caption("※ 대량 이관·재동의 스파이크(연도별 중앙값 10배 초과)는 왜곡 방지를 위해 제외")

        # 요일 패턴 — 신규추가/이탈 요일별 일평균 (기준 연도)
        st.subheader(f"요일별 평균 — 신규추가·이탈 ({ref_year}년)")
        dows = ["월", "화", "수", "목", "금", "토", "일"]
        fig_d = go.Figure()
        for nm, src, cname in [("신규추가", add, "green"), ("이탈", out, "red")]:
            s = src[(src["segment"] == "*TOTAL") & (src["dt"].dt.year == ref_year)]
            if s.empty: continue
            m = s.groupby(s["dt"].dt.dayofweek)["value"].mean().reindex(range(7))
            fig_d.add_trace(go.Bar(x=dows, y=m.values, name=nm, marker_color=clr(cname)))
        fig_d.update_layout(**base_layout(280, title="요일별 일평균 (명)"))
        st.plotly_chart(fig_d, width="stretch")

    st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
    st.subheader("월별 전년비 증감 추이표 (YoY)")

    push_metrics = ["가입자수", "앱푸시수신동의", "동의율"]
    sub = df[(df["gran"] == "일") & (df["year"].isin(cyrs)) & (df["metric"].isin(["가입자수", "앱푸시수신동의"])) & (df["segment"] == "*TOTAL") & (df["close"] == "final")].copy()

    if sub.empty:
        st.info("일자별 가입자수·앱푸시 수신동의 데이터가 없어서 월별 YoY 표를 보여줄 수 없어요.")
    else:
        sub["month"] = sub["label"].apply(lambda x: int(str(x).split('/')[0]) if '/' in str(x) else 0)
        sub["day"] = sub["label"].apply(lambda x: int(str(x).split('/')[1]) if '/' in str(x) else 0)

        # 푸시동의 스파이크 이상치 마스킹 (가입자수 행은 규칙상 제외됨)
        sub = mask_push_spikes(sub)

        # 업로드한 최신 월 및 일(MTD) 계산
        max_month = 12
        max_day = 31
        if cyrs:
            last_year = cyrs[-1]
            for m in range(12, 0, -1):
                r = sub[(sub["year"] == last_year) & (sub["month"] == m) & sub["value"].notna()]
                if not r.empty:
                    max_month = m
                    max_day = r["day"].max()
                    break

        # 당월(max_month)에 대해서는 전년도 데이터도 MTD(max_day)까지만 필터링하여 누적 비교
        if max_month > 0:
            sub = sub[~((sub["month"] == max_month) & (sub["day"] > max_day))]

        # min_count=1 — 값이 하나도 없는 달을 0으로 만들지 않는다.
        # (기본 sum()은 전부 NaN이면 0을 돌려줘서, 아직 안 올라온 당월 가입자수가
        #  '0명 · YoY △100%'로 찍히고 동의율은 288/0=inf → 마스킹돼 '-'로 빠졌다.)
        grp = (sub.groupby(["year", "month", "metric"])["value"].sum(min_count=1)
               .unstack("metric").reset_index())

        if "가입자수" in grp.columns and "앱푸시수신동의" in grp.columns:
            grp["동의율"] = grp["앱푸시수신동의"] / grp["가입자수"]
            grp.loc[grp["동의율"] > 1.0, "동의율"] = float('nan')
        else:
            grp["동의율"] = float('nan')

        out_tbl = {}
        cols = []
        for m in range(1, max_month + 1):
            if m == max_month:
                cols.append(f"{m}월 (MTD)")
            else:
                cols.append(f"{m}월")

        for met in push_metrics:
            for y in cyrs:
                row_name = f"{met} ({y})"
                out_tbl[row_name] = []
                for m in range(1, max_month + 1):
                    r = grp[(grp["year"] == y) & (grp["month"] == m)]
                    out_tbl[row_name].append(r[met].values[0] if (not r.empty and met in r.columns) else float('nan'))

            # 최근 2개 연도 간 YoY
            if len(cyrs) >= 2:
                yoy_name = f"{met} (YoY)"
                out_tbl[yoy_name] = []
                prev_y, cur_y = cyrs[-2], cyrs[-1]
                for i, m in enumerate(range(1, max_month + 1)):
                    prev_val = out_tbl[f"{met} ({prev_y})"][i]
                    cur_val = out_tbl[f"{met} ({cur_y})"][i]
                    if pd.isna(cur_val) and pd.isna(prev_val):
                        out_tbl[yoy_name].append("-")
                    else:
                        out_tbl[yoy_name].append(fmt_delta(met, cur_val, prev_val) or "-")

        tbl_df = pd.DataFrame(out_tbl, index=cols).T

        disp = tbl_df.astype(object).copy()
        for row_name in disp.index:
            if "(YoY)" in row_name:
                continue
            else:
                met = row_name.split(" ")[0]
                disp.loc[row_name] = [fmt_value(met, v) if pd.notna(v) else "-" for v in tbl_df.loc[row_name]]

        def style_yoy_rows(tbl):
            def _color(v):
                s = str(v)
                if s.startswith("△"): return "color:#dc2626;font-weight:600"
                if s.startswith("+"): return "color:#16a34a;font-weight:600"
                return ""
            def _highlight(df_):
                styles = pd.DataFrame('', index=df_.index, columns=df_.columns)
                for idx in df_.index:
                    if "(YoY)" in str(idx):
                        for col in df_.columns:
                            styles.loc[idx, col] = _color(df_.loc[idx, col])
                return styles
            return tbl.style.apply(_highlight, axis=None)

        wtable(style_yoy_rows(disp), width="stretch")

    if not res_df.empty:
        st.markdown("### 상세 데이터")
        disp_df = res_df.copy()
        disp_df["앱푸시수신동의"] = disp_df["앱푸시수신동의"].apply(lambda x: f"{int(x):,}" if not pd.isna(x) else "–")
        disp_df["가입자수"] = disp_df["가입자수"].apply(lambda x: f"{int(x):,}" if not pd.isna(x) else "–")
        disp_df["동의율"] = disp_df["동의율"].apply(lambda x: f"{x*100:.1f}%" if not pd.isna(x) else "–")
        wtable(disp_df.set_index("날짜"), width="stretch", dl_name="상세 데이터")

# ══════════════════════════════════════════════════════
# 페이지 PDF 저장 (브라우저 인쇄 → PDF, 차트 포함)
# ══════════════════════════════════════════════════════
def guard_select(key, opts):
    """옵션 목록이 바뀌면 세션에 남은 옛 선택값이 목록 밖이 된다 — 조용한 리셋·예외 방지."""
    if key in st.session_state and st.session_state[key] not in opts:
        st.session_state.pop(key, None)


def guard_multi(key, opts):
    """multiselect용 가드 — 사라진 값이 세션에 남으면 위젯이 예외로 죽는다."""
    cur = st.session_state.get(key)
    if isinstance(cur, (list, tuple)):
        keep = [v for v in cur if v in opts]
        if len(keep) != len(cur):
            st.session_state[key] = keep


def opick(odf, gran, metric, org, cat, lfms, year, label, prefer="final"):
    """조직×카테고리 값 하나 — 마스터의 pick()과 같은 규칙(final 우선)."""
    sub = odf[(odf["gran"] == gran) & (odf["metric"] == metric) &
              (odf["org"] == org) & (odf["cat"] == cat) & (odf["lfms"] == lfms) &
              (odf["year"] == year) & (odf["label"] == label)]
    if sub.empty: return np.nan
    order = ["final", "mtd"] if prefer == "final" else ["mtd", "final"]
    for c in order:
        v = sub[sub["close"] == c]["value"].dropna()
        if len(v): return v.iloc[-1]
    return np.nan


# 조직·카테고리를 더하면 전체가 되는 지표는 **거래액 하나뿐**이다. 실파일 검증에서
# 조직 합이 전체와 오차 0.00%로 일치했다. 고객수(+2.4%)·상품UV(+33%)는 유니크 값이라
# 같은 사람이 여러 조직에 잡혀 합이 전체를 넘는다 — 기여도 분해를 하면 안 된다.
ORGCAT_ADDITIVE = {"첫구매 거래액"}
# 거래액 = 상품UV × 상품CR × 객단가 (실파일에서 모든 조직·카테고리에 오차 0.000%로 성립)
ORGCAT_FACTORS = [("상품UV", "유입"), ("상품CR", "전환"), ("첫구매 객단가", "객단가")]


def _logmean(a, b):
    if a <= 0 or b <= 0: return np.nan
    if abs(a - b) < 1e-12: return float(a)
    return (a - b) / np.log(a / b)


def factor_split(prev, cur):
    """거래액 증감을 유입·전환·객단가 셋으로 쪼갠다 → (기여액 3개, 총증감). 못 쪼개면 None.

    LMDI(로그평균 디비지아)라 셋을 더하면 실제 증감과 **정확히** 같고, 어느 요인을 먼저
    대입하느냐로 답이 달라지지 않는다(순차 대입법의 순서 의존을 피한다).
    prev·cur = (상품UV, 상품CR, 객단가). 0·음수·결측이 하나라도 있으면 분해가 정의되지 않는다.
    """
    if any(v is None or not np.isfinite(v) or v <= 0 for v in list(prev) + list(cur)):
        return None
    v0 = prev[0] * prev[1] * prev[2]
    v1 = cur[0] * cur[1] * cur[2]
    L = _logmean(v1, v0)
    if not np.isfinite(L): return None
    return tuple(L * np.log(cur[i] / prev[i]) for i in range(3)), v1 - v0


def _won_m(v, digits=1):
    """백만원 표기 — 증감액은 부호를 살려 △ 대신 −를 쓴다(표의 전년비와 구분)."""
    if v is None or not np.isfinite(v): return "–"
    return f"{v / 1e6:+,.{digits}f}백만"


def render_orgcat_page(odf):
    """08. 조직·카테고리별 실적 — '어디가 원인인지' 파고드는 진단 화면.

    사이드바 기준 기간(월·주)에 기대지 않고 **자체 기간 선택**을 쓴다. 이 데이터는 일
    단위까지 오고 커버리지도 마스터와 달라서, 사이드바 값을 그대로 쓰면 빈 화면이 되기
    쉽다. 덕분에 마스터 데이터가 아직 없어도 이 화면만은 열린다.
    """
    st.markdown("## 조직·카테고리별 첫구매 실적")
    if odf is None or odf.empty:
        st.info("조직×카테고리 데이터가 없어요. MICRO 대시보드에서 받은 "
                "**구분06(BPU) × 구분07(카테고리)** 엑셀을 사이드바에 올려 주세요.")
        st.caption("일자별·주별·월별 파일을 따로 올리면 각각 누적돼요.")
        return

    # ── 필터: 집계 단위 · LFMS · 지표 ──
    grans = [g for g in ("월", "주", "일") if (odf["gran"] == g).any()]
    f1, f2, f3 = st.columns([1.2, 1, 1.6])
    with f1:
        guard_select("oc_gran", grans)
        gran = st.radio("집계 단위", grans, horizontal=True, key="oc_gran")
    # LFMS 선택지는 **고른 단위 안에서** 뽑는다. 단위마다 받아 온 export가 달라
    # (예: 일별만 LFMS=Y) 전역 목록을 쓰면 '일'로 바꿨을 때 이전 선택 'N'이 남아
    # 데이터가 있는데도 빈 화면이 된다.
    lfmss = sorted(odf[odf["gran"] == gran]["lfms"].dropna().astype(str).unique())
    with f2:
        if len(lfmss) > 1:
            guard_select("oc_lfms", lfmss)
            lfms = st.radio("LFMS 포함", lfmss, horizontal=True, key="oc_lfms",
                            help="원본 헤더의 'LFMS 포함여부' 값이에요. 포함/미포함은 "
                                 "모집단이 달라서 섞어 보면 안 돼요.")
        else:
            lfms = lfmss[0] if lfmss else "N"
            st.caption(f"LFMS 포함여부 **{lfms}** 데이터만 있어요")
    sub = odf[(odf["gran"] == gran) & (odf["lfms"] == lfms)]
    if sub.empty:
        st.info("고른 조건에 데이터가 없어요."); return
    pref = list(ORGCAT_MAP.values()) + ["상품UV", "상품CR", "거래액비중", "고객비중"]
    mets = [m for m in pref if (sub["metric"] == m).any()]
    mets += [m for m in sub["metric"].unique() if m not in mets]
    with f3:
        guard_select("oc_met", mets)
        met = st.selectbox("진단 지표", mets, key="oc_met",
                           help="표·차트·히트맵이 이 지표를 따라가요. 기여도 분해는 "
                                "조직 합이 전체와 일치하는 거래액에서만 나와요.")

    # ── 기준 기간 ──
    labs = (sub[sub["value"].notna()][["year", "label", "sortkey"]]
            .drop_duplicates().sort_values("sortkey"))
    if labs.empty:
        st.info("고른 조건에 값이 없어요."); return
    opts = [(int(r["year"]), str(r["label"])) for _, r in labs.iterrows()][::-1]
    # 진행 중인 기간은 '일마감'(mtd)만 온다 — 값이 다 일평균이라 완결 기간과 그대로 견줄 수는
    # 있지만 며칠치 표본이라 크게 흔들린다. 어느 기간이 진행 중인지 라벨에 박아 둔다.
    _fin = {(int(r["year"]), str(r["label"]))
            for _, r in sub[(sub["close"] == "final") & sub["value"].notna()]
            [["year", "label"]].drop_duplicates().iterrows()}
    part = {p for p in opts if p not in _fin}
    guard_select("oc_per", opts)
    cy, clabel = st.selectbox("기준 기간", opts, key="oc_per",
                              format_func=lambda p: f"{p[0]}년 {p[1]}"
                              + (" · 진행 중" if p in part else ""))
    py = cy - 1
    ccol, pcol = f"{cy}년 {clabel}", f"{py}년 {clabel}"
    has_prev = not sub[(sub["year"] == py) & (sub["label"] == clabel)].empty
    if (cy, clabel) in part:
        st.info(f"**{ccol}은 아직 진행 중이에요(일마감).** 값이 모두 일평균이라 완결 기간과 "
                "견줄 수는 있지만, 며칠치라 크게 흔들려요. 추세 판단은 완결된 기간으로 하세요.")
    if not has_prev:
        st.warning(f"전년({py}년) 같은 기간 데이터가 없어 전년비·기여도는 비어 있어요.")

    def V(org, cat, year, metric=None):
        return opick(sub, gran, metric or met, org, cat, lfms, year, clabel)

    orgs_all = [o for o in sub["org"].unique() if o != ORGCAT_TOTAL]
    unit, div = METRIC_UNIT.get(met, ("", 1))
    is_pct = met in PCT_METRICS
    xdiv = 0.01 if is_pct else div
    additive = met in ORGCAT_ADDITIVE

    # ── 파고들기 ──
    st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
    d1, d2 = st.columns(2)
    org_opts = ["전체"] + [o for o in orgs_all
                          if np.isfinite(V(o, ORGCAT_TOTAL, cy)) or
                          np.isfinite(V(o, ORGCAT_TOTAL, py))]
    with d1:
        guard_select("oc_dorg", org_opts)
        sel_org = st.selectbox("① 조직 파고들기", org_opts, key="oc_dorg")
    tgt_org = ORGCAT_TOTAL if sel_org == "전체" else sel_org
    cats_of = [c for c in sub[sub["org"] == tgt_org]["cat"].unique() if c != ORGCAT_TOTAL]
    cat_opts = ["전체"] + (cats_of if sel_org != "전체" else [])
    with d2:
        guard_select("oc_dcat", cat_opts)
        sel_cat = st.selectbox("② 카테고리 파고들기", cat_opts, key="oc_dcat",
                               disabled=(sel_org == "전체"),
                               help="조직을 먼저 고르면 카테고리까지 들어갈 수 있어요.")
    tgt_cat = ORGCAT_TOTAL if sel_cat == "전체" else sel_cat
    crumb = " › ".join(["전체"] + ([sel_org] if sel_org != "전체" else [])
                       + ([sel_cat] if sel_cat != "전체" else []))
    st.caption(f"보는 곳: **{esc(crumb)}** · {ccol} · {met} · LFMS {lfms}")

    # ── 진단 요약 ──
    tot_p = V(ORGCAT_TOTAL, ORGCAT_TOTAL, py)   # 히트맵 기여도의 분모
    cur_c, cur_p = V(tgt_org, tgt_cat, cy), V(tgt_org, tgt_cat, py)

    def _kids(org, cat_mode):
        """지금 레벨의 한 단계 아래 목록 — 전체면 조직, 조직이면 그 조직의 카테고리."""
        if cat_mode:
            return [(c, org, c) for c in cats_of]
        return [(o, o, ORGCAT_TOTAL) for o in orgs_all]

    drill_cat = (sel_org != "전체" and sel_cat == "전체")
    kids = _kids(tgt_org, drill_cat) if (sel_org == "전체" or drill_cat) else []
    kid_lbl = "카테고리" if drill_cat else "조직"

    rows, base = [], (cur_p if np.isfinite(cur_p) and cur_p else np.nan)
    for name, o, c in kids:
        vc, vp = V(o, c, cy), V(o, c, py)
        if not np.isfinite(vc) and not np.isfinite(vp):
            continue
        d = (vc - vp) if (np.isfinite(vc) and np.isfinite(vp)) else np.nan
        rows.append({"_n": name, "_o": o, "_c": c, "_vc": vc, "_vp": vp, "_d": d,
                     "_share": (d / base * 100) if (additive and np.isfinite(d)
                                                    and np.isfinite(base) and base) else np.nan})
    kid = pd.DataFrame(rows)

    _msg = []
    if np.isfinite(cur_c):
        _delta = f" · 전년비 {fmt_delta(met, cur_c, cur_p) or '–'}" if np.isfinite(cur_p) else ""
        _amt = (f" ({_won_m(cur_c - cur_p)})"
                if additive and np.isfinite(cur_p) else "")
        _msg.append(f"**{esc(crumb)}**의 {ccol} {met}은 **{fmt_value(met, cur_c)}**"
                    f"{_delta}{_amt}이에요.")
    if len(kid) and kid["_d"].notna().any():
        _w = kid.reindex(kid["_d"].abs().sort_values(ascending=False).index).iloc[0]
        _dir = "끌어내린" if _w["_d"] < 0 else "끌어올린"
        # 전체 변화 대비 몫. 서로 상쇄되는 구간이면 100%를 넘는 게 정상이지만,
        # 전체가 거의 안 움직인 달엔 300%·800% 같은 무의미한 수가 나와 생략한다.
        _sh = ""
        if (additive and np.isfinite(cur_p) and np.isfinite(cur_c)
                and (cur_c - cur_p) != 0 and np.isfinite(_w["_d"])):
            _r = _w["_d"] / (cur_c - cur_p)
            if abs(_r) <= 3:
                _sh = f", 전체 변화의 **{abs(_r) * 100:.0f}%**"
        _msg.append(f"가장 크게 {_dir} {kid_lbl}{josa(kid_lbl)} **{esc(str(_w['_n']))}**"
                    + (f" ({_won_m(_w['_d'])}{_sh})" if additive else
                       f" ({fmt_delta(met, _w['_vc'], _w['_vp']) or '–'})") + "예요.")
    # 요인 분해 한 줄 (거래액 기준 — 지표와 무관하게 '왜'를 말해 준다)
    fp = tuple(V(tgt_org, tgt_cat, py, m) for m, _ in ORGCAT_FACTORS)
    fc = tuple(V(tgt_org, tgt_cat, cy, m) for m, _ in ORGCAT_FACTORS)
    fs = factor_split(fp, fc)
    if fs:
        parts, _tot = fs
        _top = max(range(3), key=lambda i: abs(parts[i]))
        _nm = ORGCAT_FACTORS[_top][1]
        _msg.append(f"거래액 변화를 쪼개면 "
                    + " · ".join(f"{lb} {_won_m(parts[i])}"
                                 for i, (_, lb) in enumerate(ORGCAT_FACTORS))
                    + f" — **{_nm}**{josa(_nm, '이가')} 가장 크게 움직였어요.")
    if _msg:
        st.markdown('<div class="vg">📌 ' + "<br>".join(_msg) + '</div>',
                    unsafe_allow_html=True)

    # ── ① 기여도 ──
    st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
    st.subheader(f"① 어디가 움직였나 — {kid_lbl}별")
    if not len(kid):
        st.info("한 단계 더 들어갈 대상이 없어요. 위에서 조직을 '전체'로 바꿔 보세요.")
    else:
        show = kid.reindex(
            (kid["_d"].abs() if additive else kid["_vc"]).sort_values(ascending=False).index)
        trow = []
        for _, r in show.iterrows():
            item = {kid_lbl: r["_n"], pcol: fmt_value(met, r["_vp"]),
                    ccol: fmt_value(met, r["_vc"]),
                    "전년비": fmt_delta(met, r["_vc"], r["_vp"]) or "–"}
            if additive:
                item["증감액"] = _won_m(r["_d"]) if np.isfinite(r["_d"]) else "–"
                item["기여도"] = (f"{r['_share']:+.1f}%p"
                                if np.isfinite(r["_share"]) else "–")
            # 같은 줄에서 원인 축까지 보이게 — 유입·전환·객단가 전년비를 붙인다
            for mkey, lb in ORGCAT_FACTORS:
                item[f"{lb} 전년비"] = fmt_delta(mkey, V(r["_o"], r["_c"], cy, mkey),
                                               V(r["_o"], r["_c"], py, mkey)) or "–"
            trow.append(item)
        wtable(style_delta_cols(pd.DataFrame(trow).set_index(kid_lbl)),
               width="stretch", dl_name=f"{kid_lbl}별 {met}")
        if additive:
            st.caption(f"**기여도**는 그 {kid_lbl}의 증감액을 **{pcol} 전체**로 나눈 값이에요. "
                       f"다 더하면 전체 전년비와 같아요. 정렬은 증감액이 큰 순이에요.")
        else:
            st.caption(f"{met}은 유니크 값이라 {kid_lbl} 합이 전체와 달라요 — 기여도 분해는 "
                       "거래액에서만 나와요. 정렬은 값이 큰 순이에요.")

        # 워터폴 — 전년 전체에서 시작해 각 항목 증감을 쌓아 당년 전체로 닫는다
        if additive and np.isfinite(cur_p) and show["_d"].notna().any():
            # 상위 8개만 막대로 세우고 나머지(0원짜리 조직 등)는 '기타'로 합친다.
            # 다 세우면 눈금이 뭉개져 정작 큰 항목이 안 읽힌다.
            TOPN = 8
            wall = show[show["_d"].notna()]
            ws, rest = wall.head(TOPN), wall.iloc[TOPN:]
            resid = (cur_c - cur_p) - ws["_d"].sum()      # 나머지 + 합계 불일치
            _extra = abs(resid) > 5e4
            _rlab = f"기타 {len(rest)}곳" if len(rest) else "기타·미분류"
            names = [pcol] + [str(x) for x in ws["_n"]] + ([_rlab] if _extra else []) + [ccol]
            meas = (["absolute"] + ["relative"] * len(ws)
                    + (["relative"] if _extra else []) + ["total"])
            vals = [cur_p / 1e6] + [v / 1e6 for v in ws["_d"]] + \
                   ([resid / 1e6] if _extra else []) + [cur_c / 1e6]
            figw = go.Figure(go.Waterfall(
                orientation="v", measure=meas, x=names, y=vals,
                text=[f"{v:+,.1f}" if m == "relative" else f"{v:,.1f}"
                      for v, m in zip(vals, meas)],
                textposition="outside",
                connector=dict(line=dict(color="#cbd5e1", width=1)),
                increasing=dict(marker=dict(color=clr("green"))),
                decreasing=dict(marker=dict(color=clr("red"))),
                totals=dict(marker=dict(color=clr("slate")))))
            lay = base_layout(380, title=f"{kid_lbl}별 증감 분해 (백만원)")
            lay["showlegend"] = False
            figw.update_layout(**lay)
            st.plotly_chart(figw, width="stretch")

    # ── ② 왜 — 유입 × 전환 × 객단가 ──
    st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
    st.subheader(f"② 왜 그랬나 — 유입·전환·객단가 (거래액 기준)")
    if not fs:
        st.info("이 대상은 전년 값이 없거나 0이라 요인 분해를 할 수 없어요.")
    else:
        parts, tot_d = fs
        frow = []
        for i, (mkey, lb) in enumerate(ORGCAT_FACTORS):
            vp_, vc_ = V(tgt_org, tgt_cat, py, mkey), V(tgt_org, tgt_cat, cy, mkey)
            frow.append({"요인": lb, "지표": mkey,
                         pcol: fmt_value(mkey, vp_), ccol: fmt_value(mkey, vc_),
                         "전년비": fmt_delta(mkey, vc_, vp_) or "–",
                         "거래액 기여": _won_m(parts[i]),
                         "설명력": f"{abs(parts[i]) / sum(abs(x) for x in parts) * 100:.0f}%"})
        wtable(style_delta_cols(pd.DataFrame(frow).set_index("요인")),
               width="stretch", dl_name=f"{crumb} 요인 분해")
        figf = go.Figure(go.Waterfall(
            orientation="v",
            measure=["absolute"] + ["relative"] * 3 + ["total"],
            x=[pcol] + [lb for _, lb in ORGCAT_FACTORS] + [ccol],
            y=[(cur_p or 0) / 1e6] + [p / 1e6 for p in parts] + [(cur_c or 0) / 1e6],
            text=[f"{(cur_p or 0)/1e6:,.1f}"] + [f"{p/1e6:+,.1f}" for p in parts]
                 + [f"{(cur_c or 0)/1e6:,.1f}"],
            textposition="outside",
            connector=dict(line=dict(color="#cbd5e1", width=1)),
            increasing=dict(marker=dict(color=clr("green"))),
            decreasing=dict(marker=dict(color=clr("red"))),
            totals=dict(marker=dict(color=clr("slate")))))
        lay = base_layout(360, title=f"{crumb} · 거래액 증감을 요인으로 (백만원)")
        lay["showlegend"] = False
        figf.update_layout(**lay)
        st.plotly_chart(figf, width="stretch")
        st.caption("**거래액 = 상품UV × 상품CR × 객단가**가 원본에서 정확히 성립해요. "
                   "세 기여액을 더하면 실제 증감과 딱 맞고(LMDI), 어느 요인을 먼저 "
                   "대입하느냐로 답이 달라지지 않아요. "
                   "**유입**이 크면 사람이 덜 들어온 것, **전환**이면 들어와서 안 산 것, "
                   "**객단가**면 사긴 샀는데 덜 쓴 거예요.")

    # ── ③ 추이 ──
    st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
    st.subheader(f"③ 추이 — {crumb}")
    tsub = sub[(sub["metric"] == met) & (sub["org"] == tgt_org) & (sub["cat"] == tgt_cat)]
    if tsub.empty:
        st.info("추이를 그릴 값이 없어요.")
    else:
        order = (tsub[tsub["year"] == cy][["label", "sortkey"]]
                 .drop_duplicates().sort_values("sortkey"))
        xs = [str(x) for x in order["label"]]
        MAXP = {"일": 90, "주": 53, "월": 12}[gran]
        xs = xs[-MAXP:]
        figt = go.Figure()
        for yr, color in ((py, "slate"), (cy, "blue")):
            g = tsub[tsub["year"] == yr]
            m = {str(r["label"]): r["value"] for _, r in g.iterrows()}
            figt.add_trace(go.Scatter(
                x=xs, y=[(m[k] / xdiv) if (k in m and m[k] == m[k]) else None for k in xs],
                mode="lines+markers", name=f"{yr}년", connectgaps=False,
                line=dict(color=clr(color), width=2 if yr == cy else 1.5,
                          dash=None if yr == cy else "dot"),
                marker=dict(size=5)))
        figt.update_layout(**base_layout(
            340, ysuffix="%" if is_pct else "",
            title=f"{met} · {cy}년 vs {py}년 ({'%' if is_pct else unit})"))
        st.plotly_chart(figt, width="stretch")

    # ── ④ 히트맵 ──
    st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
    st.subheader("④ 한눈에 — 조직 × 카테고리")
    hm = sub[(sub["metric"] == met) & (sub["cat"] != ORGCAT_TOTAL) &
             (sub["org"] != ORGCAT_TOTAL) & sub["value"].notna() &
             (sub["label"] == clabel) & (sub["year"].isin([cy, py]))]
    if hm.empty:
        st.info("이 기간엔 조직×카테고리 교차 값이 없어요.")
    else:
        cur_p_ = hm[hm["year"] == cy].pivot_table(index="org", columns="cat",
                                                  values="value", aggfunc="last")
        prv_p_ = hm[hm["year"] == py].pivot_table(index="org", columns="cat",
                                                  values="value", aggfunc="last")
        prv_p_ = prv_p_.reindex(index=cur_p_.index, columns=cur_p_.columns)
        if additive and np.isfinite(tot_p) and tot_p:
            z = (cur_p_ - prv_p_) / tot_p * 100
            zttl, zsuf = "기여도(%p) — 전체 전년 대비", "%p"
        else:
            z = (cur_p_ / prv_p_ - 1) * 100 if not is_pct else (cur_p_ - prv_p_) * 100
            zttl, zsuf = ("전년비(%)" if not is_pct else "전년비(%p)"), ("%" if not is_pct else "%p")
        lim = float(np.nanmax(np.abs(z.values))) if np.isfinite(z.values).any() else 1.0
        figh = go.Figure(go.Heatmap(
            z=z.values, x=[str(c) for c in z.columns], y=[str(i) for i in z.index],
            zmid=0, zmin=-lim, zmax=lim,
            colorscale=[[0, clr("red")], [0.5, "#ffffff"], [1, clr("green")]],
            text=[[("" if not np.isfinite(v) else f"{v:+.1f}") for v in row]
                  for row in z.values],
            texttemplate="%{text}", textfont=dict(size=10),
            hovertemplate="%{y} · %{x}<br>" + zttl + ": %{z:+.2f}" + zsuf + "<extra></extra>",
            colorbar=dict(title=dict(text=zsuf, side="right"), thickness=10)))
        lay = base_layout(max(260, 30 * len(z.index) + 120), title=f"{ccol} · {zttl}")
        lay["showlegend"] = False
        lay["yaxis"]["autorange"] = "reversed"
        figh.update_layout(**lay)
        st.plotly_chart(figh, width="stretch")
        st.caption(("빨간 칸이 전체를 끌어내린 곳이에요. 숫자를 다 더하면 전체 전년비가 돼요."
                    if additive and np.isfinite(tot_p) and tot_p else
                    "빈 칸은 그 조직에 해당 카테고리 실적이 없다는 뜻이에요.")
                   + " 위 파고들기에서 조직·카테고리를 고르면 그 칸의 원인이 ②에 나와요.")

    st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
    st.caption("MICRO 대시보드의 **구분06(BPU) × 구분07(카테고리)** export예요. "
               "거래액·고객수·객단가는 **일평균**이고 비중·상품CR은 비율이에요. "
               "`*TOTAL`은 하위 항목의 단순 합이 아니라 파일이 준 값을 그대로 써요 — "
               "거래액만 합이 전체와 일치하고, 고객수·상품UV는 유니크 값이라 조금 넘어요.")

def print_button(label="이 페이지 PDF 저장 / 인쇄"):
    # st.components.v1.html은 제거 예정(2026-06-01 이후) — 후속 API인 st.iframe 사용
    st.iframe(
        f"""<button onclick="window.parent.print()"
        style="float:right;background:#2E68B0;color:#fff;border:0;border-radius:6px;
        padding:7px 14px;font-size:13px;font-weight:600;cursor:pointer;
        font-family:'Pretendard',-apple-system,sans-serif">{label}</button>
        <div style="clear:both"></div>""", height=44)
    st.caption("버튼이 안 눌리면 Ctrl+P(Mac ⌘+P)를 누르고 대상을 'PDF로 저장'으로 바꿔 주세요.")

# ══════════════════════════════════════════════════════
# 메인 앱
# ══════════════════════════════════════════════════════
def main():
    if "wr_texts" not in st.session_state:
        st.session_state.wr_texts = load_insights()

    with st.sidebar:
        st.markdown("## 📋 주간보고 통합")
        files = st.file_uploader(
            "원천 엑셀/CSV/ZIP 업로드 (복수 선택)",
            type=["xlsx", "xls", "csv", "zip"], accept_multiple_files=True, key="wr_up",
            help="주간 폴더를 zip으로 묶어 통째로 올려도 돼요. "
                 "전체관점 마스터(일·주·월)와 지표별 파일(가입율·가입자수·당일가입 첫구매율·비회원 트래픽), "
                 "MICRO 대시보드의 조직×카테고리(구분06×구분07) export를 자동으로 인식해요.")
        st.markdown("---")
        PAGES = ["01. 주간보고 요약", "02. 월별 추이", "03. 주차별 추이",
                 "04. 채널별 실적", "05. 통합 데이터·다운로드", "06. 앱푸시 동의 현황",
                 "07. 첫구매 고객 세그먼트 성과", "08. 조직·카테고리별 실적"]
        page = st.radio("페이지", PAGES, key="wr_page")

    stored = load_store()
    expanded = expand_uploads(files) if files else []
    df_new = combine_files(tuple(expanded)) if expanded else pd.DataFrame()
    # 조직×카테고리는 축이 둘이라 별도 store에 쌓는다 (같은 업로드에서 같이 갈라 담긴다)
    oc_stored = load_orgcat_store()
    oc_new = combine_orgcat(tuple(expanded)) if expanded else pd.DataFrame(columns=ORGCAT_COLS)
    odf = merge_orgcat(oc_stored, oc_new)

    has_any = not stored.empty or not df_new.empty or not odf.empty
    if files and df_new.empty and stored.empty:
        st.error("올린 파일에서 데이터를 읽지 못했어요. 파일명 형식을 확인해 주세요.")
        st.stop()
    if not has_any:
        st.markdown("## 📋 주간보고 통합 — 시작하기")
        st.caption("누적 데이터가 없어요. 원천 파일을 올리거나 예전 백업(ZIP)을 복원해 주세요.")
        cU, cR = st.columns(2)
        with cU:
            st.markdown("#### ① 원천 파일 업로드")
            source_upload_widget("wr_up_main")
        with cR:
            st.markdown("#### ② 또는 백업 복원")
            restore_widget("wr_restore_empty", label="백업 ZIP / CSV / JSON 올리기")
            st.caption("예전에 받은 `주간보고백업_*.zip`을 그대로 올리면 데이터와 메모가 통째로 복원돼요.")
        st.markdown("""
---
**인식되는 원천 파일**
- **마스터**: `전체관점 - 일자별/주별/월별 실적 (기본)`
- **지표별**: `월_가입율(일평균)`, `주_가입자수(일평균)`, `일_비회원 트래픽(일평균)`, `월_당일가입 첫구매율 (일평균)` …
- **조직×카테고리**: MICRO 대시보드 export (헤더에 `구분06`·`구분07`) — 일·주·월 각각
- **앱푸시**: 파일명에 `PUSH`/`앱푸시`/`수신동의` 포함 또는 헤더가 앱푸시 형식이면 자동 인식
- 주간 폴더를 **zip으로 묶어 통째로** 올려도 돼요.
""")
        st.stop()

    # 누적 저장소와 병합 — 미리보기(저장 전까지 영구 반영 안 함)
    df = merge_store(stored, df_new)

    has_new = not df_new.empty or not oc_new.empty
    sig = tuple(sorted((n, len(b)) for n, b in expanded))
    with st.sidebar:
        # 업로드한 파일이 각각 무엇으로 인식됐는지 (미인식 파일 가시화)
        if expanded:
            cls = classify_uploads(tuple(expanded))
            n_ok = sum(1 for _, k, _ in cls if k.startswith(("✅", "♻")))
            n_bad = len(cls) - n_ok
            head = f"📄 업로드 인식 {n_ok}/{len(cls)}" + (f" · ⚠ 미인식 {n_bad}" if n_bad else "")
            with st.expander(head, expanded=bool(n_bad)):
                for nm, kind, nrows in cls:
                    base = os.path.basename(nm)
                    st.markdown(f"- `{base}` → {kind}"
                                + (f" ({nrows:,}행)" if nrows else ""))
                if n_bad:
                    st.caption("인식 못 한 파일은 파일명(전체관점·월_가입율 등)과 형식을 확인해 주세요. "
                               "인식된 파일만 저장에 반영돼요.")
        if has_new:
            added, updated = upload_diff(stored, df_new)
            saved = st.session_state.get("wr_saved_sig") == sig
            if saved:
                st.success("저장됨 ✓ (누적 반영 완료)")
            else:
                _oc_note = (f"\n\n조직×카테고리 {len(oc_new):,}행도 함께 반영돼요."
                            if not oc_new.empty else "")
                st.warning(f"새 데이터 감지 — 추가 {added}기간 · 갱신(겹침) {updated}기간"
                           + _oc_note + "\n\n**저장**을 눌러야 누적에 반영돼요.")
                if st.button("💾 저장 (누적 반영)", key="wr_commit",
                             type="primary", width="stretch"):
                    if not df_new.empty: save_store(df)
                    if not oc_new.empty: save_orgcat_store(odf)
                    st.session_state["wr_saved_sig"] = sig
                    st.rerun()

    if df.empty:
        # 사이드바 기준 기간·차트 연도가 전부 마스터에서 나오므로 01~07은 열 수 없다.
        # 08은 자체 기간 선택을 쓰니 조직×카테고리만 올린 상태에서도 보여 준다.
        if page.startswith("08.") and not odf.empty:
            render_orgcat_page(odf)
            st.stop()
        st.warning("첫구매(전체관점·지표별) 데이터가 없어요. 원천 파일을 올려 주세요."
                   + (f" (조직×카테고리 {len(odf):,}행은 저장돼 있어요 — "
                      "「08. 조직·카테고리별 실적」에서 볼 수 있어요)" if not odf.empty else ""))
        st.stop()

    # ── 인식 결과 + 필터
    years_all = sorted(df["year"].dropna().unique().astype(int))
    ly, llabel = latest_period(df, "월")
    ref_year_default = ly or years_all[-1]
    with st.sidebar:
        st.markdown("---")
        src = f"인식된 파일 {len(expanded)}개" if expanded else "누적 데이터 사용 중"
        st.caption(f"{src} · 지표 {df['metric'].nunique()}종 · "
                   f"{years_all[0]}–{years_all[-1]}년")
        st.markdown("**기준 기간**")
        ref_year = st.selectbox("기준 연도", years_all[::-1],
                                index=years_all[::-1].index(ref_year_default), key="wr_refy")
        months_avail = sorted({int(re.match(r"(\d+)월", l).group(1))
                               for l in df[(df["gran"] == "월") & (df["year"] == ref_year)]["label"]})
        ref_month = st.selectbox("기준 월", months_avail[::-1], key="wr_refm")
        weeks_avail = (df[(df["gran"] == "주") & (df["year"] == ref_year) & df["value"].notna()]
                       [["label", "sortkey"]].drop_duplicates()
                       .sort_values("sortkey")["label"].tolist())
        if weeks_avail:
            # 최신 주차가 진행 중(일마감 데이터이거나 오늘이 속한 주차)이면 직전 주를 기본값으로
            latest_w = weeks_avail[-1]
            today = today_kst()
            cur_week_lbl = f"{today.month:02d}월 {(today.day - 1) // 7 + 1}주차"
            is_partial = not df[(df["gran"] == "주") & (df["year"] == ref_year) &
                                (df["label"] == latest_w) & (df["close"] == "mtd")].empty
            in_progress = is_partial or (ref_year == today.year and latest_w == cur_week_lbl)
            default_week = (weeks_avail[-2] if in_progress and len(weeks_avail) >= 2
                            else latest_w)
            ref_week = st.selectbox("기준 주차", weeks_avail[::-1],
                                    index=weeks_avail[::-1].index(default_week),
                                    key="wr_refw",
                                    help="주간보고 대상 주차예요. 최신 주차가 진행 중이면 직전 완료 주차가 기본으로 잡혀요.")
        else:
            ref_week = None
        st.markdown("**차트 연도**")
        default_yrs = years_all[-2:] if len(years_all) >= 2 else years_all
        chart_years = st.multiselect("비교 연도", years_all, default=default_yrs, key="wr_cyrs")
        st.markdown("**채널**")
        ch_sel = st.multiselect("채널 선택", CHANNELS, default=CHANNELS, key="wr_ch")
        if not chart_years: chart_years = default_yrs

        st.markdown("---")
        st.markdown("**AI 인사이트 모델**")
        ai_label = st.selectbox("모델 선택", list(AI_MODELS.keys()), key="wr_ai_model_label")
        st.session_state["wr_ai_model"] = AI_MODELS[ai_label]
        st.caption("✅ API 키 설정됨" if _anthropic_key()
                   else "⚠ ANTHROPIC_API_KEY가 없어요. Secrets에 추가해 주세요")

        st.markdown("---")
        st.markdown("**백업 · 복원**")
        saved_rows = len(load_store())
        pend = " · 저장 시 반영" if (has_new and not st.session_state.get("wr_saved_sig") == sig) else ""
        st.caption(f"누적 {saved_rows:,}행 저장됨 / 현재 보기 {len(df):,}행{pend} · "
                   f"조직×카테고리 {len(odf):,}행 · "
                   f"메모 {len(st.session_state.wr_texts)}개")

        # 통합 백업 — 누적 데이터 + 메모를 한 ZIP으로 (하나만 받아도 전부 보존)
        st.download_button(
            "⬇ 통합 백업 (ZIP · 데이터+메모)",
            make_backup_zip(df, st.session_state.wr_texts, odf),
            f"주간보고백업_{today_kst():%Y%m%d}.zip", "application/zip",
            width="stretch", type="primary",
            help="누적 데이터 CSV와 보고란·메모 JSON을 한 파일로 백업해요. "
                 "재배포로 초기화돼도 이 ZIP을 '백업 복원'에 올리면 그대로 되살아나요.")

        # 통합 복원 — zip/csv/json 자동 인식
        restore_widget("wr_restore")

        with st.expander("개별 백업 (데이터만 / 메모만)"):
            st.download_button("⬇ 누적 데이터 (CSV)",
                               df[STORE_COLS].to_csv(index=False).encode("utf-8-sig"),
                               f"wr_data_store_{today_kst():%Y%m%d}.csv", "text/csv",
                               width="stretch")
            if not odf.empty:
                st.download_button("⬇ 조직×카테고리 (CSV)",
                                   odf[ORGCAT_COLS].to_csv(index=False).encode("utf-8-sig"),
                                   f"wr_orgcat_store_{today_kst():%Y%m%d}.csv", "text/csv",
                                   width="stretch")
            st.download_button("⬇ 보고란·메모 (JSON)",
                               json.dumps(st.session_state.wr_texts, ensure_ascii=False,
                                          indent=2).encode("utf-8"),
                               f"wr_insights_{today_kst():%Y%m%d}.json", "application/json",
                               width="stretch")

        # 초기화 — 2단계 확인 (실수 방지)
        if st.session_state.get("wr_confirm_clear"):
            st.warning("정말 초기화할까요? 누적 데이터와 조직×카테고리가 모두 지워지고 "
                       "되돌릴 수 없어요. 메모는 남아요. 초기화 전에 통합 백업을 받아두세요.")
            cc1, cc2 = st.columns(2)
            if cc1.button("삭제 확인", key="wr_clear_yes", type="primary", width="stretch"):
                if os.path.exists(DATA_STORE): os.remove(DATA_STORE)
                if os.path.exists(ORGCAT_STORE): os.remove(ORGCAT_STORE)
                st.cache_data.clear()
                st.session_state["wr_confirm_clear"] = False
                st.session_state.pop("wr_saved_sig", None)
                st.rerun()
            if cc2.button("취소", key="wr_clear_no", width="stretch"):
                st.session_state["wr_confirm_clear"] = False; st.rerun()
        else:
            if st.button("🗑 누적 데이터 초기화", key="wr_clear_store", width="stretch"):
                st.session_state["wr_confirm_clear"] = True; st.rerun()

    texts = st.session_state.wr_texts
    print_button()

    # ════════════ 01. 주간보고 요약 ════════════
    if page == "01. 주간보고 요약":
        st.markdown(f"## 첫구매 주간보고 — {ref_year}년 {ref_month}월")
        wy, wlabel = week_ref(df, ref_year, ref_week)
        if wlabel:
            st.caption(f"기준 주차: {week_disp(wy, wlabel)}")

        # KPI 카드 (기준 주차 — 전주비·전월비·전년비 모두 주차 기준)
        if wlabel:
            cols = st.columns(4)
            kpi_metrics = ["첫구매 거래액", "첫구매 고객수", "첫구매 객단가", "가입자수"]
            py, plb = prev_label(df, "주", wy, wlabel)
            # 전월 동일 주차 (예: 06월 1주차 → 05월 1주차).
            # 5주차처럼 전월(또는 전년 동월)에 같은 주차가 없으면 그 달 마지막 주차로
            # 대체하고, 어떤 주와 비교했는지 뱃지에 같이 적는다.
            mom_y = mom_lbl = None
            yoy_lbl, mom_exact, yoy_exact = wlabel, True, True
            wm = re.match(r"(\d{1,2})월 (\d)주차", wlabel)
            if wm:
                mo, wk = int(wm.group(1)), int(wm.group(2))
                mom_y, mom_m = (wy, mo - 1) if mo > 1 else (wy - 1, 12)
                mom_lbl, mom_exact = week_like(df, mom_y, mom_m, wk)
                yoy_lbl, yoy_exact = week_like(df, wy - 1, mo, wk)
            mom_tag = "전월비" if (mom_exact or not mom_lbl) else f"전월비 · {mom_lbl}"
            yoy_tag = "전년비" if (yoy_exact or not yoy_lbl) else f"전년비 · {yoy_lbl}"
            for col, met in zip(cols, kpi_metrics):
                cur = pick(df, "주", met, "*TOTAL", wy, wlabel, "mtd")
                deltas = []
                prv = pick(df, "주", met, "*TOTAL", py, plb, "final") if plb else np.nan
                deltas.append((fmt_delta(met, cur, prv), "전주비"))
                if wm:
                    mom = (pick(df, "주", met, "*TOTAL", mom_y, mom_lbl, "final")
                           if mom_lbl else np.nan)
                    deltas.append((fmt_delta(met, cur, mom), mom_tag))
                yoy = (pick(df, "주", met, "*TOTAL", wy - 1, yoy_lbl, "final")
                       if yoy_lbl else np.nan)
                deltas.append((fmt_delta(met, cur, yoy), yoy_tag))
                pills = ""
                for d, lab in deltas:
                    if d:
                        neg = d.startswith("△")
                        cls = "down" if neg else "up"
                        prefix = "" if neg else "↑ "
                        pills += f'<div class="kpi-delta {cls}">{prefix}{d} ({lab})</div>'
                    else:
                        pills += f'<div class="kpi-delta na">– ({lab})</div>'
                col.markdown(
                    f'<div class="kpi-card"><div class="kpi-label">{met} ({week_disp(wy, wlabel)})</div>'
                    f'<div class="kpi-value">{fmt_value(met, cur)}</div>{pills}</div>',
                    unsafe_allow_html=True)
            _sub = []
            if wm and not mom_exact and mom_lbl:
                _sub.append(f"전월에 {wm.group(2)}주차가 없어 **{mom_y}년 {mom_lbl}**(전월 마지막 주)와 비교")
            if wm and not yoy_exact and yoy_lbl:
                _sub.append(f"전년 동월에 {wm.group(2)}주차가 없어 **{wy-1}년 {yoy_lbl}**와 비교")
            if _sub:
                st.caption("ℹ️ " + " · ".join(_sub) + " — 주차 수가 달마다 4~5개로 달라서예요.")

            # 전전년까지 펼쳐보기 — 올해 흐름이 작년에도 있었는지(= 구조적인지) 확인용
            with st.expander(f"🔍 전년비 · 전전년비 비교 — {wy-2}년에도 같은 추세였는지",
                             expanded=False):
                t2, _l1, _l2, _e1, _e2 = yoy2_summary_table(df, wy, wlabel, METRICS7)
                _basis = []
                if not _e1 and _l1: _basis.append(f"전년 {wy-1}년 {_l1}")
                if not _e2 and _l2: _basis.append(f"전전년 {wy-2}년 {_l2}")
                st.caption(
                    f"기준: {week_disp(wy, wlabel)}"
                    + (f" · 대응 주차가 없어 {' · '.join(_basis)} 기준으로 비교" if _basis else "")
                    + " — **전년의 전년비**는 전년 동주가 그 전해 대비 어땠는지예요. "
                      "**추세**가 '2년 연속'이면 올해만의 현상이 아니라 이어져 온 흐름이에요.")
                wtable(style_delta_cols(t2), width="stretch")
            st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)

        # 실적 요약 — 전주비 (주차 기준)
        if wlabel:
            st.subheader("실적 요약 (일평균 · 전주비)")
            st.caption(f"기준 주차: {week_disp(wy, wlabel)} — 전주·전년 동주 대비")
            wtable(style_delta_cols(wow_summary_table(df, wy, wlabel, METRICS7)),
                         width="stretch", dl_name="실적 요약 (일평균 · 전주비)")
            st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)

        # 실적 요약 YoY 표
        st.subheader("실적 요약 (일평균 · 전년비)")
        tbl, (pm_y, pm_m) = yoy_summary_table(df, ref_year, ref_month, METRICS7)
        st.caption(f"전월({pm_m}월)은 월마감, 당월({ref_month}월)은 일마감(MTD) 기준 동일기간 비교. "
                   f"맨 끝 전월비(당월)은 {ref_year}년 {ref_month}월 ↔ {pm_y}년 {pm_m}월이에요 — "
                   "값이 일평균이라 기간 길이가 달라도 맞댈 수 있어요.")
        wtable(style_delta_cols(tbl), width="stretch", dl_name="실적 요약 (일평균 · 전년비)")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)

        # 다각도 뷰 — 전환 퍼널(단계 카드) + 채널 기여. 비교 기준 전환 가능
        cmp_mode = st.radio("비교 기준", ["주간 — 전년 동주", "월누적(MTD) — 전년 동월"],
                            horizontal=True, key="wr_multi_cmp")
        weekly_mode = cmp_mode.startswith("주간")
        if weekly_mode and not wlabel:
            st.info("주차 데이터가 없어요. 월누적(MTD) 비교를 골라 주세요.")
        else:
            if weekly_mode:
                period_lbl, base_lbl = week_disp(wy, wlabel), "전년 동주"
                x_prv, x_cur = f"{wy-1}년", f"{wy}년"
                def get_cur(met, seg="*TOTAL"):
                    return pick(df, "주", met, seg, wy, wlabel, "mtd")
                def get_prv(met, seg="*TOTAL"):
                    return pick(df, "주", met, seg, wy - 1, wlabel, "final")
            else:
                period_lbl = f"{ref_year}년 {ref_month}월 누적(MTD)"
                base_lbl = "전년 동월 MTD"
                x_prv, x_cur = f"{ref_year-1}년 {ref_month}월", f"{ref_year}년 {ref_month}월"
                def get_cur(met, seg="*TOTAL"):
                    return pick(df, "월", met, seg, ref_year, month_label(ref_month), "mtd")
                def get_prv(met, seg="*TOTAL"):
                    # 전년 동월도 동일기간(MTD) 잘린 값 우선 — 실적 요약 표와 동일 기준
                    return pick(df, "월", met, seg, ref_year - 1, month_label(ref_month), "mtd")

            st.subheader("전환 퍼널 — 트래픽→가입→첫구매")
            st.caption(f"{period_lbl} vs {base_lbl} (일평균) · "
                       "첫구매 고객에는 과거 가입자도 포함되어 단계 비율은 참고용")
            stages = ["비회원트래픽", "가입자수", "첫구매 고객수"]
            cur_v = [get_cur(m) for m in stages]
            pry_v = [get_prv(m) for m in stages]
            if any(np.isnan(v) for v in cur_v):
                st.info("이 주차는 퍼널 데이터가 부족해요.")
            else:
                # 트래픽이 가입·첫구매의 수십~수백 배라 면적형 퍼널은 왜곡됨 →
                # 단계 카드 + 전환율 pill 로 표현 (전환율을 1급 정보로)
                def _stage_rate(a, b):
                    return a / b if (not np.isnan(a) and not np.isnan(b) and b > 0) else np.nan
                cells = []
                for i, (mname, cv, pv) in enumerate(zip(stages, cur_v, pry_v)):
                    d = fmt_delta(mname, cv, pv)
                    pill = (f'<div class="kpi-delta {"down" if d.startswith("△") else "up"}">'
                            f'{d} (전년동주)</div>' if d else
                            '<div class="kpi-delta na">– (전년동주)</div>')
                    cells.append(
                        f'<div class="kpi-card" style="flex:1.2;min-width:0">'
                        f'<div class="kpi-label">{mname}</div>'
                        f'<div class="kpi-value">{cv:,.0f}명</div>{pill}</div>')
                    if i < 2:
                        cr = _stage_rate(cur_v[i + 1], cv)
                        pr = _stage_rate(pry_v[i + 1], pv)
                        rate_lbl = "가입율" if i == 0 else "가입 대비 첫구매"
                        cur_s = "–" if np.isnan(cr) else f"{cr*100:.2f}%"
                        sub = ""
                        if not (np.isnan(cr) or np.isnan(pr)):
                            pp = (cr - pr) * 100
                            colr = "#dc2626" if pp < 0 else "#16a34a"
                            sign = "△" if pp < 0 else "+"
                            sub = (f'<div style="font-size:11px;color:#64748b">전년 {pr*100:.2f}% '
                                   f'<span style="color:{colr};font-weight:700">'
                                   f'{sign}{abs(pp):.2f}%p</span></div>')
                        cells.append(
                            f'<div style="flex:0.9;display:flex;flex-direction:column;'
                            f'justify-content:center;text-align:center;min-width:0">'
                            f'<div style="color:#94a3b8;font-size:16px;line-height:1">→</div>'
                            f'<div style="font-size:12px;color:#64748b;margin-top:2px">{rate_lbl}</div>'
                            f'<div style="font-size:17px;font-weight:700;color:#1e293b">{cur_s}</div>'
                            f'{sub}</div>')
                st.markdown('<div style="display:flex;gap:10px;align-items:stretch">'
                            + "".join(cells) + '</div>', unsafe_allow_html=True)

            st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
            st.subheader("채널 기여 분해 — 어디서 늘고 줄었나")
            # 채널 합 ≈ TOTAL 이 성립하는 가산 지표만 (객단가·가입율·CR은 비율이라 분해 무의미)
            ADDITIVE = ["첫구매 거래액", "첫구매 고객수", "가입자수", "비회원트래픽"]
            avail_dec = [m for m in ADDITIVE
                         if not df[(df["metric"] == m)
                                   & df["segment"].isin(CHANNELS)].empty]
            if not avail_dec:
                st.info("채널별로 나눠 볼 지표 데이터가 없어요.")
            else:
                dec_met = st.selectbox("분해 지표", avail_dec, key="wr_decomp_met",
                                       help="선택한 지표의 YoY 증감을 채널별로 나눠 봐요. "
                                            "채널 합이 전체와 맞는 가산 지표만 골라요.")
                # 거래액만 만원(더 잘게), 나머지 카운트는 명 그대로
                div, unit = (1e4, "만원") if dec_met == "첫구매 거래액" else (1, "명")
                st.caption(f"{period_lbl} {base_lbl} 대비 «{dec_met}» 증감 분해 "
                           f"(일평균·{unit}) · 세로축은 변동 구간만 확대 표시")
                base_t = get_prv(dec_met)
                cur_t = get_cur(dec_met)
                if np.isnan(base_t) or np.isnan(cur_t):
                    st.info("이 기간엔 채널 데이터가 없어요.")
                else:
                    labels, deltas = [], []
                    for chn in CHANNELS:
                        pv = get_prv(dec_met, chn)
                        cv = get_cur(dec_met, chn)
                        # 한쪽만 있으면 결측(채널분해 부재) — 0 취급하면 가짜 급락이 되므로 제외.
                        # 실제 0은 pick이 0.0으로 돌려주므로 유지된다.
                        if np.isnan(pv) or np.isnan(cv): continue
                        labels.append(chn); deltas.append((cv - pv) / div)
                    if not labels:
                        st.info(f"이 기간은 «{dec_met}» 채널별 분해 데이터가 없어요. 전체 값만 있어요. "
                                "사이드바에서 다른 기준 주차를 고르거나, 위 비교 기준을 "
                                "**월누적(MTD)** 으로 바꿔 보세요.")
                    else:
                        # 잔차(채널합↔TOTAL 차이·미분류)를 '기타'로 표시해 총계 막대가 항상
                        # 라벨(=당년 TOTAL)과 정확히 맞게 한다. 0.5단위 미만(반올림 0)만 생략.
                        resid = (cur_t - base_t) / div - sum(deltas)
                        if abs(resid) >= 0.5:
                            labels.append("기타"); deltas.append(resid)
                        fig_w = go.Figure(go.Waterfall(
                            x=[x_prv] + labels + [x_cur],
                            measure=["absolute"] + ["relative"] * len(labels) + ["total"],
                            y=[base_t / div] + deltas + [0],
                            text=[f"{base_t/div:,.0f}"] + [f"{d:+,.0f}" for d in deltas]
                                 + [f"{cur_t/div:,.0f}"],
                            textposition="outside", cliponaxis=False,
                            increasing=dict(marker=dict(color=clr("green"))),
                            decreasing=dict(marker=dict(color=clr("red"))),
                            totals=dict(marker=dict(color=clr("slate"))),
                            connector=dict(line=dict(color="#e2e8f0")),
                            hovertemplate=("%{x}<br>변동 %{delta:+,.0f}" + unit
                                           + " · 누계 %{final:,.0f}" + unit + "<extra></extra>"),
                        ))
                        # 총액 막대가 델타를 압도하지 않도록 세로축을 변동 구간 주변으로 제한
                        run, acc = [base_t / div], base_t / div
                        for d in deltas:
                            acc += d; run.append(acc)
                        run.append(cur_t / div)
                        lo, hi = min(run), max(run)
                        span = max(hi - lo, abs(hi) * 0.01, 1.0)
                        lyw = base_layout(360, title=f"일평균 {dec_met} ({unit}) — 변동 구간 확대")
                        lyw["showlegend"] = False
                        lyw["yaxis"]["range"] = [lo - span * 0.45, hi + span * 0.55]
                        fig_w.update_layout(**lyw)
                        st.plotly_chart(fig_w, width="stretch")

                    # 산출식 (접이식 첨부)
                    with st.expander("📐 산출식 · 계산 방법", expanded=False):
                        st.markdown(f"""
**비교 기준**: {period_lbl} vs {base_lbl} · 값은 모두 **일평균**, 단위 **{unit}**
{"(거래액은 가독성을 위해 원 → 만원)" if dec_met == "첫구매 거래액" else ""}

**막대 구성** (왼쪽 → 오른쪽 누적):

- **시작(전년 전체)** = 전년 동기 «{dec_met}» 전체(*TOTAL) 값
- **채널별 증감** = `당년 채널값 − 전년 채널값`
  ㄴ 전년·당년 **양쪽에 값이 있는 채널만** 포함 (한쪽만 있으면 결측으로 보고 제외 — 0으로 두면 가짜 급락이 되므로)
- **기타** = `당년 전체 − 전년 전체 − Σ(채널별 증감)`
  ㄴ 채널 합과 전체의 차이·채널 미분류분을 담는 잔차 (0.5{unit} 미만이면 생략)
- **끝(당년 전체)** = 당년 «{dec_met}» 전체(*TOTAL) 값
  ㄴ 항상 `시작 + Σ(채널별 증감) + 기타` 와 정확히 일치

**대상 지표**: 채널 합이 전체와 일치하는 **가산 지표만** 제공
(거래액·고객수·가입자수·비회원트래픽). 객단가·가입율·당일가입CR 등 **비율 지표는
채널 합 ≠ 전체** 라 분해가 성립하지 않아 뺐어요.
""")
            st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)

        # 보고란
        draft = auto_draft(df, ref_year, ref_month, ref_week=wlabel)
        ai_model = st.session_state.get("wr_ai_model", DEFAULT_AI_MODEL)
        cL, cR = st.columns(2)
        with cL:
            report_text_block("wr_metrics_summary", "전주 주요 지표 현황",
                              default=draft, regen=draft,
                              ai_fn=lambda memo: ai_generate_insight(df, ref_year, ref_month,
                                                                wlabel, ai_model, memo=memo))
        with cR:
            report_text_block("wr_exec_summary", "금주 집행 내용 요약")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)

        # 핵심 차트 (주차별 YoY 3종)
        st.subheader("주차별 추이 — 전년 비교")
        c1, c2, c3 = st.columns(3)
        for col, met in zip((c1, c2, c3), ["첫구매 거래액", "첫구매 고객수", "첫구매 객단가"]):
            with col:
                st.plotly_chart(yoy_chart(df, "주", met, chart_years, h=280),
                                width="stretch")

    # ════════════ 02. 월별 추이 ════════════
    elif page == "02. 월별 추이":
        st.markdown("## 월별 추이")
        st.subheader("월별 추이 차트 — 전년 비교")
        c1, c2, c3 = st.columns(3)
        for col, met in zip((c1, c2, c3), ["첫구매 거래액", "첫구매 고객수", "첫구매 객단가"]):
            with col:
                st.plotly_chart(yoy_chart(df, "월", met, chart_years, h=280),
                                width="stretch")
        c4, c5, c6 = st.columns(3)
        for col, met in zip((c4, c5, c6), ["비회원트래픽", "가입자수", "가입율"]):
            with col:
                st.plotly_chart(yoy_chart(df, "월", met, chart_years, h=280),
                                width="stretch")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.subheader("월별 추이표 (일평균)")
        tbl = trend_table(df, "월", METRICS7, chart_years)
        wtable(style_trend(tbl, METRICS7), width="stretch", dl_name="월별 추이표 (일평균)")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        ai_model = st.session_state.get("wr_ai_model", DEFAULT_AI_MODEL)
        report_text_block(
            f"wr_month_memo_{ref_year}_{ref_month}",
            f"{ref_year}년 {ref_month}월 액션·이슈사항",
            ai_fn=lambda memo: ai_generate_insight(df, ref_year, ref_month, None, ai_model,
                                              focus=f"{ref_year}년 {ref_month}월 액션·이슈 및 인사이트",
                                              memo=memo))

    # ════════════ 03. 주차별 추이 ════════════
    elif page == "03. 주차별 추이":
        st.markdown("## 주차별 추이")
        st.subheader("주차별 추이 차트 — 전년 비교")
        c1, c2, c3 = st.columns(3)
        for col, met in zip((c1, c2, c3), ["첫구매 거래액", "첫구매 고객수", "첫구매 객단가"]):
            with col:
                st.plotly_chart(yoy_chart(df, "주", met, chart_years, h=280),
                                width="stretch")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.subheader(f"주차별 추이표 — {ref_year}년")
        tbl = trend_table(df, "주", METRICS7, [ref_year])
        if not tbl.empty:
            recent = tbl.columns[-16:]
            wtable(style_trend(tbl[recent], METRICS7), width="stretch", dl_name="주차별 추이 차트 — 전년 비교")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.subheader("전주비(WoW)·전년비(YoY) 증감")
        wy, wlabel = week_ref(df, ref_year, ref_week)
        if wlabel:
            st.caption(f"기준 주차: {week_disp(wy, wlabel)}")
            wtable(style_delta_cols(wow_summary_table(df, wy, wlabel, METRICS7)),
                         width="stretch", dl_name="전주비(WoW)·전년비(YoY) 증감")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        ai_model = st.session_state.get("wr_ai_model", DEFAULT_AI_MODEL)
        report_text_block(
            f"wr_week_memo_{wy}_{wlabel}",
            f"{wy}년 {wlabel} 액션·이슈사항" if wlabel else "주차별 액션·이슈사항",
            ai_fn=lambda memo: ai_generate_insight(df, ref_year, ref_month, wlabel, ai_model,
                                              focus=f"{wlabel} 주차 액션·이슈 및 인사이트",
                                              memo=memo))

    # ════════════ 04. 채널별 실적 ════════════
    elif page == "04. 채널별 실적":
        st.markdown("## 채널별 실적")
        avail = [m for m in METRICS7 if (df["metric"] == m).any()]
        met = st.selectbox("지표 선택", avail, key="wr_chmet")

        st.subheader(f"{met} — {ref_year}년 {ref_month}월 채널별 전년비")
        rows = []
        for seg in ["*TOTAL"] + [c for c in CHANNELS if c in ch_sel]:
            pv = pick(df, "월", met, seg, ref_year - 1, month_label(ref_month), "mtd")
            cv = pick(df, "월", met, seg, ref_year, month_label(ref_month), "mtd")
            rows.append({"채널": seg,
                         f"{ref_year-1}년 {ref_month}월": fmt_value(met, pv),
                         f"{ref_year}년 {ref_month}월": fmt_value(met, cv),
                         "전년비": fmt_delta(met, cv, pv) or "–"})
        wtable(style_delta_cols(pd.DataFrame(rows).set_index("채널")),
                     width="stretch", dl_name="채널별 실적")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.subheader(f"{met} — 채널별 월 추이 ({ref_year}년)")
        unit, div = METRIC_UNIT.get(met, ("", 1))
        if met in PCT_METRICS: div, unit = 0.01, "%"
        fig = go.Figure()
        x = [month_label(i) for i in range(1, 13)]
        for seg in [c for c in CHANNELS if c in ch_sel]:
            s = series_by_label(df, "월", met, seg, ref_year).reindex(x).dropna()
            if s.empty: continue
            fig.add_trace(go.Scatter(
                x=s.index.tolist(), y=(s / div).tolist(), mode="lines+markers", name=seg,
                line=dict(color=clr(CHANNEL_PAL.get(seg, "blue")), width=1.8),
                marker=dict(size=4)))
        ly = base_layout(340, ysuffix=unit if unit == "%" else "",
                         title=f"{met} 채널별 ({unit})")
        fig.update_layout(**ly)
        st.plotly_chart(fig, width="stretch")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.subheader(f"{met} — 채널 × 월 표 ({ref_year}년)")
        rows = []
        for seg in ["*TOTAL"] + [c for c in CHANNELS if c in ch_sel]:
            s = series_by_label(df, "월", met, seg, ref_year)
            row = {"채널": seg}
            for lb in [month_label(i) for i in range(1, 13)]:
                if lb in s.index and not np.isnan(s[lb]):
                    row[lb] = fmt_value(met, s[lb])
            rows.append(row)
        wtable(pd.DataFrame(rows).set_index("채널"), width="stretch")

    # ════════════ 05. 통합 데이터·다운로드 ════════════
    elif page == "05. 통합 데이터·다운로드":
        st.markdown("## 통합 데이터 · 다운로드")
        st.caption("올린 파일을 모두 합친 통합 long 데이터예요.")
        wtable(df.sort_values(["gran", "metric", "segment", "sortkey"]).head(2000),
                     width="stretch", height=420, dl_name="통합 데이터 · 다운로드")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.subheader("통합 워크북 다운로드")
        st.caption("`첫구매_요약`(요약표·보고란·YoY 차트·채널표) + `월`·`주`(통합 데이터) 3개 시트")
        if st.button("📥 엑셀 워크북 생성", key="wr_build"):
            with st.spinner("워크북 생성 중…"):
                xls = build_workbook(df, st.session_state.wr_texts,
                                     ref_year, ref_month, chart_years)
            st.download_button(
                "다운로드 — 첫구매_주간보고.xlsx", xls,
                file_name=f"첫구매_주간보고_{ref_year}{ref_month:02d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        csv = df.to_csv(index=False).encode("utf-8-sig")
        st.download_button("통합 long 데이터 CSV", csv,
                           f"통합데이터_{today_kst():%Y%m%d}.csv", "text/csv")

    # ════════════ 06. 앱푸시 동의 현황 ════════════
    elif page == "06. 앱푸시 동의 현황":
        render_push_page(df, ref_year, chart_years)

    # ════════════ 07. 첫구매 고객 세그먼트 성과 ════════════
    elif page == "07. 첫구매 고객 세그먼트 성과":
        st.markdown("## 첫구매 고객 세그먼트 성과")
        
        # 세그먼트 선택 필터 추가
        all_segs = ["1_신규", "1_당월신규", "2_기가입신규", "3_기존"]
        
        # 데이터프레임에 존재하는 세그먼트만 추출
        available_segs = []
        for s in all_segs:
            if not df[df["segment"] == s].empty:
                available_segs.append(s)
                
        if not available_segs:
            available_segs = all_segs
            
        sel_seg = st.selectbox("세그먼트 선택", available_segs)
        segs = [sel_seg]
        
        def wow_segment_table(wy, wlabel):
            df_gran = df[df["gran"] == "주"]
            # 해당 세그먼트에 존재하는 모든 지표 추출 (순서 유지)
            metrics = df_gran[df_gran["segment"] == sel_seg]["metric"].dropna().unique().tolist()
            if not metrics:
                metrics = ["거래액", "고객수", "객단가", "CR"]
                
            py, plb = prev_label(df, "주", wy, wlabel)
            cols = [week_disp(py, plb), week_disp(wy, wlabel), "전주비",
                    week_disp(wy - 1, wlabel), "전년비"]
            rows = []
            
            def _get_val(yr, lb, c, seg, met):
                v = pick(df, "주", met, seg, yr, lb, c)
                if pd.isna(v) and "객단가" in met:
                    # Fallback
                    rev_m = next((m for m in metrics if "거래액" in m and "비중" not in m), None)
                    cus_m = next((m for m in metrics if "고객수" in m and "비중" not in m), None)
                    if rev_m and cus_m:
                        rev = pick(df, "주", rev_m, seg, yr, lb, c)
                        cus = pick(df, "주", cus_m, seg, yr, lb, c)
                        if pd.notna(rev) and pd.notna(cus) and cus > 0:
                            v = rev / cus
                return v

            for seg in segs:
                if not df_gran[df_gran["segment"] == seg].empty:
                    for met in metrics:
                        cur = _get_val(wy, wlabel, "mtd", seg, met)
                        prv = _get_val(py, plb, "final", seg, met) if plb else np.nan
                        yoy = _get_val(wy - 1, wlabel, "final", seg, met)
                        rows.append({
                            "지표": met,
                            cols[0]: fmt_value(met, prv), cols[1]: fmt_value(met, cur),
                            cols[2]: fmt_delta(met, cur, prv) or "-",
                            cols[3]: fmt_value(met, yoy), cols[4]: fmt_delta(met, cur, yoy) or "-",
                        })
            if rows:
                return pd.DataFrame(rows).set_index("지표")
            return pd.DataFrame(columns=["지표"] + cols).set_index("지표")
            
        def yoy_segment_table(ry, rm):
            df_gran = df[df["gran"] == "월"]
            metrics = df_gran[df_gran["segment"] == sel_seg]["metric"].dropna().unique().tolist()
            if not metrics:
                metrics = ["거래액", "고객수", "객단가", "CR"]
                
            pm_y, pm_m = (ry, rm - 1) if rm > 1 else (ry - 1, 12)
            cols = [f"{pm_y-1}년 {pm_m}월", f"{pm_y}년 {pm_m}월", "전년비 (전월)",
                    f"{ry-1}년 {rm}월", f"{ry}년 {rm}월", "전년비 (당월)"]
            rows = []
            
            def _get_val(yr, m, c, seg, met):
                lb = month_label(m)
                v = pick(df, "월", met, seg, yr, lb, c)
                if pd.isna(v) and "객단가" in met:
                    rev_m = next((m2 for m2 in metrics if "거래액" in m2 and "비중" not in m2), None)
                    cus_m = next((m2 for m2 in metrics if "고객수" in m2 and "비중" not in m2), None)
                    if rev_m and cus_m:
                        rev = pick(df, "월", rev_m, seg, yr, lb, c)
                        cus = pick(df, "월", cus_m, seg, yr, lb, c)
                        if pd.notna(rev) and pd.notna(cus) and cus > 0:
                            v = rev / cus
                return v
                
            for seg in segs:
                if not df_gran[df_gran["segment"] == seg].empty:
                    for met in metrics:
                        pm_prev = _get_val(pm_y - 1, pm_m, "final", seg, met)
                        pm_cur  = _get_val(pm_y, pm_m, "final", seg, met)
                        cm_prev = _get_val(ry - 1, rm, "mtd", seg, met)
                        cm_cur  = _get_val(ry, rm, "mtd", seg, met)
                        rows.append({
                            "지표": met,
                            cols[0]: fmt_value(met, pm_prev), cols[1]: fmt_value(met, pm_cur),
                            cols[2]: fmt_delta(met, pm_cur, pm_prev) or "-",
                            cols[3]: fmt_value(met, cm_prev), cols[4]: fmt_value(met, cm_cur),
                            cols[5]: fmt_delta(met, cm_cur, cm_prev) or "-",
                        })
            if rows:
                return pd.DataFrame(rows).set_index("지표"), (pm_y, pm_m)
            return pd.DataFrame(columns=["지표"] + cols).set_index("지표"), (pm_y, pm_m)

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        wy, wlabel = week_ref(df, ref_year, ref_week)
        
        if wlabel:
            st.subheader(f"실적 요약 (전주비)")
            st.caption(f"기준 주차: {week_disp(wy, wlabel)} — 전주·전년 동주 대비")
            tbl1 = wow_segment_table(wy, wlabel)
            if not tbl1.empty:
                wtable(style_delta_cols(tbl1), width="stretch", dl_name="실적 요약 (전주비)")
            else:
                st.info("이 주차 데이터가 없어요.")
            st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
            
        st.subheader(f"실적 요약 (전년비)")
        tbl2, (pm_y, pm_m) = yoy_segment_table(ref_year, ref_month)
        st.caption(f"전월({pm_m}월) 마감 및 당월({ref_month}월·MTD) 기준 동일기간 비교")
        if not tbl2.empty:
            wtable(style_delta_cols(tbl2), width="stretch", dl_name="실적 요약 (전년비)")
        else:
            st.info("이 달 데이터가 없어요.")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        with st.expander("📊 지표 산출식 및 용어 설명 (클릭하여 펼치기)", expanded=False):
            st.markdown("""
            **산출식**
            * **거래액비중**: 첫구매 거래액 ÷ 전체 거래액
            * **고객비중**: 첫구매 고객수 ÷ 전체 고객수
            * **첫구매 객단가**: 첫구매 거래액 ÷ 첫구매 고객수
            * **유입율**: DAU ÷ 유효회원수
            * **CR (전환율)**: 첫구매 고객수 ÷ DAU
            
            **용어 설명**
            * **유효회원수**: 서비스에 정상적으로 가입해서 활동할 수 있는 전체 회원 수예요.
            * **DAU (Daily Active Users)**: 하루 동안 서비스에 한 번 이상 방문해서 활동한 사용자 수예요.
            """)

    # ════════════ 08. 조직·카테고리별 실적 ════════════
    elif page == "08. 조직·카테고리별 실적":
        render_orgcat_page(odf)

if st.runtime.exists():
    main()
