# -*- coding: utf-8 -*-
"""
LF몰 CRM 발송성과 대시보드
─────────────────────────────────────────────────────────────
발송기획(문구포함) 시트 + 발송성과(실적) 시트를 (발송일 + AF코드)로 머지하여
"어떤 문구·오퍼·타이밍 패턴이 성과를 만드는가"를 도출하는 대시보드.

· 조인 방향: 실적(성과) 기준 — 기획에만 있고 실제 발송 안 된 건은 제외
· 문구 자동 태깅(규칙 기반) + Gemini AI 인사이트/처방
· 디자인: 기존 발송 피로도 / 첫구매 주간보고 대시보드 슬레이트 팔레트 계승

데이터 로직(parse_perf_bytes / parse_plan_bytes / merge_perf_plan / tag_copy)은
Streamlit 의존이 없는 순수 함수이며 모듈 import 만으로 테스트 가능하다.
앱 UI 는 main() 안에 있고 `python -m streamlit run` 시에만 실행된다.
"""
import io, os, re, json, hashlib, datetime, time
import numpy as np
import pandas as pd


def today_kst():
    """'오늘'을 한국 시간 기준으로 — 서버(Streamlit Cloud)가 UTC라 KST 자정~09시 사이에
    date.today()를 쓰면 '이번 주/금주' 경계가 하루 전으로 밀리는 문제 방지."""
    from zoneinfo import ZoneInfo
    return datetime.datetime.now(ZoneInfo("Asia/Seoul")).date()

try:
    from streamlit_quill import st_quill
    HAS_QUILL = True
except Exception:
    HAS_QUILL = False

# ══════════════════════════════════════════════════════════════════════
# 1. 순수 데이터 로직 (Streamlit 비의존)
# ══════════════════════════════════════════════════════════════════════

AF_RE   = re.compile(r'^(AP|PB)\d+$', re.I)          # PUSH AF코드 형식
WEEK_RE = re.compile(r'\d{1,2}\s*월\s*\d\s*주차')    # 기획 주차 시트명

# 실적 '소재별 실적(당주)' 시트 컬럼 → 표준 키
PERF_COLMAP = {
    "날짜": "date", "요일": "dow_k", "시간대": "hour", "타겟 구분": "target",
    "발송유형": "stype", "BPU": "bpu", "우선순위": "prio", "카테고리": "cat",
    "속성": "attr", "담당자": "owner", "브랜드": "brand", "AF코드": "af",
    "기획전": "promo", "발송": "send", "UV": "uv", "VISIT": "visit",
    "고객수": "cust", "주문건수": "oc", "주문금액": "amt",
    "유입전환율": "infl_cr", "주문전환율": "ord_cr", "효율": "eff",
}
NUM_COLS = ["hour", "send", "uv", "visit", "cust", "oc", "amt", "infl_cr", "ord_cr", "eff"]

# 속성(attr) 표기 오타 통합 — 같은 의미인데 오타/표기가 다른 값을 하나로 머지
ATTR_ALIASES = {"마켸팅": "마케팅"}


def _norm_attr(x):
    """속성 표기 정규화: 공백 정리 + 오타 머지(마켸팅 등 '마○팅' 3글자는 마케팅으로)."""
    if pd.isna(x):
        return x
    s = str(x).strip()
    if not s:
        return s
    if s in ATTR_ALIASES:
        return ATTR_ALIASES[s]
    if len(s) == 3 and s[0] == "마" and s[-1] == "팅":
        return "마케팅"
    return s


def _norm_date(v):
    """엑셀 셀 값 → 'YYYYMMDD' 문자열 또는 None.
    8자리(YYYYMMDD)는 그대로, 6자리(YYMMDD, 예: '250111')는 '20'을 붙여 복구한다.
    구분자형(2025-01-11·2025.1.11·2025/1/11)도 구분자를 떼고 표준 8자리로 복구한다.
    (기획 시트에 'YYYY' 앞 두 자리가 빠진 날짜 오타·셀 표시형식 변경으로 매칭에서
     누락되던 문제 대응)"""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y%m%d")
    s = re.sub(r'\.0$', '', str(v).strip())
    if re.match(r'^\d{8}$', s):                        # YYYYMMDD — 6자리와 대칭으로 유효성 검증
        try:
            datetime.datetime.strptime(s, '%Y%m%d')
            return s
        except ValueError:
            return None
    if re.match(r'^\d{6}$', s):                       # YYMMDD → 20YYMMDD (실제 날짜인지 검증)
        cand = '20' + s
        try:
            datetime.datetime.strptime(cand, '%Y%m%d')
            return cand
        except ValueError:
            return None
    # 구분자형(-, ., /, 공백) — YYYY[sep]M[sep]D. 구분자를 떼고 zero-pad 후 검증.
    md = re.match(r'^(\d{4})[.\-/\s]+(\d{1,2})[.\-/\s]+(\d{1,2})$', s)
    if md:
        cand = f"{int(md.group(1)):04d}{int(md.group(2)):02d}{int(md.group(3)):02d}"
        try:
            datetime.datetime.strptime(cand, '%Y%m%d')
            return cand
        except ValueError:
            return None
    return None


def parse_perf_bytes(file_bytes):
    """실적 엑셀 → 캠페인 단위 DataFrame. '소재별 실적(당주)' 시트(또는 AF코드 헤더 시트)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    target_sheet = None
    if "소재별 실적(당주)" in wb.sheetnames:
        target_sheet = "소재별 실적(당주)"
    else:                                   # 헤더에 AF코드+주문금액 있는 첫 시트
        for s in wb.sheetnames:
            ws = wb[s]
            hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            hdr = [str(x).strip() if x is not None else "" for x in hdr]
            if "AF코드" in hdr and ("주문금액" in hdr or "발송" in hdr):
                target_sheet = s
                break
    if target_sheet is None:
        wb.close()
        raise ValueError("실적 시트(소재별 실적/AF코드 헤더)를 찾지 못했습니다.")

    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame()
    hdr = [str(x).strip() if x is not None else "" for x in rows[0]]
    # 헤더를 공백 제거 키로 인덱싱 — 'AF 코드'·'주문 금액'처럼 내부 공백이 섞이면
    # 정확일치 조회('AF코드')가 None이 돼 열 전체(또는 전 행)가 유실되던 문제 방지.
    # 중복 헤더는 첫 등장 우선(setdefault) — dict 컴프리헨션의 '마지막 승리'로 앞 열이 묻히던 문제.
    idx = {}
    for i, h in enumerate(hdr):
        idx.setdefault(h, i)
        _hns = h.replace(" ", "")
        if _hns:
            idx.setdefault(_hns, i)
    # 날짜 컬럼 헤더 변형 대응. PERF_COLMAP은 '날짜' 키로 날짜를 읽으므로, '날짜'가 없으면
    # '일자'/'날짜'를 포함하거나 '발송일'류인 헤더를 '날짜'로 매핑한다.
    #   예) '일자', '일자(8자리)', '발송일', '발송일자' … (← 이게 없으면 날짜가 전부 비어 매칭 0%)
    if "날짜" not in idx:
        for h in hdr:
            hs = h.replace(" ", "")
            if hs and ("일자" in hs or "날짜" in hs or hs in ("발송일", "발송일자")):
                idx["날짜"] = idx[h]
                break
    recs = []
    for r in rows[1:]:
        afi = idx.get("AF코드")
        if afi is None or afi >= len(r) or r[afi] is None:
            continue
        af = str(r[afi]).strip().upper()             # 대문자 정규화 — 실적↔기획 대소문자
        if not AF_RE.match(af):                        # 불일치로 인한 미매칭·저장소 중복 방지
            continue
        rec = {}
        for kcol, key in PERF_COLMAP.items():
            i = idx.get(kcol)
            rec[key] = r[i] if (i is not None and i < len(r)) else None
        rec["af"] = af
        rec["date"] = _norm_date(rec.get("date"))
        recs.append(rec)
    return _finalize(pd.DataFrame(recs))


def _finalize(df):
    """숫자 변환 + 파생지표(rps/aov/dt/dow) 계산. 신규 파싱·저장소 로드 양쪽에서 재사용."""
    if df is None or df.empty:
        return df if df is not None else pd.DataFrame()
    df = df.copy()
    df["date"] = df["date"].astype(str).str.replace(r'\.0$', '', regex=True)
    # 속성(attr) 오타·표기흔들림 머지 (예: 마켸팅→마케팅). 공백도 정리.
    if "attr" in df:
        df["attr"] = df["attr"].map(_norm_attr)
    # promo(기획전번호)는 실적 파싱에서 int(100328)와 str('장바구니')가 섞여 들어온다 —
    # 문자열로 통일하지 않으면 st.dataframe의 Arrow 직렬화가 매번 실패(자동 복구로 성능
    # 저하+로그 오염)한다. norm_promo로 기획전 매칭 키와도 같은 형식으로 맞춘다.
    if "promo" in df:
        df["promo"] = df["promo"].map(norm_promo)
    for c in NUM_COLS:
        if c in df:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    # 분모(발송/주문)가 0·결측이면 0이 아니라 NaN — 0으로 넣으면 평균·BOTTOM 순위가
    # '가짜 0원 캠페인'에 체계적으로 끌려 내려간다 (표시부는 NaN을 '–'로 처리)
    df["rps"] = np.where(df["send"].fillna(0) > 0, df["amt"] / df["send"], np.nan)  # 발송건당 거래액
    df["aov"] = np.where(df["oc"].fillna(0) > 0, df["amt"] / df["oc"], np.nan)      # 객단가
    df["dt"] = pd.to_datetime(df["date"], format="%Y%m%d", errors="coerce")
    df["dow"] = df["dt"].dt.dayofweek
    return df


def _parse_plan_sheet(rows, lookup):
    """한 주차 시트의 rows(2차원 리스트/튜플) → lookup 갱신.

    헤더가 2줄로 쪼개져 있어도(예: 45행=AF코드, 46행=제목/내용) 컬럼 매핑을
    행마다 '누적' 갱신하며 AF코드 행을 파싱한다. 엑셀/구글시트 공용.
    """
    c_af = c_date = c_title = c_body = None
    _last_date = None                                # 병합셀 forward-fill용 직전 유효 날짜
    for row in rows:
        cells = [("" if v is None else str(v)) for v in row[:45]]
        for j, c in enumerate(cells):
            cs = c.strip()
            if cs == "AF코드":
                c_af = j
            elif cs == "제목":
                c_title = j
            elif cs == "내용":
                c_body = j
            elif (("일자" in cs or cs == "날짜") and len(cs) < 16):
                c_date = j
        if c_af is None or c_af >= len(cells):
            continue
        af = cells[c_af].strip().upper()             # 대문자 정규화 (실적 파서와 대칭)
        if not AF_RE.match(af):
            continue
        d = _norm_date(row[c_date]) if (c_date is not None and c_date < len(row)) else None
        if d is None:                                # 날짜열이 비면 보통 B열(index 1)에 있음
            d = _norm_date(row[1]) if len(row) > 1 else None
        if d is None:
            # 병합셀(첫 행에만 날짜, 이하 None) → 직전 유효 날짜를 이어받는다.
            # 이게 없으면 (None, af) 키로 저장돼 실적의 실제 날짜와 절대 매칭 안 됨(문구 유실).
            d = _last_date
        else:
            _last_date = d
        title = cells[c_title].strip() if (c_title is not None and c_title < len(cells)) else ""
        body = cells[c_body].strip() if (c_body is not None and c_body < len(cells)) else ""
        if title or body:
            if d is None:
                continue                             # 끝내 날짜를 못 구하면 오염 대신 스킵
            lookup[(d, af)] = (title, body)


def parse_plan_bytes(file_bytes):
    """기획 엑셀(통합본) → {(date, af): (title, body)} 사전.

    165개 주차 시트를 순회하며, 'AF코드' 가 들어간 헤더 행을 만날 때마다
    컬럼 매핑(일자/AF코드/제목/내용)을 갱신해 하위 표가 여러 개여도 견고하게 파싱한다.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    lookup = {}
    for sname in wb.sheetnames:
        if not WEEK_RE.search(sname):
            continue
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True, max_col=45))
        _parse_plan_sheet(rows, lookup)
    wb.close()
    return lookup


def _week_sheet_end_date(title):
    """시트명에서 종료 날짜를 추출한다.
    슬래시형 '(6/15-6/21)' 과 무슬래시형 '(615~621)'·'(1229~14)' 둘 다 지원한다.
    무슬래시형은 구분자가 없어 'N월' 접두어를 힌트로 끝 토큰을 월/일로 분해한다.
    시트명엔 연도가 없으므로 '미래가 아닌 가장 가까운 과거 연도'를 택한다(최대 2년 전까지)."""
    from datetime import date, timedelta
    today = today_kst()

    def _pick_year(em, ed):
        for yr in (today.year, today.year - 1, today.year - 2):
            try:
                d = date(yr, em, ed)
            except ValueError:
                continue
            if d <= today + timedelta(days=14):
                return d
        return None

    # 1) 슬래시형: (m/d - m/d) → 끝 날짜 사용
    m = re.search(r'\((\d{1,2})/(\d{1,2})\s*[-~]\s*(\d{1,2})/(\d{1,2})\)', title)
    if m:
        return _pick_year(int(m.group(3)), int(m.group(4)))

    # 2) 무슬래시형: (start~end). end 토큰(2~4자리)을 월/일로 분해.
    #    접두어 'N월'을 힌트로, 끝 월이 접두월 또는 다음달(주 경계 spill)인 분해를 우선.
    m = re.search(r'\(\s*\d{1,4}\s*[-~]\s*(\d{1,4})\s*\)', title)
    if m:
        end = m.group(1)
        pm = re.search(r'(\d{1,2})\s*월', title)
        prefix_m = int(pm.group(1)) if pm else None
        cands = []
        for cut in (2, 1):                       # 월 자릿수 2 → 1 순으로 분해 시도
            if len(end) > cut:
                try:
                    em, ed = int(end[:cut]), int(end[cut:])
                except ValueError:
                    continue
                if 1 <= em <= 12 and 1 <= ed <= 31:
                    cands.append((em, ed))
        if cands:
            if prefix_m is not None:
                allowed = {prefix_m, prefix_m % 12 + 1}
                pref = [c for c in cands if c[0] in allowed]
                if pref:
                    cands = pref
            return _pick_year(*cands[0])
    return None


def parse_plan_gsheet(sh, recent=None, progress_cb=None):
    """구글시트(gspread Spreadsheet) → 기획 lookup. 주차(WEEK_RE) 시트만 순회.

    · 시트를 종료 날짜 기준 최신순 정렬하고, 금주까지 포함하되 미래 주차는 제외한다.
      (주간보고의 '금주 집행 내용 요약'이 이번 주 기획 시트를 읽어야 해서 금주는 포함)
    · recent=N 이면 최신 N개만 읽어 API 호출/속도를 통제한다.
    · progress_cb(i, total, title) 콜백으로 진행 상황을 외부에 알린다.
    반환: (lookup, 읽은_시트명_리스트)
    """
    from datetime import date, timedelta
    all_ws = [ws for ws in sh.worksheets() if WEEK_RE.search(ws.title)]
    dated, undated = [], []
    for ws in all_ws:
        d = _week_sheet_end_date(ws.title)
        if d is not None:
            dated.append((d, ws))
        else:
            undated.append(ws)                       # 날짜를 못 읽은 시트도 버리지 않는다
    dated.sort(key=lambda x: x[0], reverse=True)
    _today = today_kst()
    this_week_end = _today + timedelta(days=6 - _today.weekday())
    week_ws = [ws for d, ws in dated if d <= this_week_end]
    if recent and recent > 0:
        week_ws = week_ws[:recent]
    else:
        week_ws = week_ws + undated                  # 전부 불러올 땐 미파싱 시트까지 포함
    lookup = {}
    read = []
    total = len(week_ws)
    for i, ws in enumerate(week_ws):
        if progress_cb:
            progress_cb(i + 1, total, ws.title)
        try:
            rows = ws.get_all_values()
        except Exception:
            continue
        _parse_plan_sheet(rows, lookup)
        read.append(ws.title)
    return lookup, read


def merge_perf_plan(perf_df, plan_lookup, keep_unmatched=False):
    """실적 기준 조인 — 문구(제목/내용) 부착. keep_unmatched=False면 문구 없는 행 제외."""
    if perf_df.empty:
        return perf_df.copy()
    df = perf_df.copy()
    titles, bodies, matched = [], [], []
    for key in zip(df["date"].tolist(), df["af"].tolist()):
        tb = plan_lookup.get(key)
        if tb:
            titles.append(tb[0]); bodies.append(tb[1]); matched.append(True)
        else:
            titles.append(""); bodies.append(""); matched.append(False)
    df["title"] = titles
    df["body"] = bodies
    df["matched"] = matched
    if not keep_unmatched:
        df = df[df["matched"]].reset_index(drop=True)
    return df


# ── 누적 저장소 — 업로드할 때마다 (발송일+AF코드) 키로 병합·영속 ──────────
DATA_STORE = "send_perf_store.csv"
STORE_COLS = list(PERF_COLMAP.values()) + ["title", "body", "matched"]


def _to_bool(v):
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return (not (isinstance(v, float) and np.isnan(v))) and bool(v)
    return str(v).strip().lower() in ("true", "1", "1.0", "yes", "y")


def load_store():
    """누적 CSV → DataFrame. 없으면 빈 프레임."""
    if os.path.exists(DATA_STORE):
        try:
            d = pd.read_csv(DATA_STORE, encoding="utf-8-sig", dtype={"date": str, "af": str})
            return d
        except Exception:
            pass
    return pd.DataFrame(columns=STORE_COLS)


def save_store(df):
    df[[c for c in STORE_COLS if c in df]].to_csv(DATA_STORE, index=False, encoding="utf-8-sig")


def expand_uploads(files):
    """업로드 목록 → [(표시이름, xlsx바이트)]. zip 은 내부 .xlsx 를 모두 풀어낸다(연도별 폴더 OK)."""
    import zipfile
    out = []
    for f in files:
        data = f.getvalue()
        nm = f.name
        if nm.lower().endswith(".zip"):
            try:
                zf = zipfile.ZipFile(io.BytesIO(data))
                for zi in zf.infolist():
                    if zi.is_dir():
                        continue
                    inner = zi.filename
                    base = inner.split("/")[-1]
                    if "__MACOSX" in inner or base.startswith(".") or not base.lower().endswith(".xlsx"):
                        continue
                    try:                                      # 한글 파일명 복원(cp437→cp949)
                        disp = base.encode("cp437").decode("cp949")
                    except Exception:
                        disp = base
                    out.append((disp, zf.read(zi)))
            except Exception as e:
                out.append((f"{nm} (zip 오류: {e})", None))
        else:
            out.append((nm, data))
    return out


class _UF:
    """file_uploader 객체 호환 shim — 통합 업로더에서 분류한 (이름, 바이트)를
    기존 처리 코드(.name / .getvalue())가 그대로 쓰게 한다."""
    def __init__(self, name, data):
        self.name = name
        self._data = data

    def getvalue(self):
        return self._data


def classify_upload(name, file_bytes):
    """업로드 xlsx 한 개의 종류 자동 판별:
    perf(발송실적) / plan(발송기획 통합본) / promo(기획전 성과시트) / mtd(전사 MTD) / push(앱푸시 동의 현황) / unknown."""
    nm_lower = (name or "").lower()
    if any(k in nm_lower for k in ("push", "consent", "앱푸시", "동의현황")):
        return "push"
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return "unknown"
    try:
        names = [str(s) for s in wb.sheetnames]
        if any(k in "".join(names) for k in ("앱푸시", "수신동의", "동의현황")):
            return "push"
        # 1) 발송기획 통합본 — 주차 시트가 여러 개
        if sum(1 for s in names if WEEK_RE.search(s)) >= 3:
            return "plan"
        # 2) 발송실적 — '소재별 실적' 시트
        if any("소재별 실적" in s for s in names):
            return "perf"
        first = wb[wb.sheetnames[0]]
        head = []
        for i, r in enumerate(first.iter_rows(min_row=1, max_row=14, values_only=True)):
            head.append([str(x).strip() if x is not None else "" for x in r])
        
        # 첫 14행 헤더 중 앱푸시 지표 관련 컬럼 키워드 감지
        flat_head = []
        for r in head:
            flat_head.extend(r)
        flat_head = "".join(flat_head)
        if any(k in flat_head for k in ("기존 이탈", "기존이탈", "신규추가")):
            return "push"

        # 3) 기획전 성과시트 — 첫 칸이 '기획전 번호'
        for r in head:
            if r and r[0].replace(" ", "") == "기획전번호":
                return "promo"
        # 4) AF코드 헤더가 있으면 perf/plan 구분
        for s in wb.sheetnames[:4]:
            hdr = next(wb[s].iter_rows(min_row=1, max_row=1, values_only=True), ())
            hdr = [str(x).strip() if x is not None else "" for x in hdr]
            if "AF코드" in hdr:
                if "제목" in hdr or "내용" in hdr:
                    return "plan"
                if "주문금액" in hdr or "발송" in hdr or "주문전환율" in hdr:
                    return "perf"
        # 5) 전사 MTD — 2번째 행에 날짜 다수 또는 1열에 지표명
        r1 = head[1] if len(head) > 1 else []
        date_like = 0
        for v in r1[1:]:
            if v in (None, ""):
                continue
            try:
                pd.Timestamp(v); date_like += 1
            except Exception:
                pass
        col0 = [r[0] for r in head if r]
        mtd_kw = any(any(k in c for k in ("인당발송", "발송건당", "유니크발송", "M당거래", "적립M"))
                     for c in col0)
        if mtd_kw or date_like >= 5:
            return "mtd"
        # 6) 주차 시트가 1~2개라도 있으면 plan
        if any(WEEK_RE.search(s) for s in names):
            return "plan"
        return "unknown"
    finally:
        wb.close()


# 발송 1건 식별 키 — 같은 날 같은 AF라도 시간대/타겟 다르면 별개 발송으로 본다
STORE_KEY_COLS = ["date", "af", "hour", "target"]


def store_key_frame(d):
    """중복 판정용 정규화 키(문자열) DataFrame. gsheets(문자열)·신규(숫자) 혼재에도 일치하게 맞춘다."""
    k = pd.DataFrame(index=d.index)
    for c in STORE_KEY_COLS:
        if c in d.columns:
            s = d[c].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            # 결측을 한 토큰('')으로 통일 — 컬럼 부재('')와 값 결측이 서로 다른 키가 돼
            # 같은 발송이 중복 누적(이중 집계)되던 비대칭 제거. isna()를 명시적으로 OR하는
            # 이유는 astype(str)이 일부 pandas 버전에서 NaN을 'nan' 문자열로 안 바꾸기 때문.
            _blank = d[c].isna() | s.str.lower().isin(["nan", "none", "nat", ""])
            k[c] = np.where(_blank, "", s)
        else:
            k[c] = ""
    return k


def merge_store(old, new):
    """기존 누적 + 신규 업로드 병합 — 같은 발송키(STORE_KEY_COLS)는 신규 우선.

    단, 신규가 문구 미매칭(matched=False)이고 기존이 이미 매칭돼 있으면 기존 행을 유지한다.
    (기획 lookup을 최근 N주만 적재한 상태에서 옛 실적 파일을 재업로드하면, 이미 매칭돼
     저장된 문구가 빈 값으로 덮어써져 영구 유실되던 문제 방지)"""
    def _pick(d):
        if d is None or len(d) == 0:
            return pd.DataFrame(columns=STORE_COLS)
        return d[[c for c in STORE_COLS if c in d]].copy()
    o, n = _pick(old), _pick(new)
    if len(o) and len(n) and "matched" in o.columns and "matched" in n.columns:
        o_m = o["matched"].map(_to_bool)
        matched_old_keys = set(map(tuple, store_key_frame(o).loc[o_m.values].values))
        if matched_old_keys:
            n_m = n["matched"].map(_to_bool)
            n_keys = list(map(tuple, store_key_frame(n).values))
            drop = [(k in matched_old_keys) and (not m) for k, m in zip(n_keys, n_m.values)]
            if any(drop):
                n = n.loc[[not x for x in drop]]
    both = pd.concat([o, n], ignore_index=True)
    if both.empty:
        return both
    both["date"] = both["date"].astype(str).str.replace(r'\.0$', '', regex=True)
    keep = ~store_key_frame(both).duplicated(keep="last")
    return both.loc[keep].reset_index(drop=True)


# ── 전사 MTD 발송상세 (인당발송 피로도·CTR) — send_dashboard 분석 이식 ──────
MTD_METRICS = {
    "perSend": ["인당발송건수"], "revenue": ["거래액"], "rps": ["발송건당거래액"],
    "totalSend": ["총발송건수"], "customers": ["유니크발송고객수"], "ctr": ["CTR"],
    "uniqueInflow": ["유니크유입"], "totalInflow": ["총유입"], "visitPerPerson": ["인당방문횟수"],
    "purchaseCust": ["구매고객수"], "purchaseCnt": ["구매건수"], "purchasePerPerson": ["인당구매건수"],
    "avgOrderVal": ["객단가"], "unitPrice": ["건단가"], "mRevenue": ["M당거래액"], "pointM": ["적립M"],
}
MTD_LABELS = {
    "perSend": "인당 발송 건수", "revenue": "거래액", "rps": "발송건당 거래액",
    "totalSend": "총 발송 건수", "customers": "유니크 발송 고객수", "ctr": "CTR",
    "uniqueInflow": "유니크 유입", "totalInflow": "총 유입", "visitPerPerson": "인당 방문 횟수",
    "purchaseCust": "구매 고객수", "purchaseCnt": "구매 건수", "purchasePerPerson": "인당 구매 건수",
    "avgOrderVal": "객단가", "unitPrice": "건단가", "mRevenue": "M당 거래액", "pointM": "적립M",
    "purchaseRate": "구매전환율(CR)", "rpc": "고객당 매출",
}
MTD_DERIVED = ["purchaseRate", "rpc"]
MTD_STORE = "send_perf_mtd_store.csv"
MTD_STORE_COLS = ["date"] + list(MTD_METRICS)


def _linreg(x, y):
    import scipy.stats as ss
    x = np.asarray(x, float); y = np.asarray(y, float)
    m = ~np.isnan(x) & ~np.isnan(y)
    if m.sum() < 5:
        return dict(slope=np.nan, intercept=np.nan, r2=np.nan, p=np.nan)
    sl, ic, r, p, _ = ss.linregress(x[m], y[m])
    return dict(slope=sl, intercept=ic, r2=r ** 2, p=p)


def _dow_residual(df, col):
    """요일 평균을 뺀 잔차 — 요일 효과 통제용."""
    vals = df[col].values.astype(float); res = vals.copy()
    for d in range(7):
        idx = df["dow"].values == d
        if idx.sum() > 0:
            res[idx] -= np.nanmean(res[idx])
    return res


def parse_mtd_bytes(file_bytes):
    """전사 MTD 발송상세(날짜=열, 지표=행) → 일자별 DataFrame. 시작 열 자동 탐지."""
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    ws = pd.read_excel(xls, header=None, sheet_name=xls.sheet_names[0])
    # 날짜행(2행)·값행이 없는 시트(오분류된 1~2행짜리 파일)는 iloc[1] IndexError 대신 빈 반환
    if ws.shape[0] < 3 or ws.shape[1] < 2:
        return pd.DataFrame()
    date_row = ws.iloc[1, :]
    start = None
    for j in range(1, ws.shape[1]):
        v = date_row.iloc[j]
        if pd.isnull(v):
            continue
        try:
            (pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(v))) if isinstance(v, (int, float)) else pd.Timestamp(v)
            start = j; break
        except Exception:
            continue
    if start is None:
        return pd.DataFrame()
    dates = []
    for v in ws.iloc[1, start:]:
        if pd.isnull(v):
            dates.append(pd.NaT); continue
        try:
            d = (pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(v))) if isinstance(v, (int, float)) else pd.Timestamp(v)
            dates.append(d)
        except Exception:
            dates.append(pd.NaT)
    metric_col = ws.iloc[3:, 0].astype(str).str.strip()

    def get_m(keys):
        for kw in keys:
            match = metric_col[metric_col.str.replace(" ", "") == kw.replace(" ", "")]
            if not match.empty:
                vals = ws.iloc[match.index[0], start:start + len(dates)].values
                return pd.to_numeric(vals, errors="coerce")
        return np.full(len(dates), np.nan)

    data = {k: get_m(v) for k, v in MTD_METRICS.items()}
    data["date"] = dates
    df = pd.DataFrame(data).dropna(subset=["perSend", "revenue"]).copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.dropna(subset=["date"]).sort_values("date").reset_index(drop=True)
    return df


def load_mtd_store():
    if os.path.exists(MTD_STORE):
        try:
            return pd.read_csv(MTD_STORE, encoding="utf-8-sig")
        except Exception:
            pass
    return pd.DataFrame(columns=MTD_STORE_COLS)


def save_mtd_store(df):
    out = df[[c for c in MTD_STORE_COLS if c in df]].copy()
    out["date"] = pd.to_datetime(out["date"], errors="coerce").dt.strftime("%Y-%m-%d")
    out.to_csv(MTD_STORE, index=False, encoding="utf-8-sig")


def merge_mtd_store(old, new):
    """일자 기준 병합 — 같은 날짜는 신규 우선."""
    def _p(d):
        if d is None or len(d) == 0:
            return pd.DataFrame(columns=MTD_STORE_COLS)
        d = d[[c for c in MTD_STORE_COLS if c in d]].copy()
        d["date"] = pd.to_datetime(d["date"], errors="coerce").dt.strftime("%Y-%m-%d")
        return d
    both = pd.concat([_p(old), _p(new)], ignore_index=True).dropna(subset=["date"])
    if both.empty:
        return both
    return both.drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)


# ── 기획전 성과시트 (기획전번호별 유입/총매출) — 발송 promo 와 조인 ──────────
PROMO_STORE = "send_perf_promo_store.csv"
PROMO_STORE_COLS = ["promo", "pname", "pstart", "pend", "p_pv", "p_uv",
                    "inf_amt", "inf_sales", "inf_oc", "inf_ocust", "inf_qty",
                    "tot_amt", "tot_sales", "tot_oc", "tot_ocust", "tot_qty"]
PROMO_NUM_COLS = ["p_pv", "p_uv", "inf_amt", "inf_sales", "inf_oc", "inf_ocust", "inf_qty",
                  "tot_amt", "tot_sales", "tot_oc", "tot_ocust", "tot_qty"]


def norm_promo(v):
    """기획전번호 정규화 — 발송 promo·기획전시트 번호를 같은 문자열 키로 맞춘다."""
    if v is None:
        return ""
    s = str(v).strip()
    s = re.sub(r"\.0$", "", s)
    if s in ("", "nan", "NaN", "None", "0"):
        return ""
    return s


def _promo_num(v):
    if v is None:
        return None
    s = str(v).replace(",", "").strip()
    if s in ("", "-", "nan", "None"):
        return None
    try:
        return float(s)
    except Exception:
        return None


def parse_promo_bytes(file_bytes):
    """기획전 성과시트 → 기획전 단위 DataFrame. 2단 헤더(공통 / 유입 / 총매출).
    '거래액'·'매출액' 등이 유입·총매출 블록에 중복되므로 그룹 헤더(유입/총매출) 위치로 구분 파싱."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame(columns=PROMO_STORE_COLS)

    # 헤더(컬럼명) 행 탐색: 첫 칸이 '기획전 번호'
    hr = None
    for i, r in enumerate(rows[:12]):
        c0 = str(r[0]).strip() if (r and r[0] is not None) else ""
        if c0.replace(" ", "") == "기획전번호":
            hr = i
            break
    if hr is None:
        raise ValueError("기획전 성과시트 헤더('기획전 번호')를 찾지 못했습니다.")

    grp = rows[hr - 1] if hr > 0 else ()
    names = [str(x).strip() if x is not None else "" for x in rows[hr]]
    ncols = len(names)
    gi = ti = None
    for j, gv in enumerate(grp):
        gs = str(gv).strip() if gv is not None else ""
        if gs == "유입":
            gi = j
        elif gs.replace(" ", "") == "총매출":
            ti = j
    g_end = gi if gi is not None else ncols
    i_lo, i_hi = (gi, ti) if gi is not None else (ncols, ncols)
    t_lo, t_hi = (ti, ncols) if ti is not None else (ncols, ncols)

    def find(nm, lo, hi):
        key = nm.replace(" ", "")
        for j in range(lo, min(hi, ncols)):
            if names[j].replace(" ", "") == key:
                return j
        return None

    idx = {
        "promo": 0, "pname": find("기획전명", 0, g_end),
        "pstart": find("기획전 시작일", 0, g_end), "pend": find("기획전 종료일", 0, g_end),
        "p_pv": find("기획전 PV", 0, g_end), "p_uv": find("기획전 UV", 0, g_end),
        "inf_amt": find("거래액", i_lo, i_hi), "inf_sales": find("매출액", i_lo, i_hi),
        "inf_oc": find("주문건수", i_lo, i_hi), "inf_ocust": find("주문고객수", i_lo, i_hi),
        "inf_qty": find("총결제 수량", i_lo, i_hi),
        "tot_amt": find("거래액", t_lo, t_hi), "tot_sales": find("매출액", t_lo, t_hi),
        "tot_oc": find("주문건수", t_lo, t_hi), "tot_ocust": find("주문고객수", t_lo, t_hi),
        "tot_qty": find("총결제 수량", t_lo, t_hi),
    }

    def cell(r, key):
        j = idx.get(key)
        return r[j] if (j is not None and j < len(r)) else None

    recs = []
    for r in rows[hr + 1:]:
        if not r or r[0] is None:
            continue
        promo = norm_promo(r[0])
        if not promo:
            continue
        rec = {"promo": promo}
        nm = cell(r, "pname")
        rec["pname"] = str(nm).strip() if nm is not None else ""
        for k in ("pstart", "pend"):
            v = cell(r, k)
            rec[k] = str(v)[:10] if v is not None else ""
        for k in PROMO_NUM_COLS:
            rec[k] = _promo_num(cell(r, k))
        recs.append(rec)
    df = pd.DataFrame(recs, columns=PROMO_STORE_COLS)
    if len(df):
        df = df.drop_duplicates(subset=["promo"], keep="last").reset_index(drop=True)
    return df


def load_promo_store():
    if os.path.exists(PROMO_STORE):
        try:
            return pd.read_csv(PROMO_STORE, encoding="utf-8-sig", dtype={"promo": str})
        except Exception:
            pass
    return pd.DataFrame(columns=PROMO_STORE_COLS)


def save_promo_store(df):
    df[[c for c in PROMO_STORE_COLS if c in df]].to_csv(PROMO_STORE, index=False, encoding="utf-8-sig")


def merge_promo_store(old, new):
    """기획전번호 기준 병합 — 같은 번호는 신규 우선."""
    def _pick(d):
        if d is None or len(d) == 0:
            return pd.DataFrame(columns=PROMO_STORE_COLS)
        return d[[c for c in PROMO_STORE_COLS if c in d]].copy()
    both = pd.concat([_pick(old), _pick(new)], ignore_index=True)
    if both.empty:
        return both
    both["promo"] = both["promo"].map(norm_promo)
    both = both[both["promo"] != ""]
    return both.drop_duplicates(subset=["promo"], keep="last").reset_index(drop=True)


# ── 앱푸시 수신동의 현황 로컬 저장소 ──────────────────────────────────────
PUSH_STORE = "send_perf_push_store.csv"
PUSH_STORE_COLS = ["date", "group", "consent", "added", "removed", "diff", "is_outlier"]


def load_push_store():
    if os.path.exists(PUSH_STORE):
        try:
            df = pd.read_csv(PUSH_STORE, encoding="utf-8-sig", dtype={"group": str})
            if "date" in df.columns:
                df["date"] = pd.to_datetime(df["date"], errors="coerce")
            return df
        except Exception:
            pass
    return pd.DataFrame(columns=PUSH_STORE_COLS)


def save_push_store(df):
    if df is None:
        return
    out = df[[c for c in PUSH_STORE_COLS if c in df]].copy()
    if "date" in out.columns:
        out["date"] = pd.to_datetime(out["date"], errors="coerce").dt.strftime("%Y-%m-%d")
    out.to_csv(PUSH_STORE, index=False, encoding="utf-8-sig")


def finalize_push(df):
    """push 저장소 로드 공통 — 표준 dtype 복구 (campaign의 _finalize와 같은 역할).

    구글시트 라운드트립은 모든 값을 문자열로 되돌리는데 push만 복구 함수가 없어서,
    저장 후 새 세션에서 주간보고·앱푸시 페이지가 date 비교/`~is_outlier`/int(consent)
    에서 크래시하던 문제 방지. 어떤 소스(gsheets/CSV/파싱 직후)든 통과시켜도 안전하다."""
    if df is None or len(df) == 0:
        return pd.DataFrame(columns=PUSH_STORE_COLS)
    d = df.copy()
    if "date" in d:
        d["date"] = pd.to_datetime(d["date"], errors="coerce")
    for c in ("consent", "diff", "added", "removed"):
        if c in d:
            d[c] = pd.to_numeric(d[c], errors="coerce")
    if "group" in d:
        d["group"] = d["group"].astype(str).str.strip()
    d["is_outlier"] = d["is_outlier"].map(_to_bool) if "is_outlier" in d else False
    d = d.dropna(subset=["date"]) if "date" in d else d
    return d.reset_index(drop=True)


# ── 주간보고 보고란(전주 지표 현황·금주 집행) 영속 저장 — 주차별 key/text ──
NOTES_STORE = "send_perf_notes.csv"
NOTES_STORE_COLS = ["key", "text"]


def load_notes_store():
    if os.path.exists(NOTES_STORE):
        try:
            return pd.read_csv(NOTES_STORE, encoding="utf-8-sig", dtype=str).fillna("")
        except Exception:
            pass
    return pd.DataFrame(columns=NOTES_STORE_COLS)


def save_notes_store(df):
    if df is None:
        return
    df[[c for c in NOTES_STORE_COLS if c in df]].to_csv(
        NOTES_STORE, index=False, encoding="utf-8-sig")


def notes_df_to_dict(df):
    """저장소 DataFrame([key,text]) → {key: text} 사전."""
    if df is None or len(df) == 0 or "key" not in df:
        return {}
    d = {}
    for _, r in df.iterrows():
        k = str(r.get("key", "")).strip()
        if k and k.lower() != "nan":
            d[k] = "" if pd.isna(r.get("text")) else str(r.get("text"))
    return d


def notes_dict_to_df(d):
    """{key: text} → 저장소 DataFrame([key,text])."""
    if not d:
        return pd.DataFrame(columns=NOTES_STORE_COLS)
    return pd.DataFrame([{"key": k, "text": v} for k, v in d.items()],
                        columns=NOTES_STORE_COLS)


def finalize_promo(df):
    """저장소/파싱 로드 공통 — 키 정규화 + 숫자 컬럼 형변환."""
    if df is None or len(df) == 0:
        return pd.DataFrame(columns=PROMO_STORE_COLS)
    d = df.copy()
    if "promo" in d:
        d["promo"] = d["promo"].map(norm_promo)
        d = d[d["promo"] != ""]
    for c in PROMO_NUM_COLS:
        if c in d:
            d[c] = pd.to_numeric(d[c], errors="coerce")
    return d.reset_index(drop=True)


def compute_mtd(df):
    """일자별 MTD → 피로도/효율 집계 (수신동의 제외)."""
    df = df.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.dropna(subset=["date"]).sort_values("date").reset_index(drop=True)
    for c in list(MTD_METRICS):
        df[c] = pd.to_numeric(df[c], errors="coerce")
    _cust = df["customers"].replace(0, np.nan)
    df["purchaseRate"] = (df["purchaseCust"] / _cust).clip(0, 1)
    df["rpc"] = df["revenue"] / _cust
    df["t"] = np.arange(len(df))
    df["dow"] = df["date"].dt.dayofweek
    df["month"] = df["date"].dt.to_period("M").astype(str)
    df["quarter"] = df["date"].dt.to_period("Q").astype(str)
    df["year"] = df["date"].dt.year
    allk = list(MTD_METRICS) + MTD_DERIVED
    agg = {k: pd.NamedAgg(k, "mean") for k in allk}
    # 거래액·유입은 '기간 합계'(총거래액·총유입)도 함께 — 차트 선지표 옵션에서 사용
    _sum_keys = [k for k in ("revenue", "totalInflow", "uniqueInflow", "totalSend") if k in df]
    agg_sum = {f"{k}_sum": pd.NamedAgg(k, "sum") for k in _sum_keys}
    monthly = df.groupby("month", sort=True).agg(n=("revenue", "count"), **agg, **agg_sum).reset_index()
    quarterly = df.groupby("quarter", sort=True).agg(n=("revenue", "count"), **agg, **agg_sum).reset_index()

    BINS = [0, 2.0, 2.5, 3.0, 3.5, 4.0, 4.5, 99]
    LBLS = ["~2.0건", "2.0~2.5", "2.5~3.0", "3.0~3.5", "3.5~4.0", "4.0~4.5", "4.5건+"]
    df["bucket"] = pd.cut(df["perSend"], bins=BINS, labels=LBLS)
    buckets = df.groupby("bucket", observed=True).agg(n=("revenue", "count"), **agg).reset_index()
    buckets = buckets[buckets["n"] >= 30].reset_index(drop=True)

    df_s = df.sort_values("totalSend").reset_index(drop=True)
    qidx = np.array_split(np.arange(len(df_s)), 5) if len(df_s) >= 5 else []
    quintile = pd.DataFrame([
        dict(label=["Q1 최소", "Q2", "Q3", "Q4", "Q5 최대"][i], n=len(qidx[i]),
             **{k: df_s.iloc[qidx[i]][k].mean() for k in allk})
        for i in range(len(qidx))]) if qidx else pd.DataFrame()

    DOW = {0: "월", 1: "화", 2: "수", 3: "목", 4: "금", 5: "토", 6: "일"}
    dow_mean = df.groupby("dow").agg(**agg).reset_index()
    dow_mean["요일"] = dow_mean["dow"].map(DOW)
    dow_comp = []
    for d in [0, 1, 2, 3, 4]:
        sub = df[df["dow"] == d]
        if len(sub) < 20:
            continue
        med = sub["perSend"].median()
        lo, hi = sub[sub["perSend"] <= med], sub[sub["perSend"] > med]
        dow_comp.append(dict(요일=DOW[d], lowRps=lo["rps"].mean(), highRps=hi["rps"].mean(),
                             lowCtr=lo["ctr"].mean(), highCtr=hi["ctr"].mean()))
    t = df["t"].values.astype(float)
    reg = {k: _linreg(t, _dow_residual(df, k) if k != "perSend" else df[k].values)
           for k in ["perSend", "ctr", "purchaseRate", "rps", "revenue"]}
    meta = dict(start=str(df["date"].min().date()), end=str(df["date"].max().date()), days=len(df))
    return dict(df=df, monthly=monthly, quarterly=quarterly, buckets=buckets,
                quintile=quintile, dow_mean=dow_mean, dow_comp=dow_comp, reg=reg, meta=meta)


# ══════════════════════════════════════════════════════════════════════
# 앱푸시 동의 현황 파싱 (PUSH (1).xlsx 형태)
# 구조: 행1=Date헤더, 행2=날짜(m/d), 행3~6=기존(수신동의/증감/신규추가/기존이탈),
#        행7~10=신규(수신동의/증감/신규추가/기존이탈), 행11~14=Total(수신동의/증감/신규추가/기존이탈)
# ══════════════════════════════════════════════════════════════════════

PUSH_OUTLIER_THRESHOLD = 10000   # 신규추가 또는 기존이탈 > 이 값이면 배치/이관 이벤트로 판단, 제외


def parse_push_consent_bytes(file_bytes, start_year=None):
    """앱푸시 수신동의 현황 Excel → 일별 DataFrame.

    · 열 구조: A=그룹(기존/신규/Total), B=지표(수신동의/증감/신규추가/기존이탈), C열부터=날짜별 값
    · 날짜는 'm/d' 형식이며 연도는 start_year 기준 순서대로 할당(연말 넘어가면 +1년)
    · start_year=None 이면 자동 추론 — '마지막 날짜가 미래로 가지 않는 가장 최근 연도'를
      택한다 (예전 2024 하드코딩은 새해 시작 파일을 전부 과거 연도로 오배정했음)
    · 이상치(신규추가>PUSH_OUTLIER_THRESHOLD 또는 기존이탈>PUSH_OUTLIER_THRESHOLD)는 is_outlier=True 플래그
    반환 컬럼: date(datetime), group(기존/신규/Total), consent(수신동의수), diff(증감),
               added(신규추가), removed(기존이탈), is_outlier(bool)
    """
    import openpyxl, datetime as _dt
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 14:
        return pd.DataFrame()

    # 날짜 행 (행2, index 1) — col C(index 2)부터
    date_strs = rows[1][2:]
    n = len(date_strs)

    # 행 인덱스 → (group, metric)
    ROW_MAP = {
        2:  ("기존",  "consent"),
        3:  ("기존",  "diff"),
        4:  ("기존",  "added"),
        5:  ("기존",  "removed"),
        6:  ("신규",  "consent"),
        7:  ("신규",  "diff"),
        8:  ("신규",  "added"),
        9:  ("신규",  "removed"),
        10: ("Total", "consent"),
        11: ("Total", "diff"),
        12: ("Total", "added"),
        13: ("Total", "removed"),
    }

    # 그룹별 값 배열 — 반드시 길이 n으로 패딩. 뒤쪽 빈칸이 trim된 짧은(ragged) 지표행이면
    # 슬라이스가 n보다 짧아지고, 아래 레코드 조합의 [i] 접근이 IndexError로 파싱 전체 실패.
    data = {(g, m): [None] * n for (g, m) in ROW_MAP.values()}
    for ri, (g, m) in ROW_MAP.items():
        if ri < len(rows):
            vals = (list(rows[ri][2:2 + n]) + [None] * n)[:n]
            data[(g, m)] = [v if isinstance(v, (int, float)) else None for v in vals]

    # 날짜 파싱 — m/d 형식, start_year 기준 순서 할당 (월이 줄면 연도 +1 롤오버)
    md_pairs = []
    for ds in date_strs:
        if ds is None:
            md_pairs.append(None)
            continue
        try:
            parts = str(ds).strip().split("/")
            md_pairs.append((int(parts[0]), int(parts[1])))
        except Exception:
            md_pairs.append(None)

    def _assign_years(start_y):
        cur_year, prev_month, out = start_y, None, []
        for md in md_pairs:
            if md is None:
                out.append(None)
                continue
            m_val, d_val = md
            if prev_month is not None and m_val < prev_month:
                cur_year += 1
            prev_month = m_val
            try:
                out.append(_dt.date(cur_year, m_val, d_val))
            except ValueError:
                out.append(None)
        return out

    if start_year is None:
        # 연도 추론: 롤오버까지 시뮬레이션했을 때 마지막 날짜가 미래(+7일 여유)로
        # 가지 않는 가장 최근 시작 연도를 택한다. 다년치 파일도 올바르게 잡힌다.
        _today = today_kst()
        start_year = _today.year - 3
        for cand in range(_today.year, _today.year - 4, -1):
            _dates = [d for d in _assign_years(cand) if d is not None]
            if _dates and max(_dates) <= _today + _dt.timedelta(days=7):
                start_year = cand
                break
    parsed_dates = _assign_years(start_year)

    # 레코드 조합
    records = []
    for i, dt in enumerate(parsed_dates):
        if dt is None:
            continue
        for g in ("기존", "신규", "Total"):
            rec = {
                "date":    pd.Timestamp(dt),
                "group":   g,
                "consent": data.get((g, "consent"), [None] * n)[i],
                "diff":    data.get((g, "diff"),    [None] * n)[i],
                "added":   data.get((g, "added"),   [None] * n)[i],
                "removed": data.get((g, "removed"), [None] * n)[i],
            }
            # 이상치 플래그: 배치 이관 이벤트 (신규추가·기존이탈·순증감 급락 > 임계값)
            added_v  = rec["added"]  if rec["added"]  is not None else 0
            removed_v = rec["removed"] if rec["removed"] is not None else 0
            diff_v   = rec["diff"]   if rec["diff"]   is not None else 0
            rec["is_outlier"] = bool(abs(added_v) > PUSH_OUTLIER_THRESHOLD or
                                      abs(removed_v) > PUSH_OUTLIER_THRESHOLD or
                                      abs(diff_v) > PUSH_OUTLIER_THRESHOLD)
            records.append(rec)

    df = pd.DataFrame(records)
    if df.empty:
        return df
    for c in ("consent", "diff", "added", "removed"):
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df["date"] = pd.to_datetime(df["date"])
    return df.sort_values(["date", "group"]).reset_index(drop=True)


def push_weekly(df, group="Total"):
    """일별 동의 DataFrame → 주간 집계 (주 시작=월요일).

    이상치 제외 후 주간 평균 동의수, 주간 순증감, 주간 신규추가합, 주간 탈퇴합 반환.
    """
    sub = df[(df["group"] == group) & (~df["is_outlier"])].copy()
    if sub.empty:
        return pd.DataFrame()
    sub["week"] = sub["date"].dt.to_period("W-SUN").apply(lambda p: p.start_time)
    agg = sub.groupby("week").agg(
        일수=("date", "count"),
        평균동의수=("consent", "mean"),
        기말동의수=("consent", "last"),
        순증감합=("diff", "sum"),
        신규추가합=("added", "sum"),
        탈퇴합=("removed", "sum"),
    ).reset_index()
    agg.rename(columns={"week": "주시작"}, inplace=True)
    agg["주차"] = agg["주시작"].apply(
        lambda d: f"{d.year}년 {d.isocalendar()[1]}주차 ({d.strftime('%m/%d')}~{(d + pd.Timedelta(days=6)).strftime('%m/%d')})"
    )
    return agg


# ── 구글시트 영속 저장 (선택) — 미설정 시 로컬 CSV 폴백 ──────────────────
GS_TITLES = {"campaign": "campaign_store", "mtd": "mtd_store", "promo": "promo_store",
             "push": "push_store", "notes": "notes_store"}


def _fix_pem(pk):
    """Secrets에서 깨진 private_key 복구 — '\\n' 글자·공백·무구분 등 어떤 형태든
    base64 본문을 추출해 64자씩 표준 PEM으로 재구성한다 ('Unable to load PEM file' 방지)."""
    if not isinstance(pk, str):
        return pk
    s = pk.strip().strip('"').strip("'")
    s = s.replace("\\r\\n", "\n").replace("\\n", "\n").replace("\r\n", "\n").replace("\r", "\n")
    mt = re.search(r"-----BEGIN ([A-Z0-9 ]+?)-----(.*?)-----END \1-----", s, re.S)
    if mt:
        header, body = mt.group(1).strip(), re.sub(r"\s+", "", mt.group(2))
    else:
        header = "PRIVATE KEY"
        body = re.sub(r"\s+", "", s)
    if not body:
        return s + "\n"
    wrapped = "\n".join(body[i:i + 64] for i in range(0, len(body), 64))
    return f"-----BEGIN {header}-----\n{wrapped}\n-----END {header}-----\n"


def _pem_diag(pk):
    """private_key 진단 문자열 — 어디가 잘못됐는지 알려준다(키 값은 노출 안 함)."""
    if not isinstance(pk, str) or not pk.strip():
        return "키가 비어있음(미입력)"
    has_b = "-----BEGIN" in pk
    has_e = "-----END" in pk
    body = re.sub(r"\s+", "", re.sub(r"-----[A-Z0-9 ]+-----", "", pk))
    n = len(body)
    tip = "정상 길이" if n >= 1500 else "너무 짧음 → 잘린 듯"
    return f"BEGIN:{'있음' if has_b else '없음'} END:{'있음' if has_e else '없음'} 본문:{n}자({tip})"


def gs_open(creds_dict, spreadsheet):
    """서비스 계정 자격으로 스프레드시트 열기 (URL/키/제목 모두 허용)."""
    import gspread
    from google.oauth2.service_account import Credentials
    scopes = ["https://www.googleapis.com/auth/spreadsheets",
              "https://www.googleapis.com/auth/drive"]
    info = dict(creds_dict)
    info["private_key"] = _fix_pem(info.get("private_key"))
    try:
        creds = Credentials.from_service_account_info(info, scopes=scopes)
    except Exception as e:
        raise ValueError(f"서비스계정 private_key 문제 — {_pem_diag(info.get('private_key'))}. "
                         "값이 잘렸거나 마커가 빠졌을 수 있어요. 새 JSON의 private_key를 통째로 다시 넣어주세요. "
                         f"(원본오류: {str(e)[:50]})")
    gc = gspread.authorize(creds)
    sp = str(spreadsheet).strip()
    if sp.startswith("http"):
        return gc.open_by_url(sp)
    if "/" not in sp and " " not in sp and len(sp) >= 30:
        return gc.open_by_key(sp)
    try:
        return gc.open(sp)
    except Exception:
        return gc.open_by_key(sp)


def _is_ws_not_found(e):
    """gspread WorksheetNotFound 판별 — 시트가 진짜 없는 것과 API 오류(쿼터·네트워크)를
    구분한다. API 오류까지 '빈 저장소'로 위장하면 이후 저장 때 데이터가 유실된다."""
    return type(e).__name__ == "WorksheetNotFound"


def _gs_retry(fn, tries=3):
    """일시 오류(쿼터 429·네트워크)용 지수 백오프 재시도."""
    last = None
    for i in range(tries):
        if i:
            time.sleep(2 ** i)
        try:
            return fn()
        except Exception as e:
            last = e
    raise last


def gs_read_ws(sh, title, cols):
    """워크시트 → DataFrame (모든 값 문자열). 시트가 없으면 빈 프레임(cols).
    읽기 API 오류는 삼키지 않고 올린다 — 호출부(storage_load)가 실패 플래그로 처리."""
    try:
        ws = sh.worksheet(title)
    except Exception as e:
        if _is_ws_not_found(e):
            return pd.DataFrame(columns=cols)
        raise
    vals = _gs_retry(ws.get_all_values)
    if not vals or len(vals) < 2:
        return pd.DataFrame(columns=cols)
    df = pd.DataFrame(vals[1:], columns=vals[0])
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    return df[cols]


def gs_write_ws(sh, title, df, cols):
    """DataFrame 으로 워크시트 전체 덮어쓰기.

    clear→update 순서가 아니라 update→resize 순서 — clear 후 update가 실패하면
    시트가 빈 채로 남아 누적 데이터가 통째로 날아가던 문제 방지. update가 실패해도
    기존 데이터는 그대로 남고, 성공 후에만 그리드를 새 크기로 줄여 옛 잔여 행을 지운다."""
    out = df.copy()
    for c in cols:
        if c not in out.columns:
            out[c] = ""
    out = out[cols].fillna("").astype(str)
    try:
        ws = sh.worksheet(title)
    except Exception as e:
        if not _is_ws_not_found(e):
            raise
        ws = sh.add_worksheet(title=title, rows=max(len(out) + 10, 100), cols=max(len(cols), 10))
    values = [cols] + out.values.tolist()
    if ws.row_count < len(values) or ws.col_count < len(cols):
        _gs_retry(lambda: ws.resize(rows=max(ws.row_count, len(values)),
                                    cols=max(ws.col_count, len(cols))))
    _gs_retry(lambda: ws.update(values=values, range_name="A1"))
    if ws.row_count > len(values):                     # 새 데이터보다 아래에 남은 옛 행 제거
        _gs_retry(lambda: ws.resize(rows=len(values)))


def gs_clear_ws(sh, title, cols):
    try:
        ws = sh.worksheet(title)
        ws.clear()
        ws.update(values=[cols], range_name="A1")
    except Exception:
        pass


# ── 문구 자동 태깅 ─────────────────────────────────────────────────────
EMOJI_RE = re.compile(
    "[" "\U0001F300-\U0001FAFF" "\U00002600-\U000027BF" "\U0001F000-\U0001F0FF"
    "\U0000FE00-\U0000FE0F" "\U00002190-\U000021FF" "\U00002B00-\U00002BFF"
    "\U000023E0-\U000023FF" "✅❌✨⚠⏰⏳⌚❤❗❓✌☝"
    "]", flags=re.UNICODE)

# 할인율(%)·가격(원) 소구는 숫자 패턴으로 판별 — 날짜/시간 숫자(2일·1시간)는 제외된다
PCT_RE   = re.compile(r'\d+\s*[%％]')
PRICE_RE = re.compile(r'\d[\d,]*\s*원|\d+\s*만\s*원|\d+\s*만\b')

# CRM PUSH 카피 실무형 소구(訴求) 분류 — 제목+내용 조합으로 판별.
# 실문구 1.5만 건 코퍼스 + 실캠페인 1,277건 성과로 재검증한 체계(v2):
#  · '지금'(CTA)·'최대' 같은 과대매칭 키워드 제거 → 마감임박·할인율의 변별력 회복
#  · 무속성 문구의 주류였던 신상컬렉션·베스트인기·추천큐레이션·시즌환절기 신설
#  · 개인화에 섞여 있던 장바구니·찜은 '행동리타겟'으로 분리(+0.18%p, p=0.001로 유의)
#  · 무료배송(0.3%)은 가격소구에 흡수, 신상입고는 신상컬렉션으로 확장(런칭·발매·룩북 등)
KW = {
    "할인율소구": ["할인", "세일", "반값", "오프", "OFF", "off", "클리어런스", "균일"],
    "가격소구":   ["특가", "단돈", "균일가", "최저가", "초특가", "땡처리", "최저", "득템", "가성비",
                "무료배송", "무배", "배송비", "무료 배송", "인하"],
    "쿠폰적립":   ["쿠폰", "적립", "페이백", "캐시백", "포인트", "마일리지", "코드입력", "코드 입력",
                "무이자", "즉시할인"],
    "사은품증정": ["증정", "사은품", "기프트", "1+1", "더블", "덤", "추가증정", "사은", "드려요", "드립니다", "받아가", "선물"],
    "마감임박":   ["오늘", "마지막", "마감", "종료", "임박", "단 ", "남았", "까지", "D-", "곧 ",
                "놓치", "마지막날", "단하루", "오늘만", "막차", "오늘까지", "내일까지",
                "타임세일", "타임딜", "분만", "막판", "마지막 기회"],
    "한정희소":   ["한정", "단독", "선착순", "수량", "품절", "리미티드", "한정판", "독점", "단 한",
                "한정수량", "소진", "조기"],
    "신상컬렉션": ["신상", "신규 입고", "출시", "입고", "재입고", "새로", "론칭", "런칭", "발매",
                "신제품", "컬렉션", "룩북", "공개", "드롭", "Drop", "NEW", "new", "New",
                "최초", "예약판매", "예판", "선공개"],
    "베스트인기": ["베스트", "BEST", "Best", "랭킹", "TOP", "Top", "1위", "인기", "사랑받은",
                "어워즈", "화제", "난리", "리뷰", "후기", "재구매"],
    "추천큐레이션": ["추천", "큐레이션", "엄선", "MD", "모았", "골라", "가이드", "제안", "셋업",
                 "코디", "스타일링", "어울리", "고민 끝", "맛집", "PICK", "Pick", "픽 "],
    "시즌환절기": ["여름", "겨울", "봄", "가을", "한파", "추위", "더위", "무더위", "장마", "환절기",
                "쌀쌀", "간절기", "서머", "윈터", "S/S", "F/W", "신년", "설날", "명절",
                "크리스마스", "연말", "새해", "휴가", "바캉스"],
    "행동리타겟": ["장바구니", "찜", "위시", "담아두신", "담은", "담으신", "보신", "눈여겨",
                "주문서", "최근 본", "놓고 가신", "다시 만나", "기다리"],
    "이벤트응모": ["이벤트", "응모", "당첨", "럭키", "추첨", "참여", "인증", "댓글", "좋아요",
                "퀴즈", "출석", "룰렛"],
    "라이브방송": ["라이브", "LIVE", "Live", "방송", "생방"],
    "개인화":     ["고객님", "회원님", "님,", "님!", "님 ", "님의", "님을", "님께", "#{", "고객명",
                "관심", "맞춤", "VIP", "등급", "멤버십", "전용"],
    "알림안내":   ["알림", "안내", "도착", "공지", "리마인드", "확인하", "체크", "소식", "업데이트", "리마인"],
}


def _has(s, words):
    return any(w in s for w in words)


# 태그별 키워드 목록을 정규식 1개로 사전 컴파일 — 캠페인 1건당 부분문자열 검사를
# 250여 회 돌리던 것을 태그당 1회 검색으로 줄인다 (의미는 _has와 동일: '아무 키워드나 포함')
KW_RE = {k: re.compile("|".join(re.escape(w) for w in ws)) for k, ws in KW.items()}


def _s(v):
    """NaN/None/숫자 → 안전한 문자열."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ""
    return str(v)


def esc(v):
    """사용자 업로드 문구를 unsafe_allow_html로 렌더하기 전 HTML 이스케이프.
    엑셀 셀의 <, >, & (예: '크리스F&C', '침구<특가>')가 태그/엔티티로 오해석돼 레이아웃이
    깨지거나 주입되는 것을 막는다 (CLAUDE.md esc() 원칙). 줄바꿈은 이스케이프 후 <br>로."""
    import html as _html
    return _html.escape(_s(v)).replace("\n", "<br>")


def safe_ai_html(s):
    """AI 응답을 unsafe_allow_html로 렌더하기 전 위험 태그/속성만 제거.
    AI 프롬프트에는 사용자 업로드 문구가 들어가므로, 악의적 문구가 AI를 조종해 <script>·
    onerror= 등을 출력하면 iframe 내에서 실행될 여지가 있다. 의도된 span/b/br 색상 마크업은
    보존하고 실행 벡터(script·iframe·style·이벤트 핸들러·javascript:)만 걷어낸다."""
    s = _s(s)
    s = re.sub(r'(?is)<\s*(script|iframe|style|object|embed|link|meta)\b.*?(</\s*\1\s*>|$)', '', s)
    s = re.sub(r'(?is)<\s*/?\s*(script|iframe|style|object|embed|link|meta)\b[^>]*>', '', s)
    s = re.sub(r'(?i)\son\w+\s*=\s*("[^"]*"|\'[^\']*\'|[^\s>]+)', '', s)   # onclick·onerror 등
    s = re.sub(r'(?i)(href|src)\s*=\s*(["\']?)\s*javascript:', r'\1=\2#', s)
    return s


def tag_copy(title, body=""):
    """제목+내용 → CRM 실무형 문구 속성 dict. 분석의 핵심 축. (제목·본문 조합 기준)"""
    t = _s(title)
    b = _s(body)
    full = (t + " " + b).strip()
    d = {k: bool(KW_RE[k].search(full)) for k in KW}
    
    has_pct = False
    for m in PCT_RE.finditer(full):
        match_str = m.group()
        # 100% 당첨, 룰렛, 페이백 등은 할인율이 아님
        if "100" in match_str and any(w in full for w in ["당첨", "룰렛", "추첨", "페이백", "캐시백", "지급"]):
            continue
        has_pct = True
        
    d["할인율소구"] = d["할인율소구"] or has_pct
    d["가격소구"] = d["가격소구"] or bool(PRICE_RE.search(full))     # 9,900원 같은 가격 언급
    d["질문형"] = ("?" in full or "？" in full)
    d["이모지"] = bool(EMOJI_RE.search(full))
    d["이모지수"] = len(EMOJI_RE.findall(full))     # 최적 개수 분석용 (보유 여부와 별개)
    d["제목길이"] = len(t)
    d["본문길이"] = len(b)
    return d


TAG_BOOLS = list(KW.keys()) + ["질문형", "이모지"]


def add_tags(df):
    """머지 DataFrame 에 문구 속성 컬럼 추가.

    df.apply(axis=1)는 행마다 Series를 만들어 태깅 자체보다 오버헤드가 크다 —
    리스트 zip 순회로 대체 (결과 동일, 캐시 미스 재태깅 시 수 배 빠름)."""
    if df.empty:
        return df
    titles = df["title"].tolist() if "title" in df else [""] * len(df)
    bodies = df["body"].tolist() if "body" in df else [""] * len(df)
    tdf = pd.DataFrame([tag_copy(t, b) for t, b in zip(titles, bodies)], index=df.index)
    return pd.concat([df, tdf], axis=1)


# ── 키워드(단어)·이모지 단위 성과 분석 ────────────────────────────────────
TOKEN_RE = re.compile(r'[가-힣]{2,}|[A-Za-z]{2,}')
STOPWORDS = set((
    "그리고 그러나 하지만 또는 그 이 저 것 수 등 더 안 못 잘 좀 또 및 의 가 은 는 를 을 에 와 과 도 "
    "으로 한 할 있 너무 정말 지금 바로 모든 위한 위해 통해 에서 에게 부터 까지 처럼 보다 만약 우리 "
    "여기 거기 이런 저런 그런 어떤 무슨 많은 모두 각 약 총 단 매 본 전 후 중 시 분 일 월 년 개 건 명"
).split())


def _word_sig(out, word_col, positions, arr):
    """단어/이모지 표에 보유 vs 미보유 Welch p + FDR 보정 유의성 컬럼을 붙인다.
    소표본 요행 단어가 상위를 채우는 문제의 판별 근거 — '유의' 없는 차이는 참고만."""
    import scipy.stats as ss
    pvals = []
    for tk in out[word_col]:
        pos = positions.get(tk, set())
        yes = arr[list(pos)]
        no = np.delete(arr, list(pos))
        if len(yes) < 3 or len(no) < 3:
            pvals.append(np.nan)
            continue
        try:
            pvals.append(float(ss.ttest_ind(yes, no, equal_var=False).pvalue))
        except Exception:
            pvals.append(np.nan)
    adj = fdr_bh(pvals)
    out = out.copy()
    out["_p_adj"] = adj
    out["유의성"] = ["✱ 유의" if (pd.notna(p) and p < 0.1) else "—" for p in adj]
    return out


def keyword_perf(df, metric_col, min_n=5, top=30, with_sig=False):
    """제목+내용을 단어로 토큰화 → 단어별 (캠페인 단위) 평균 성과·표본·전체평균 대비 차이.

    한 캠페인에서 같은 단어가 여러 번 나와도 1회로만 집계(set). 불용어·숫자 제외.
    반환: DataFrame[단어, 캠페인수, 평균, 차이] (평균 내림차순).
    with_sig=True 면 보유 vs 미보유 Welch+FDR 유의성(_p_adj, 유의성) 컬럼 추가.
    """
    from collections import defaultdict
    if df is None or len(df) == 0 or metric_col not in df:
        return pd.DataFrame(columns=["단어", "캠페인수", "평균", "차이"])
    bucket = defaultdict(list)
    positions = defaultdict(set)
    metvals = []
    for _, r in df.iterrows():
        m = r.get(metric_col)
        if m is None or (isinstance(m, float) and np.isnan(m)):
            continue
        m = float(m)
        i = len(metvals)
        metvals.append(m)
        text = _s(r.get("title", "")) + " " + _s(r.get("body", ""))
        for tk in set(TOKEN_RE.findall(text)):
            if tk in STOPWORDS:
                continue
            bucket[tk].append(m)
            positions[tk].add(i)
    if not metvals:
        return pd.DataFrame(columns=["단어", "캠페인수", "평균", "차이"])
    base_mean = float(np.mean(metvals))
    rows = [dict(단어=tk, 캠페인수=len(v), 평균=float(np.mean(v)), 차이=float(np.mean(v)) - base_mean)
            for tk, v in bucket.items() if len(v) >= min_n]
    if not rows:
        return pd.DataFrame(columns=["단어", "캠페인수", "평균", "차이"])
    out = pd.DataFrame(rows).sort_values("평균", ascending=False).head(top).reset_index(drop=True)
    if with_sig:
        out = _word_sig(out, "단어", positions, np.asarray(metvals, float))
    return out


def emoji_perf(df, metric_col, min_n=3, top=30, with_sig=False):
    """제목+내용의 이모지(기호)별 (캠페인 단위) 평균 성과·표본·전체평균 대비 차이."""
    from collections import defaultdict
    if df is None or len(df) == 0 or metric_col not in df:
        return pd.DataFrame(columns=["이모지", "캠페인수", "평균", "차이"])
    bucket = defaultdict(list)
    positions = defaultdict(set)
    metvals = []
    for _, r in df.iterrows():
        m = r.get(metric_col)
        if m is None or (isinstance(m, float) and np.isnan(m)):
            continue
        m = float(m)
        i = len(metvals)
        metvals.append(m)
        text = _s(r.get("title", "")) + " " + _s(r.get("body", ""))
        for em in set(EMOJI_RE.findall(text)):
            bucket[em].append(m)
            positions[em].add(i)
    if not metvals:
        return pd.DataFrame(columns=["이모지", "캠페인수", "평균", "차이"])
    base_mean = float(np.mean(metvals))
    rows = [dict(이모지=em, 캠페인수=len(v), 평균=float(np.mean(v)), 차이=float(np.mean(v)) - base_mean)
            for em, v in bucket.items() if len(v) >= min_n]
    if not rows:
        return pd.DataFrame(columns=["이모지", "캠페인수", "평균", "차이"])
    out = pd.DataFrame(rows).sort_values("평균", ascending=False).head(top).reset_index(drop=True)
    if with_sig:
        out = _word_sig(out, "이모지", positions, np.asarray(metvals, float))
    return out


# ── 통계 헬퍼: 효과크기 · 다중비교 보정 · 다변량 회귀 ──────────────────────
def cohen_d(a, b):
    """두 그룹 평균차의 표준화 효과크기(Cohen's d). |d|≈0.2 작음/0.5 중간/0.8 큼."""
    a = np.asarray(a, float); b = np.asarray(b, float)
    a = a[~np.isnan(a)]; b = b[~np.isnan(b)]
    if len(a) < 2 or len(b) < 2:
        return np.nan
    na, nb = len(a), len(b)
    sp2 = ((na - 1) * a.var(ddof=1) + (nb - 1) * b.var(ddof=1)) / (na + nb - 2)
    if sp2 <= 0:
        return np.nan
    return float((a.mean() - b.mean()) / np.sqrt(sp2))


def fdr_bh(pvals):
    """Benjamini-Hochberg FDR 보정 p값 (NaN은 그대로 통과)."""
    p = np.asarray(pvals, float)
    out = np.full(len(p), np.nan)
    mask = ~np.isnan(p)
    pv = p[mask]; n = len(pv)
    if n == 0:
        return out
    order = np.argsort(pv)
    ranked = pv[order]
    adj = ranked * n / np.arange(1, n + 1)
    adj = np.minimum.accumulate(adj[::-1])[::-1]
    adj = np.clip(adj, 0, 1)
    res = np.empty(n); res[order] = adj
    out[mask] = res
    return out


# 순위 보정용 지표 → (분자, 분모) 매핑 — 캠페인/슬롯 레벨 공용
RANK_ND = {"ord_cr": ("oc", "uv"), "infl_cr": ("uv", "send"),
           "rps": ("amt", "send"), "aov": ("amt", "oc")}


def rank_adjusted(d, col, ascending=False):
    """순위용 표본 보정 점수 — UV 수십 건짜리 캠페인(또는 슬롯)이 우연 하나로
    TOP/BOTTOM을 점령하는 것을 막는다. 표시 값은 그대로 두고 '정렬 순서'에만 쓴다.

    · 비율 지표(주문CR=주문/UV, CTR=UV/발송): Jeffreys 신뢰구간 경계
      - TOP(내림차순)은 하한 — 표본이 작을수록 점수가 깎여 순위에서 밀린다
      - BOTTOM(오름차순)은 상한 — 소표본의 '가짜 0%'가 바닥을 점령하지 않는다
    · 금액 지표(RPS=거래액/발송, 객단가=거래액/주문): 전체 가중평균으로의 수축
      (분자 + k·전체평균) / (분모 + k), k=분모 중앙값(통상 캠페인 1건의 무게)
    · 그 외(거래액 등 절대값): 원값 그대로 반환.
    """
    nd = RANK_ND.get(col)
    if nd is None or nd[0] not in d or nd[1] not in d:
        return pd.to_numeric(d[col], errors="coerce") if col in d else pd.Series(np.nan, index=d.index)
    num = pd.to_numeric(d[nd[0]], errors="coerce").fillna(0.0).clip(lower=0)
    den = pd.to_numeric(d[nd[1]], errors="coerce").fillna(0.0).clip(lower=0)
    if col in ("ord_cr", "infl_cr"):
        import scipy.stats as ss
        x = np.minimum(num, den)                     # 분자>분모 데이터 오류 방어
        q = 0.95 if ascending else 0.05
        return pd.Series(ss.beta.ppf(q, x + 0.5, (den - x) + 0.5), index=d.index)
    gden = den.sum()
    prior = (num.sum() / gden) if gden else 0.0
    pos = den[den > 0]
    k = float(np.median(pos)) if len(pos) else 1.0
    return (num + k * prior) / (den + k)


def ols_effects(df, attr_cols, ctrl_cols, ycol):
    """속성(0/1) + 통제변수(카테고리·시간대 등 더미) 다중회귀 → 속성별 '순효과' 계수·표준오차·p값.

    단순 평균비교가 잡아내지 못하는 교란(예: 특정 속성이 특정 카테고리에 몰림)을 통제한다.
    반환: DataFrame[속성, 순효과(계수), 표준오차, p] — attr_cols 에 대해서만.
    """
    from scipy import stats as _stats
    if df is None or len(df) == 0 or ycol not in df:
        return pd.DataFrame(columns=["속성", "순효과", "표준오차", "p"])
    d = df.dropna(subset=[ycol]).copy()
    use_attrs = [a for a in attr_cols if a in d.columns and d[a].nunique() > 1]
    if not use_attrs or len(d) < len(use_attrs) + 5:
        return pd.DataFrame(columns=["속성", "순효과", "표준오차", "p"])
    parts = [np.ones((len(d), 1))]
    names = ["intercept"]
    for a in use_attrs:
        parts.append(d[a].astype(float).values.reshape(-1, 1)); names.append(a)
    for c in ctrl_cols:
        if c in d.columns and d[c].astype(str).nunique() > 1:
            dum = pd.get_dummies(d[c].astype(str), prefix=c, drop_first=True)
            if dum.shape[1] > 0:
                parts.append(dum.values.astype(float)); names.extend(list(dum.columns))
    X = np.hstack(parts).astype(float)
    y = d[ycol].astype(float).values
    n, k = X.shape
    if n - k <= 0:
        return pd.DataFrame(columns=["속성", "순효과", "표준오차", "p"])
    beta, _, _, _ = np.linalg.lstsq(X, y, rcond=None)
    resid = y - X @ beta
    dof = n - k
    sigma2 = float(resid @ resid) / dof
    XtX_inv = np.linalg.pinv(X.T @ X)
    se = np.sqrt(np.clip(np.diag(sigma2 * XtX_inv), 0, None))
    with np.errstate(divide="ignore", invalid="ignore"):
        tvals = np.where(se > 0, beta / se, np.nan)
    pvals = 2 * (1 - _stats.t.cdf(np.abs(tvals), dof))
    idx = {nm: i for i, nm in enumerate(names)}
    rows = []
    for a in use_attrs:
        i = idx[a]
        rows.append(dict(속성=a, 순효과=float(beta[i]), 표준오차=float(se[i]), p=float(pvals[i])))
    return pd.DataFrame(rows).sort_values("순효과", ascending=False).reset_index(drop=True)


def fmt_hhmm(h):
    """발송 시간대(HHMM 또는 H 정수) → '12시' / '08시' / '10시 50분'. 잘못된 값은 '–'."""
    try:
        v = int(float(h))
    except Exception:
        return "–"
    if v < 0:
        return "–"
    hour, minute = (v, 0) if v <= 23 else divmod(v, 100)
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return "–"
    return f"{hour:02d}시" + (f" {minute:02d}분" if minute else "")


def hhmm_to_minutes(h):
    """HHMM(또는 H) → 자정 기준 분(정렬용). 잘못된 값은 0."""
    try:
        v = int(float(h))
    except Exception:
        return 0
    if v < 0:
        return 0
    hour, minute = (v, 0) if v <= 23 else divmod(v, 100)
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return 0
    return hour * 60 + minute


def build_report_html(title, blocks):
    """페이지에서 캡처한 차트/표 블록 → 인쇄용 자립형 HTML (브라우저 Ctrl+P로 PDF 저장)."""
    import plotly.io as pio

    def _e(s):
        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    parts, first_fig = [], True
    for kind, obj in blocks:
        try:
            if kind == "fig":
                parts.append(pio.to_html(obj, include_plotlyjs=("cdn" if first_fig else False),
                                         full_html=False, default_width="100%"))
                first_fig = False
            elif kind == "table":
                data = getattr(obj, "data", obj)
                if hasattr(data, "shape") and getattr(data, "shape", (0,))[0] > 300:
                    parts.append(data.head(300).to_html(index=False) + "<p>(상위 300행만 표시)</p>")
                elif hasattr(obj, "to_html"):
                    try:
                        parts.append(obj.to_html())
                    except Exception:
                        if hasattr(data, "to_html"):
                            parts.append(data.to_html(index=False))
        except Exception:
            pass
    css = ("@import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/variable/pretendardvariable-dynamic-subset.min.css');"
           "body{font-family:'Pretendard Variable',Pretendard,'Malgun Gothic','Apple SD Gothic Neo','Nanum Gothic',sans-serif;"
           "color:#1e293b;margin:24px;} h1{font-size:20px;margin:0 0 4px;} "
           "table{border-collapse:collapse;font-size:12px;margin:10px 0;} "
           "th,td{border:1px solid #e2e8f0;padding:4px 8px;text-align:right;} "
           "th{background:#f1f5f9;} .meta{color:#64748b;font-size:12px;margin-bottom:14px;}")
    when = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    return (f"<!doctype html><html><head><meta charset='utf-8'><title>{_e(title)}</title>"
            f"<style>{css}</style></head><body><h1>{_e(title)}</h1>"
            f"<div class='meta'>LF몰 발송성과 대시보드 · 생성 {when}</div>"
            f"{''.join(parts)}</body></html>")


# 기획전 비교분석 표 — 발송 성과 / 기획전 성과 2단 그룹 헤더 + 서식
PROMO_TBL_SPECS = [
    ("", "promo", "기획전번호", None),
    ("", "pname", "기획전명", None),
    ("", "발송일자", "발송일자", None),
    ("", "발송시간", "발송시간", None),
    ("", "n_camp", "캠페인수", "{:,.0f}"),
    ("발송 성과", "send", "발송", "{:,.0f}"),
    ("발송 성과", "s_uv", "UV", "{:,.0f}"),
    ("발송 성과", "s_visit", "VISIT", "{:,.0f}"),
    ("발송 성과", "s_ctr", "CTR", "{:.2%}"),
    ("발송 성과", "s_cr", "CR", "{:.2%}"),
    ("발송 성과", "s_amt", "발송추적거래액", "{:,.0f}"),
    ("발송 성과", "s_rps", "RPS", "{:,.0f}"),
    ("기획전 성과", "p_pv", "PV", "{:,.0f}"),
    ("기획전 성과", "p_uv", "UV", "{:,.0f}"),
    ("기획전 성과", "inf_amt", "유입거래액", "{:,.0f}"),
    ("", "기여율", "기여율", "{:.1%}"),
]


def promo_perf_table(d):
    """per-기획전 DataFrame → 발송/기획전 성과를 2단 헤더로 묶은 Styler."""
    specs = [s for s in PROMO_TBL_SPECS if s[1] in d.columns]
    sub = d[[c for _, c, _, _ in specs]].copy()
    sub.columns = pd.MultiIndex.from_tuples([(g, l) for g, _, l, _ in specs])
    fmtmap = {(g, l): f for g, _, l, f in specs if f}
    return sub.style.format(fmtmap)


# ══════════════════════════════════════════════════════════════════════
# 2. Streamlit 앱
# ══════════════════════════════════════════════════════════════════════
def get_push_stats(df, start_d, end_d, grp):
    sd = pd.to_datetime(start_d)
    ed = pd.to_datetime(end_d)
    sub = df[(df["group"] == grp) & (df["date"] >= sd) & (df["date"] <= ed) & (~df["is_outlier"])].copy()
    if sub.empty:
        return None
    return {
        "last_consent": sub.sort_values("date").iloc[-1]["consent"],
        "tot_added": sub["added"].sum(),
        "tot_removed": sub["removed"].sum(),
        "tot_diff": sub["diff"].sum() if "diff" in sub else 0
    }

def main():
    import streamlit as st
    import plotly.graph_objects as go
    import streamlit.components.v1 as components
    from scipy import stats

    # ── 페이지 리포트 캡처: st.plotly_chart / st.dataframe 호출을 가로채 기록 ──
    if not hasattr(st, "_orig_plotly_chart"):
        st._orig_plotly_chart = st.plotly_chart
    if not hasattr(st, "_orig_dataframe"):
        st._orig_dataframe = st.dataframe
    # 캡처 버퍼는 session_state에 — st 모듈 속성은 프로세스(멀티유저) 공유라 클로저
    # 리스트에 쌓으면 동시 접속한 다른 사용자의 차트/표가 내 리포트에 섞인다.
    # session_state 프록시는 '호출 시점' 세션 컨텍스트를 따라가므로 항상 자기 세션에 쌓인다.
    st.session_state["_page_report"] = []
    _REPORT = st.session_state["_page_report"]

    def _cap_plotly(*a, **k):
        try:
            f = a[0] if a else k.get("figure_or_data")
            if f is not None:
                st.session_state.setdefault("_page_report", []).append(("fig", f))
        except Exception:
            pass
        return st._orig_plotly_chart(*a, **k)

    def _cap_df(*a, **k):
        try:
            d = a[0] if a else k.get("data")
            if d is not None:
                st.session_state.setdefault("_page_report", []).append(("table", d))
        except Exception:
            pass
        return st._orig_dataframe(*a, **k)
    st.plotly_chart = _cap_plotly
    st.dataframe = _cap_df

    st.set_page_config(page_title="LF몰 발송성과 대시보드", layout="wide",
                       initial_sidebar_state="expanded")
    st.markdown("""
    <style>
    /* 한글 가독성 폰트: Pretendard (동적 서브셋 — 쓰인 글자만 내려받아 빠름) */
    @import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/variable/pretendardvariable-dynamic-subset.min.css');
    html, body, .stApp, .stApp p, .stApp div, .stApp span, .stApp label, .stApp li,
    .stApp td, .stApp th, .stApp button, .stApp input, .stApp textarea, .stApp select,
    h1, h2, h3, h4, h5, h6 {
      font-family: 'Pretendard Variable', Pretendard, -apple-system, BlinkMacSystemFont,
                   'Malgun Gothic', 'Apple SD Gothic Neo', 'Noto Sans KR', sans-serif;
    }
    /* Streamlit UI 아이콘(Material Symbols)은 아이콘 폰트를 유지해야 깨지지 않는다 */
    [data-testid="stIconMaterial"], .material-symbols-rounded, .material-symbols-outlined {
      font-family: 'Material Symbols Rounded', 'Material Symbols Outlined' !important;
    }
    .stApp code, .stApp pre { font-family: 'Source Code Pro', ui-monospace, monospace !important; }
    /* 숫자 자릿수 정렬(테이블·지표 가독성): 고정폭 숫자 */
    [data-testid="stMetricValue"], [data-testid="stMetricDelta"] { font-feature-settings: 'tnum' 1; }
    [data-testid="stAppViewContainer"]{background:#f8f9fc}
    [data-testid="stSidebar"]{background:#ffffff;border-right:1px solid #e2e8f0}
    [data-testid="stMetric"]{background:#ffffff;border-radius:8px;padding:12px 16px;border:1px solid #e2e8f0;box-shadow:0 1px 3px rgba(0,0,0,0.06)}
    [data-testid="stMetricLabel"]{color:#64748b!important;font-size:12px!important}
    [data-testid="stMetricValue"]{color:#1e293b!important;font-size:20px!important}
    .vg{border:1px solid #e2e8f0;border-radius:8px;padding:14px 18px;margin:8px 0;line-height:1.7;background:#ffffff}
    .sdiv{border-top:1px solid #e2e8f0;margin:22px 0}
    .stat-label{font-size:11px;color:#545c6a;margin-bottom:3px;font-weight:500;letter-spacing:.04em;text-transform:uppercase}
    .appendix{background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:14px 18px;margin-top:12px;font-size:13px;color:#475569}
    .tag{display:inline-block;font-size:11px;padding:1px 7px;border-radius:10px;margin:1px 2px;border:1px solid}
    h1,h2,h3{color:#1e293b}
    </style>""", unsafe_allow_html=True)

    PALETTE = {
        "blue": "rgba(79,143,255,1)", "red": "rgba(245,101,101,1)", "green": "rgba(72,187,120,1)",
        "amber": "rgba(237,137,54,1)", "purple": "rgba(159,122,234,1)", "teal": "rgba(56,178,172,1)",
        "slate": "rgba(100,116,139,1)",
    }

    def base_layout(h=300, ysuffix="", title="", hover=None):
        # 인-차트 제목은 왼쪽 상단에 고정하고, 상단 가로 범례(y≈1.12)와 겹치지 않도록
        # 제목 영역(top margin)을 충분히 확보한다. (각 차트 위 마크다운 헤더와도 분리)
        # hover="x" 이면 시계열용 통합 툴팁(hovermode=x unified) + 세로 크로스헤어를 켠다.
        has_title = bool(title)
        _FONT = "Pretendard Variable, Pretendard, Malgun Gothic, Apple SD Gothic Neo, sans-serif"
        xaxis = dict(gridcolor="rgba(0,0,0,0)", linecolor="#e2e8f0",
                     tickfont=dict(color="#64748b", size=11))
        lay = dict(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                   font=dict(color="#475569", size=11, family=_FONT),
                   margin=dict(l=10, r=10, t=(58 if has_title else 30), b=10),
                   height=h, showlegend=False,
                   barcornerradius=4,
                   hoverlabel=dict(bgcolor="#ffffff", bordercolor="#e2e8f0",
                                   font=dict(size=12, color="#1e293b", family=_FONT)),
                   title=dict(text=title, font=dict(color="#94a3b8", size=13),
                              x=0, xanchor="left", y=0.99, yanchor="top"),
                   legend=dict(orientation="h", yanchor="bottom", y=1.02,
                               xanchor="left", x=0, bgcolor="rgba(0,0,0,0)"),
                   xaxis=xaxis,
                   yaxis=dict(gridcolor="#f1f5f9", linecolor="#e2e8f0",
                              tickfont=dict(color="#64748b", size=11), ticksuffix=ysuffix))
        if hover == "x":
            lay["hovermode"] = "x unified"
            xaxis.update(showspikes=True, spikemode="across", spikethickness=1,
                         spikedash="dot", spikecolor="#cbd5e1")
        return lay

    # 시계열 팔레트 순서 — 소구 속성별 고정색(필터가 바뀌어도 같은 속성은 늘 같은 색)
    SERIES_SEQ = [PALETTE["purple"], PALETTE["green"], PALETTE["amber"], PALETTE["blue"],
                  PALETTE["red"], PALETTE["teal"], PALETTE["slate"]]

    def tag_color(t):
        try:
            return SERIES_SEQ[TAG_BOOLS.index(t) % len(SERIES_SEQ)]
        except ValueError:
            return PALETTE["slate"]

    def stacked_panels(x, bar_y, bar_name, line_y, line_name, bar_color, line_color,
                       h=430, bar_suffix="", line_suffix="", title=""):
        """이중축 대신 X축을 공유하는 상(막대)/하(선) 패널 — 두 지표의 스케일을 섞지 않으면서
        같은 시점끼리 비교할 수 있게 한다. 통합 툴팁·크로스헤어 포함."""
        from plotly.subplots import make_subplots
        fig = make_subplots(rows=2, cols=1, shared_xaxes=True,
                            row_heights=[0.42, 0.58], vertical_spacing=0.07)
        fig.add_trace(go.Bar(x=x, y=bar_y, name=bar_name, marker_color=bar_color,
                             opacity=0.75), row=1, col=1)
        fig.add_trace(go.Scatter(x=x, y=line_y, name=line_name, mode="lines+markers",
                                 line=dict(color=line_color, width=2),
                                 marker=dict(size=7)), row=2, col=1)
        lay = base_layout(h=h, title=title)
        lay.pop("xaxis", None); lay.pop("yaxis", None)
        lay["showlegend"] = True
        lay["hovermode"] = "x unified"
        fig.update_layout(**lay)
        fig.update_xaxes(gridcolor="rgba(0,0,0,0)", linecolor="#e2e8f0",
                         tickfont=dict(color="#64748b", size=11),
                         showspikes=True, spikemode="across", spikethickness=1,
                         spikedash="dot", spikecolor="#cbd5e1")
        fig.update_yaxes(gridcolor="#f1f5f9", linecolor="#e2e8f0",
                         tickfont=dict(color="#64748b", size=11))
        fig.update_yaxes(ticksuffix=bar_suffix, row=1, col=1)
        fig.update_yaxes(ticksuffix=line_suffix, row=2, col=1)
        return fig

    def overlay_dual(x, bar_y, bar_name, line_y, line_name, bar_color, line_color,
                     h=430, bar_suffix="", line_suffix="", title=""):
        """한 차트 이중축 오버레이 — 막대(좌축)+선(우축)을 같은 X에 겹쳐 그린다.
        스케일이 크게 다른 '발송 강도 ↔ 효율/증감' 관계를 한 눈에 보려는 용도.
        어느 축인지 헷갈리지 않게 축 눈금·제목 색을 각 시리즈 색과 맞춘다.
        통합 툴팁(x unified)+세로 크로스헤어 포함."""
        fig = go.Figure()
        fig.add_trace(go.Bar(x=x, y=bar_y, name=bar_name, marker_color=bar_color,
                             opacity=0.55))
        fig.add_trace(go.Scatter(x=x, y=line_y, name=line_name, mode="lines+markers",
                                 line=dict(color=line_color, width=2.5),
                                 marker=dict(size=6), yaxis="y2"))
        lay = base_layout(h=h, title=title, hover="x")
        lay["showlegend"] = True
        lay["legend"] = dict(orientation="h", yanchor="bottom", y=1.02,
                             xanchor="left", x=0, bgcolor="rgba(0,0,0,0)")
        lay["yaxis"]["ticksuffix"] = bar_suffix
        lay["yaxis"]["title"] = dict(text=bar_name, font=dict(color=bar_color, size=11))
        lay["yaxis"]["tickfont"] = dict(color=bar_color, size=11)
        lay["yaxis2"] = dict(overlaying="y", side="right", showgrid=False,
                             ticksuffix=line_suffix,
                             title=dict(text=line_name, font=dict(color=line_color, size=11)),
                             tickfont=dict(color=line_color, size=11),
                             zeroline=True, zerolinecolor="#e2e8f0")
        fig.update_layout(**lay)
        return fig

    def sig_label(p):
        if p is None or (isinstance(p, float) and np.isnan(p)):
            return "–"
        if p < 0.001: return "p<0.001 · 매우 확실"
        if p < 0.01:  return "p<0.01 · 신뢰 가능"
        if p < 0.05:  return "p<0.05 · 유의함"
        return f"p={p:.2f} · 유의하지 않음(우연일 수 있음)"

    def welch(a, b):
        a = np.asarray(a, float); b = np.asarray(b, float)
        a = a[~np.isnan(a)]; b = b[~np.isnan(b)]
        if len(a) < 3 or len(b) < 3:
            return np.nan
        try:
            return stats.ttest_ind(a, b, equal_var=False).pvalue
        except Exception:
            return np.nan

    # ── 캐시 래퍼 (무거운 파싱 1회만) ──
    @st.cache_data(show_spinner=False)
    def cached_perf(b): return parse_perf_bytes(b)

    @st.cache_data(show_spinner=False)
    def cached_plan(b): return parse_plan_bytes(b)

    @st.cache_data(show_spinner=False)
    def cached_mtd(b): return parse_mtd_bytes(b)

    @st.cache_data(show_spinner=False)
    def cached_promo(b): return parse_promo_bytes(b)

    # (기획 시트 파싱은 캐시하지 않는다 — 「기획 문구 가져오기」는 '지금 시트를 다시 읽어라'는
    #  명시적 사용자 액션인데, 캐시하면 시트 문구를 수정하고 다시 눌러도 최대 1시간 옛 값이
    #  반환돼 기능 의도를 정면으로 해친다. 매 클릭 fresh 로드.)

    @st.cache_data(show_spinner=False)
    def cached_classify(b): return classify_upload("", b)

    @st.cache_data(show_spinner=False)
    def cached_compute_mtd(mtd_df):
        """compute_mtd 캐시 — 어느 페이지에서든 매 rerun마다 groupby 4회+회귀 5회를
        다시 돌던 것을 데이터가 같으면 건너뛴다."""
        return compute_mtd(mtd_df)

    # 키워드/이모지 성과 캐시 — iterrows 토큰화가 rerun마다 돌지 않게.
    # _ver 에 불용어·토큰 규칙 시그니처를 넣어 규칙 변경 시 캐시가 무효화되게 한다.
    _KWVER = hashlib.md5((TOKEN_RE.pattern + "|".join(sorted(STOPWORDS))).encode()).hexdigest()[:10]

    @st.cache_data(show_spinner=False)
    def cached_keyword_perf(d, metric_col, min_n, top, _ver=_KWVER):
        return keyword_perf(d, metric_col, min_n=min_n, top=top, with_sig=True)

    @st.cache_data(show_spinner=False)
    def cached_emoji_perf(d, metric_col, min_n, top, _ver=_KWVER):
        return emoji_perf(d, metric_col, min_n=min_n, top=top, with_sig=True)

    # 태깅 체계 버전 키 — 속성 구성(TAG_BOOLS)뿐 아니라 키워드 '내용'(KW)이 바뀌어도
    # prepare_raw 캐시를 무효화한다. (st.cache_data는 내부에서 호출하는 tag_copy의 변경을
    # 감지하지 못해, 태그명이 그대로면 구버전 태깅 결과가 캐시로 반환되던 구멍 방지)
    TAGSET_VER = (hashlib.md5(json.dumps(KW, ensure_ascii=False, sort_keys=True).encode())
                  .hexdigest()[:12] + "|" + "|".join(TAG_BOOLS) + "|이모지수v1")

    @st.cache_data(show_spinner=False)
    def prepare_raw(work_df, tagset_ver):
        """파생 재계산 + 타입정리 + 문구 태깅을 1회만 — 필터·페이지 이동 때 재계산 방지(성능 핵심).
        입력 work_df 가 동일하면(저장 데이터만 볼 때) 캐시 히트하여 무거운 태깅을 건너뛴다."""
        r = _finalize(work_df.copy())
        r["title"] = r["title"].fillna("").astype(str) if "title" in r else ""
        r["body"] = r["body"].fillna("").astype(str) if "body" in r else ""
        r["matched"] = r["matched"].map(_to_bool) if "matched" in r else False
        return add_tags(r)

    # ── 저장소 백엔드: 구글시트(설정 시) ↔ 로컬 CSV(폴백) ──
    @st.cache_resource(show_spinner=False)
    def _get_sh(_email, spreadsheet):
        return gs_open(st.secrets["gcp_service_account"], spreadsheet)

    def init_storage():
        try:
            has = "gcp_service_account" in st.secrets
        except Exception:
            has = False
        if not has:
            return {"mode": "local", "status": "💾 로컬 CSV로 저장해요 (구글시트 미설정)"}
        try:
            sp = None
            if "gsheets" in st.secrets:
                sp = st.secrets["gsheets"].get("spreadsheet")
            sp = sp or st.secrets.get("gsheets_spreadsheet")
            if not sp:
                return {"mode": "local", "status": "⚠️ gsheets.spreadsheet 미설정 → 로컬에 저장해요"}
            sh = _get_sh(st.secrets["gcp_service_account"].get("client_email", ""), sp)
            return {"mode": "gsheets", "sh": sh, "status": "☁️ 구글시트에 연결됐어요"}
        except Exception as e:
            return {"mode": "local", "status": f"⚠️ 구글시트 연결 실패 → 로컬에 저장해요 ({str(e)[:50]})"}

    _STORE_COLS_BY_KIND = {"campaign": STORE_COLS, "mtd": MTD_STORE_COLS, "promo": PROMO_STORE_COLS,
                           "push": PUSH_STORE_COLS, "notes": NOTES_STORE_COLS}
    _LOCAL_LOAD = {"campaign": load_store, "mtd": load_mtd_store, "promo": load_promo_store,
                   "push": load_push_store, "notes": load_notes_store}
    _LOCAL_SAVE = {"campaign": save_store, "mtd": save_mtd_store, "promo": save_promo_store,
                   "push": save_push_store, "notes": save_notes_store}
    _LOCAL_FILE = {"campaign": DATA_STORE, "mtd": MTD_STORE, "promo": PROMO_STORE,
                   "push": PUSH_STORE, "notes": NOTES_STORE}

    def storage_load(bk, kind):
        cols = _STORE_COLS_BY_KIND[kind]
        if bk["mode"] == "gsheets":
            try:
                out = gs_read_ws(bk["sh"], GS_TITLES[kind], cols)
                st.session_state.pop(f"_load_failed_{kind}", None)
                return out
            except Exception as e:
                # 읽기 실패(쿼터·네트워크)를 '빈 저장소'로 위장하면, 이후 저장 때 빈
                # 스냅샷으로 시트 전체를 덮어써 누적 데이터가 유실된다 — 실패 플래그를
                # 남겨 storage_save 가 저장을 차단하게 한다.
                st.session_state[f"_load_failed_{kind}"] = str(e)[:120]
                return pd.DataFrame(columns=cols)
        return _LOCAL_LOAD[kind]()

    _KIND_LABEL = {"campaign": "캠페인", "mtd": "MTD", "promo": "기획전",
                   "push": "앱푸시", "notes": "보고란"}

    def storage_save(bk, kind, df):
        """저장 성공 여부(bool) 반환 — 클라우드 읽기가 실패한 상태면 덮어쓰기 방지를 위해 차단."""
        err = st.session_state.get(f"_load_failed_{kind}")
        if err:
            st.error(f"⛔ {_KIND_LABEL.get(kind, kind)} 저장을 차단했어요 — 기존 클라우드 데이터를 "
                     f"읽지 못한 상태에서 저장하면 누적분이 덮어써져요. 사이드바 「🔄 새로 불러오기」로 "
                     f"다시 읽은 뒤 저장해 주세요. (읽기 실패 원인: {err})")
            return False
        cols = _STORE_COLS_BY_KIND[kind]
        if bk["mode"] == "gsheets":
            gs_write_ws(bk["sh"], GS_TITLES[kind], df, cols)
        else:
            _LOCAL_SAVE[kind](df)
        return True

    def storage_clear(bk, kind):
        cols = _STORE_COLS_BY_KIND[kind]
        if bk["mode"] == "gsheets":
            gs_clear_ws(bk["sh"], GS_TITLES[kind], cols)
        else:
            f = _LOCAL_FILE[kind]
            if os.path.exists(f):
                os.remove(f)

    BK = init_storage()

    # (병렬 프리로드는 단일 gspread 클라이언트를 여러 스레드가 공유해 토큰 refresh·세션
    #  경합 위험이 있고, 얻는 2~4초보다 간헐 로드 실패 리스크가 커서 제거 — 아래 각
    #  storage_load 개별(순차) 로드 경로가 그대로 처리한다.)

    AI_MODELS = {
        "Gemini 2.5 Pro (최고 품질)": "gemini-2.5-pro",
        "Gemini 2.5 Flash (균형)": "gemini-2.5-flash",
        "Gemini 2.5 Flash-Lite (빠름·저렴)": "gemini-2.5-flash-lite",
    }

    def gemini_key():
        for k in ("GEMINI_API_KEY", "GOOGLE_API_KEY"):
            try:
                if k in st.secrets:
                    return st.secrets[k]
            except Exception:
                pass
            v = os.environ.get(k)
            if v:
                return v
        return None

    @st.cache_resource(show_spinner=False)
    def _gemini_client(key):
        """클라이언트 1회 생성 재사용 — 호출마다 신규 생성하던 오버헤드 제거. 120초 타임아웃."""
        from google import genai
        return genai.Client(api_key=key, http_options={"timeout": 120_000})

    def _ai_transient(e):
        s = str(e).lower()
        return any(t in s for t in ("429", "500", "503", "quota", "rate", "timeout",
                                    "deadline", "unavailable", "overloaded"))

    @st.cache_data(ttl=3600, show_spinner=False)
    def _ai_gen_cached(system, user, model, _key):
        """같은 (지시문·데이터·모델) 재클릭 시 1시간 내 재과금·재대기 없이 재사용.
        데이터가 바뀌면 user 텍스트가 달라져 자동으로 새로 생성된다.
        실패·빈 응답은 예외로 올려 캐시에 남기지 않는다 (재클릭 시 재시도 가능)."""
        from google.genai import types
        client = _gemini_client(_key)
        last_err = None
        for wait in (0, 3):                       # 일시 오류(429·5xx·타임아웃) 1회 백오프 재시도
            if wait:
                time.sleep(wait)
            try:
                resp = client.models.generate_content(
                    model=model,
                    contents=user,
                    config=types.GenerateContentConfig(
                        system_instruction=system,
                        max_output_tokens=8000,
                    ),
                )
                text = (resp.text or "").strip()
                if not text:
                    raise RuntimeError("빈 응답 — 잠시 후 다시 시도해 주세요.")
                # 길이 제한으로 잘린 응답을 '완전한 결과'처럼 보여주지 않게 감지·표시
                try:
                    _fr = str(resp.candidates[0].finish_reason or "")
                except Exception:
                    _fr = ""
                if "MAX_TOKENS" in _fr.upper():
                    text += ("\n\n(⚠️ 응답이 길이 제한으로 중간에 잘렸어요 — 필터로 데이터 범위를 "
                             "좁히거나 다시 생성해 보세요.)")
                return text
            except Exception as e:
                last_err = e
                if not _ai_transient(e):
                    break
        raise RuntimeError(str(last_err)[:200])

    def ai_generate(system, user, model):
        key = gemini_key()
        if not key:
            return None, "GEMINI_API_KEY 미설정 — Streamlit Secrets 또는 환경변수에 추가하세요."
        try:
            import google.genai                     # noqa: F401 — 설치 여부만 확인
        except ImportError:
            return None, "google-genai 패키지가 없습니다. requirements.txt 반영 후 재배포하세요."
        try:
            return _ai_gen_cached(system, user, model, _key=key), None
        except Exception as e:
            # 예외 문자열에 API 키가 섞여(요청 URL·설정 repr 등) 화면에 노출되지 않게 마스킹
            _msg = str(e).replace(str(key), "***KEY***")[:300]
            return None, f"생성 오류: {_msg}"

    # 숫자 포맷 헬퍼
    def won(v):
        if v is None or (isinstance(v, float) and np.isnan(v)): return "–"
        if abs(v) >= 1e8: return f"{v/1e8:.2f}억"
        if abs(v) >= 1e4: return f"{v/1e4:,.0f}만"
        return f"{v:,.0f}"

    def bar_label(v, col, pct):
        """막대 텍스트 라벨 — 비율은 %, 통화 지표는 원화 축약.
        (RPS·거래액 라벨이 '1234567.89'처럼 콤마·단위 없이 찍히던 문제 통일)"""
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "–"                                # 결측 버킷 라벨이 'nan%'로 노출되던 것 방지
        if pct:
            return f"{v:.2f}%"
        if col in ("rps", "aov", "amt", "revenue", "avgOrderVal", "unitPrice"):
            return won(v)
        return f"{v:,.1f}"

    def bar_xrange(yv, mult=1.18):
        """수평 막대 x축 상한 — 값이 전부 0/NaN이면 None(자동)으로 둬 축이 [0,0]·[0,NaN]으로
        붕괴해 막대·라벨이 안 보이는 빈 차트가 되는 것을 막는다."""
        arr = np.asarray(yv, float)
        if not len(arr):
            return None
        mx = float(np.nanmax(arr)) if np.isfinite(arr).any() else 0.0
        return [0, mx * mult] if (np.isfinite(mx) and mx > 0) else None

    def guard_select(key, opts):
        """지표·필터 변경으로 옵션 목록이 바뀌면 세션에 남은 선택값이 목록 밖이 될 수 있다 —
        무효 선택은 버려서 조용한 리셋/예외를 막는다 (개발 규칙의 pop 가드 일원화)."""
        if key in st.session_state and st.session_state[key] not in opts:
            st.session_state.pop(key, None)

    def legend_h():
        """차트 상단 수평 범례 — 페이지마다 복붙되던 dict를 한 곳으로."""
        return dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0,
                    bgcolor="rgba(0,0,0,0)")

    # ══════════════════════════════════════════════════════════════
    # 사이드바 — 업로드 / 필터 / 페이지
    # ══════════════════════════════════════════════════════════════
    st.sidebar.markdown("### 📨 발송성과 대시보드")
    st.sidebar.caption("문구와 성과를 합쳐서 분석해요")

    def _has_sa():
        try:
            return "gcp_service_account" in st.secrets
        except Exception:
            return False

    # ── 통합 업로더: 한 곳에 올리면 자동 인식·분류 ──
    uni_files = st.sidebar.file_uploader(
        "📂 파일 올리기 (xlsx/zip · 한 번에 여러 개 가능)",
        type=["xlsx", "zip"], accept_multiple_files=True, key="uni_up",
        help="발송실적·기획·기획전성과·전사MTD·앱푸시 수신동의 파일을 한 번에 올리면 "
             "자동으로 분류돼요. ZIP 파일도 돼요.")

    if "push_consent_df" not in st.session_state:
        st.session_state.push_consent_df = finalize_push(storage_load(BK, "push"))
    push_consent_df = st.session_state.push_consent_df

    # ── 발송기획(문구) 소스: 업로드 파일 ↔ 구글시트 직접연결 ──
    _PLAN_SHEET_URL = "https://docs.google.com/spreadsheets/d/1xqlaRnHa5HMLz3ASUn-H7AvMkUsvQw6l7aw1rWLqfjw"
    plan_file = None
    st.sidebar.markdown("---")
    recent_n = st.sidebar.number_input("가져올 최근 주차 수 (0이면 전부)", value=12, min_value=0, step=1,
                                       help="시트가 많으면 전부(0)는 느리고 API 한도에 걸릴 수 있어요. 보통 12주면 충분해요.")
    if st.sidebar.button("📥 기획 문구 가져오기", width="stretch"):
        if not _has_sa():
            st.sidebar.error("서비스계정이 없어요. Secrets에 gcp_service_account를 추가해 주세요.")
        else:
            try:
                sh_plan = gs_open(st.secrets["gcp_service_account"], _PLAN_SHEET_URL)
                prog = st.sidebar.progress(0.0, text="기획 시트를 읽고 있어요…")
                def _cb(i, total, title):
                    prog.progress(i / max(total, 1), text=f"가져오는 중 {i}/{total} — {title[:16]}")
                lk, read = parse_plan_gsheet(sh_plan, recent=(recent_n or None), progress_cb=_cb)
                prog.empty()
                st.session_state.plan_lookup_gs = lk
                st.session_state.plan_lookup_meta = f"{len(read)}개 주차 · 문구 {len(lk):,}건"
                st.sidebar.success(f"기획 가져오기 완료 — {st.session_state.plan_lookup_meta}")
            except Exception as e:
                st.sidebar.error(f"가져오기에 실패했어요: {str(e)[:90]}")
    if st.session_state.get("plan_lookup_gs") is not None:
        st.sidebar.caption(f"✓ 기획 불러옴: {st.session_state.get('plan_lookup_meta','')}")

    # ── 통합 업로드 자동 분류 → 기존 처리 변수로 라우팅 및 앱푸시 즉시 파싱 ──
    perf_files, promo_files, mtd_files = [], None, []
    if uni_files:
        _LBL = {"perf": "발송실적", "plan": "발송기획", "promo": "기획전성과",
                "mtd": "전사MTD", "push": "📱앱푸시", "unknown": "❓미인식"}
        _perf_b, _promo_b, _mtd_b, _cls = [], [], [], []
        with st.spinner("파일 종류를 파악하고 있어요…"):
            for nm, b in expand_uploads(uni_files):
                if b is None:
                    _cls.append((nm, "unknown")); continue
                k = cached_classify(b)
                _cls.append((nm, k))
                if k == "perf":
                    _perf_b.append((nm, b))
                elif k == "plan":
                    if plan_file is None:
                        plan_file = _UF(nm, b)
                elif k == "promo":
                    _promo_b.append((nm, b))
                elif k == "mtd":
                    _mtd_b.append((nm, b))
                elif k == "push":
                    try:
                        _push_hash = hashlib.md5(b).hexdigest()
                        if st.session_state.get("push_consent_hash") != _push_hash:
                            parsed_df = finalize_push(parse_push_consent_bytes(b))
                            # 기존 누적과 병합 (date+group 키, 신규 우선) — 기간 짧은 새 파일을
                            # 올리면 과거 일자 데이터가 통째로 사라지던 '교체 저장' 문제 방지
                            _cur = st.session_state.get("push_consent_df")
                            if _cur is not None and len(_cur):
                                merged_push = pd.concat([finalize_push(_cur), parsed_df],
                                                        ignore_index=True)
                                merged_push = (merged_push
                                               .drop_duplicates(subset=["date", "group"], keep="last")
                                               .sort_values(["date", "group"]).reset_index(drop=True))
                            else:
                                merged_push = parsed_df
                            st.session_state["push_consent_df"] = merged_push
                            st.session_state["push_consent_hash"] = _push_hash
                            push_consent_df = merged_push
                            if storage_save(BK, "push", merged_push):
                                st.sidebar.success("📱 앱푸시 동의 데이터를 병합 저장했어요 "
                                                   f"(총 {len(merged_push)//3:,}일치).")
                    except Exception as _e:
                        st.sidebar.error(f"앱푸시 파일 읽기 실패: {str(_e)[:80]}")

        perf_files = [_UF(n, b) for n, b in _perf_b]
        promo_files = _UF(_promo_b[0][0], _promo_b[0][1]) if _promo_b else None
        mtd_files = [_UF(n, b) for n, b in _mtd_b]
        from collections import Counter
        _cnt = Counter(k for _, k in _cls)
        st.sidebar.success("자동 분류 — " + " · ".join(
            f"{_LBL[k]} {_cnt[k]}" for k in ("perf", "plan", "promo", "mtd", "push", "unknown") if _cnt.get(k)))
        with st.sidebar.expander("분류 상세"):
            for nm, k in _cls:
                st.caption(f"**{_LBL[k]}** ← {nm[:34]}")
            if _cnt.get("unknown"):
                st.caption("❓인식 못 한 파일이 있어요. 헤더를 확인해 주세요 (실적=’AF코드’ · "
                           "기획전=’기획전 번호’ · 기획=주차 시트 · MTD=날짜행 · 앱푸시=기존 이탈).")

    st.sidebar.caption(BK["status"])
    if st.sidebar.button("🔄 새로 불러오기", width="stretch"):
        for k in ("camp_store", "mtd_store_df", "promo_store_df", "push_consent_df", "push_consent_hash",
                  "_merge_sig", "_merge_new_raw", "_merge_parse_log", "wr_notes"):
            st.session_state.pop(k, None)
        for k in [k for k in st.session_state if str(k).startswith("_load_failed_")]:
            st.session_state.pop(k, None)
        st.cache_data.clear()
    if "camp_store" not in st.session_state:
        st.session_state.camp_store = storage_load(BK, "campaign")
    stored = st.session_state.camp_store
    parse_log = []
    new_raw = None

    if perf_files:
        # 기획 lookup 확보: 업로드 파일(자동 인식) 또는 구글시트 직접연결(세션 적재분)
        plan_lookup = None
        if plan_file:
            try:
                with st.spinner("기획 파일을 읽고 있어요… 처음에만 수십 초 걸려요."):
                    plan_lookup = cached_plan(plan_file.getvalue())
            except Exception as e:
                st.error(f"기획 파일을 읽지 못했어요: {e}"); st.stop()
        if plan_lookup is None:
            plan_lookup = st.session_state.get("plan_lookup_gs")
        if plan_lookup is None:
            st.sidebar.warning("기획 문구가 없어요. 「📥 기획 문구 가져오기」를 눌러 주세요.\n(없으면 기존 데이터만 보여요)")
        else:
            # 같은 파일·기획 조합이면 재파싱/재머지를 건너뛴다.
            # (Streamlit은 페이지 이동마다 스크립트를 재실행 → 저장 후에도 매번 합치던 문제 방지)
            try:
                _perf_sig = tuple((getattr(f, "name", ""), len(f.getvalue())) for f in perf_files)
            except Exception:
                _perf_sig = None
            # 기획 lookup은 건수가 아니라 '내용'으로 시그니처를 만든다 — 문구를 고쳐서
            # 다시 가져와도 건수가 같으면 옛 문구로 머지 캐시가 재사용되던 문제 방지
            try:
                _plan_sig = hash(frozenset(plan_lookup.items()))
            except Exception:
                _plan_sig = len(plan_lookup)
            _merge_sig = (_perf_sig, _plan_sig)
            if (_perf_sig is not None and st.session_state.get("_merge_sig") == _merge_sig
                    and "_merge_new_raw" in st.session_state):
                new_raw = st.session_state["_merge_new_raw"]
                parse_log = st.session_state.get("_merge_parse_log", [])
            else:
                frames = []
                expanded = expand_uploads(perf_files)
                prog = st.sidebar.progress(0.0, text="실적 파일을 합치고 있어요…")
                for k, (nm, b) in enumerate(expanded):
                    prog.progress((k + 1) / max(len(expanded), 1), text=f"합치는 중… {nm[:24]}")
                    if b is None:
                        parse_log.append(f"· {nm}"); continue
                    try:
                        pdf = cached_perf(b)
                        mdf = merge_perf_plan(pdf, plan_lookup, keep_unmatched=True)
                        frames.append(mdf[[c for c in STORE_COLS if c in mdf]])
                        mr = mdf["matched"].mean() * 100 if len(mdf) else 0
                        parse_log.append(f"· {nm[:26]} — {len(mdf)}건 (매칭 {mr:.0f}%)")
                        # 매칭이 저조하면 원인 힌트: 기획에 '날짜 자체'가 없는지(=커버리지) vs
                        # 날짜는 있는데 'AF코드'가 안 맞는지(=키 불일치)를 구분해 로그에 남긴다.
                        if len(mdf) and mr < 50:
                            un = mdf[~mdf["matched"]]
                            plan_dates = {d for (d, a) in plan_lookup}
                            plan_afs = {a for (d, a) in plan_lookup}
                            un_dates = {d for d in un["date"] if d and str(d).lower() not in ("nan", "none")}
                            d_hit = len(un_dates & plan_dates)
                            af_hit = int(un["af"].isin(plan_afs).sum())
                            parse_log.append(
                                f"   ↳ 진단: 미매칭 {len(un)}건 · 기획에 있는 날짜 {d_hit}/{len(un_dates)} · "
                                f"AF코드 기획존재 {af_hit}/{len(un)}")
                    except Exception as e:
                        parse_log.append(f"· {nm[:26]} — 실패: {e}")
                prog.empty()
                if frames:
                    new_raw = pd.concat(frames, ignore_index=True)
                st.session_state["_merge_sig"] = _merge_sig
                st.session_state["_merge_new_raw"] = new_raw
                st.session_state["_merge_parse_log"] = parse_log

    # 누적 병합 — 저장 전까지는 미리보기(영구 반영 X)
    if new_raw is not None and len(new_raw):
        work = merge_store(stored, new_raw)
        ko = set(map(tuple, store_key_frame(stored).values)) if len(stored) else set()
        kn = set(map(tuple, store_key_frame(new_raw).values))
        st.sidebar.success(f"신규 {len(new_raw)}건 → 누적 {len(work)}건 "
                           f"(추가 {len(kn - ko)} · 갱신 {len(kn & ko)})")
        if st.sidebar.button("💾 저장하기", width="stretch"):
            # 낙관적 병합: 저장 직전 시트를 다시 읽어 합친다 — 두 세션이 각자 스냅샷으로
            # 저장하면 나중 저장이 먼저 저장분을 지우던 last-write-wins 유실 방지
            _tosave = merge_store(storage_load(BK, "campaign"), work)
            if storage_save(BK, "campaign", _tosave):
                st.session_state.camp_store = _tosave
                tgt = "구글시트" if BK["mode"] == "gsheets" else "로컬"
                st.sidebar.success(f"저장했어요 ✓ ({tgt}, 총 {len(_tosave):,}건) — 다음에도 유지돼요.")
        st.sidebar.caption("※ 저장을 눌러야 반영돼요. 안 누르면 이번 세션에서만 볼 수 있어요.")
    else:
        work = stored

    def _apply_backup(file):
        """통합 백업(ZIP) 또는 단일 캠페인 백업(CSV)을 복원. 복원 요약 문자열 반환."""
        name = (getattr(file, "name", "") or "").lower()
        done = []
        if name.endswith(".zip"):
            import zipfile as _zf
            z = _zf.ZipFile(file)
            for nm in z.namelist():
                base = nm.split("/")[-1].lower()
                if not base.endswith(".csv"):
                    continue
                data = io.BytesIO(z.read(nm))
                if "mtd" in base:
                    df = pd.read_csv(data, encoding="utf-8-sig")
                    _cur = st.session_state.get("mtd_store_df")
                    if _cur is None:                      # 빈 화면 복원 경로 — 아직 로드 전이면
                        _cur = storage_load(BK, "mtd")    # 기존 저장분을 읽어와 병합(덮어쓰기 방지)
                    m = merge_mtd_store(_cur, df)
                    if storage_save(BK, "mtd", m):
                        st.session_state.mtd_store_df = m
                        done.append(f"MTD {len(m):,}")
                elif "promo" in base:
                    df = pd.read_csv(data, encoding="utf-8-sig", dtype={"promo": str})
                    _cur = st.session_state.get("promo_store_df")
                    if _cur is None:
                        _cur = storage_load(BK, "promo")
                    m = merge_promo_store(_cur, df)
                    if storage_save(BK, "promo", m):
                        st.session_state.promo_store_df = m
                        done.append(f"기획전 {len(m):,}")
                elif "note" in base:
                    df = pd.read_csv(data, encoding="utf-8-sig", dtype=str).fillna("")
                    try:
                        _cur_n = notes_df_to_dict(storage_load(BK, "notes"))
                    except Exception:
                        _cur_n = {}
                    merged_n = {**_cur_n, **notes_df_to_dict(df)}      # 같은 키는 백업 우선
                    if storage_save(BK, "notes", notes_dict_to_df(merged_n)):
                        st.session_state.wr_notes = merged_n
                        done.append(f"보고란 {len(merged_n):,}")
                elif "push" in base or "consent" in base:
                    # dtype 정규화 후 병합 — 문자열/Timestamp 혼재로 dedupe가 실패해
                    # 같은 날이 두 벌 남던 문제 방지
                    df = finalize_push(pd.read_csv(data, encoding="utf-8-sig", dtype={"group": str}))
                    cur_push = st.session_state.get("push_consent_df")
                    if cur_push is not None and not cur_push.empty:
                        m = pd.concat([finalize_push(cur_push), df], ignore_index=True)
                        m = m.drop_duplicates(subset=["date", "group"], keep="last")
                    else:
                        m = df
                    if storage_save(BK, "push", m):
                        st.session_state.push_consent_df = m
                        done.append(f"앱푸시 {len(m)//3:,}일")
                else:
                    df = pd.read_csv(data, encoding="utf-8-sig", dtype={"date": str, "af": str})
                    m = merge_store(stored, df)
                    if storage_save(BK, "campaign", m):
                        st.session_state.camp_store = m
                        done.append(f"캠페인 {len(m):,}")
        else:
            df = pd.read_csv(file, encoding="utf-8-sig", dtype={"date": str, "af": str})
            m = merge_store(stored, df)
            if storage_save(BK, "campaign", m):
                st.session_state.camp_store = m
                done.append(f"캠페인 {len(m):,}")
        st.cache_data.clear()
        return " · ".join(done) if done else "복원할 데이터가 없어요"

    if work is None or len(work) == 0:
        st.title("LF몰 CRM 발송성과 대시보드")
        st.markdown("""
        <div class="vg">
        <b>기획 문구</b>와 <b>발송 실적</b>을 합쳐서 <b>어떤 문구·오퍼·타이밍이 성과를 만드는지</b> 알 수 있어요.<br><br>
        왼쪽 <b>📂 파일 올리기</b>에 <b>발송실적</b>과 <b>기획 통합본</b>을 같이 올리면 자동으로 인식돼요.
        <b>「저장하기」</b>를 누르면 쌓여요.<br>
        · 실제로 발송한 건만 분석해요. 기획만 있고 발송 안 한 건은 빠져요.<br>
        · 한 번 저장하면 다음부터는 <b>새 주차 실적만</b> 올리면 돼요.
        </div>""", unsafe_allow_html=True)
        st.markdown("##### 💾 백업으로 바로 불러오기")
        st.caption("이전에 받아둔 ‘통합 백업(ZIP)’ 또는 ‘누적 백업(CSV)’을 올리면 재업로드·머지 없이 바로 대시보드가 떠요.")
        _up = st.file_uploader("백업 올리기 (ZIP/CSV)", type=["zip", "csv"], key="restore_empty")
        if _up is not None:
            _sig = (_up.name, getattr(_up, "size", None))
            if st.session_state.get("_restored_sig") != _sig:
                try:
                    summary = _apply_backup(_up)
                    st.session_state["_restored_sig"] = _sig
                    st.success(f"불러왔어요 ✓ {summary}")
                    st.rerun()
                except Exception as e:
                    st.error(f"불러오지 못했어요: {e}")
        st.stop()

    # 작업 데이터 확정: 파생 재계산 + 타입 정리 + 문구 태깅 (캐시 — 성능 핵심)
    raw = prepare_raw(work, TAGSET_VER)

    # ── 전사 MTD (발송피로도) 누적 처리 ──
    if "mtd_store_df" not in st.session_state:
        st.session_state.mtd_store_df = storage_load(BK, "mtd")
    mtd_stored = st.session_state.mtd_store_df
    mtd_new = None
    if mtd_files:
        mframes = []
        for nm, b in expand_uploads(mtd_files):
            if b is None:
                continue
            try:
                md = cached_mtd(b)
                if len(md):
                    mframes.append(md[[c for c in MTD_STORE_COLS if c in md]])
                parse_log.append(f"· [MTD] {nm[:22]} — {len(md)}일")
            except Exception as e:
                parse_log.append(f"· [MTD] {nm[:22]} — 실패: {e}")
        if mframes:
            mtd_new = pd.concat(mframes, ignore_index=True)
    if mtd_new is not None and len(mtd_new):
        mtd_work = merge_mtd_store(mtd_stored, mtd_new)
        st.sidebar.success(f"MTD 새로 {len(mtd_new)}일 → 합쳐서 {len(mtd_work)}일 (미리보기)")
        if st.sidebar.button("💾 MTD 저장하기", width="stretch", key="save_mtd"):
            _tosave = merge_mtd_store(storage_load(BK, "mtd"), mtd_work)   # 낙관적 병합
            if storage_save(BK, "mtd", _tosave):
                st.session_state.mtd_store_df = _tosave
                tgt = "구글시트" if BK["mode"] == "gsheets" else "로컬"
                st.sidebar.success(f"MTD 저장했어요 ✓ ({tgt})")
    else:
        mtd_work = mtd_stored
    mtd_data = cached_compute_mtd(mtd_work) if (mtd_work is not None and len(mtd_work) >= 10) else None

    # ── 기획전 성과시트 (기획전번호별 매출) 누적 처리 ──
    if "promo_store_df" not in st.session_state:
        st.session_state.promo_store_df = storage_load(BK, "promo")
    promo_stored = st.session_state.promo_store_df
    promo_work = promo_stored
    if promo_files is not None:
        try:
            pnew = cached_promo(promo_files.getvalue())
            promo_work = merge_promo_store(promo_stored, pnew)
            st.sidebar.success(f"기획전 시트 {len(pnew):,}건 → 합쳐서 {len(promo_work):,}건 (미리보기)")
            if st.sidebar.button("💾 기획전 저장하기", width="stretch", key="save_promo"):
                _tosave = merge_promo_store(storage_load(BK, "promo"), promo_work)   # 낙관적 병합
                if storage_save(BK, "promo", _tosave):
                    st.session_state.promo_store_df = _tosave
                    tgt = "구글시트" if BK["mode"] == "gsheets" else "로컬"
                    st.sidebar.success(f"기획전 저장했어요 ✓ ({tgt})")
            st.sidebar.caption("※ 저장을 눌러야 반영돼요.")
        except Exception as e:
            st.sidebar.error(f"기획전 시트를 읽지 못했어요: {str(e)[:90]}")
    promo_df = finalize_promo(promo_work)

    # ── 필터 (카테고리별) ──
    st.sidebar.markdown("---")
    st.sidebar.markdown("#### 조건 설정")

    def _opts(d, col):
        if col not in d:
            return []
        v = d[col].dropna().astype(str).str.replace(r'\.0$', '', regex=True)
        return sorted([x for x in v.unique() if x.strip() not in ("", "nan", "NaN", "None")])

    def _apply_in(d, col, sel):
        if not sel or col not in d:
            return d
        return d[d[col].astype(str).str.replace(r'\.0$', '', regex=True).isin(sel)]

    # 필터 위젯 키 목록 — 「필터 전체 초기화」가 한 번에 리셋할 대상
    _FLT_KEYS = ("flt_q", "flt_matched", "flt_minsend", "flt_date", "flt_st", "flt_target",
                 "flt_bpu", "flt_prio", "flt_hour", "flt_dow", "flt_cat", "flt_attr",
                 "flt_owner", "flt_tags", "tags_mode")
    if st.sidebar.button("↩️ 필터 전체 초기화", width="stretch", key="flt_reset",
                         help="검색·기간·속성 등 아래 모든 필터를 기본값으로 되돌려요"):
        for _k in _FLT_KEYS:
            st.session_state.pop(_k, None)
        st.rerun()

    search = st.sidebar.text_input("🔎 검색", "", key="flt_q",
                                   help="제목·내용·브랜드·AF코드·카테고리에서 찾아요")
    only_matched = st.sidebar.checkbox("문구 매칭된 것만", value=True, key="flt_matched",
                                       help="문구(제목·내용)를 못 찾은 건 빼요")
    min_send = st.sidebar.number_input(
        "최소 발송수", value=5000, step=1000, min_value=0, key="flt_minsend",
        help="이 숫자 미만의 캠페인은 분석에서 빠져요. 발송이 너무 적으면 우연에 흔들리거든요. "
             "낮추면 더 많은 캠페인이 포함돼요.")

    base_opt = raw[raw["matched"]] if only_matched else raw

    date_sel = None
    with st.sidebar.expander("📅 기간"):
        dts = raw["dt"].dropna()
        if len(dts):
            dmin, dmax = dts.min().date(), dts.max().date()
            if dmin < dmax:
                date_sel = st.date_input("발송일 범위", value=(dmin, dmax),
                                         min_value=dmin, max_value=dmax, key="flt_date")
                if date_sel is not None and not (isinstance(date_sel, (tuple, list))
                                                 and len(date_sel) == 2):
                    st.caption("⚠️ 종료일까지 골라야 기간 필터가 적용돼요 (지금은 미적용).")
            else:
                st.caption(f"단일 일자: {dmin}")

    with st.sidebar.expander("📤 발송 속성"):
        sel_st = st.multiselect("발송유형", _opts(base_opt, "stype"), key="flt_st")
        sel_target = st.multiselect("타겟 구분", _opts(base_opt, "target"), key="flt_target",
                                    help="신규·휴면·전체 등 누구에게 보냈는지")
        sel_bpu = st.multiselect("BPU(사업부)", _opts(base_opt, "bpu"), key="flt_bpu")
        sel_prio = st.multiselect("우선순위", _opts(base_opt, "prio"), key="flt_prio",
                                  help="같은 시간에 몇 번째로 보냈는지 (1=가장 먼저)")

    def _hm_label(v):  # 시간대 HHMM 문자열 → 'HH시MM분' (예: 1030→10시30분, 800→08시00분)
        try:
            n = int(float(v))
        except Exception:
            return str(v)
        hh, mm = (n, 0) if n <= 23 else divmod(n, 100)
        return f"{hh:02d}시{mm:02d}분" if (0 <= hh <= 23 and 0 <= mm <= 59) else str(v)

    with st.sidebar.expander("🕒 발송 시점"):
        _hour_opts = sorted(_opts(base_opt, "hour"),
                            key=lambda v: int(float(v)) if str(v).replace(".", "", 1).isdigit() else 10 ** 9)
        sel_hour = st.multiselect("시간대", _hour_opts, format_func=_hm_label, key="flt_hour")
        sel_dow = st.multiselect("요일", _opts(base_opt, "dow_k"), key="flt_dow")

    with st.sidebar.expander("🏷️ 상품·담당"):
        sel_cat = st.multiselect("카테고리", _opts(base_opt, "cat"), key="flt_cat")
        sel_attr = st.multiselect("대상 속성", _opts(base_opt, "attr"), key="flt_attr",
                                  help="통합·정상·이월·입점·마케팅 등")
        sel_owner = st.multiselect("담당자", _opts(base_opt, "owner"), key="flt_owner")

    with st.sidebar.expander("✍️ 문구 속성"):
        sel_tags = st.multiselect("소구 속성", TAG_BOOLS, key="flt_tags",
                                  help="할인율·마감임박 등 제목+내용에서 자동으로 분류한 속성이에요")
        tags_and = True
        if sel_tags:
            tags_and = st.radio("조건", ["모두 충족(AND)", "하나라도(OR)"], horizontal=True,
                                key="tags_mode",
                                help="AND: 선택한 속성을 다 가진 것만 / OR: 하나라도 가진 것") \
                == "모두 충족(AND)"

    CATSEL = {"stype": sel_st, "target": sel_target, "bpu": sel_bpu, "hour": sel_hour,
              "dow_k": sel_dow, "prio": sel_prio, "cat": sel_cat,
              "attr": sel_attr, "owner": sel_owner}

    def apply_filters(d):
        d = d.copy()
        if date_sel and isinstance(date_sel, (tuple, list)) and len(date_sel) == 2 and "dt" in d:
            lo, hi = pd.Timestamp(date_sel[0]), pd.Timestamp(date_sel[1]) + pd.Timedelta(days=1)
            d = d[(d["dt"] >= lo) & (d["dt"] < hi)]
        for col, sel in CATSEL.items():
            d = _apply_in(d, col, sel)
        present_tags = [t for t in sel_tags if t in d.columns]
        if present_tags:
            if tags_and:
                for t in present_tags:
                    d = d[d[t]]
            else:
                mask = np.logical_or.reduce([d[t].values for t in present_tags])
                d = d[mask]
        if search.strip():
            q = search.strip().lower()
            hay = (d["title"].astype(str) + " " + d["body"].astype(str) + " " +
                   d["brand"].astype(str) + " " + d["af"].astype(str) + " " +
                   d["cat"].astype(str)).str.lower()
            d = d[hay.str.contains(q, na=False, regex=False)]
        return d

    dff_all = apply_filters(raw)
    df = dff_all[dff_all["matched"]] if only_matched else dff_all
    fdf = df[df["send"].fillna(0) >= min_send].reset_index(drop=True)

    # 활성 필터 요약 — 결과가 0건일 때 '왜 0건인지'를 바로 알 수 있게
    _active_flt = []
    if search.strip():
        _active_flt.append(f"검색 '{search.strip()[:12]}'")
    for _lab, _sel in (("발송유형", sel_st), ("타겟", sel_target), ("BPU", sel_bpu),
                       ("우선순위", sel_prio), ("시간대", sel_hour), ("요일", sel_dow),
                       ("카테고리", sel_cat), ("대상속성", sel_attr), ("담당자", sel_owner),
                       ("소구", sel_tags)):
        if _sel:
            _active_flt.append(f"{_lab} {len(_sel)}개")
    _dsel = st.session_state.get("flt_date")
    if (isinstance(_dsel, (tuple, list)) and len(_dsel) == 2 and len(dts)
            and (pd.Timestamp(_dsel[0]).date(), pd.Timestamp(_dsel[1]).date()) != (dmin, dmax)):
        _active_flt.append("기간 지정")
    if _active_flt:
        st.sidebar.caption("🎛 활성 필터: " + " · ".join(_active_flt))
    if len(dff_all) == 0:
        st.warning("현재 필터 조합에 맞는 캠페인이 **0건**이에요"
                   + (f" — 활성 필터: {' · '.join(_active_flt)}" if _active_flt else "")
                   + ". 조건을 풀거나 사이드바 「↩️ 필터 전체 초기화」를 눌러 주세요.")

    st.sidebar.markdown("---")
    # 연관 주제를 그룹으로 묶고(사이드바), 그룹 안 여러 주제는 본문 상단 하위탭으로 전환.
    # ▸ value 리스트의 각 항목 문자열은 아래 페이지 분기(if "..." in page)의 매칭 키를 포함해야 함.
    # 단일 내비게이션 — 피로도 진단(구 F1~F4)을 '6. 효율·피로도' 하위탭으로 통합,
    # BPU·우선순위는 성격이 같은 '4. 성과 진단'(차원별 랭킹)으로 이동
    CAMPAIGN_GROUPS = {
        "0. 주간보고":           ["주간보고"],
        "1. 종합 요약":          ["종합 요약"],
        "2. 문구 분석":          ["문구 속성별 성과", "키워드·이모지 성과", "소구 추세·마모"],
        "3. 캠페인 리더보드":    ["캠페인 리더보드"],
        "4. 성과 진단":          ["전환·AOV 진단", "발송유형·브랜드 랭킹", "BPU·우선순위 효율"],
        "5. 맥락·타이밍":        ["카테고리·시간대", "타이밍·발송슬롯"],
        "6. 효율·피로도":        ["전체 효율·추이", "피로도 시계열", "발송 빈도·한계수익", "요일 패턴"],
        "7. 기획전 비교분석":    ["기획전 비교분석"],
        "8. 액션":               ["다음주 발송 플레이북", "AI 처방·카피"],
        "9. 데이터·다운로드":    ["데이터·다운로드"],
        "10. 앱푸시 동의 현황":  ["앱푸시 동의 현황"],
    }
    _grp = st.sidebar.radio("페이지", list(CAMPAIGN_GROUPS))
    _subs = CAMPAIGN_GROUPS[_grp]
    if len(_subs) > 1:
        _gname = _grp.split(". ", 1)[-1]
        # 본문 상단 하위탭 (기획전 비교분석 페이지의 하위탭과 동일한 위치/역할)
        page = st.radio(_gname, _subs, horizontal=True, key=f"subtab_{_grp}",
                        label_visibility="collapsed")
    else:
        page = _subs[0]
    if _grp.startswith("6.") and mtd_data is None:
        st.sidebar.info("피로도 하위탭(시계열·빈도·요일)은 전사 MTD 파일을 올리면 볼 수 있어요.")
    _model_keys = list(AI_MODELS.keys())
    _default_model = "Gemini 2.5 Flash (균형)"
    model_name = st.sidebar.selectbox(
        "AI 모델", _model_keys,
        index=_model_keys.index(_default_model) if _default_model in _model_keys else 0)
    model = AI_MODELS[model_name]

    st.sidebar.markdown("---")
    st.sidebar.markdown("#### 현황")
    drange = "—"
    _ds = [x for x in raw["date"].dropna().astype(str).str.strip().unique()
           if x and x.lower() not in ("nan", "none")]
    if _ds:
        _ds = sorted(_ds)
        drange = f"{_ds[0]} ~ {_ds[-1]}"
    st.sidebar.caption(f"전체 {len(raw)}건 · 매칭 {raw['matched'].mean()*100:.0f}% · 분석 {len(fdf)}건\n\n{drange}")
    # 최신 주 매칭률 경보 — 전체 %는 과거 누적에 묻혀서 최근 붕괴를 못 보여준다
    _rw = raw.dropna(subset=["dt"]) if "dt" in raw else raw.iloc[0:0]
    if len(_rw):
        _lastwk = _rw["dt"].max().to_period("W").start_time
        _lw = _rw[_rw["dt"] >= _lastwk]
        if len(_lw) >= 5 and _lw["matched"].mean() < 0.7:
            st.sidebar.warning(f"⚠️ 최신 주 매칭률 {_lw['matched'].mean()*100:.0f}% — "
                               "기획 시트 적재·형식(날짜/AF코드)을 확인해 주세요. "
                               "상세는 「9. 데이터·다운로드」의 매칭 품질 참고.")
    if parse_log:
        with st.sidebar.expander("파싱 로그"):
            st.text("\n".join(parse_log))

    with st.sidebar.expander("데이터 관리"):
        _n_saved = len(stored) if stored is not None else 0
        _n_new = len(new_raw) if new_raw is not None else 0
        st.caption(f"전체 {len(work):,}건 = 저장됨 {_n_saved:,} + 이번 세션 {_n_new:,}")
        if _n_new and not _n_saved:
            st.warning("저장된 데이터가 0건이에요. 「💾 저장하기」를 눌러야 다음에도 유지돼요.")

        # ── 통합 백업: 캠페인+MTD+기획전+앱푸시 데이터를 한 ZIP 파일로 ──
        # ZIP 생성은 버튼 클릭 시에만 — expander가 접혀 있어도 매 rerun마다 전체 데이터를
        # CSV 직렬화+압축하던 것이 모든 페이지 상호작용을 느리게 만들던 문제 방지
        st.markdown("##### 📁 통합 백업 (단일 파일)")
        import zipfile as _zf
        _push_df = st.session_state.get("push_consent_df")
        # 백업 신선도 시그니처 — 행 '개수'만 보면 dedup 덮어쓰기(개수 불변)나 노트 수정을
        # 놓쳐 옛 ZIP을 최신인 양 내려준다. 내용 해시 + notes 포함으로 실제 변경을 감지.
        def _dfsig(d):
            if d is None or len(d) == 0:
                return 0
            try:
                return int(pd.util.hash_pandas_object(d.astype(str), index=False).sum())
            except Exception:
                return len(d)
        try:
            _notes_now = st.session_state.get("wr_notes") or {}
            _notes_sig = hash(frozenset((k, str(v)) for k, v in _notes_now.items()))
        except Exception:
            _notes_sig = 0
        _bak_sig = (len(work), _dfsig(work),
                    _dfsig(mtd_work), _dfsig(promo_work), _dfsig(_push_df), _notes_sig)
        if st.button("📦 백업 파일 만들기", width="stretch", key="bak_make"):
            _zbuf = io.BytesIO()
            _has_any = False
            with _zf.ZipFile(_zbuf, "w", _zf.ZIP_DEFLATED) as _z:
                if len(work):
                    _z.writestr("campaign.csv",
                        work[[c for c in STORE_COLS if c in work]].to_csv(index=False).encode("utf-8-sig"))
                    _has_any = True
                if mtd_work is not None and len(mtd_work):
                    _z.writestr("mtd.csv",
                        mtd_work[[c for c in MTD_STORE_COLS if c in mtd_work]].to_csv(index=False).encode("utf-8-sig"))
                    _has_any = True
                if promo_work is not None and len(promo_work):
                    _z.writestr("promo.csv",
                        promo_work[[c for c in PROMO_STORE_COLS if c in promo_work]].to_csv(index=False).encode("utf-8-sig"))
                    _has_any = True
                if _push_df is not None and not _push_df.empty:
                    _z.writestr("push_consent.csv",
                        _push_df.to_csv(index=False).encode("utf-8-sig"))
                    _has_any = True
                # 보고란(notes)도 백업에 포함 — 재배포(로컬 CSV 초기화) 시 유일한 복구 수단
                try:
                    _nd = st.session_state.get("wr_notes")
                    _notes_df = notes_dict_to_df(_nd) if _nd else storage_load(BK, "notes")
                    if _notes_df is not None and len(_notes_df):
                        _z.writestr("notes.csv", _notes_df.to_csv(index=False).encode("utf-8-sig"))
                        _has_any = True
                except Exception:
                    pass
            st.session_state["_bak_zip"] = (_bak_sig, _zbuf.getvalue()) if _has_any else None
        _bak_cached = st.session_state.get("_bak_zip")
        if _bak_cached:
            if _bak_cached[0] != _bak_sig:
                st.caption("⚠️ 데이터가 바뀌었어요 — 「📦 백업 파일 만들기」를 다시 눌러 최신본을 받으세요.")
            st.download_button(
                "📥 통합 백업 (전체 ZIP)", _bak_cached[1],
                file_name=f"lf_dashboard_backup_{today_kst():%Y%m%d}.zip", mime="application/zip",
                width="stretch", key="bak_all")
        st.caption("캠페인·MTD·기획전·앱푸시 수신동의 데이터를 모두 포함하여 백업합니다. 이 ZIP 파일을 아래에 다시 올리면 원클릭으로 일괄 복원됩니다.")
        _rest_all = st.file_uploader("통합 백업 복원하기 (ZIP/CSV)", type=["zip", "csv"], key="restore_all",
                                     help="통합 백업(ZIP) 또는 예전 캠페인 백업(CSV) 모두 올릴 수 있어요.")
        if _rest_all is not None:
            _sig = (_rest_all.name, getattr(_rest_all, "size", None))
            if st.session_state.get("_restored_sig") != _sig:
                try:
                    summary = _apply_backup(_rest_all)
                    st.session_state["_restored_sig"] = _sig
                    st.success(f"복원 완료 ✓ {summary} — 바로 표시합니다.")
                    st.rerun()
                except Exception as e:
                    st.error(f"복원하지 못했어요: {e}")

        st.markdown("---")
        # ── 지저분한 개별 제어기는 서브 expander 안으로 숨겨서 정돈 ──
        with st.expander("🛠️ 개별 데이터 설정 및 삭제", expanded=False):
            st.markdown("##### 개별 백업 (선택)")
            if len(work):
                st.download_button(
                    f"📥 캠페인 백업 (CSV · {len(work):,}건)",
                    work[[c for c in STORE_COLS if c in work]].to_csv(index=False).encode("utf-8-sig"),
                    file_name=f"send_perf_store_backup_{datetime.date.today():%Y%m%d}.csv", mime="text/csv", width="stretch")
            if st.button("🧹 캠페인 저장소 초기화", width="stretch", key="clear_store"):
                storage_clear(BK, "campaign")
                st.session_state.camp_store = pd.DataFrame(columns=STORE_COLS)
                st.cache_data.clear()
                st.success("지웠어요. 새로고침해 주세요.")

            st.markdown("---")
            st.caption(f"전사 MTD {0 if mtd_work is None else len(mtd_work)}일치")
            if mtd_work is not None and len(mtd_work):
                st.download_button(
                    "📥 MTD 백업 (CSV)",
                    mtd_work[[c for c in MTD_STORE_COLS if c in mtd_work]].to_csv(index=False).encode("utf-8-sig"),
                    file_name="send_perf_mtd_backup.csv", mime="text/csv",
                    width="stretch", key="mtd_bak")
            if st.button("🧹 MTD 저장소 초기화", width="stretch", key="clear_mtd"):
                storage_clear(BK, "mtd")
                st.session_state.mtd_store_df = pd.DataFrame(columns=MTD_STORE_COLS)
                st.cache_data.clear()
                st.success("MTD를 지웠어요. 새로고침해 주세요.")

            st.markdown("---")
            st.caption(f"기획전 성과 {0 if promo_work is None else len(promo_work):,}건")
            if promo_work is not None and len(promo_work):
                st.download_button(
                    "📥 기획전 백업 (CSV)",
                    promo_work[[c for c in PROMO_STORE_COLS if c in promo_work]].to_csv(index=False).encode("utf-8-sig"),
                    file_name="send_perf_promo_backup.csv", mime="text/csv",
                    width="stretch", key="promo_bak")
            rest_p = st.file_uploader("기획전 백업 CSV로 복원하기", type=["csv"], key="restore_promo")
            if rest_p is not None:
                try:
                    dp = pd.read_csv(rest_p, encoding="utf-8-sig", dtype={"promo": str})
                    merged_p = merge_promo_store(st.session_state.get("promo_store_df"), dp)
                    if storage_save(BK, "promo", merged_p):
                        st.session_state.promo_store_df = merged_p
                        st.cache_data.clear()
                        st.success("기획전 복원했어요 ✓ 새로고침해 주세요")
                except Exception as e:
                    st.error(f"기획전 복원하지 못했어요: {e}")
            if st.button("🧹 기획전 저장소 초기화", width="stretch", key="clear_promo"):
                storage_clear(BK, "promo")
                st.session_state.promo_store_df = pd.DataFrame(columns=PROMO_STORE_COLS)
                st.cache_data.clear()
                st.success("기획전을 지웠어요. 새로고침해 주세요.")

            st.markdown("---")
            _p_df = st.session_state.get("push_consent_df")
            st.caption(f"앱푸시 동의 {0 if _p_df is None else len(_p_df):,}건 (약 {0 if _p_df is None else len(_p_df)//3:,}일치)")
            if _p_df is not None and not _p_df.empty:
                st.download_button(
                    "📥 앱푸시 백업 (CSV)",
                    _p_df.to_csv(index=False).encode("utf-8-sig"),
                    file_name="send_perf_push_backup.csv", mime="text/csv",
                    width="stretch", key="push_bak"
                )
            if st.button("🧹 앱푸시 저장소 초기화", width="stretch", key="clear_push"):
                storage_clear(BK, "push")
                st.session_state.push_consent_df = pd.DataFrame(columns=PUSH_STORE_COLS)
                st.session_state.pop("push_consent_hash", None)
                st.cache_data.clear()
                st.success("앱푸시 데이터를 영구히 지웠어요. 새로고침해 주세요.")
        st.markdown("---")
        st.caption(f"저장 위치: {BK['status']}")
        if BK["mode"] != "gsheets":
            st.caption("구글시트를 쓰려면 Secrets에 `gcp_service_account`와 "
                       "`[gsheets] spreadsheet`를 넣고, 시트를 서비스계정 이메일에 **편집자**로 공유해 주세요.")
        st.markdown("##### 📥 기획 문구 구글시트 연결")
        try:
            _sa_email = st.secrets["gcp_service_account"].get("client_email", "(secrets 확인)")
        except Exception:
            _sa_email = "(서비스계정 미설정)"
        st.markdown(
            "★APP PUSH 시트에서 기획 문구를 바로 읽어와요:<br>"
            f"1) 구글시트가 <code>{_sa_email}</code>에 **뷰어**로 공유되어 있어야 해요<br>"
            "2) 사이드바에서 <b>📥 기획 문구 가져오기</b> 버튼만 누르면 끝!<br>"
            "<span style='color:#94a3b8'>※ 시트가 많으면 '최근 주차 수'로 줄여 주세요. "
            "보통 새 주차만 가져오면 충분해요.</span>",
            unsafe_allow_html=True)

    # 지표 메타
    METRIC_OPTS = {
        "주문전환율": ("ord_cr", "%", PALETTE["purple"]),
        "CTR(유입전환율)": ("infl_cr", "%", PALETTE["blue"]),
        "발송건당거래액(RPS)": ("rps", "원", PALETTE["green"]),
        "객단가(AOV)": ("aov", "원", PALETTE["amber"]),
        "거래액": ("amt", "원", PALETTE["teal"]),
    }

    # 전환율은 캠페인별 단순 평균으로 비교(표본은 사이드바 '최소 발송수'로 통제)

    def glossary(which="full"):
        """비전문가용 지표·발송·통계 용어 주석 (접이식) — 모든 페이지 하단에 호출."""
        metrics_md = (
            "**📊 성과 지표**\n"
            "- **발송**: 메시지를 보낸 건수예요.\n"
            "- **UV(유입)**: 메시지를 눌러 들어온 사람 수예요(중복 제거).\n"
            "- **VISIT**: 들어온 사람들이 발생시킨 방문 횟수예요.\n"
            "- **CTR(유입전환율)** = UV ÷ 발송. 보낸 것 중 몇 %가 들어왔는지예요.\n"
            "- **주문전환율(CR)** = 주문 ÷ UV. 들어온 사람 중 몇 %가 샀는지예요.\n"
            "- **RPS** = 거래액 ÷ 발송. 1건 보냈을 때 평균 매출이에요.\n"
            "- **객단가(AOV)** = 거래액 ÷ 주문. 주문 1건에 평균 얼마를 썼는지예요.\n"
            "- **거래액**: 발송을 통해 발생한 주문 금액 합계예요.\n")
        biz_md = (
            "**📨 발송·캠페인 용어**\n"
            "- **AF코드**: 캠페인(발송 소재) 하나를 식별하는 코드예요 (예: AP101).\n"
            "- **BPU**: 사업부 단위예요. 1BPU·2BPU처럼 발송 주체를 구분해요.\n"
            "- **우선순위(순번)**: 같은 시간대에 몇 번째로 나간 발송인지예요. 1이 가장 먼저예요.\n"
            "- **소구(訴求)**: 문구가 어필하는 포인트예요 — 할인율·마감임박·쿠폰 같은 혜택형부터 "
            "신상컬렉션·베스트인기·추천큐레이션·시즌환절기·행동리타겟(장바구니·찜)·이벤트응모·"
            "라이브방송 같은 비혜택형까지 자동 분류돼요.\n"
            "- **세그먼트(타겟)**: 누구에게 보냈는지 그룹이에요 — 신규·휴면·전체 등.\n"
            "- **퍼널**: 발송→유입→주문으로 좁아지는 단계 흐름이에요. 급락 단계가 개선 포인트예요.\n"
            "- **기여율** = 발송 추적 거래액 ÷ 기획전 유입 거래액. 기획전 매출 중 발송이 끌어온 비중이에요.\n"
            "- **리프트**: 어떤 속성이 있을 때와 없을 때의 성과 차이예요.\n"
            "- **마모**: 같은 소구를 반복할수록 반응이 무뎌지는 현상이에요.\n"
            "- **MTD·인당 발송 건수**: 전사 기준 일별 집계예요. 인당 발송 건수는 고객 1명이 "
            "하루에 받은 평균 메시지 수(발송 강도)예요.\n"
            "- **기말(期末)**: '주말(weekend)'이 아니라 **그 기간의 마지막 날**이라는 뜻이에요. "
            "'기말 동의수'는 그 주 마지막 날 기준 누적 값(스냅샷)이고, 신규추가·기존이탈·순증감처럼 "
            "그 주를 다 더한 '합계'와는 달라요.\n"
            "- **전주비·전월비·전년비**: 기준 기간을 직전 주·전월 동주(한 달 전 같은 주)·작년 같은 "
            "기간과 비교한 증감이에요. 비율 지표(CTR·CR)는 %p 차이, 나머지는 증감률(%)로 표시해요.\n"
            "- **△ 표기**: 마이너스(감소)를 뜻해요 — 회사 보고 양식이에요. 예: △8.9% = 8.9% 감소.\n")
        stats_md = (
            "**🔬 통계 용어**\n"
            "- **p값·유의성**: 이 차이가 우연일 가능성이에요. 0.05 미만이면 보통 '진짜 차이'로 봐요.\n"
            "- **효과크기(d)**: 차이가 실제로 얼마나 큰지예요. 0.2 작음 · 0.5 중간 · 0.8 큼.\n"
            "- **보정(FDR)**: 여러 개를 한꺼번에 비교하면 우연히 '유의'가 나오기 쉬워서 "
            "더 엄격하게 걸러낸 값이에요.\n"
            "- **순효과**: 카테고리·시간대 등 다른 조건을 맞춘 뒤 그 요소만의 진짜 기여예요.\n"
            "- **상관 r**: 둘이 같이 움직이는 정도예요(−1\\~+1). +면 같이, −면 반대로 움직여요. "
            "상관은 인과(원인→결과)와 달라요.\n"
            "- **R²(결정계수)**: 추세선이 데이터를 얼마나 잘 설명하는지예요(0~1). 1에 가까울수록 뚜렷해요.\n"
            "- **중앙값**: 크기순으로 줄 세웠을 때 딱 가운데 값이에요. 극단값에 안 휘둘려요.\n"
            "- **±2σ(시그마)**: 평균에서 아주 많이 벗어났다는 뜻이에요. 상·하위 약 2.5%에 해당해요.\n"
            "- **5분위(Q1\\~Q5)**: 데이터를 크기순으로 5등분한 거예요. Q1이 가장 작고 Q5가 가장 커요.\n"
            "- **가중 평균 vs 단순 평균**: 가중은 발송 많은 캠페인에 비중을 더 두고(실제 효율에 가까움), "
            "단순은 캠페인 1건을 1표로 봐요.\n"
            "- **최소 발송수**: 표본이 너무 적으면 우연에 흔들려서, 일정 발송수 이상만 분석에 넣어요.\n")
        with st.expander("📖 용어 주석 — 어려운 용어는 여기서 확인하세요"):
            g1, g2 = st.columns(2)
            g1.markdown(metrics_md)
            g1.markdown(biz_md)
            g2.markdown(stats_md)

    @st.fragment
    def render_messages(d, mcol, key, n=200):
        """선택 구간/속성에 해당하는 실제 발송 메시지 + 성과 표 + 원문 보기.

        fragment — 행 클릭·원문 선택 때 전체 스크립트(필터·사이드바·전 페이지)가
        다시 돌지 않고 이 블록만 다시 그린다. 자기 완결 블록(외부 컨테이너 미사용)."""
        if d is None or len(d) == 0:
            st.info("해당 조건에 맞는 메시지가 없어요."); return
        dd = d.sort_values(mcol, ascending=False).head(n).reset_index(drop=True).copy()
        if "body" in dd.columns:
            dd["_bprev"] = dd["body"].map(lambda x: " ".join(_s(x).split())[:60])
        if "date" in dd.columns:                      # 'YYYYMMDD' 원시 문자열 → 'MM/DD' 표기 통일
            _dts = pd.to_datetime(dd["date"], format="%Y%m%d", errors="coerce")
            dd["date"] = _dts.dt.strftime("%m/%d").where(_dts.notna(), dd["date"])
        all_cols = ["date", "cat", "brand", "title", "_bprev", "send", "infl_cr", "ord_cr", "rps", "amt"]
        cols = [c for c in all_cols if c in dd.columns]
        ren = {"date": "날짜", "cat": "카테고리", "brand": "브랜드", "title": "제목", "_bprev": "내용",
               "send": "발송", "infl_cr": "CTR", "ord_cr": "주문CR", "rps": "RPS", "amt": "거래액"}
        fmts = {"발송": "{:,.0f}", "CTR": "{:.2%}", "주문CR": "{:.2%}", "RPS": "{:,.0f}", "거래액": "{:,.0f}"}
        show = dd[cols].rename(columns=ren)
        fmts = {k: v for k, v in fmts.items() if k in show.columns}
        # 표의 행을 클릭하면 그 행 원문을 아래에 보여준다(on_select). 미지원 버전이면 일반 표로 폴백.
        try:
            _ev = st.dataframe(show.style.format(fmts), hide_index=True, width="stretch",
                               height=340, key=f"tbl_{key}", on_select="rerun",
                               selection_mode="single-row")
        except TypeError:
            _ev = None
            st.dataframe(show.style.format(fmts), hide_index=True, width="stretch", height=340)
        if "title" not in dd.columns:
            return
        opts = {}
        for i, r in dd.iterrows():
            cr = r["ord_cr"] if ("ord_cr" in dd.columns and pd.notna(r["ord_cr"])) else 0
            # 순번 프리픽스 — 같은 날짜·제목·CR 행이 있어도 라벨이 겹쳐 사라지지 않게
            opts[f"{i+1}. [{r.get('date', '–')}] {str(r['title'])[:44]} (주문CR {cr*100:.2f}%)"] = i
        if opts:
            keys_list = list(opts.keys())
            # 클릭한 행 위치 파악
            picked = None
            try:
                _rows = _ev.selection["rows"] if _ev is not None else []
                if _rows:
                    picked = int(_rows[0])
            except Exception:
                picked = None
            # 새로 클릭한 행이 있을 때만 셀렉트박스 기본값을 그 행으로 맞춘다(수동 선택은 그대로 유지).
            _last = f"_lastpick_{key}"
            if picked is not None and 0 <= picked < len(keys_list) and picked != st.session_state.get(_last):
                st.session_state[_last] = picked
                st.session_state[f"msg_{key}"] = keys_list[picked]
            # 지표/정렬이 바뀌면 목록이 달라져 저장된 선택값이 무효가 될 수 있다 — 무효면 버린다
            if st.session_state.get(f"msg_{key}") not in opts:
                st.session_state.pop(f"msg_{key}", None)
            sel = st.selectbox("문구 원문 보기 (표에서 행을 클릭해도 돼요)", keys_list, key=f"msg_{key}")
            r = dd.loc[opts.get(sel, list(opts.values())[0])]
            body = esc(r["body"]) if ("body" in dd.columns and pd.notna(r["body"]) and str(r["body"]).strip()) else "—"
            st.markdown(f'<div class="vg"><b>제목</b><br>{esc(r["title"])}<br><br>'
                        f'<b>내용</b><br>{body}</div>', unsafe_allow_html=True)

    # ── 전 페이지 공통: PDF 저장 버튼 (브라우저 인쇄 → PDF, 인쇄 시 사이드바·툴바 자동 숨김) ──
    components.html("""
    <script>
    (function(){
      var doc = window.parent.document;
      if (!doc.getElementById('lf-pdf-css')) {
        var s = doc.createElement('style'); s.id = 'lf-pdf-css';
        s.textContent = '@media print{[data-testid="stSidebar"],[data-testid="stToolbar"],[data-testid="stHeader"],header,#lf-pdf-btn{display:none!important} [data-testid="stAppViewContainer"] .block-container{max-width:100%!important;padding-top:0!important} .stApp{background:#fff!important}}';
        doc.head.appendChild(s);
      }
      if (!doc.getElementById('lf-pdf-btn')) {
        var b = doc.createElement('button'); b.id = 'lf-pdf-btn';
        b.textContent = '📄 PDF로 저장';
        b.style.cssText = 'position:fixed;top:10px;right:18px;z-index:100000;padding:6px 12px;background:#2E68B0;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px;box-shadow:0 1px 4px rgba(0,0,0,.2);';
        b.onclick = function(){ window.parent.focus(); window.parent.print(); };
        doc.body.appendChild(b);
      }
    })();
    </script>
    """, height=0)

    _REPORT.clear()   # 이 지점부터(페이지 본문) 생성되는 차트/표만 리포트에 담는다

    # ══════════════════════════════════════════════════════════════
    # PAGE 01 — 종합 요약
    # ══════════════════════════════════════════════════════════════
    if "종합 요약" in page:
        st.title("종합 요약")
        st.caption(f"발송 {min_send:,}건 이상 · {len(fdf)}개 캠페인 · {drange}")
        base = fdf if len(fdf) else df
        if not len(fdf) and len(df):
            st.caption("⚠️ '최소 발송수' 기준을 충족하는 캠페인이 없어 **전체(발송수 무관)** 기준으로 "
                       "표시 중이에요 — 위 캡션의 기준과 다르니 참고하세요.")
        # 주차 선택 필터 (년도 포함 · 최신순). 선택 시 이 페이지 전체가 해당 주차로 좁혀짐.
        scope_label = "최근 7일"
        _bp = base.dropna(subset=["dt"]) if "dt" in base else base.iloc[0:0]
        if len(_bp):
            _wks = sorted(_bp["dt"].dt.to_period("W").apply(lambda p: p.start_time).unique(),
                          reverse=True)

            def _wlab(ws):
                we = ws + pd.Timedelta(days=6)
                iy, iw, _ = ws.isocalendar()
                return f"{iy}년 {iw}주차 ({ws.strftime('%m/%d')}~{we.strftime('%m/%d')})"
            _wopts = ["전체 기간"] + [_wlab(w) for w in _wks]
            guard_select("p01_week", _wopts)
            _wsel = st.selectbox("📅 주차 선택", _wopts, index=0, key="p01_week",
                                 help="특정 주차만 보려면 선택해 주세요. 최신 주차가 위에 있어요.")
        base_prev = None                              # 주차 선택 시 '직전 주차' — 전주 대비 증감 표시용
        if len(_bp):
            if _wsel != "전체 기간":
                _ws = _wks[_wopts.index(_wsel) - 1]
                _b0 = base
                base = _b0[(_b0["dt"] >= _ws) & (_b0["dt"] < _ws + pd.Timedelta(days=7))]
                base_prev = _b0[(_b0["dt"] >= _ws - pd.Timedelta(days=7)) & (_b0["dt"] < _ws)]
                if len(base_prev) == 0:
                    base_prev = None
                scope_label = _wsel

        def _wow(cur, prev, pct=False):
            """전주 대비 증감 문자열 (주차를 선택했고 직전 주가 있을 때만)."""
            if base_prev is None or prev is None or pd.isna(prev) or pd.isna(cur):
                return None
            if pct:
                return f"{(cur - prev) * 100:+.2f}%p 전주 대비"
            return f"{(cur / prev - 1) * 100:+.1f}% 전주 대비" if prev else None
        _pv = base_prev if base_prev is not None else base.iloc[0:0]
        c = st.columns(4)
        c[0].metric("발송 캠페인 수", f"{len(base):,}", _wow(len(base), len(_pv) or None),
                    help="현재 조건(필터·최소 발송수)에서 분석에 포함된 캠페인 수예요.")
        c[1].metric("총 발송 건수", won(base["send"].sum()), _wow(base["send"].sum(), _pv["send"].sum() if len(_pv) else None),
                    help="보낸 메시지의 총 건수예요.")
        c[2].metric("총 거래액", won(base["amt"].sum()), _wow(base["amt"].sum(), _pv["amt"].sum() if len(_pv) else None),
                    help="발송으로 발생한 주문 금액 합계예요.")
        c[3].metric("총 주문건수", won(base["oc"].sum()) if "oc" in base else "–",
                    _wow(base["oc"].sum(), _pv["oc"].sum() if len(_pv) else None) if "oc" in base else None,
                    help="발송으로 발생한 주문 건수 합계예요.")
        c = st.columns(4)
        c[0].metric("평균 CTR", f"{base['infl_cr'].mean()*100:.2f}%",
                    _wow(base["infl_cr"].mean(), _pv["infl_cr"].mean() if len(_pv) else None, pct=True),
                    help="UV ÷ 발송 — 보낸 것 중 들어온 비율이에요 (캠페인 단순평균).")
        c[1].metric("평균 주문전환율", f"{base['ord_cr'].mean()*100:.2f}%",
                    _wow(base["ord_cr"].mean(), _pv["ord_cr"].mean() if len(_pv) else None, pct=True),
                    help="주문 ÷ UV — 들어온 사람 중 구매한 비율이에요 (캠페인 단순평균).")
        c[2].metric("평균 RPS(발송건당)", won(base["rps"].mean()),
                    _wow(base["rps"].mean(), _pv["rps"].mean() if len(_pv) else None),
                    help="거래액 ÷ 발송 — 1건 보냈을 때 평균 매출이에요.")
        c[3].metric("평균 객단가", won(base["aov"].mean()),
                    _wow(base["aov"].mean(), _pv["aov"].mean() if len(_pv) else None),
                    help="거래액 ÷ 주문 — 주문 1건당 평균 금액이에요.")
        # 단순평균(위 카드) vs 가중평균 — 페이지마다 '평균 CTR' 정의가 달라 어긋나 보이던 문제 명시
        _st_, _ut_, _ot_ = base["send"].sum(), base["uv"].sum(), base["oc"].sum()
        if _st_ and _ut_:
            st.caption(f"가중(합산) 기준 — CTR {_ut_/_st_*100:.2f}% · 주문CR {_ot_/_ut_*100:.2f}% "
                       "(주간보고 페이지와 같은 정의). 위 카드는 캠페인 단순평균이라 값이 다를 수 있어요 — "
                       "단순평균은 '캠페인 한 건의 전형적 성과', 가중평균은 '발송 전체의 실제 효율'이에요.")

        # ── 최근 13주 미니 추이 — 첫 화면에서 '지금이 좋은 흐름인지' 맥락 제공 ──
        _tr = (fdf if len(fdf) else df).dropna(subset=["dt"])
        if len(_tr):
            _trw = _tr.copy()
            _trw["주"] = _trw["dt"].dt.to_period("W").dt.start_time
            _tw = (_trw.groupby("주")
                   .agg(발송=("send", "sum"), 거래액=("amt", "sum"), UV=("uv", "sum"))
                   .reset_index().sort_values("주").tail(13))
            if len(_tw) >= 3:
                _tw["CTR"] = np.where(_tw["발송"] > 0, _tw["UV"] / _tw["발송"] * 100, np.nan)
                figt = stacked_panels(_tw["주"], _tw["거래액"], "거래액(주간 합)",
                                      _tw["CTR"], "가중 CTR", PALETTE["slate"], PALETTE["blue"],
                                      h=340, line_suffix="%",
                                      title="최근 13주 흐름 — 거래액(위) · 가중 CTR(아래)")
                st.plotly_chart(figt, width="stretch")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        _rk_lab = st.selectbox("TOP/BOTTOM 순위 기준", list(METRIC_OPTS.keys()), key="p01_rank",
                               help="선택한 지표 기준으로 아래 TOP/BOTTOM 순위가 바뀌어요.")
        _rk_col = METRIC_OPTS[_rk_lab][0]
        base_cr = base[base["uv"].fillna(0) >= 100] if "uv" in base else base
        # 최근 7일 (데이터 최신 발송일 기준)
        recent7 = base_cr
        _last = base_cr["dt"].max() if "dt" in base_cr and base_cr["dt"].notna().any() else None
        if _last is not None:
            _r = base_cr[base_cr["dt"] > (_last - pd.Timedelta(days=7))]
            if len(_r):
                recent7 = _r

        def _date_only(r):
            t = r.get("dt")
            return f"{t.year}년 {t.month}월 {t.day}일" if (t is not None and not pd.isna(t)) else "–"
        # 순위는 표본 보정 점수(rank_adjusted) 기준 — UV 겨우 넘긴 캠페인이 주문 한두 건으로
        # TOP/BOTTOM을 점령하는 것을 막는다. 표에 보이는 값은 원래 지표 그대로.
        # mergesort(안정 정렬)로 동점 순서를 결정적으로. BOTTOM은 TOP에 든 행을 제외해
        # 점수가 전부 동점일 때 같은 캠페인이 '최고'이자 '최악'으로 뜨는 것을 막는다.
        win = (recent7.assign(_rk=rank_adjusted(recent7, _rk_col, ascending=False))
               .sort_values("_rk", ascending=False, kind="mergesort").head(8).copy())
        _los_pool = recent7.drop(index=win.index) if len(recent7) > 8 else recent7
        los = (_los_pool.assign(_rk=rank_adjusted(_los_pool, _rk_col, ascending=True))
               .sort_values("_rk", kind="mergesort").head(8).copy())
        for _w in (win, los):
            _w["발송일자"] = _w.apply(_date_only, axis=1)
            _w["발송시간"] = _w["hour"].map(fmt_hhmm) if "hour" in _w else "–"
            _w["_body"] = _w["body"].map(lambda x: " ".join(str(x).split())[:60]) if "body" in _w else ""
            # 표시는 발송 일시 오름차순(먼 일자 → 가까운 일자)
            _w["_sortdt"] = _w["dt"] + pd.to_timedelta(
                _w["hour"].map(hhmm_to_minutes) if "hour" in _w else 0, unit="m")
        win = win.sort_values("_sortdt", na_position="last")
        los = los.sort_values("_sortdt", na_position="last")
        _rng = f"{(_last - pd.Timedelta(days=7)).date()} ~ {_last.date()}" if _last is not None else drange
        _scope_txt = f"최근 7일({_rng})" if scope_label == "최근 7일" else scope_label
        st.caption(f"{_scope_txt} 기준 · UV가 적으면 1건만 주문해도 전환율이 크게 튀어서, "
                   "UV 100 이상인 캠페인만 순위에 넣고, 순위 자체도 표본 크기를 반영한 "
                   "보정 점수(작은 표본일수록 전체 평균 쪽으로)로 매겨요. 표의 값은 원래 지표예요.")
        _tbcols = ["발송일자", "발송시간", "title", "_body", "cat", "send", "infl_cr", "ord_cr", "rps", "aov", "amt"]
        _tbren = {"발송일자": "발송일자", "발송시간": "발송시간", "title": "제목", "_body": "내용",
                  "cat": "카테고리", "send": "발송", "infl_cr": "CTR",
                  "ord_cr": "주문CR", "rps": "RPS", "aov": "객단가", "amt": "거래액"}
        _tbfmt = {"발송": "{:,.0f}", "CTR": "{:.2%}", "주문CR": "{:.2%}",
                  "RPS": "{:,.0f}", "객단가": "{:,.0f}", "거래액": "{:,.0f}"}
        # 반폭 2단 대신 전폭 탭 — 11개 컬럼 표가 좁은 화면에서 가로 스크롤로 잘리던 문제
        _tab_top, _tab_bot = st.tabs([f"🏆 {_rk_lab} TOP ({scope_label})",
                                      f"🧊 {_rk_lab} BOTTOM ({scope_label})"])
        with _tab_top:
            st.dataframe(win[_tbcols].rename(columns=_tbren).style.format(_tbfmt),
                         hide_index=True, width="stretch")
        with _tab_bot:
            st.dataframe(los[_tbcols].rename(columns=_tbren).style.format(_tbfmt),
                         hide_index=True, width="stretch")

        # ── 주목 캠페인 자동 탐지 (이상치) ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 🚩 눈여겨볼 캠페인")
        st.caption("전환율이 평균에서 크게 벗어난 캠페인을 자동으로 찾았어요. "
                   "급등은 '성공 공식', 급락은 '점검 대상'이에요.")
        # TOP/BOTTOM과 같은 UV 가드 — 저UV 캠페인이 '급등' 목록을 요행으로 채우지 않게.
        # 기준도 평균±σ 대신 중앙값±robust σ(MAD) — 전환율은 우측으로 길게 치우친 분포라
        # 평균·표준편차가 이상치 자체에 오염되어 탐지가 무뎌진다.
        _bo = base[base["uv"].fillna(0) >= 100] if "uv" in base else base
        av = _bo["ord_cr"].dropna()
        _med = float(av.median()) if len(av) else np.nan
        _rsd = float((av - _med).abs().median() * 1.4826) if len(av) else 0.0
        if _rsd == 0 and len(av):
            _rsd = float(av.std(ddof=0))              # MAD가 0(값 절반 이상 동일)이면 표준편차 폴백
        if len(av) >= 8 and _rsd > 0:
            mu, sd = _med, _rsd
            bz = _bo.copy()
            bz["_z"] = (bz["ord_cr"] - mu) / sd
            hi = bz[bz["_z"] >= 2].sort_values("_z", ascending=False)
            lo = bz[bz["_z"] <= -2].sort_values("_z")
            oc1, oc2 = st.columns(2)

            def _show_outliers(d, container, emoji, label):
                container.markdown(f"**{emoji} {label} ({len(d)}건)**")
                if len(d) == 0:
                    container.caption("해당 없음 (±2σ 벗어난 건 없음)"); return
                v = d.head(8).copy()
                v["편차σ"] = v["_z"].map(lambda z: f"{z:+.1f}σ")
                vv = v[["date", "cat", "title", "ord_cr", "send", "편차σ"]].rename(
                    columns={"date": "날짜", "cat": "카테고리", "title": "제목", "ord_cr": "주문CR", "send": "발송"})
                container.dataframe(vv.style.format({"주문CR": "{:.2%}", "발송": "{:,.0f}"}),
                                    hide_index=True, width="stretch")
            _show_outliers(hi, oc1, "🔼", "급등 (성공 공식)")
            _show_outliers(lo, oc2, "🔽", "급락 (점검 대상)")
            st.markdown(f'<div class="appendix">기준: UV 100 이상 캠페인의 중앙값 주문CR {mu*100:.2f}% '
                        f'± 2 robust σ({sd*100:.2f}%p, MAD 기반 — 치우친 분포에서도 안정적). '
                        f'급등 건의 문구·오퍼·타이밍을 다음에 써 보고, 급락 건은 원인을 확인해 보세요.</div>',
                        unsafe_allow_html=True)
        else:
            st.info("8건 이상 있어야 찾을 수 있어요.")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 🤖 AI가 찾은 핵심 포인트")
        with st.expander("🔍 AI에 전달되는 데이터 미리보기 (외부 API로 전송돼요)"):
            st.text(build_facts(base))
        if st.button("AI 인사이트 만들기", key="ai_sum"):
            facts = build_facts(base)
            system = ("당신은 LF몰 CRM PUSH 발송 분석가입니다. 주어진 데이터만 근거로 "
                      "핵심 인사이트 3~5개를 한국어 불릿으로 작성하세요. 수치를 지어내지 말고 "
                      "문구 패턴과 성과의 관계에 집중하세요. 출력은 HTML, 불릿은 <br>로 구분, "
                      "긍정은 <span style='color:#16a34a;font-weight:700'>…</span>, "
                      "부정/주의는 <span style='color:#dc2626;font-weight:700'>…</span>로 감싸세요.")
            with st.spinner("생성 중…"):
                txt, err = ai_generate(system, facts, model)
            if err: st.warning(err)
            else: st.session_state["ai_sum_txt"] = txt
        if st.session_state.get("ai_sum_txt"):
            st.markdown(f'<div class="vg">{safe_ai_html(st.session_state["ai_sum_txt"])}</div>', unsafe_allow_html=True)

        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 00 — 주간보고 (전주·전월·전년·MTD 대비) — weekly_report.py 구조 계승
    # ══════════════════════════════════════════════════════════════
    elif "주간보고" in page:
        import calendar
        st.title("주간보고")
        st.caption("기준 주차(월\\~일) 실적을 전주·전월 동주·전년 동주와 비교하고, "
                   "월 누계(MTD)는 전월·전년 같은 기간과 비교해요. "
                   "모든 값은 합산(가중) 기준이에요. 사이드바 필터는 반영되고 '최소 발송수'는 제외돼요.")
        g0 = dff_all.dropna(subset=["dt"]).copy()
        g0 = g0[g0["send"].fillna(0) > 0]
        if len(g0) < 3:
            st.info("데이터가 부족해요. 실적 파일을 더 올려 주세요."); st.stop()
        g0["주"] = g0["dt"].dt.to_period("W").apply(lambda p: p.start_time)

        _wks_all = sorted(g0["주"].unique(), reverse=True)

        def _wklab(ws):
            ws = pd.Timestamp(ws)
            iy, iw, _ = ws.isocalendar()
            return (f"{iy}년 {iw}주차 ({ws.strftime('%m/%d')}~"
                    f"{(ws + pd.Timedelta(days=6)).strftime('%m/%d')})")
        _wlabs = [_wklab(w) for w in _wks_all]
        guard_select("wr_week", _wlabs)
        ref_sel = st.selectbox("기준 주차", _wlabs, index=0, key="wr_week",
                               help="보고 기준이 되는 주(월~일)예요. 최신 주가 위에 있어요.")
        ref_ws = pd.Timestamp(_wks_all[_wlabs.index(ref_sel)])
        ref_we = ref_ws + pd.Timedelta(days=6)

        # ── 기간 집계 헬퍼 (합산 기준 — 주간보고 관행) ──
        def _slice(d0, d1):
            return g0[(g0["dt"] >= pd.Timestamp(d0)) &
                      (g0["dt"] < pd.Timestamp(d1) + pd.Timedelta(days=1))]

        def _agg(d):
            s, u, o, a = d["send"].sum(), d["uv"].sum(), d["oc"].sum(), d["amt"].sum()
            return {"캠페인수": float(len(d)), "발송": s, "UV": u, "주문건수": o, "거래액": a,
                    "CTR": (u / s if s else np.nan), "주문CR": (o / u if u else np.nan),
                    "RPS": (a / s if s else np.nan), "객단가": (a / o if o else np.nan)}

        RATE = {"CTR", "주문CR"}
        METS = ["캠페인수", "발송", "UV", "주문건수", "거래액", "CTR", "주문CR", "RPS", "객단가"]

        def _fmt(met, v):
            if v is None or pd.isna(v):
                return "–"
            if met in RATE:
                return f"{v*100:.2f}%"
            if met in ("거래액", "RPS", "객단가"):
                return won(v)
            return f"{v:,.0f}"

        def _dlt(met, cur, prev):
            """증감 문자열 — 비율 지표는 %p 차이, 그 외 증감율.
            마이너스는 회사 보고 양식에 맞춰 △ 로 표기한다 (예: △8.9%)."""
            if cur is None or prev is None or pd.isna(cur) or pd.isna(prev):
                return "–"
            if met in RATE:
                d = (cur - prev) * 100
                return f"△{abs(d):.2f}%p" if d < 0 else f"+{d:.2f}%p"
            if not prev:
                return "–"
            d = (cur / prev - 1) * 100
            return f"△{abs(d):.1f}%" if d < 0 else f"+{d:.1f}%"

        def _prop_z_p(met, a, b):
            """CTR/주문CR 증감의 두 비율 z검정 p값 — a·b는 _agg 결과 dict."""
            if b is None:
                return np.nan
            if met == "CTR":
                x1, n1, x2, n2 = a["UV"], a["발송"], b["UV"], b["발송"]
            elif met == "주문CR":
                x1, n1, x2, n2 = a["주문건수"], a["UV"], b["주문건수"], b["UV"]
            else:
                return np.nan
            try:
                x1, n1, x2, n2 = float(x1), float(n1), float(x2), float(n2)
            except Exception:
                return np.nan
            if (not n1) or (not n2) or any(pd.isna(v) for v in (x1, n1, x2, n2)):
                return np.nan
            p = (x1 + x2) / (n1 + n2)
            se = np.sqrt(p * (1 - p) * (1 / n1 + 1 / n2))
            if (not se) or pd.isna(se):
                return np.nan
            z = (x1 / n1 - x2 / n2) / se
            return float(2 * (1 - stats.norm.cdf(abs(z))))

        def _dlt_sig(met, a, b):
            """_dlt에 유의성 마크를 더한 버전 — 비율 지표(CTR·주문CR)의 ±가 통계적으로
            유의(✱, p<0.05)한 변화인지 우연 변동인지 구분해 준다. a·b는 _agg dict."""
            s = _dlt(met, a[met], (b[met] if b else np.nan))
            if s == "–" or met not in RATE:
                return s
            pv = _prop_z_p(met, a, b)
            return s + (" ✱" if (pd.notna(pv) and pv < 0.05) else "")

        def _rng_short(ws):
            """짧은 기간 라벨 — 컬럼 헤더용. 예: '26년 6/22~6/28'."""
            ws = pd.Timestamp(ws); we_ = ws + pd.Timedelta(days=6)
            return f"{ws.year % 100}년 {ws.month}/{ws.day}~{we_.month}/{we_.day}"

        def _md(s):
            """마크다운 안전 문자열 — '~'가 취소선으로 해석되지 않게 이스케이프."""
            return str(s).replace("~", "\\~")

        # 기준주가 '부분 주'(아직 일요일까지 데이터가 안 참)면 동요일 누계 비교 — 화요일에
        # 열면 기준주 2일치가 전주 7일치와 비교되어 전주비 △70%대의 가짜 급락이 뜨는 착시
        # 방지. 비교 기간(전주·전월 동주·전년 동주)도 전부 기준주와 같은 요일까지만 잘라 집계.
        # 부분 주 = 달력상 진행 중 주(_is_live_week) 또는 데이터상 가장 최근 주인데 일요일까지
        # 안 온 경우(_is_latest_partial). 후자를 안 보면 '지난 주지만 실적이 목요일까지만
        # 업로드된' 최신 주가 완결로 처리돼 전주 전체와 비교되어 가짜 하락이 뜬다.
        _today_k = today_kst()
        _is_live_week = ref_ws.date() <= _today_k <= ref_we.date()
        _latest_data = g0["dt"].max()
        _is_latest_week = bool(ref_ws <= _latest_data <= ref_we)
        _is_latest_partial = _is_latest_week and (_latest_data.normalize() < ref_we.normalize())
        _partial = _is_live_week or _is_latest_partial
        _elapsed = 6
        if _partial:
            _ref_rows = _slice(ref_ws, ref_we)
            if len(_ref_rows):
                _elapsed = min(int((_ref_rows["dt"].max().normalize() - ref_ws).days), 6)
            elif _is_live_week:
                _elapsed = min((pd.Timestamp(_today_k) - ref_ws).days, 6)
            _elapsed = max(_elapsed, 0)

        def _aggw(ws):
            """주 시작 ws부터 기준주와 같은 경과 요일까지 집계 (완결 주면 월~일 전체)."""
            return _agg(_slice(pd.Timestamp(ws), pd.Timestamp(ws) + pd.Timedelta(days=_elapsed)))

        cur_w = _aggw(ref_ws)
        prev_ws = ref_ws - pd.Timedelta(days=7)
        prev_w = _aggw(prev_ws)
        # 전월 동주 — 기준주 '한가운데(목요일)'의 한 달 전 날짜가 속한 주(월~일).
        # 시작일(월요일) 앵커를 쓰면 한 달 전 날짜가 그 주의 꼬리(토·일)에 걸릴 때 한 주
        # 이른 주가 선택된다 (예: 기준주 7/6~7/12 → 6/6(토) → 6/1~6/7, 35일 전).
        # 목요일 앵커는 '4주 전'·'월내 같은 주차 순번' 두 관례와 일치 (→ 6/8~6/14).
        # ISO 주차가 목요일로 주의 소속을 정하는 것과 같은 원리.
        pm_ws = pd.Timestamp(((ref_ws + pd.Timedelta(days=3)) - pd.DateOffset(months=1))
                             .to_period("W").start_time)
        pm_w = _aggw(pm_ws)
        try:                                             # 전년 동주 — ISO 주차 번호 기준
            _iy, _iw, _ = ref_ws.isocalendar()
            yo_ws = pd.Timestamp(datetime.date.fromisocalendar(_iy - 1, _iw, 1))
            yoy_w = _aggw(yo_ws)
            yo_lab = _wklab(yo_ws)
        except ValueError:                               # 53주차 등 전년에 없는 주
            yo_ws, yoy_w, yo_lab = None, None, "–"
        if _elapsed < 6:
            _dowk = ["월", "화", "수", "목", "금", "토", "일"][_elapsed]
            _why = "진행 중이라" if _is_live_week else "실적이 아직 다 안 들어와서"
            st.info(f"⏳ 기준주가 {_why} **월~{_dowk} 동요일 누계**로 비교해요 — "
                    "전주·전월 동주·전년 동주도 같은 요일까지만 집계해 부분 주 착시를 없앴어요. "
                    "주가 끝나고 데이터가 다 차면 자동으로 월~일 전체 비교로 돌아가요.")

        # ── KPI 카드 (전주비·전년비) — △ 표기 방향이 st.metric 화살표와 어긋나서 커스텀 카드 사용 ──
        def _delta_line(d, label):
            if d == "–":
                return (f'<div style="font-size:12px;color:#94a3b8;margin-top:3px">'
                        f'{label} 데이터 없음</div>')
            _neg = d.startswith("△") or d.startswith("-")
            return (f'<div style="font-size:12px;font-weight:600;margin-top:3px;'
                    f'color:{"#dc2626" if _neg else "#16a34a"}">'
                    f'{d} {label} 대비</div>')
        k = st.columns(6)
        for col, met in zip(k, ["발송", "UV", "CTR", "주문CR", "거래액", "RPS"]):
            _d = _dlt_sig(met, cur_w, prev_w)
            _y = _dlt_sig(met, cur_w, yoy_w)
            col.markdown(
                f'<div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:8px;'
                f'padding:12px 16px;box-shadow:0 1px 3px rgba(0,0,0,0.06)">'
                f'<div style="font-size:12px;color:#64748b">{met}</div>'
                f'<div style="font-size:20px;color:#1e293b;font-weight:600;'
                f'font-feature-settings:\'tnum\' 1">{_fmt(met, cur_w[met])}</div>'
                f'{_delta_line(_d, "전주")}{_delta_line(_y, "전년")}</div>',
                unsafe_allow_html=True)

        # ── 주요 지표 현황 표 (전주·전월 동주·전년 동주 — 컬럼에 기간 일자 표기) ──
        st.markdown("##### 📋 주요 지표 현황")
        col_prev = f"전주 ({_rng_short(prev_ws)})"
        col_cur = f"기준주 ({_rng_short(ref_ws)})"
        col_pm = f"전월 동주 ({_rng_short(pm_ws)})"
        col_yoy = f"전년 동주 ({_rng_short(yo_ws)})" if yo_ws is not None else "전년 동주 (–)"
        rows = []
        for met in METS:
            yv = yoy_w[met] if yoy_w else np.nan
            rows.append({"지표": met,
                         col_prev: _fmt(met, prev_w[met]), col_cur: _fmt(met, cur_w[met]),
                         "전주비": _dlt_sig(met, cur_w, prev_w),
                         col_pm: _fmt(met, pm_w[met]),
                         "전월비": _dlt_sig(met, cur_w, pm_w),
                         col_yoy: _fmt(met, yv), "전년비": _dlt_sig(met, cur_w, yoy_w)})
        wr_tbl = pd.DataFrame(rows)

        def _clr(v):
            s = str(v)
            if s.startswith("+"):
                return "color:#16a34a;font-weight:600"
            if s.startswith("△") or s.startswith("-"):    # △ = 마이너스 (회사 보고 양식)
                return "color:#dc2626;font-weight:600"
            return ""
        st.dataframe(wr_tbl.style.map(_clr, subset=["전주비", "전월비", "전년비"]),
                     hide_index=True, width="stretch", height=38 + 35 * len(wr_tbl))
        st.caption(f"기준주 {_md(_wklab(ref_ws))} · 전년 동주 {_md(yo_lab)} — "
                   "해당 기간에 데이터가 없으면 '–'로 표시돼요. "
                   "전월비는 전월 동주(기준주의 정확히 4주 전 주)와 비교해요. "
                   "✱ = CTR·주문CR의 증감이 통계적으로 유의(p<0.05) — 표본 크기를 감안해도 "
                   "우연 변동 범위를 벗어났다는 뜻이에요. 마크가 없으면 노이즈일 수 있어요.")

        # ── 보고란 (weekly_report.py 동일 구성 · 접이식) — 주차별로 저장 백엔드에 영속 ──
        # 구글시트(설정 시) 또는 로컬 CSV에 {key:text}를 저장 → 세션이 끊겨도 유지된다.
        def _notes_save(d):
            try:
                # 저장 직전 시트를 다시 읽어 병합 — 캠페인/MTD/기획전 저장과 동일한 낙관적 병합.
                # 이게 없으면 A세션이 노트를 저장할 때 B세션이 추가한 '다른 주차' 노트가
                # A의 스냅샷으로 덮어써져 사라진다(같은 키는 방금 편집한 A값 우선).
                try:
                    _fresh = notes_df_to_dict(storage_load(BK, "notes"))
                except Exception:
                    _fresh = {}
                merged = {**_fresh, **d}
                st.session_state.wr_notes = merged
                if storage_save(BK, "notes", notes_dict_to_df(merged)):
                    d.clear(); d.update(merged)         # 호출부의 store 참조도 병합본으로 동기화
            except Exception as e:
                # 실패를 삼키면 사용자는 저장된 줄 알고 떠난다 — 명시적으로 알린다
                st.error(f"보고란 저장에 실패했어요 — 잠시 후 다시 시도해 주세요. ({str(e)[:80]})")

        if "wr_notes" not in st.session_state:
            try:
                st.session_state.wr_notes = notes_df_to_dict(storage_load(BK, "notes"))
            except Exception:
                st.session_state.wr_notes = {}

        def _auto_kpi_note():
            """기준주 실적으로 보고 문구 자동 생성 — weekly_report 템플릿 형식."""
            lines = []
            for met in ["발송", "거래액", "CTR", "주문CR", "RPS", "캠페인수"]:
                yv = yoy_w[met] if yoy_w else np.nan
                lines.append(f"- {met} — {_fmt(met, cur_w[met])}, "
                             f"전주비 {_dlt(met, cur_w[met], prev_w[met])}, "
                             f"전년비 {_dlt(met, cur_w[met], yv)}")

            # 앱푸시 수신동의 데이터가 세션에 적재되어 있는 경우 요약 한 줄 추가
            _push_df = st.session_state.get("push_consent_df")
            if _push_df is not None and not _push_df.empty:
                
                c_sum = get_push_stats(_push_df, ref_ws, ref_we, "Total")
                if c_sum:
                    c_con = c_sum["last_consent"]
                    c_add = c_sum["tot_added"]
                    c_rem = c_sum["tot_removed"]
                    
                    p_sum = get_push_stats(_push_df, prev_ws, ref_we - pd.Timedelta(days=7), "Total")
                    p_con = p_sum["last_consent"] if p_sum else None
                    
                    y_sum = get_push_stats(_push_df, yo_ws, yo_ws + pd.Timedelta(days=6), "Total") if yo_ws is not None else None
                    y_con = y_sum["last_consent"] if y_sum else None
                    
                    def _d_pct(cur, prev):
                        if cur is None or prev is None or not prev:
                            return "–"
                        pct = (cur / prev - 1) * 100
                        return f"△{abs(pct):.2f}%" if pct < 0 else f"+{pct:.2f}%"
                    
                    p_diff = _d_pct(c_con, p_con)
                    y_diff = _d_pct(c_con, y_con)
                    
                    lines.append(f"- 앱푸시 수신동의 — {c_con:,.0f}명 (신규: +{c_add:,.0f}명, 이탈: △{c_rem:,.0f}명), "
                                 f"전주비 {p_diff}, 전년비 {y_diff}")
            return "\n".join(lines)

        def _note_render(text):
            """보고란 표시 — △는 빨강, +는 초록 (회사 양식). HTML 태그를 보존하면서 텍스트 노드만 색칠합니다."""
            s = (text or "").strip() or "내용을 입력하세요."
            parts = re.split(r'(<[^>]+>)', s)
            for i in range(len(parts)):
                if i % 2 == 0:
                    txt = parts[i]
                    txt = re.sub(
                        r"(△[\d.,]+%?p?)",
                        r'<span style="color:#dc2626;font-weight:700">\1</span>',
                        txt
                    )
                    txt = re.sub(
                        r"(\+[\d.,]+%?p?)",
                        r'<span style="color:#16a34a;font-weight:700">\1</span>',
                        txt
                    )
                    parts[i] = txt
            res = "".join(parts)
            if "<p>" not in res and "<li>" not in res and "<br>" not in res:
                res = res.replace("\n", "<br>")
            return f'<div class="vg">{res}</div>'

        def _ai_kpi_note():
            system = ("당신은 LF몰 CRM 발송 주간보고 작성자입니다. 아래 지표 현황을 바탕으로 "
                      "보고서 '전주 주요 지표 현황' 란에 넣을 요약을 한국어 플레인 텍스트로 작성하세요. "
                      "긴 문장형 요약을 피하고, 반드시 '- '(메인 요약)와 'ㄴ '(세부 수치/해석)를 사용하는 "
                      "계층형 불릿 구조로 출력하세요.\n"
                      "'-' 불릿에는 지표별 핵심 요약(예: 발송량 증가 및 효율 개선)을 적고, "
                      "'ㄴ' 불릿에는 구체적인 수치 팩트(값, 전주비 ..., 전년비 ...)나 원인 분석을 적어주세요. "
                      "마지막에 전체를 아우르는 '핵심 코멘트' 불릿을 1~2개 추가하세요. "
                      "마이너스는 △ 표기를 유지하고 수치를 지어내지 마세요. HTML 태그 없이 순수 텍스트로 출력하세요.")
            return ai_generate(system, _auto_kpi_note(), model)

        def _this_week_range():
            """오늘(실제 달력 기준) 이 속한 주의 월~일 date 쌍. '금주 집행' 은 선택된 기준주(과거 실적
            비교용)가 아니라 항상 '오늘' 기준 이번 주 — 아직 실적이 없는 발송 예정 주라서 실적 대신
            기획(문구) 시트에서 읽어야 한다."""
            _today = today_kst()
            _mon = _today - datetime.timedelta(days=_today.weekday())
            return _mon, _mon + datetime.timedelta(days=6)

        def _this_week_plan_rows():
            """'기획 문구 가져오기'로 세션에 적재된 lookup에서 이번 주(월~일) 항목만 추출.
            반환: (월요일, 일요일, [(date_str, af, title, body), ...] 날짜순 정렬) — lookup 없으면 rows=None."""
            lk = st.session_state.get("plan_lookup_gs")
            mon, sun = _this_week_range()
            if not lk:
                return mon, sun, None
            lo, hi = mon.strftime("%Y%m%d"), sun.strftime("%Y%m%d")
            seen, rows = set(), []
            for (d, af), (title, body) in lk.items():
                if not d or not (lo <= d <= hi):
                    continue
                key = (d, title.strip(), body.strip())
                if key in seen:                          # 같은 문구가 여러 AF코드(세그먼트)로 중복 등록된 경우 1건만
                    continue
                seen.add(key)
                rows.append((d, af, title, body))
            rows.sort(key=lambda r: (r[0], r[1]))
            return mon, sun, rows

        def _auto_exec_note():
            """이번 주 기획 문구를 날짜순 그대로 나열 — AI 없이 즉시 채우는 템플릿."""
            mon, sun, rows = _this_week_plan_rows()
            _rng = f"{mon.month}/{mon.day}~{sun.month}/{sun.day}"
            if rows is None:
                return (f"[이번 주 {_rng}] 기획 문구가 아직 없어요 — 사이드바 "
                        "「📥 기획 문구 가져오기」를 눌러 이번 주 시트를 불러온 뒤 다시 시도해 주세요.")
            if not rows:
                return f"[이번 주 {_rng}] 기획 시트에 등록된 캠페인이 없어요."
            lines = [f"[이번 주 {_rng}] 집행 예정 {len(rows)}건"]
            for d, af, title, body in rows:
                dd = f"{int(d[4:6])}/{int(d[6:8])}"
                b = " ".join(_s(body).split())[:40]
                lines.append(f"- [{dd}] {af} · {title.strip()}" + (f" — {b}" if b else ""))
            return "\n".join(lines)

        def _ai_exec_note():
            mon, sun, rows = _this_week_plan_rows()
            _rng = f"{mon.month}/{mon.day}~{sun.month}/{sun.day}"
            if rows is None:
                return None, ("이번 주 기획 문구가 없어요. 사이드바 「📥 기획 문구 가져오기」를 "
                              "눌러 이번 주 시트를 먼저 불러와 주세요.")
            if not rows:
                return None, f"이번 주({_rng}) 기획 시트에 등록된 캠페인이 없어요."
            lines = [f"[이번 주 {_rng} 집행 예정 — 총 {len(rows)}건]"]
            for d, af, title, body in rows[:300]:          # 과도한 프롬프트 길이 방지용 안전 상한
                dd = f"{int(d[4:6])}/{int(d[6:8])}"
                b = " ".join(_s(body).split())[:50]
                lines.append(f"- [{dd}] {af} · {title.strip()}" + (f" │ {b}" if b else ""))
            system = ("당신은 LF몰 CRM 발송 주간보고 작성자입니다. 아래는 이번 주(월~일) 발송 기획에 "
                      "등록된 캠페인 문구 목록(날짜·AF코드·제목·내용)입니다. 아직 발송 전이라 실적은 "
                      "없고 계획만 있는 상태예요. 이 목록만 근거로 '금주 집행 내용 요약' 란에 넣을 "
                      "요약을 한국어 플레인 텍스트로 작성하세요. "
                      "단, 긴 문장을 피하고 반드시 '- '(주요 요약/하이라이트)와 'ㄴ '(상세 내용/캠페인 예시) 형태의 "
                      "계층형 불릿 구조로 정리해주세요: \n"
                      "1) 이번 주 집행 볼륨·전반적 톤에 대한 총평, \n"
                      "2) 주요 소구·오퍼 유형별 하이라이트, \n"
                      "3) 눈에 띄는 반복 프로모션 패턴. \n"
                      "목록에 없는 내용(수치·할인율 등)은 지어내지 말고, 목록에 실제로 있는 문구만 "
                      "근거로 쓰세요. HTML 태그 없이 순수 텍스트로 출력하세요.")
            return ai_generate(system, "\n".join(lines), model)

        @st.fragment
        def _note_block_body(nkey, title, regen=None, ai_fn=None):
            """편집/자동 생성/AI 생성 버튼이 달린 보고란 본문 (weekly_report.report_text_block 계승).

            fragment — 편집 토글·quill 입력 등 상호작용 시 전체 스크립트가 아니라 이 보고란만
            다시 그린다. 저장/생성 버튼은 st.rerun()으로 전체 갱신(다른 표시 영역 반영)."""
            store = st.session_state.wr_notes
            ekey = f"_wr_note_edit_{nkey}"
            st.markdown(f"**{title}**")
            if st.session_state.get(ekey, False):
                val = store.get(nkey, "")
                if val and not (val.startswith("<p>") or val.startswith("<ul>") or val.startswith("<li>") or "<div" in val):
                    if val.strip().startswith("-"):
                        # '-'로 시작하는 줄만 접두어를 떼고, 'ㄴ'(세부) 줄은 텍스트를
                        # 그대로 보존한다 — 무조건 첫 글자를 자르면 계층 구조가 깨진다
                        items = []
                        for item in val.strip().split("\n"):
                            t = item.strip()
                            if not t:
                                continue
                            if t.startswith("-"):
                                items.append(f"<li>{t[1:].strip()}</li>")
                            else:
                                items.append(f"<li>{t}</li>")
                        val = f"<ul>{''.join(items)}</ul>"
                    else:
                        val = "".join(f"<p>{line}</p>" for line in val.split("\n"))
                
                if HAS_QUILL:
                    toolbar = [
                        [{"size": ["small", False, "large", "huge"]}],
                        ["bold", "italic", "underline", "strike"],
                        [{"color": []}, {"background": []}],
                        [{"list": "ordered"}, {"list": "bullet"}],
                        [{"align": []}], ["clean"],
                    ]
                    new = st_quill(value=val, html=True, toolbar=toolbar,
                                   key=f"quill_{nkey}")
                else:
                    new = st.text_area("내용", val, key=f"ta_{nkey}",
                                       height=200, label_visibility="collapsed")
                
                if st.button("저장", key=f"btn_s_{nkey}", type="primary", width="stretch"):
                    store[nkey] = new if new is not None else val
                    _notes_save(store)
                    st.session_state[ekey] = False
                    st.rerun()
            else:
                st.markdown(_note_render(store.get(nkey, "")), unsafe_allow_html=True)
            
            bcols = st.columns(1 + (regen is not None) + (ai_fn is not None))
            bi = 0
            editing = st.session_state.get(ekey, False)
            if bcols[bi].button("보기" if editing else "편집", key=f"btn_e_{nkey}", width="stretch"):
                st.session_state[ekey] = not editing
                st.rerun()
            bi += 1
            if regen is not None:
                if bcols[bi].button("자동 생성", key=f"btn_r_{nkey}", width="stretch",
                                    help="기준주 실적으로 지표 문구를 자동으로 채워요 (기존 내용 대체)"):
                    # callable을 받아 클릭 시에만 평가 — 값으로 받으면 매 rerun마다
                    # 자동 생성 로직(기획 lookup 전수 순회 등)이 실행된다
                    store[nkey] = regen() if callable(regen) else regen
                    _notes_save(store)
                    st.session_state[ekey] = False
                    st.rerun()
                bi += 1
            if ai_fn is not None:
                if bcols[bi].button("AI 생성", key=f"btn_a_{nkey}", width="stretch",
                                    help="AI가 데이터를 보고 요약 문구를 작성해요 (기존 내용 대체)"):
                    with st.spinner("AI 작성 중…"):
                        text, err = ai_fn()
                    if err:
                        st.error(err)
                    else:
                        store[nkey] = text
                        _notes_save(store)
                        st.session_state[ekey] = False
                        st.rerun()


        def _note_block(col, nkey, title, regen=None, ai_fn=None):
            with col:                     # 컨테이너 진입은 fragment 밖에서 (외부 컨테이너 제약)
                _note_block_body(nkey, title, regen, ai_fn)

        _wkkey = ref_ws.strftime("%Y%m%d")
        # '금주 집행'은 기준 주차 선택과 무관하게 항상 '오늘' 기준 이번 주를 가리키므로
        # 저장 키도 ref_ws가 아니라 실제 이번 주 월요일로 고정한다(기준 주차를 바꿔도 안 사라짐).
        _this_wkkey = _this_week_range()[0].strftime("%Y%m%d")
        with st.expander("📝 보고란 — 전주 주요 지표 현황 · 금주 집행 내용 요약", expanded=True):
            nb1, nb2 = st.columns(2)
            _note_block(nb1, f"kpi_{_wkkey}", "전주 주요 지표 현황",
                        regen=_auto_kpi_note, ai_fn=_ai_kpi_note)
            _note_block(nb2, f"exec_{_this_wkkey}", "금주 집행 내용 요약",
                        regen=_auto_exec_note, ai_fn=_ai_exec_note)
            st.caption("내용은 주차별로 저장돼요 — 기준 주차를 바꾸면 그 주차의 보고란이 열려요. "
                       "'금주 집행 내용 요약'은 선택한 기준 주차와 무관하게 **오늘 기준 이번 주** "
                       "기획 시트를 읽어요(아직 실적이 없는 발송 예정 주라서).")
            # 지난 주차의 '금주 집행' 노트도 열람 가능하게 — 주가 넘어가면 키가 바뀌어
            # 저장소에는 남는데 UI에서 영영 접근 불가하던 문제 방지 (기준 주차 선택과 연동)
            _past_exec = st.session_state.wr_notes.get(f"exec_{_wkkey}")
            if _wkkey != _this_wkkey and _past_exec:
                with st.expander(f"🗂 {_md(ref_sel)} 당시의 '금주 집행 내용 요약' 보기"):
                    st.markdown(_note_render(_past_exec), unsafe_allow_html=True)

        # ── 월 누계(MTD) — 기준주 일요일 마감 기준 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📆 월 누계(MTD) — 전월·전년 같은 기간 대비")
        ref_end = ref_we.date()
        m_first = ref_end.replace(day=1)

        def _mtd_range(y, m, day):
            last = calendar.monthrange(y, m)[1]
            return datetime.date(y, m, 1), datetime.date(y, m, min(day, last))
        cur_mtd = _agg(_slice(m_first, ref_end))
        pm_y, pm_m = (ref_end.year, ref_end.month - 1) if ref_end.month > 1 else (ref_end.year - 1, 12)
        pm0, pm1 = _mtd_range(pm_y, pm_m, ref_end.day)
        prev_mtd = _agg(_slice(pm0, pm1))
        py0, py1 = _mtd_range(ref_end.year - 1, ref_end.month, ref_end.day)
        yoy_mtd = _agg(_slice(py0, py1))
        _cur_lab = f"당월 MTD ({m_first.month}/1~{ref_end.month}/{ref_end.day})"
        _pm_lab = f"전월 MTD ({pm0.month}/1~{pm1.month}/{pm1.day})"
        _py_lab = f"전년 MTD ({py0.year}년 {py0.month}월)"
        rows = []
        for met in METS:
            rows.append({"지표": met,
                         _pm_lab: _fmt(met, prev_mtd[met]), _cur_lab: _fmt(met, cur_mtd[met]),
                         "전월비": _dlt_sig(met, cur_mtd, prev_mtd),
                         _py_lab: _fmt(met, yoy_mtd[met]),
                         "전년비": _dlt_sig(met, cur_mtd, yoy_mtd)})
        st.dataframe(pd.DataFrame(rows).style.map(_clr, subset=["전월비", "전년비"]),
                     hide_index=True, width="stretch", height=38 + 35 * len(rows))
        st.markdown('<div class="appendix">MTD는 <b>기준주 일요일까지의 월 누계</b>예요. '
                    '전월·전년은 같은 일수(1일~같은 날짜)로 맞춰 비교해요. '
                    '월초 주차일수록 누계 일수가 짧아 값이 작게 보이는 게 정상이에요.</div>',
                    unsafe_allow_html=True)

        # ── 📱 앱푸시 수신동의 주간 요약 (주간보고용 연동) ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📱 앱푸시 수신동의 주간 요약")

        _push_df = st.session_state.get("push_consent_df")
        if _push_df is None or _push_df.empty:
            st.info("👈 사이드바에서 **앱푸시 동의 현황 xlsx** 파일을 올려주시면, "
                    "이 주간보고 페이지에서도 기준 주차의 수신동의 추이와 전주 대비 증감 요약을 바로 볼 수 있어요.")
        else:
            # 기준주 및 전주의 날짜 필터 정의
            _cur_ws, _cur_we = pd.Timestamp(ref_ws), pd.Timestamp(ref_we)
            _prev_ws, _prev_we = _cur_ws - pd.Timedelta(days=7), _cur_we - pd.Timedelta(days=7)

            push_summary = []
            for g in ["Total", "기존", "신규"]:
                c_sum = get_push_stats(_push_df, _cur_ws, _cur_we, g)
                p_sum = get_push_stats(_push_df, _prev_ws, _prev_we, g)
                
                if c_sum:
                    c_con = c_sum["last_consent"]
                    c_add = c_sum["tot_added"]
                    c_rem = c_sum["tot_removed"]
                    c_dif = c_sum["tot_diff"]
                    
                    p_con = p_sum["last_consent"] if p_sum else None
                    p_add = p_sum["tot_added"] if p_sum else None
                    p_rem = p_sum["tot_removed"] if p_sum else None
                    p_dif = p_sum["tot_diff"] if p_sum else None

                    def _d_str(cur, prev, is_con=False):
                        if cur is None or prev is None:
                            return "–"
                        diff_val = cur - prev
                        if is_con:
                            # 동의수 차이값 그대로
                            return f"△{abs(diff_val):,}명" if diff_val < 0 else f"+{diff_val:,}명"
                        else:
                            # 추가/탈퇴량 증감율 (%)
                            if not prev: return "–"
                            pct = (cur / prev - 1) * 100
                            return f"△{abs(pct):.1f}%" if pct < 0 else f"+{pct:.1f}%"

                    def _d_val_str(cur, prev):
                        if cur is None or prev is None: return "–"
                        d = cur - prev
                        return f"△{abs(d):,}명" if d < 0 else f"+{d:,}명"

                    # 날짜 정보를 컬럼명에 동적으로 기입하기 위해 포맷팅 문자열 생성
                    _cur_range = _rng_short(_cur_ws)
                    _prev_range = _rng_short(_prev_ws)
                    col_consent = f"기말 동의수 ({_cur_range})"
                    col_added = f"주간 신규추가 ({_cur_range})"
                    col_removed = f"주간 기존이탈 ({_cur_range})"
                    col_diff = f"주간 순증감 ({_cur_range})"
                    col_con_diff = "동의수 증감(전주비)"
                    col_add_pct = "신규추가 전주비"
                    col_rem_pct = "기존이탈 전주비"
                    col_dif_diff = "순증감 전주비"

                    push_summary.append({
                        "구분": g,
                        col_consent: f"{c_con:,.0f}명" if pd.notna(c_con) else "–",
                        col_con_diff: _d_val_str(c_con, p_con),
                        col_added: f"{c_add:,.0f}명" if pd.notna(c_add) else "–",
                        col_add_pct: _d_str(c_add, p_add),
                        col_removed: f"{c_rem:,.0f}명" if pd.notna(c_rem) else "–",
                        col_rem_pct: _d_str(c_rem, p_rem),
                        col_diff: f"{c_dif:+,.0f}명" if pd.notna(c_dif) else "–",
                        col_dif_diff: _d_val_str(c_dif, p_dif),
                    })

            if push_summary:
                push_sum_df = pd.DataFrame(push_summary)
                st.dataframe(
                    push_sum_df.style.map(_clr, subset=[col_con_diff, col_add_pct, col_rem_pct, col_dif_diff]),
                    hide_index=True, width="stretch"
                )
                st.caption(f"기준주 ({_md(_rng_short(_cur_ws))}) vs 전주 ({_md(_rng_short(_prev_ws))}) 앱푸시 수신동의 지표 비교 데이터예요. "
                           f"마이너스 수치는 **△** 로 표기되며 붉은색으로 강조돼요.\n\n"
                           "**기말 동의수**는 그 주 마지막 날 기준 누적 동의자 수(스냅샷)이고, "
                           "**신규추가·기존이탈·순증감**은 그 주 전체를 더한 값(합계)이에요 — "
                           "주말/평일로 나뉜 게 아니라 '시점값 하나 + 합계값 셋'이에요.")
            else:
                st.info("해당 기간의 앱푸시 동의 현황 데이터가 부재합니다.")

        # ── 심화 분석 탭: 증감 기여 분해 · 하이라이트 · 월말 마감 예상 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        tabW, tabH, tabP = st.tabs(["📉 증감 기여 분해", "🏆 하이라이트·로우라이트", "🎯 월말 마감 예상"])

        def _damt(v):
            """증감액 문자열 — 마이너스는 △ (회사 보고 양식)."""
            if v is None or pd.isna(v):
                return "–"
            return f"△{won(abs(v))}" if v < 0 else f"+{won(v)}"

        # ① 거래액 전주 대비 — 카테고리 기여 분해 (워터폴)
        with tabW:
            st.markdown("##### 거래액 전주 대비 — 어느 카테고리가 끌어올리고/깎아먹었나")
            def _catfill(d):
                """cat 결측/공백을 '(미분류)'로 채운 사본 — groupby('cat')는 NaN 행을 드롭하고
                아래 union/성분합은 공백 카테고리를 제외하는데, 총합(prev_tot·s1t·ΔCTR)은 전체
                행 기준이라 무카테고리 행이 있으면 워터폴·믹스분해 항등식이 깨진다(합≠총증감).
                하나의 카테고리로 포함시켜 총합을 보존한다."""
                dd = d.copy()
                if "cat" in dd.columns:
                    _c = dd["cat"]
                    # isna()를 명시적으로 OR — astype(str)이 일부 pandas 버전에서 NaN을
                    # 'nan' 문자열로 안 바꿔(nullable dtype) isin이 결측을 놓치는 것 방지
                    _blank = _c.isna() | _c.astype(str).str.strip().isin(["", "nan", "None", "NaN"])
                    dd["cat"] = np.where(_blank, "(미분류)", _c.astype(str))
                return dd
            cwd = _catfill(_slice(ref_ws, ref_we))
            pwd = _catfill(_slice(ref_ws - pd.Timedelta(days=7), ref_we - pd.Timedelta(days=7)))
            if "cat" not in cwd.columns or len(pwd) == 0:
                st.info("전주 데이터가 없어 분해할 수 없어요.")
            else:
                cur_g = cwd.groupby("cat").agg(
                    send=("send", "sum"),
                    uv=("uv", "sum"),
                    oc=("oc", "sum"),
                    amt=("amt", "sum")
                )
                prv_g = pwd.groupby("cat").agg(
                    send=("send", "sum"),
                    uv=("uv", "sum"),
                    oc=("oc", "sum"),
                    amt=("amt", "sum")
                )
                
                curS = cur_g["amt"]
                prvS = prv_g["amt"]
                union = curS.index.union(prvS.index)
                union = [c for c in union if str(c).strip() not in ("", "nan", "None")]
                dif = (curS.reindex(union).fillna(0) - prvS.reindex(union).fillna(0))
                dif = dif[dif != 0].sort_values(ascending=False)
                # 카테고리별 증감이 전부 0이면 워터폴은 건너뛰고 안내만 (LMDI·믹스는 계속).
                # dif가 비면 아래 워터폴/표는 자연히 빈 값이 되지만, 안내로 오해를 막는다.
                if dif.empty:
                    st.info("전주 대비 카테고리별 거래액 증감이 없어요 (동일하거나 데이터 없음).")
                # 기여 큰 8개만 개별 표시, 나머지는 '기타'로 합산
                if len(dif) > 8:
                    top8 = dif.reindex(dif.abs().sort_values(ascending=False).head(8).index)
                    etc = float(dif.drop(top8.index).sum())
                    dif = top8.sort_values(ascending=False)
                    if etc != 0:
                        dif = pd.concat([dif, pd.Series({"기타": etc})])
                prev_tot = float(pwd["amt"].sum()); cur_tot = float(cwd["amt"].sum())
                labels = [str(i) for i in dif.index]
                diffs = [float(v) for v in dif.values]
                colors = [PALETTE["green"] if v >= 0 else PALETTE["red"] for v in diffs]
                figw = go.Figure(go.Bar(
                    x=labels,
                    y=diffs,
                    text=[_damt(d) for d in diffs],
                    textposition="outside",
                    marker_color=colors,
                    hovertemplate="%{x}<br>증감액: %{text}<extra></extra>"
                ))
                figw.update_layout(**base_layout(
                    h=410, title=f"거래액 {won(prev_tot)} → {won(cur_tot)} "
                                 f"({_dlt('거래액', cur_tot, prev_tot)})"
                ))
                st.plotly_chart(figw, width="stretch")
                
                wrows = []
                top8_cats = [x for x in dif.index if x != "기타"]
                for c in dif.index:
                    if c == "기타":
                        p_send = float(prv_g.drop(index=top8_cats, errors="ignore")["send"].sum())
                        p_uv = float(prv_g.drop(index=top8_cats, errors="ignore")["uv"].sum())
                        p_oc = float(prv_g.drop(index=top8_cats, errors="ignore")["oc"].sum())
                        p_amt = float(prv_g.drop(index=top8_cats, errors="ignore")["amt"].sum())
                        
                        c_send = float(cur_g.drop(index=top8_cats, errors="ignore")["send"].sum())
                        c_uv = float(cur_g.drop(index=top8_cats, errors="ignore")["uv"].sum())
                        c_oc = float(cur_g.drop(index=top8_cats, errors="ignore")["oc"].sum())
                        c_amt = float(cur_g.drop(index=top8_cats, errors="ignore")["amt"].sum())
                    else:
                        p_send = float(prv_g.loc[c, "send"]) if c in prv_g.index else 0.0
                        p_uv = float(prv_g.loc[c, "uv"]) if c in prv_g.index else 0.0
                        p_oc = float(prv_g.loc[c, "oc"]) if c in prv_g.index else 0.0
                        p_amt = float(prv_g.loc[c, "amt"]) if c in prv_g.index else 0.0
                        
                        c_send = float(cur_g.loc[c, "send"]) if c in cur_g.index else 0.0
                        c_uv = float(cur_g.loc[c, "uv"]) if c in cur_g.index else 0.0
                        c_oc = float(cur_g.loc[c, "oc"]) if c in cur_g.index else 0.0
                        c_amt = float(cur_g.loc[c, "amt"]) if c in cur_g.index else 0.0
                    
                    p_ctr = p_uv / p_send if p_send > 0 else 0.0
                    c_ctr = c_uv / c_send if c_send > 0 else 0.0
                    p_cr = p_oc / p_uv if p_uv > 0 else 0.0
                    c_cr = c_oc / c_uv if c_uv > 0 else 0.0
                    p_rps = p_amt / p_send if p_send > 0 else 0.0
                    c_rps = c_amt / c_send if c_send > 0 else 0.0
                    
                    v = dif.get(c, 0.0)
                    send_diff = c_send - p_send
                    send_diff_str = f"△{abs(send_diff):,.0f}" if send_diff < 0 else f"+{send_diff:,.0f}"
                    wrows.append({
                        "카테고리": str(c),
                        "전주 거래액": won(p_amt),
                        "기준주 거래액": won(c_amt),
                        "거래액 증감": _damt(v),
                        "거래액 전주비": _dlt("거래액", c_amt, p_amt if p_amt > 0 else np.nan),
                        "전주 발송": f"{p_send:,.0f}",
                        "기준주 발송": f"{c_send:,.0f}",
                        "발송 증감": send_diff_str,
                        "발송 전주비": _dlt("발송", c_send, p_send if p_send > 0 else np.nan),
                        "전주 CTR": f"{p_ctr:.2%}",
                        "기준주 CTR": f"{c_ctr:.2%}",
                        "전주 UV": f"{p_uv:,.0f}",
                        "기준주 UV": f"{c_uv:,.0f}",
                        "전주 CR": f"{p_cr:.2%}",
                        "기준주 CR": f"{c_cr:.2%}",
                        "전주 RPS": f"{p_rps:,.0f}원",
                        "기준주 RPS": f"{c_rps:,.0f}원",
                    })
                if wrows:
                    st.dataframe(pd.DataFrame(wrows).style.map(_clr, subset=["거래액 증감", "거래액 전주비", "발송 증감", "발송 전주비"]),
                                 hide_index=True, width="stretch", height=min(38 + 35 * len(wrows), 640))
                st.markdown('<div class="appendix">카테고리별 전주 대비 거래액 증감 기여도입니다. '
                            '녹색은 매출 상승 기여, 적색(△)은 매출 감소 기여를 의미하며, 기여도가 큰 8개만 표시하고 나머지는 기타로 합산했습니다.</div>', unsafe_allow_html=True)

                # ── 지표 체인 분해(LMDI) — '어느 카테고리'가 아니라 '어느 지표'가 만들었나 ──
                st.markdown("##### 거래액 전주 대비 — 어느 지표(발송·CTR·CR·객단가)가 만들었나")

                def _chain(d):
                    s, u, o, a = (float(d["send"].sum()), float(d["uv"].sum()),
                                  float(d["oc"].sum()), float(d["amt"].sum()))
                    if min(s, u, o, a) <= 0:
                        return None
                    return {"발송량": s, "CTR": u / s, "주문CR": o / u, "객단가": a / o, "_tot": a}
                f1, f0 = _chain(cwd), _chain(pwd)
                if f1 and f0:
                    import math as _math
                    v1, v0 = f1["_tot"], f0["_tot"]
                    # LMDI-I: 각 요인 기여 = L(V1,V0)·ln(f1/f0), L=로그평균 → 기여 합 = ΔV (정확)
                    _L = (v1 - v0) / (_math.log(v1) - _math.log(v0)) if v1 != v0 else v1
                    contrib = {k: _L * _math.log(f1[k] / f0[k])
                               for k in ("발송량", "CTR", "주문CR", "객단가")}
                    wf = go.Figure(go.Waterfall(
                        x=["전주 거래액"] + list(contrib.keys()) + ["기준주 거래액"],
                        measure=["absolute"] + ["relative"] * 4 + ["total"],
                        y=[v0] + [contrib[k] for k in contrib] + [0],
                        text=[won(v0)] + [_damt(contrib[k]) for k in contrib] + [won(v1)],
                        textposition="outside",
                        connector=dict(line=dict(color="#cbd5e1")),
                        increasing=dict(marker=dict(color=PALETTE["green"])),
                        decreasing=dict(marker=dict(color=PALETTE["red"])),
                        totals=dict(marker=dict(color=PALETTE["slate"])),
                    ))
                    wf.update_layout(**base_layout(h=380, title="지표 체인 기여 분해 (LMDI) — 전주 → 기준주"))
                    st.plotly_chart(wf, width="stretch")
                    _tot_d = v1 - v0
                    _big = max(contrib, key=lambda k: abs(contrib[k]))
                    _shr = (abs(contrib[_big]) / abs(_tot_d) * 100) if _tot_d else 0.0
                    st.markdown('<div class="appendix">거래액 = 발송량 × CTR × 주문CR × 객단가 항등식을 '
                                'LMDI(로그평균 디비지아 지수)로 분해한 거예요 — 네 기여의 합이 정확히 총 증감'
                                f'({_damt(_tot_d)})과 일치해요. 이번 증감의 최대 요인은 <b>{_big}</b>'
                                f'({_damt(contrib[_big])}, 총 증감 대비 크기 {_shr:.0f}%)예요.</div>',
                                unsafe_allow_html=True)
                else:
                    st.caption("지표 체인 분해는 전주·기준주 모두 발송·UV·주문·거래액이 0보다 커야 계산돼요.")

                # ── 가중 CTR 증감의 믹스 분해 — 진짜 효율 악화 vs 저효율 카테고리 비중 증가 ──
                _cats_mx = [c for c in cur_g.index.union(prv_g.index)
                            if str(c).strip() not in ("", "nan", "None")]
                s1t, s0t = float(cwd["send"].sum()), float(pwd["send"].sum())
                if s1t > 0 and s0t > 0 and _cats_mx:
                    _real = _mix = 0.0
                    for c in _cats_mx:
                        s1 = float(cur_g["send"].get(c, 0) or 0); s0 = float(prv_g["send"].get(c, 0) or 0)
                        u1 = float(cur_g["uv"].get(c, 0) or 0); u0 = float(prv_g["uv"].get(c, 0) or 0)
                        ctr1 = (u1 / s1) if s1 else 0.0
                        ctr0 = (u0 / s0) if s0 else 0.0
                        _real += (s0 / s0t) * (ctr1 - ctr0)
                        _mix += (s1 / s1t - s0 / s0t) * ctr1
                    _dctr = float(cwd["uv"].sum()) / s1t - float(pwd["uv"].sum()) / s0t

                    def _pp(v):
                        return f"△{abs(v)*100:.2f}%p" if v < 0 else f"+{v*100:.2f}%p"
                    st.markdown(f'<div class="appendix"><b>가중 CTR 전주 대비 {_pp(_dctr)}</b> = '
                                f'실질 효율 {_pp(_real)} + 카테고리 믹스 {_pp(_mix)} '
                                '(두 성분의 합은 총 증감과 일치). 믹스 성분이 크면 CTR 변화가 '
                                '문구·타깃 효율 문제가 아니라 카테고리 발송 비중 변화 때문이에요. '
                                '단, 이번 주 새로 시작하거나 중단한 카테고리의 효과는 실질 효율 '
                                '쪽에 섞일 수 있어요.</div>',
                                unsafe_allow_html=True)

        # ② 금주 하이라이트 · 로우라이트
        with tabH:
            st.markdown("##### 금주 주요 성과 지표 (Top 10 / Bottom 10)")
            hlab = st.radio("기준 지표", ["CTR", "주문CR", "거래액", "RPS"], horizontal=True, key="wr_hl_met")
            hcol = {"CTR": "infl_cr", "주문CR": "ord_cr", "거래액": "amt", "RPS": "rps"}[hlab]
            hw = _slice(ref_ws, ref_we)
            if "uv" in hw.columns:
                hw = hw[hw["uv"].fillna(0) >= 100]        # UV 적으면 전환율이 튀어서 제외
            if len(hw) < 3:
                st.info("기준주에 UV 100 이상 캠페인이 3건 미만이라 표시할 수 없어요.")
            else:
                _hc = ["date", "cat", "title", "send", "infl_cr", "ord_cr", "rps", "amt"]
                _hc = [c for c in _hc if c in hw.columns]
                _hrn = {"date": "날짜", "cat": "카테고리", "title": "제목", "send": "발송",
                        "infl_cr": "CTR", "ord_cr": "주문CR", "rps": "RPS", "amt": "거래액"}
                _hft = {"발송": "{:,.0f}", "CTR": "{:.2%}", "주문CR": "{:.2%}",
                        "RPS": "{:,.0f}", "거래액": "{:,.0f}"}
                hc1, hc2 = st.columns(2)
                with hc1:
                    st.markdown(f"**🏆 하이라이트 — {hlab} 상위 10**")
                    st.dataframe(hw.sort_values(hcol, ascending=False).head(10)[_hc]
                                 .rename(columns=_hrn).style.format(_hft),
                                 hide_index=True, width="stretch")
                with hc2:
                    st.markdown(f"**🧊 로우라이트 — {hlab} 하위 10**")
                    st.dataframe(hw.sort_values(hcol).head(10)[_hc]
                                 .rename(columns=_hrn).style.format(_hft),
                                 hide_index=True, width="stretch")
                st.caption("UV 100 미만 캠페인은 전환율 변동성이 커 제외하였습니다. "
                           "성과 우수(Top) 소구 패턴은 기획에 재활용하고, 저성과(Bottom) 요인은 발송 조건을 재점검하십시오.")

        # ③ 월말 마감 예상 (run-rate)
        with tabP:
            st.markdown("##### 월말 예상 실적 (Run-rate 기준)")
            elapsed = ref_end.day
            total_days = calendar.monthrange(ref_end.year, ref_end.month)[1]
            pm_full = _agg(_slice(datetime.date(pm_y, pm_m, 1),
                                  datetime.date(pm_y, pm_m, calendar.monthrange(pm_y, pm_m)[1])))
            _pyy = ref_end.year - 1
            py_full = _agg(_slice(datetime.date(_pyy, ref_end.month, 1),
                                  datetime.date(_pyy, ref_end.month,
                                                calendar.monthrange(_pyy, ref_end.month)[1])))

            # ── 요일 가중 run-rate — 경과 구간의 요일 구성(주말이 많았는지)을 보정 ──
            # 직전 12주 일별 실적으로 지표별 '요일 평균'을 만들고, 경과분/전체월을 그 가중치로
            # 환산해 예상한다. 이력이 부족하면(4주 미만) 단순 선형 run-rate로 폴백.
            _dcolmap = {"캠페인수": ("af", "size"), "발송": ("send", "sum"), "UV": ("uv", "sum"),
                        "주문건수": ("oc", "sum"), "거래액": ("amt", "sum")}
            _hist0 = m_first - datetime.timedelta(days=84)
            _hist = g0[(g0["dt"] >= pd.Timestamp(_hist0)) & (g0["dt"] < pd.Timestamp(m_first))]
            _dowm, _amt_sd = None, np.nan
            if len(_hist) >= 20 and _hist["dt"].dt.date.nunique() >= 28:
                _daily = _hist.groupby(_hist["dt"].dt.normalize()).agg(
                    **{k: v for k, v in _dcolmap.items()})
                _daily = _daily.reindex(
                    pd.date_range(_hist0, m_first - datetime.timedelta(days=1)), fill_value=0)
                _dowm = _daily.groupby(_daily.index.dayofweek).mean()
                _amt_res = _daily["거래액"] - _daily.index.dayofweek.map(_dowm["거래액"])
                _amt_sd = float(_amt_res.std(ddof=1))
            _mdays = pd.date_range(m_first, datetime.date(ref_end.year, ref_end.month, total_days))

            def _project(met, cv):
                """마감 예상값 — 요일 가중(이력 충분 시) 또는 선형 run-rate."""
                if not elapsed or pd.isna(cv):
                    return np.nan
                if _dowm is not None and met in _dowm.columns:
                    w = _dowm[met].reindex(range(7)).fillna(0.0)
                    exp_month = float(w.loc[_mdays.dayofweek].sum())
                    exp_sofar = float(w.loc[_mdays[:elapsed].dayofweek].sum())
                    if exp_sofar > 0 and exp_month > 0:
                        return cv * exp_month / exp_sofar
                return cv / elapsed * total_days
            _method = "요일 가중" if _dowm is not None else "단순 선형"
            st.caption(f"{ref_end.year}년 {ref_end.month}월 · 총 {total_days}일 중 "
                       f"{elapsed}일 경과({ref_end.month}/1~{ref_end.month}/{elapsed}) 기준 · "
                       f"추정 방식: **{_method} run-rate**"
                       + (" (직전 12주 요일별 평균으로 경과 구간의 요일 구성을 보정)" if _dowm is not None
                          else " (이력이 12주 미만이라 일평균×잔여일)")
                       + ". 월초일수록 경과일이 적어 추정 오차가 커요.")
            prow = []
            for met in ["캠페인수", "발송", "UV", "주문건수", "거래액"]:
                cv = cur_mtd[met]
                land = _project(met, cv)
                prow.append({"지표": met, "당월 MTD": _fmt(met, cv),
                             "일평균": _fmt(met, cv / elapsed if elapsed else np.nan),
                             "마감 예상": _fmt(met, land),
                             "전월 실적": _fmt(met, pm_full[met]),
                             "전월비(마감)": _dlt(met, land, pm_full[met]),
                             "전년 동월": _fmt(met, py_full[met]),
                             "전년비(마감)": _dlt(met, land, py_full[met])})
            st.dataframe(pd.DataFrame(prow).style.map(_clr, subset=["전월비(마감)", "전년비(마감)"]),
                         hide_index=True, width="stretch", height=240)
            remain = total_days - elapsed
            land_amt = _project("거래액", float(cur_mtd["거래액"]))
            _se = _amt_sd * np.sqrt(max(remain, 0)) if pd.notna(_amt_sd) else np.nan
            if pd.notna(land_amt) and pd.notna(_se) and _se > 0 and remain > 0:
                st.caption(f"거래액 마감 예상 **{won(land_amt)}** · 95% 구간 "
                           f"{won(max(land_amt - 1.96 * _se, 0))} ~ {won(land_amt + 1.96 * _se)} "
                           "(과거 일별 변동성 기준 — 대형 프로모션 등 특이 이벤트는 반영 못 해요)")
            tgt = st.number_input("월 거래액 목표 (억원 · 선택)", min_value=0.0, value=0.0,
                                  step=0.5, key="wr_target",
                                  help="목표를 입력하면 진척률·필요 일평균·달성 확률을 계산해 드려요.")
            if tgt > 0:
                tgt_won = tgt * 1e8
                cur_amt = float(cur_mtd["거래액"])
                st.progress(min(cur_amt / tgt_won, 1.0),
                            text=f"목표 대비 진척 {cur_amt / tgt_won * 100:.1f}% "
                                 f"({won(cur_amt)} / {won(tgt_won)})")
                _prob_txt = ""
                if pd.notna(land_amt) and pd.notna(_se) and _se > 0 and remain > 0:
                    _prob = float(stats.norm.cdf((land_amt - tgt_won) / _se)) * 100
                    _prob_txt = f" · 달성 확률 약 <b>{_prob:.0f}%</b>"
                if pd.notna(land_amt) and land_amt >= tgt_won:
                    _vd = (f"이 속도면 <b>달성 예상</b> ✅ (마감 예상 {won(land_amt)} ≥ 목표 "
                           f"{won(tgt_won)}){_prob_txt}")
                elif remain > 0:
                    need = (tgt_won - cur_amt) / remain
                    _vd = (f"이 속도면 마감 {won(land_amt)}로 <b>미달 예상</b> — 남은 {remain}일 동안 "
                           f"일평균 <b>{won(need)}</b> 필요 (현재 일평균 {won(cur_amt / elapsed)})"
                           f"{_prob_txt}")
                else:
                    _vd = f"월 마감 — 최종 {won(cur_amt)} / 목표 {won(tgt_won)}"
                st.markdown(f'<div class="appendix">{_vd}</div>', unsafe_allow_html=True)
            st.markdown('<div class="appendix">run-rate 추정은 월말 프로모션 같은 특이 이벤트를 '
                        '반영하지 않아요. 요일 가중 방식은 경과 구간에 주말이 몰렸을 때의 왜곡을 '
                        '줄여 주지만, 월초(경과 일수가 적을 때)일수록 오차가 커요.</div>',
                        unsafe_allow_html=True)

        # ── 최근 13주 추이 (발송량·CTR·주문CR·거래액) ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📈 최근 13주 추이")
        # 기준주가 부분 주면 추이에서 제외 — 마지막 점만 부분 데이터라 꼬리가 인위적으로
        # 급락하는 착시 방지 (KPI/표는 동요일 누계로 공정 비교하지만 추이는 풀주 창이라 불일치)
        _wk_upto = [w for w in sorted(g0["주"].unique()) if pd.Timestamp(w) <= ref_ws]
        # 2주 이상 있을 때만 부분 기준주를 제외 (1주뿐이면 빈 추이가 되므로 부분이라도 표시)
        _drop_ref = (_partial and len(_wk_upto) >= 2 and pd.Timestamp(_wk_upto[-1]) == ref_ws)
        if _drop_ref:
            _wk_upto = _wk_upto[:-1]
        wk13 = _wk_upto[-13:]
        trows = []
        for w in wk13:
            a = _agg(_slice(pd.Timestamp(w), pd.Timestamp(w) + pd.Timedelta(days=6)))
            a["주"] = pd.Timestamp(w)
            trows.append(a)
        tdf = pd.DataFrame(trows)
        if _drop_ref:
            st.caption("※ 진행 중(또는 실적 미완결)인 기준주는 부분 데이터라 추이에서 제외했어요 "
                       "— 완결된 주만 표시해요.")
        if len(tdf) < 2:
            st.info("추이를 그리려면 완결된 주가 2주 이상 필요해요. 실적을 더 쌓아 주세요.")
            st.stop()
        WRT_BAR = {"발송량": "발송", "거래액": "거래액", "캠페인수": "캠페인수"}
        WRT_LINE = {"CTR": ("CTR", "%"), "주문CR": ("주문CR", "%"), "RPS": ("RPS", "")}
        tsel1, tsel2 = st.columns(2)
        _bl = tsel1.selectbox("막대 지표 (위)", list(WRT_BAR), index=1, key="wr_t_bar")
        _ll = tsel2.selectbox("선 지표 (아래)", list(WRT_LINE), index=1, key="wr_t_line")
        _lc, _ls = WRT_LINE[_ll]
        fig = stacked_panels(tdf["주"], tdf[WRT_BAR[_bl]], _bl,
                             tdf[_lc] * (100 if _ls == "%" else 1), _ll,
                             PALETTE["slate"], PALETTE["purple"], h=430, line_suffix=_ls,
                             title=f"주차별 {_bl}(위) · {_ll}(아래) — 기준주까지 13주")
        st.plotly_chart(fig, width="stretch")
        # 13주 표 — 발송량·CTR·주문CR·거래액 한눈에 (기준주는 ★ 표시)
        tv = pd.DataFrame({
            "주차": tdf["주"].map(lambda w: ("★ " if pd.Timestamp(w) == ref_ws else "") + _wklab(w)),
            "캠페인수": tdf["캠페인수"].map(lambda v: f"{v:,.0f}"),
            "발송량": tdf["발송"].map(lambda v: f"{v:,.0f}"),
            "CTR": tdf["CTR"].map(lambda v: "–" if pd.isna(v) else f"{v*100:.2f}%"),
            "주문CR": tdf["주문CR"].map(lambda v: "–" if pd.isna(v) else f"{v*100:.2f}%"),
            "RPS": tdf["RPS"].map(won),
            "거래액": tdf["거래액"].map(won),
        }).iloc[::-1]                                     # 최신 주가 위로
        st.dataframe(tv, hide_index=True, width="stretch", height=min(38 + 35 * len(tv), 500))

        # ── 카테고리별 기준주 실적 (전주 대비) — 행 클릭 시 하단에 메시지 상세 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 🗂 카테고리별 기준주 실적 — 전주 대비")
        st.caption("행을 클릭하면 아래에 그 카테고리의 기준주 메시지별 효율 상세가 떠요.")
        cw = _slice(ref_ws, ref_we)
        pw = _slice(ref_ws - pd.Timedelta(days=7), ref_we - pd.Timedelta(days=7))
        if "cat" in cw.columns and len(cw):
            def _bycat(d):
                return d.groupby("cat").agg(캠페인수=("cat", "size"), 발송=("send", "sum"),
                                            UV=("uv", "sum"), 주문=("oc", "sum"), 거래액=("amt", "sum"))
            ca, cb = _bycat(cw), _bycat(pw)
            crows, cat_names = [], []
            for cname in ca.sort_values("거래액", ascending=False).index:
                if str(cname).strip() in ("", "nan", "None"):
                    continue
                s_c, u_c, o_c = ca.loc[cname, "발송"], ca.loc[cname, "UV"], ca.loc[cname, "주문"]
                amt_c = ca.loc[cname, "거래액"]
                s_p = cb.loc[cname, "발송"] if cname in cb.index else np.nan
                amt_p = cb.loc[cname, "거래액"] if cname in cb.index else np.nan
                ctr_c = (u_c / s_c) if s_c else np.nan
                cr_c = (o_c / u_c) if u_c else np.nan
                cat_names.append(str(cname))
                crows.append({"카테고리": cname,
                              "캠페인수": f"{ca.loc[cname, '캠페인수']:,.0f}",
                              "발송량": f"{s_c:,.0f}",
                              "발송 전주비": _dlt("발송", s_c, s_p),
                              "CTR": (f"{ctr_c*100:.2f}%" if pd.notna(ctr_c) else "–"),
                              "주문CR": (f"{cr_c*100:.2f}%" if pd.notna(cr_c) else "–"),
                              "거래액": won(amt_c),
                              "거래액 전주비": _dlt("거래액", amt_c, amt_p)})
            if crows:
                _cstyled = pd.DataFrame(crows).style.map(_clr, subset=["발송 전주비", "거래액 전주비"])
                try:
                    _evc = st.dataframe(_cstyled, hide_index=True, width="stretch", height=330,
                                        key="wr_cat_tbl", on_select="rerun",
                                        selection_mode="single-row")
                except TypeError:
                    _evc = None
                    st.dataframe(_cstyled, hide_index=True, width="stretch", height=330)
                # 클릭한 행 → 카테고리 선택값에 반영 (수동 선택도 유지)
                _cpick = None
                try:
                    _crows_sel = _evc.selection["rows"] if _evc is not None else []
                    if _crows_sel:
                        _cpick = int(_crows_sel[0])
                except Exception:
                    _cpick = None
                if (_cpick is not None and 0 <= _cpick < len(cat_names)
                        and _cpick != st.session_state.get("_lastpick_wrcat")):
                    st.session_state["_lastpick_wrcat"] = _cpick
                    st.session_state["wr_cat_sel"] = cat_names[_cpick]
                if st.session_state.get("wr_cat_sel") not in cat_names:
                    st.session_state.pop("wr_cat_sel", None)
                sel_cat_wr = st.selectbox("카테고리 선택 (표에서 행을 클릭해도 돼요)",
                                          cat_names, key="wr_cat_sel")
                if sel_cat_wr:
                    sub_wr = cw[cw["cat"].astype(str) == sel_cat_wr]
                    st.markdown(f"##### 📋 '{sel_cat_wr}' — 기준주 메시지별 효율 상세")
                    st.caption(f"{_md(_wklab(ref_ws))} 발송 {len(sub_wr)}건 — 주문CR 높은 순. "
                               "표의 행을 클릭하면 문구 원문도 볼 수 있어요.")
                    render_messages(sub_wr, "ord_cr", f"wrcat_{sel_cat_wr}")
        else:
            st.caption("카테고리 데이터가 없어요.")

        # ── AI 주간보고 코멘트 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 🤖 AI 주간보고 코멘트")

        def _wr_ai_lines():
            lines = [f"[기준주] {_wklab(ref_ws)}"]
            for met in METS:
                yv = yoy_w[met] if yoy_w else np.nan
                lines.append(f"- {met}: {_fmt(met, cur_w[met])} "
                             f"(전주비 {_dlt(met, cur_w[met], prev_w[met])}, "
                             f"전월비(전월 동주) {_dlt(met, cur_w[met], pm_w[met])}, "
                             f"전년비 {_dlt(met, cur_w[met], yv)})")
            lines.append(f"[MTD] 당월 거래액 {_fmt('거래액', cur_mtd['거래액'])} "
                         f"(전월비 {_dlt('거래액', cur_mtd['거래액'], prev_mtd['거래액'])}, "
                         f"전년비 {_dlt('거래액', cur_mtd['거래액'], yoy_mtd['거래액'])})")
            return lines
        with st.expander("🔍 AI에 전달되는 데이터 미리보기 (외부 API로 전송돼요)"):
            st.text("\n".join(_wr_ai_lines()))
        if st.button("코멘트 생성", key="wr_ai"):
            lines = _wr_ai_lines()
            system = ("당신은 LF몰 CRM 발송 주간보고 작성자입니다. 주어진 수치만 근거로 임원 보고용 "
                      "요약 코멘트를 한국어로 작성하세요: 1) 금주 총평 2줄, "
                      "2) 좋았던 점·주의할 점 각 2개(수치 인용), 3) 다음 주 액션 제안 2개. "
                      "수치를 지어내지 마세요. 출력은 HTML, 항목은 <br>로 구분, "
                      "긍정은 <span style='color:#16a34a;font-weight:700'>…</span>, "
                      "부정/주의는 <span style='color:#dc2626;font-weight:700'>…</span>로 감싸세요.")
            with st.spinner("생성 중…"):
                txt, err = ai_generate(system, "\n".join(lines), model)
            if err:
                st.warning(err)
            else:
                st.session_state["wr_ai_txt"] = txt
        if st.session_state.get("wr_ai_txt"):
            st.markdown(f'<div class="vg">{safe_ai_html(st.session_state["wr_ai_txt"])}</div>', unsafe_allow_html=True)

        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 02 — 문구 속성별 성과 (핵심)
    # ══════════════════════════════════════════════════════════════
    elif "문구 속성별" in page:
        st.title("문구 속성별 성과")
        st.caption("각 소구 속성이 있을 때와 없을 때 성과가 얼마나 다른지 비교해요.")
        mlabel = st.selectbox("성과 지표", list(METRIC_OPTS.keys()))
        mcol, msuf, mclr = METRIC_OPTS[mlabel]
        base = fdf
        if len(base) < 6:
            st.info("데이터가 부족해요. '최소 발송수'를 낮춰 보세요."); st.stop()

        rows = []
        for tag in TAG_BOOLS:
            if tag not in base.columns:               # 구버전 캐시 등으로 태그 컬럼이 없으면 건너뜀
                continue
            yes = base[base[tag]][mcol].dropna().values
            no = base[~base[tag]][mcol].dropna().values
            if len(yes) == 0 or len(no) == 0:
                continue
            my, mn = float(np.mean(yes)), float(np.mean(no))
            rows.append(dict(속성=tag, 보유평균=my, 보유n=len(yes), 미보유평균=mn, 미보유n=len(no),
                             차이=my - mn, p=welch(yes, no), d=cohen_d(yes, no)))
        adf = pd.DataFrame(rows)
        if len(adf) == 0:
            st.info("속성 비교에 필요한 표본이 없어요. 필터를 넓히거나 '최소 발송수'를 낮춰 보세요."); st.stop()
        adf = adf.sort_values("차이", ascending=False)
        adf["p_adj"] = fdr_bh(adf["p"].values)  # 다중비교(FDR) 보정

        # Δ 막대
        is_pct = mcol in ("ord_cr", "infl_cr")
        delta_disp = adf["차이"] * (100 if is_pct else 1)
        fig = go.Figure(go.Bar(
            x=delta_disp, y=adf["속성"], orientation="h",
            marker_color=[PALETTE["green"] if v >= 0 else PALETTE["red"] for v in adf["차이"]],
            text=[f"{v:+.2f}{'%p' if is_pct else ''}" for v in delta_disp], textposition="outside"))
        fig.update_layout(**base_layout(h=360, title=f"{mlabel} — 속성 보유 시 평균 차이 (보유 − 미보유)"))
        fig.update_yaxes(autorange="reversed")        # 아래 순효과 차트와 같은 방향(큰 값이 위)
        st.plotly_chart(fig, width="stretch")

        def fmtv(v):
            return f"{v*100:.2f}%" if is_pct else (won(v) if mcol in ("rps", "aov", "amt") else f"{v:,.1f}")
        def _dlabel(d):
            if d is None or (isinstance(d, float) and np.isnan(d)):
                return "–"
            mag = "큼" if abs(d) >= 0.8 else ("중간" if abs(d) >= 0.5 else ("작음" if abs(d) >= 0.2 else "미미"))
            return f"{d:+.2f} ({mag})"
        show = adf.copy()
        show["보유평균"] = show["보유평균"].map(fmtv)
        show["미보유평균"] = show["미보유평균"].map(fmtv)
        show["차이"] = [f"{v:+.2f}%p" if is_pct else f"{v:+,.0f}" for v in delta_disp]
        show["효과크기"] = show["d"].map(_dlabel)
        show["유의성(보정)"] = show["p_adj"].map(sig_label)
        st.dataframe(show[["속성", "보유평균", "보유n", "미보유평균", "미보유n", "차이", "효과크기", "유의성(보정)"]],
                     hide_index=True, width="stretch")
        st.markdown('<div class="appendix">속성은 제목·본문에서 자동으로 분류한 거예요. '
                    '<b>효과크기</b>(d)는 차이가 실제로 얼마나 큰지(0.2 작음·0.5 중간·0.8 큼), '
                    '<b>유의성(보정)</b>은 여러 개를 한꺼번에 비교할 때 우연을 걸러낸 값이에요. '
                    '건수(n)가 적으면 우연일 수 있으니 함께 봐 주세요.</div>',
                    unsafe_allow_html=True)

        # ── 다변량 회귀: 교란(카테고리·시간대) 통제 후 속성 순효과 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 🧮 다른 조건을 맞춘 뒤 속성의 진짜 효과")
        st.caption("단순 비교는 '특정 카테고리에 몰린' 착시가 섞일 수 있어요. "
                   "카테고리·발송유형·시간대를 맞춘 뒤 각 속성만의 순수 효과를 봐요.")
        ctrl_opts = [c for c in ["cat", "stype", "hour", "dow_k", "bpu"] if c in base.columns]
        ctrl_label = {"cat": "카테고리", "stype": "발송유형", "hour": "시간대", "dow_k": "요일", "bpu": "BPU"}
        sel_ctrl = st.multiselect("통제할 변수", ctrl_opts, default=[c for c in ["cat", "stype", "hour"] if c in ctrl_opts],
                                  format_func=lambda c: ctrl_label.get(c, c), key="p02_ctrl")
        eff = ols_effects(base, TAG_BOOLS, sel_ctrl, mcol)
        if len(eff) == 0:
            st.info("데이터가 부족해서 분석할 수 없어요. '최소 발송수'를 낮춰 보세요.")
        else:
            eff["p_adj"] = fdr_bh(eff["p"].values)
            ev = eff["순효과"] * (100 if is_pct else 1)
            figr = go.Figure(go.Bar(
                x=ev, y=eff["속성"], orientation="h",
                marker_color=[PALETTE["green"] if v >= 0 else PALETTE["red"] for v in eff["순효과"]],
                text=[f"{v:+.3f}{'%p' if is_pct else ''}" for v in ev], textposition="outside"))
            figr.update_layout(**base_layout(h=380, title=f"{mlabel} — 속성 순효과 (교란 통제)"))
            figr.update_yaxes(autorange="reversed")
            st.plotly_chart(figr, width="stretch")
            er = eff.copy()
            er["순효과"] = [f"{v*100:+.3f}%p" if is_pct else (f"{v:+,.0f}" if mcol in ("rps", "aov", "amt") else f"{v:+,.3f}")
                          for v in eff["순효과"]]
            er["유의성(보정)"] = er["p_adj"].map(sig_label)
            st.dataframe(er[["속성", "순효과", "유의성(보정)"]], hide_index=True, width="stretch")
            st.markdown('<div class="appendix">순효과가 +이고 유의하면, 다른 조건이 같아도 그 속성이 성과를 끌어올려요. '
                        '단순 비교에선 좋아 보였는데 여기서 약하면 다른 요인이 섞인 거예요.</div>', unsafe_allow_html=True)

        # ── 속성 조합 패턴 분석 (할인율소구+마감임박 등) ──
        import itertools
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 🔗 속성 조합 — 어떤 조합이 잘 먹힐까")
        st.caption("소구 속성 2~3개를 동시에 쓴 캠페인의 평균 성과예요. "
                   "건수가 적으면 우연일 수 있으니 캠페인 수도 같이 봐 주세요.")
        cc1, cc2 = st.columns([2, 1])
        cmin = cc1.number_input("조합 최소 표본 (캠페인 수)", value=5, min_value=2, step=1, key="p02_cmin")
        ksize = cc2.radio("조합 크기", [2, 3], horizontal=True, key="p02_ksize")
        base_mean = float(base[mcol].mean())
        crows = []
        for combo in itertools.combinations(TAG_BOOLS, ksize):
            if any(t not in base for t in combo):
                continue
            mask = np.logical_and.reduce([base[t].values for t in combo])
            sub = base.loc[mask, mcol].dropna()
            if len(sub) < cmin:
                continue
            rest = base.loc[~mask, mcol].dropna()
            p = welch(sub.values, rest.values) if len(rest) > 1 else None
            row = dict(조합="+".join(combo), 캠페인수=len(sub),
                       평균=float(sub.mean()), 차이=float(sub.mean()) - base_mean, p=p)
            if ksize == 2:
                # 시너지(DID) — 'A도 좋고 B도 좋아서'가 아니라 '같이 써서 더 좋은지'를 분리.
                # 시너지 = m(A∩B) − [m(A만) + m(B만) − m(둘다없음)] (주효과 합 대비 초과분)
                _a, _b = combo
                va = base[_a].fillna(False).astype(bool).values
                vb = base[_b].fillna(False).astype(bool).values
                mv = base[mcol].values.astype(float)

                def _cellmean(cm):
                    v = mv[cm]
                    v = v[~np.isnan(v)]
                    return float(v.mean()) if len(v) >= 3 else np.nan
                m10 = _cellmean(va & ~vb)
                m01 = _cellmean(~va & vb)
                m00 = _cellmean(~va & ~vb)
                if not any(pd.isna(v) for v in (m10, m01, m00)):
                    row["기대치"] = m10 + m01 - m00
                    row["시너지"] = float(sub.mean()) - row["기대치"]
            crows.append(row)
        if not crows:
            st.info("조건에 맞는 조합이 없어요. 기준을 낮추거나 '최소 발송수'를 줄여 보세요.")
        else:
            cdf = pd.DataFrame(crows).sort_values("평균", ascending=False).reset_index(drop=True)
            topn = cdf.head(12)
            yv = topn["평균"] * (100 if is_pct else 1)
            figc = go.Figure(go.Bar(
                x=yv, y=topn["조합"], orientation="h", marker_color=mclr,
                customdata=topn["캠페인수"],
                text=[f"{v:.2f}{'%' if is_pct else ''} (n={int(n)})" for v, n in zip(yv, topn["캠페인수"])],
                textposition="outside",
                hovertemplate="%{y}<br>평균 " + mlabel + ": %{x:.2f}<br>캠페인수: %{customdata}<extra></extra>"))
            lay = base_layout(h=440, title=f"속성 {ksize}개 조합별 평균 {mlabel} (상위 12 · 전체평균 점선)")
            lay["xaxis"]["range"] = bar_xrange(yv, 1.18)
            figc.update_layout(**lay)
            figc.update_yaxes(autorange="reversed")
            bm = base_mean * (100 if is_pct else 1)
            figc.add_vline(x=bm, line_dash="dot", line_color="#94a3b8",
                           annotation_text=f"전체평균 {bm:.2f}", annotation_position="top")
            st.plotly_chart(figc, width="stretch")

            cshow = cdf.copy()
            if is_pct:
                cshow["평균"] = cshow["평균"].map(lambda v: f"{v*100:.2f}%")
                cshow["차이"] = cshow["차이"].map(lambda v: f"{v*100:+.2f}%p")
            elif mcol in ("rps", "aov", "amt"):
                cshow["평균"] = cshow["평균"].map(won)
                cshow["차이"] = cshow["차이"].map(lambda v: f"{v:+,.0f}")
            else:
                cshow["평균"] = cshow["평균"].map(lambda v: f"{v:,.1f}")
                cshow["차이"] = cshow["차이"].map(lambda v: f"{v:+,.1f}")
            cshow["유의성"] = cdf["p"].map(sig_label)
            _combo_cols = ["조합", "캠페인수", "평균", "차이", "유의성"]
            if "시너지" in cdf.columns and cdf["시너지"].notna().any():
                def _fm_signed(v):
                    if pd.isna(v):
                        return "–"
                    if is_pct:
                        return f"{v*100:+.2f}%p"
                    return f"{v:+,.0f}" if mcol in ("rps", "aov", "amt") else f"{v:+,.1f}"
                cshow["기대치(주효과 합)"] = cdf["기대치"].map(
                    lambda v: "–" if pd.isna(v)
                    else (f"{v*100:.2f}%" if is_pct
                          else (won(v) if mcol in ("rps", "aov", "amt") else f"{v:,.1f}")))
                cshow["시너지"] = cdf["시너지"].map(_fm_signed)
                _combo_cols = ["조합", "캠페인수", "평균", "기대치(주효과 합)", "시너지", "유의성"]
            st.dataframe(cshow[_combo_cols].style.format({"캠페인수": "{:,.0f}"}),
                         hide_index=True, width="stretch", height=320)
            st.markdown('<div class="appendix">유의성은 해당 조합이 있는 그룹과 없는 그룹의 차이를 검정한 결과예요. '
                        '건수가 적으면 우연일 수 있어요. <b>시너지</b>(2개 조합)는 조합 평균에서 '
                        "'주효과 합(각자 단독 효과의 합)'을 뺀 값 — <b>+면 같이 써서 더 좋아지는 진짜 궁합</b>, "
                        '0 근처면 각자 좋은 걸 합친 것에 불과하다는 뜻이에요.</div>',
                        unsafe_allow_html=True)

            guard_select("p02_combo", list(cdf["조합"]))
            sel_combo = st.selectbox("조합을 골라 실제 메시지를 확인해 보세요", list(cdf["조합"]), key="p02_combo")
            if sel_combo:
                parts = sel_combo.split("+")
                mask = np.logical_and.reduce([base[t].values for t in parts])
                subc = base[mask]
                st.caption(f"'{sel_combo}' 동시 보유 {len(subc)}건 — {mlabel} 높은 순")
                render_messages(subc, mcol, f"combo_{sel_combo}")

        # ── 문구 길이 최적 구간 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📏 문구 길이별 성과")
        st.caption("제목/본문 글자수 구간별 평균 성과예요. 어느 길이가 가장 잘 먹히는지 확인해 보세요.")
        lc1, lc2 = st.columns(2)

        def len_bins(colname, label, container):
            if colname not in base:
                return
            b = base[base[colname].fillna(0) > 0].copy()
            if len(b) < 8:
                container.info(f"{label} 표본 부족"); return
            try:
                b["_bin"] = pd.qcut(b[colname], q=4, duplicates="drop")
            except Exception:
                container.info(f"{label} 구간화 불가"); return
            g = b.groupby("_bin", observed=True)[mcol].agg(["mean", "count", "std"])
            xs = [f"{int(iv.left)}~{int(iv.right)}자" for iv in g.index]
            yvv = g["mean"].values * (100 if is_pct else 1)
            _se = (g["std"] / np.sqrt(g["count"].clip(lower=1))).fillna(0).values
            _err = _se * 1.96 * (100 if is_pct else 1)          # 95% 신뢰구간 에러바
            fig = go.Figure(go.Bar(x=xs, y=yvv, marker_color=mclr,
                                   error_y=dict(type="data", array=_err,
                                                color="#94a3b8", thickness=1.2, width=4),
                                   text=[f"{v:.2f}<br>(n={int(n)})" for v, n in zip(yvv, g["count"])],
                                   textposition="outside"))
            fig.update_layout(**base_layout(h=300, ysuffix=("%" if is_pct else ""),
                                            title=f"{label} 길이 구간별 평균 {mlabel}"))
            container.plotly_chart(fig, width="stretch")

        len_bins("제목길이", "제목", lc1)
        len_bins("본문길이", "본문", lc2)

        # ── 최적점 추정 (2차항 회귀) — 구간 평균으론 '몇 자가 피크인지' 안 보인다 ──
        def _quad_opt(colname, label, unit="자"):
            if colname not in base:
                return None
            b = base[[colname, mcol]].dropna()
            b = b[b[colname] > 0]
            if len(b) < 30 or b[colname].nunique() < 8:
                return None
            x = b[colname].values.astype(float)
            y = b[mcol].values.astype(float)
            X = np.column_stack([np.ones_like(x), x, x ** 2])
            try:
                beta, *_ = np.linalg.lstsq(X, y, rcond=None)
                yhat = X @ beta
                dof = len(x) - 3
                if dof <= 0:
                    return None
                s2 = float(((y - yhat) ** 2).sum() / dof)
                cov = s2 * np.linalg.inv(X.T @ X)
                se2 = float(np.sqrt(cov[2, 2]))
                p2 = float(2 * (1 - stats.t.cdf(abs(beta[2] / se2), dof))) if se2 > 0 else 1.0
            except Exception:
                return None
            if beta[2] < 0 and p2 < 0.1:
                xstar = float(-beta[1] / (2 * beta[2]))
                lo, hi = float(np.percentile(x, 5)), float(np.percentile(x, 95))
                if lo <= xstar <= hi:
                    return f"**{label}**: 최적점 ≈ **{xstar:.0f}{unit}** (역U형 확인, 곡률 p={p2:.3f})"
                return f"**{label}**: 역U형이지만 최적점({xstar:.0f}{unit})이 관측 범위 밖 — 참고만"
            return f"**{label}**: 뚜렷한 최적점 없음 (단조 또는 무관)"

        _opt_msgs = [m_ for m_ in (_quad_opt("제목길이", "제목 길이"),
                                   _quad_opt("본문길이", "본문 길이"),
                                   _quad_opt("이모지수", "이모지 개수", unit="개")) if m_]
        if _opt_msgs:
            st.markdown("🎯 **최적점 추정 (2차항 회귀)** — " + " · ".join(_opt_msgs))
            st.caption("'많을수록 좋다가 어느 지점부터 꺾이는' 역U형이 통계적으로 확인될 때만 "
                       "최적점을 표시해요. 카테고리 구성이 섞인 단순 회귀라 참고 지표예요.")

        # ── 속성별 드릴다운: 실제 발송 메시지 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📋 속성별 실제 메시지 보기")
        avail_tags = [r["속성"] for _, r in adf.iterrows()] if len(adf) else TAG_BOOLS
        guard_select("p02_tag", avail_tags)
        sel_tag = st.selectbox("속성 선택", avail_tags, key="p02_tag")
        if sel_tag and sel_tag in base.columns:
            sub = base[base[sel_tag]].copy()
            st.caption(f"'{sel_tag}' 속성 보유 캠페인 {len(sub)}건 — {mlabel} 높은 순")
            render_messages(sub, mcol, f"p02_{sel_tag}")

        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 03 — 캠페인 리더보드
    # ══════════════════════════════════════════════════════════════
    elif "캠페인 리더보드" in page:
        st.title("캠페인 리더보드")
        # ── 담당자 벤치마크: 현재 필터 평균 vs 전체 평균 ──
        bench_pop = (raw[raw["matched"]] if only_matched else raw)
        bench_pop = bench_pop[bench_pop["send"].fillna(0) >= min_send]

        def _bm(col, pct=True):
            cur = float(fdf[col].mean()) if (col in fdf.columns and len(fdf)) else np.nan
            allv = float(bench_pop[col].mean()) if (col in bench_pop.columns and len(bench_pop)) else np.nan
            if np.isnan(cur):
                return "–", None
            if pct:
                d = None if np.isnan(allv) else f"{(cur-allv)*100:+.2f}%p vs 전체"
                return f"{cur*100:.2f}%", d
            d = None if np.isnan(allv) else f"{cur-allv:+,.0f} vs 전체"
            return won(cur), d
        st.caption(f"현재 조건 {len(fdf):,}건 vs 전체 {len(bench_pop):,}건 평균 — "
                   "사이드바에서 **담당자**(또는 카테고리·브랜드)로 필터하면, 본인 성과가 전체 대비 "
                   "높은지(초록)·낮은지(빨강) 한눈에 볼 수 있어요.")
        bc = st.columns(4)
        v, d = _bm("infl_cr"); bc[0].metric("평균 CTR", v, d)
        v, d = _bm("ord_cr"); bc[1].metric("평균 주문CR", v, d)
        v, d = _bm("rps", pct=False); bc[2].metric("평균 RPS", v, d)
        bc[3].metric("대상 캠페인 수", f"{len(fdf):,}")
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        mlabel = st.selectbox("정렬 지표", list(METRIC_OPTS.keys()))
        mcol = METRIC_OPTS[mlabel][0]
        asc = st.radio("정렬", ["높은순", "낮은순"], horizontal=True) == "낮은순"
        base = fdf.sort_values(mcol, ascending=asc)
        tagcols = [t for t in TAG_BOOLS if t in base.columns]
        base_r = base.reset_index(drop=True)
        view = base_r.copy()
        view["속성"] = view[tagcols].apply(lambda r: " ".join(t for t in tagcols if r[t]), axis=1)
        view["_bprev"] = view["body"].map(lambda x: " ".join(_s(x).split())[:60]) if "body" in view else ""
        cols = ["date", "cat", "brand", "title", "_bprev", "send", "infl_cr", "ord_cr", "rps", "amt", "속성"]
        ren = {"date": "날짜", "cat": "카테고리", "brand": "브랜드", "title": "제목", "_bprev": "내용",
               "send": "발송", "infl_cr": "CTR", "ord_cr": "주문CR", "rps": "RPS", "amt": "거래액"}
        _styled = view[cols].rename(columns=ren).style.format(
            {"발송": "{:,.0f}", "CTR": "{:.2%}", "주문CR": "{:.2%}", "RPS": "{:,.0f}", "거래액": "{:,.0f}"})
        try:
            _ev = st.dataframe(_styled, hide_index=True, width="stretch", height=560,
                               key="p03_tbl", on_select="rerun", selection_mode="single-row")
        except TypeError:
            _ev = None
            st.dataframe(_styled, hide_index=True, width="stretch", height=560)

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 🔍 실제 문구 확인")
        bb = base_r.head(40)
        opts = {f"{i+1}. [{r['date']}] {r['title'][:40]} (CR {r['ord_cr']*100:.2f}%)": i
                for i, r in bb.iterrows()}
        keys_list = list(opts.keys())
        picked = None
        try:
            _rows = _ev.selection["rows"] if _ev is not None else []
            if _rows:
                picked = int(_rows[0])
        except Exception:
            picked = None
        # 새로 클릭한 행이 상위 40건 안이면 셀렉트박스 기본값도 맞춘다(수동 선택 유지).
        if picked is not None and 0 <= picked < len(keys_list) and picked != st.session_state.get("_lastpick_p03"):
            st.session_state["_lastpick_p03"] = picked
            st.session_state["p03_msg"] = keys_list[picked]
        if st.session_state.get("p03_msg") not in opts:      # 정렬/지표 변경으로 무효가 된 선택값 제거
            st.session_state.pop("p03_msg", None)
        sel = st.selectbox("캠페인 선택 (표에서 행을 클릭해도 돼요)", keys_list, key="p03_msg")
        if picked is not None and 0 <= picked < len(base_r):
            r = base_r.iloc[picked]                       # 클릭한 행(전체 범위)
        elif sel is not None:
            r = bb.loc[opts[sel]]
        else:
            r = None
        if r is not None:
            _body = (esc(r["body"])
                     if ("body" in base_r.columns and pd.notna(r["body"]) and str(r["body"]).strip()) else "—")
            st.markdown(f'<div class="vg"><b>제목</b><br>{esc(r["title"])}<br><br>'
                        f'<b>내용</b><br>{_body}</div>', unsafe_allow_html=True)
        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 04 — 카테고리·시간대 매트릭스
    # ══════════════════════════════════════════════════════════════
    elif "카테고리" in page:
        st.title("카테고리·시간대")
        mlabel = st.selectbox("지표", list(METRIC_OPTS.keys()))
        mcol = METRIC_OPTS[mlabel][0]
        is_pct = mcol in ("ord_cr", "infl_cr")
        base = fdf

        def heat(df, idx, col, title, min_cell_n=3):
            pv = df.pivot_table(index=idx, columns=col, values=mcol, aggfunc="mean")
            if pv.empty:
                st.info("데이터가 부족해요"); return
            cnt = df.pivot_table(index=idx, columns=col, values=mcol, aggfunc="count")
            # 요일 축은 가나다순(금·목·수…)이 아니라 월~일 순으로 정렬
            _DOW = ["월", "화", "수", "목", "금", "토", "일"]
            if col == "dow_k":
                pv = pv.reindex(columns=[d for d in _DOW if d in pv.columns])
            if idx == "dow_k":
                pv = pv.reindex([d for d in _DOW if d in pv.index])
            cnt = cnt.reindex(index=pv.index, columns=pv.columns)
            # n<min_cell_n 셀 마스킹 — 캠페인 1건짜리 우연값이 가장 진한 색으로 강조되는 것 방지
            pv = pv.where(cnt >= min_cell_n)
            z = pv.values * (100 if is_pct else 1)
            if not np.isfinite(z).any():              # 전 셀이 마스킹되면 빈 히트맵이 '데이터 있는 척'
                st.info(f"표본이 충분한(캠페인 {min_cell_n}건 이상) 조합이 없어 히트맵을 그릴 수 없어요.")
                return
            n_arr = cnt.fillna(0).values
            txt_arr = np.where(np.isnan(z), "", np.round(z, 2).astype(str))
            fig = go.Figure(go.Heatmap(
                z=z, x=[str(c) for c in pv.columns], y=[str(i) for i in pv.index],
                colorscale="Blues", text=txt_arr, texttemplate="%{text}",
                textfont=dict(size=10), colorbar=dict(thickness=10),
                customdata=n_arr,
                hovertemplate="%{y} × %{x}<br>" + mlabel + ": %{z:.2f}"
                              + ("%" if is_pct else "") + " · n=%{customdata:.0f}<extra></extra>"))
            fig.update_layout(**base_layout(h=420, title=title))
            st.plotly_chart(fig, width="stretch")
            st.caption(f"빈 셀은 캠페인 {min_cell_n}건 미만이라 가린 구간이에요 (우연값 강조 방지) · "
                       "hover에 표본수(n)가 함께 떠요.")

        def _hod(h):  # 시간대 HHMM(800·1050·2200) → '08시'(시 단위로 묶기)
            try:
                v = int(float(h))
            except Exception:
                return "기타"
            hh = v if v <= 23 else v // 100
            return f"{hh:02d}시" if 0 <= hh <= 23 else "기타"

        heat(base, "cat", "stype", f"카테고리 × 발송유형 — 평균 {mlabel}")
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        _bh = base.assign(_hod=base["hour"].map(_hod)) if "hour" in base else base
        heat(_bh, "_hod", "dow_k", f"시간대(시 단위) × 요일 — 평균 {mlabel}")

        # ── 카테고리별 최적 문구 전략: 카테고리 × 문구속성 히트맵 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 🎯 카테고리별 잘 먹히는 소구")
        st.caption("카테고리마다 효과 좋은 소구가 달라요. 색이 진할수록 성과가 높아요.")
        present_tags = [t for t in TAG_BOOLS if t in base.columns]
        cat_rows = []
        cat_list = [c for c in base["cat"].dropna().unique() if str(c).strip() not in ("", "nan", "None")]
        cat_ns = []
        for c in cat_list:
            sub = base[base["cat"] == c]
            if len(sub) < 3:
                continue
            row = {"카테고리": str(c)}
            nrow = {"카테고리": str(c)}
            for t in present_tags:
                vv = sub[sub[t]][mcol].dropna()
                # 셀 n<3 은 마스킹 — 캠페인 1건짜리 우연값이 가장 진한 색이 되는 것 방지
                row[t] = float(vv.mean()) if len(vv) >= 3 else np.nan
                nrow[t] = len(vv)
            cat_rows.append(row)
            cat_ns.append(nrow)
        if cat_rows:
            cmat = pd.DataFrame(cat_rows).set_index("카테고리")
            nmat = pd.DataFrame(cat_ns).set_index("카테고리")
            z = cmat.values * (100 if is_pct else 1)
            txt_arr = np.where(np.isnan(z), "", np.round(z, 2).astype(str))
            fig = go.Figure(go.Heatmap(
                z=z, x=list(cmat.columns), y=list(cmat.index), colorscale="Blues",
                text=txt_arr, texttemplate="%{text}", textfont=dict(size=9),
                colorbar=dict(thickness=10), hoverongaps=False,
                customdata=nmat.values,
                hovertemplate="%{y} · %{x}<br>" + mlabel + ": %{z:.2f}"
                              + ("%" if is_pct else "") + " · n=%{customdata}<extra></extra>"))
            fig.update_layout(**base_layout(h=max(320, 60 + 34 * len(cmat)),
                                            title=f"카테고리 × 문구속성 — 평균 {mlabel}"))
            st.plotly_chart(fig, width="stretch")
            st.caption("빈 셀은 해당 조합 캠페인 3건 미만(우연값 방지) · hover에 표본수(n) 표시.")
            # 카테고리별 베스트 속성 추천표 — n 병기
            recs = []
            for c in cmat.index:
                rowv = cmat.loc[c].dropna()
                if len(rowv) == 0:
                    continue
                best = rowv.idxmax()
                bv = rowv.max() * (100 if is_pct else 1)
                bstr = f"{bv:.2f}%" if is_pct else (won(rowv.max()) if mcol in ("rps", "aov", "amt") else f"{bv:,.1f}")
                recs.append(dict(카테고리=c, 추천속성=best, 평균성과=bstr,
                                 표본수=int(nmat.loc[c, best])))
            if recs:
                st.markdown("**카테고리별 추천 소구** (표본 3건 이상만)")
                st.dataframe(pd.DataFrame(recs), hide_index=True, width="stretch")
        else:
            st.info("카테고리별 데이터가 부족해요.")

        # ── 드릴다운: 카테고리 / 시간대 / 요일 선택 → 메시지 목록 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📋 조건별 실제 메시지 보기")
        dc1, dc2, dc3 = st.columns(3)
        cat_opts = sorted(base["cat"].dropna().unique()) if "cat" in base else []
        hour_opts = sorted(base["hour"].dropna().unique()) if "hour" in base else []
        dow_opts = [d for d in ["월", "화", "수", "목", "금", "토", "일"] if d in base["dow_k"].values] if "dow_k" in base else []
        for _k, _o in (("p04_cat", ["전체"] + [str(c) for c in cat_opts]),
                       ("p04_hour", ["전체"] + [str(h) for h in hour_opts]),
                       ("p04_dow", ["전체"] + dow_opts)):
            guard_select(_k, _o)
        sel_cat_d = dc1.selectbox("카테고리", ["전체"] + [str(c) for c in cat_opts], key="p04_cat")
        sel_hour_d = dc2.selectbox("시간대", ["전체"] + [str(h) for h in hour_opts], key="p04_hour",
                                   format_func=lambda v: v if v == "전체" else _hm_label(v))
        sel_dow_d = dc3.selectbox("요일", ["전체"] + dow_opts, key="p04_dow")
        sub = base.copy()
        if sel_cat_d != "전체" and "cat" in sub:
            sub = sub[sub["cat"].astype(str) == sel_cat_d]
        if sel_hour_d != "전체" and "hour" in sub:
            sub = sub[sub["hour"].astype(str) == sel_hour_d]
        if sel_dow_d != "전체" and "dow_k" in sub:
            sub = sub[sub["dow_k"] == sel_dow_d]
        st.caption(f"조건 일치 {len(sub)}건 — {mlabel} 높은 순")
        render_messages(sub, mcol, "p04_drill")
        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 05 — 타이밍·피로도
    # ══════════════════════════════════════════════════════════════
    elif "타이밍" in page:
        st.title("타이밍·발송슬롯")
        mlabel = st.selectbox("지표", list(METRIC_OPTS.keys()))
        mcol = METRIC_OPTS[mlabel][0]
        is_pct = mcol in ("ord_cr", "infl_cr")
        base = fdf

        def barby(key, title, order=None):
            g = base.groupby(key)[mcol].mean()
            if order: g = g.reindex([o for o in order if o in g.index])
            y = g.values * (100 if is_pct else 1)
            xlab = [fmt_hhmm(i) for i in g.index] if key == "hour" else [str(i) for i in g.index]
            fig = go.Figure(go.Bar(x=xlab, y=y,
                                   marker_color=METRIC_OPTS[mlabel][2],
                                   text=[bar_label(v, mcol, is_pct) for v in y], textposition="outside"))
            fig.update_layout(**base_layout(h=320, ysuffix=("%" if is_pct else ""), title=title))
            st.plotly_chart(fig, width="stretch")

        cc = st.columns(2)
        with cc[0]: barby("hour", f"시간대별 평균 {mlabel}")
        with cc[1]: barby("dow_k", f"요일별 평균 {mlabel}", order=["월", "화", "수", "목", "금", "토", "일"])

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 발송량 구간별 성과")
        b = base[base["send"] > 0].copy()
        if len(b) >= 10:
            b["bin"] = pd.qcut(b["send"], q=5, labels=False, duplicates="drop")
            g = b.groupby("bin", observed=True)[mcol].mean().sort_index()
            g.index = [f"Q{i+1}" for i in range(len(g))]
            y = g.values * (100 if is_pct else 1)
            fig = go.Figure(go.Bar(x=[str(i) for i in g.index], y=y, marker_color=PALETTE["amber"],
                                   text=[bar_label(v, mcol, is_pct) for v in y], textposition="outside"))
            fig.update_layout(**base_layout(h=320, ysuffix=("%" if is_pct else ""),
                                            title=f"발송 규모 5분위(Q1 소량→Q5 대량)별 평균 {mlabel}"))
            st.plotly_chart(fig, width="stretch")
            st.markdown('<div class="appendix">대량 발송 구간(Q5)에서 효율이 떨어지면 더 보내는 게 오히려 '
                        '낮다는 신호예요. 단, 발송유형/카테고리 구성 차이가 섞일 수 있어요.</div>',
                        unsafe_allow_html=True)

        # ── 발송 슬롯 최적 추천 (요일 × 시간) ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 🏅 언제 보내면 좋을까 (요일 × 시간)")
        st.caption(f"요일·시간 조합별 평균 {mlabel}이에요. 충분한 건수가 있는 슬롯 중 효율이 높은 순이에요.")
        if "dow_k" in base and "hour" in base:
            smin = st.number_input("슬롯 최소 표본(캠페인 수)", value=3, min_value=2, step=1, key="p05_smin")
            g = base.dropna(subset=["hour"]).copy()
            g["hour"] = pd.to_numeric(g["hour"], errors="coerce")
            g = g.dropna(subset=["hour"])
            slot = g.groupby(["dow_k", "hour"]).agg(
                캠페인수=("af", "size"), 평균=(mcol, "mean"),
                발송=("send", "sum"), 거래액=("amt", "sum"),
                _uv=("uv", "sum"), _oc=("oc", "sum")).reset_index()
            # 정렬은 슬롯 합산 기준 표본 보정 점수 — n을 겨우 넘긴 슬롯의 요행 1위 방지
            _sf = slot.rename(columns={"_uv": "uv", "_oc": "oc", "발송": "send", "거래액": "amt"})
            slot["_rk"] = rank_adjusted(_sf, mcol, ascending=False).values
            slot = slot[slot["캠페인수"] >= smin].sort_values("_rk", ascending=False)
            if len(slot) == 0:
                st.info("조건에 맞는 슬롯이 없어요. 기준을 낮춰 보세요.")
            else:
                dow_order = ["월", "화", "수", "목", "금", "토", "일"]
                top = slot.head(10).copy()
                top["슬롯"] = top["dow_k"].astype(str) + " " + top["hour"].map(fmt_hhmm)
                yv = top["평균"] * (100 if is_pct else 1)
                figs = go.Figure(go.Bar(
                    x=yv, y=top["슬롯"], orientation="h", marker_color=METRIC_OPTS[mlabel][2],
                    text=[f"{v:.2f}{'%' if is_pct else ''} (n={int(n)})" for v, n in zip(yv, top["캠페인수"])],
                    textposition="outside"))
                lay = base_layout(h=380, title=f"효율 상위 발송 슬롯 — 평균 {mlabel}")
                lay["xaxis"]["range"] = bar_xrange(yv, 1.2)
                figs.update_layout(**lay)
                figs.update_yaxes(autorange="reversed")
                st.plotly_chart(figs, width="stretch")
                disp = slot.copy()
                disp["요일"] = disp["dow_k"]
                disp["시간"] = disp["hour"].map(fmt_hhmm)
                disp["_o"] = disp["요일"].map(lambda d: dow_order.index(d) if d in dow_order else 99)
                # 숫자를 문자열로 바꾸지 않고 column_config로 포맷 — 문자열화하면
                # 헤더 클릭 정렬이 사전순("9,900" > "10,000")으로 깨진다
                disp["평균"] = disp["평균"] * (100 if is_pct else 1)
                _avg_cfg = (st.column_config.NumberColumn("평균", format="%.2f%%") if is_pct
                            else st.column_config.NumberColumn("평균", format="localized"))
                st.dataframe(disp.sort_values(["_o", "hour"])[["요일", "시간", "캠페인수", "평균", "발송", "거래액"]],
                             hide_index=True, width="stretch", height=300,
                             column_config={
                                 "평균": _avg_cfg,
                                 "캠페인수": st.column_config.NumberColumn(format="localized"),
                                 "발송": st.column_config.NumberColumn(format="localized"),
                                 "거래액": st.column_config.NumberColumn(format="localized")})
                best = slot.iloc[0]
                bv = best["평균"] * (100 if is_pct else 1)
                bstr = f"{bv:.2f}%" if is_pct else (won(best["평균"]) if mcol in ("rps", "aov", "amt") else f"{bv:,.1f}")
                st.markdown(f'<div class="appendix">💡 추천: <b>{best["dow_k"]}요일 {fmt_hhmm(best["hour"])}</b> 슬롯 — '
                            f'평균 {mlabel} <b>{bstr}</b>(n={int(best["캠페인수"])}), 표본 크기까지 반영한 '
                            f'보정 순위 1위예요. 카테고리·상품 구성 차이가 섞일 수 있으니 표본수와 함께 보세요.</div>',
                            unsafe_allow_html=True)

        # ── 드릴다운: 시간대·요일 선택 → 메시지 목록 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📋 시간대·요일별 실제 메시지 보기")
        tc1, tc2 = st.columns(2)
        hour_opts5 = sorted(base["hour"].dropna().unique()) if "hour" in base else []
        dow_opts5 = [d for d in ["월", "화", "수", "목", "금", "토", "일"] if d in base["dow_k"].values] if "dow_k" in base else []
        guard_select("p05_hour", ["전체"] + [str(h) for h in hour_opts5])
        guard_select("p05_dow", ["전체"] + dow_opts5)
        sel_h5 = tc1.selectbox("시간대", ["전체"] + [str(h) for h in hour_opts5], key="p05_hour",
                               format_func=lambda v: v if v == "전체" else _hm_label(v))
        sel_d5 = tc2.selectbox("요일", ["전체"] + dow_opts5, key="p05_dow")
        sub5 = base.copy()
        if sel_h5 != "전체" and "hour" in sub5:
            sub5 = sub5[sub5["hour"].astype(str) == sel_h5]
        if sel_d5 != "전체" and "dow_k" in sub5:
            sub5 = sub5[sub5["dow_k"] == sel_d5]
        st.caption(f"조건 일치 {len(sub5)}건 — {mlabel} 높은 순")
        render_messages(sub5, mcol, "p05_drill")
        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 06 — AI 처방
    # ══════════════════════════════════════════════════════════════
    elif "AI 처방" in page:
        st.title("AI 처방·카피")
        st.caption("성과 데이터를 바탕으로 AI가 다음 캠페인 가이드를 만들어 드려요.")
        base = fdf
        if st.button("AI 처방 만들기", key="ai_rx"):
            facts = build_facts(base, with_attr=True)
            system = ("당신은 LF몰 CRM PUSH 카피라이팅 전략가입니다. 주어진 '문구 속성별 성과'와 "
                      "'상·하위 캠페인 문구'를 근거로 다음을 한국어로 작성하세요: "
                      "1) 성과를 가른 핵심 패턴 3가지, 2) 다음 캠페인에 권장하는 카피 공식(Do) 5개, "
                      "3) 피해야 할 패턴(Don't) 3개, 4) 바로 쓸 수 있는 제목 예시 3개. "
                      "수치를 지어내지 말고 데이터 근거를 함께 제시하세요. 출력은 HTML, 소제목은 <b>, "
                      "항목은 <br>로 구분.")
            with st.spinner("생성 중…"):
                txt, err = ai_generate(system, facts, model)
            if err: st.warning(err)
            else: st.session_state["ai_rx_txt"] = txt
        if st.session_state.get("ai_rx_txt"):
            st.markdown(f'<div class="vg">{safe_ai_html(st.session_state["ai_rx_txt"])}</div>', unsafe_allow_html=True)
        with st.expander("AI에 전달되는 데이터 미리보기"):
            st.text(build_facts(base, with_attr=True))

        # ── AI 카피 초안 생성 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### ✨ AI 카피 제안 — 성과 기반 메시징")
        st.caption("성과 좋았던 속성·조합을 참고해서 다음 PUSH 문구 초안을 만들어요.")
        dc1, dc2, dc3 = st.columns(3)
        cat_opts_ai = ["(전체)"] + [str(c) for c in sorted(base["cat"].dropna().unique())
                                    if str(c).strip() not in ("", "nan", "None")]
        attr_opts_ai = ["(전체)"] + ([str(x) for x in sorted(base["attr"].dropna().unique())
                                      if str(x).strip() not in ("", "nan", "None")] if "attr" in base else [])
        guard_select("ai_draft_cat_sel", cat_opts_ai)
        guard_select("ai_draft_attr_sel", attr_opts_ai)
        draft_cat_sel = dc1.selectbox("대상 카테고리", cat_opts_ai, key="ai_draft_cat_sel")
        draft_attr_sel = dc2.selectbox("대상 속성", attr_opts_ai, key="ai_draft_attr_sel",
                                       help="발송 속성(통합·정상·이월·입점·BPU 등)으로 범위를 좁힙니다.")
        draft_goal = dc3.selectbox("목표 지표", list(METRIC_OPTS.keys()), key="ai_draft_goal")
        lc1, lc2, lc3 = st.columns(3)
        title_len = lc1.number_input("제목 글자수(내외)", min_value=5, max_value=60, value=20, step=1,
                                     key="ai_draft_tlen")
        body_len = lc2.number_input("내용 글자수(내외)", min_value=10, max_value=200, value=45, step=5,
                                    key="ai_draft_blen")
        draft_n = lc3.slider("초안 개수", 3, 10, 5, key="ai_draft_n")
        st.caption("💡 추천: 제목 15~25자 · 내용 40~60자 — 모바일 PUSH에서 안 잘리는 길이예요.")
        draft_extra = st.text_area(
            "소구 내용·기획전 특성 (선택)", key="ai_draft_brand", height=70,
            placeholder="예: 헤리스 여름 린넨 30% / 한정수량 / 오늘 마감 — 적을수록 자유롭게, "
                        "여기 내용을 카피의 핵심 소재로 반영해요.",
            help="기획전 특성·혜택·소구 포인트를 적으면 그 내용을 바탕으로 카피를 만들어요.")
        if st.button("✨ 카피 초안 생성", key="ai_draft_btn"):
            gcol = METRIC_OPTS[draft_goal][0]
            scope = base
            if draft_cat_sel != "(전체)":
                scope = scope[scope["cat"].astype(str) == draft_cat_sel]
            if draft_attr_sel != "(전체)" and "attr" in scope:
                scope = scope[scope["attr"].astype(str) == draft_attr_sel]
            facts = build_facts(scope, with_attr=True, metric_col=gcol)
            ctx = f"대상 카테고리: {draft_cat_sel} / 대상 속성: {draft_attr_sel} / 목표 지표: {draft_goal}"
            extra_line = ""
            if draft_extra.strip():
                extra_line = (f"사용자가 제공한 소구 내용·기획전 특성: '{draft_extra.strip()}' "
                              "— 이 내용을 각 카피의 핵심 소재로 반드시 반영하세요. ")
            system = (
                "당신은 LF몰 CRM PUSH 카피라이터입니다. 주어진 '문구 속성별 성과'와 상·하위 문구 "
                "데이터를 근거로, 성과가 높았던 소구·속성 조합을 적용한 새 PUSH 문구 초안을 작성하세요. "
                f"요청 맥락: {ctx}. {extra_line}"
                f"다음을 한국어로: 1) 이 맥락에 권장하는 카피 전략 2~3줄(근거 속성 명시), "
                f"2) 바로 쓸 수 있는 PUSH 문구 초안 {draft_n}개 — 각 초안은 '제목'과 '내용'을 모두 포함하고 "
                f"제목은 약 {int(title_len)}자, 내용은 약 {int(body_len)}자 내외(공백 포함)로 길이를 맞추며, "
                "사용한 소구 속성을 [할인율소구+마감임박]처럼 태그로 표기. "
                "실제 데이터에 없는 구체 수치(가격·할인율)는 〇〇로 비워두세요. "
                "출력은 HTML, 소제목 <b>, 항목 <br> 구분.")
            with st.spinner("카피 초안 생성 중…"):
                txt, err = ai_generate(system, facts, model)
            if err:
                st.warning(err)
            else:
                st.session_state["ai_draft_txt"] = txt
        if st.session_state.get("ai_draft_txt"):
            st.markdown(f'<div class="vg">{safe_ai_html(st.session_state["ai_draft_txt"])}</div>', unsafe_allow_html=True)

        with st.expander("ℹ️ 카피 초안은 어떻게 만들어지나요? (참고)", expanded=False):
            st.markdown(
                "1. **선택값으로 분석 범위를 좁혀요** — `대상 카테고리`+`대상 속성`에 맞는 캠페인만 분석하고, "
                "`목표 지표` 기준으로 성과 좋았던/나빴던 문구를 가려내요.\n"
                "2. **성과 패턴을 정리해요** — 기간·평균 지표, **상위/하위 문구**, "
                "**소구 속성별 성과**(할인율소구·마감임박·쿠폰적립 등)를 요약해 AI에 전달해요.\n"
                "3. **소구 내용·기획전 특성을 반영해요** — 입력란에 적은 혜택·소구 포인트를 "
                "카피의 핵심 소재로 반영해요(비워두면 성과 패턴만으로 작성).\n"
                "4. **성과 높았던 소구 조합으로 새 PUSH 문구 초안**을 만들어요 "
                "(전략 + 제목/내용 + 소구 태그). 글자수에 맞춰 길이를 조정하고, "
                "데이터에 없는 가격·할인율은 〇〇로 비워둬요.\n\n"
                "> **구분** · `대상 속성`(통합·정상·이월 등)은 분석 **범위를 거르는 필터**이고, "
                "실제로 조합되는 건 **문구 소구 속성**(자동 태깅된 할인율소구·마감임박 등)이에요. "
                "’이 범위에서 목표를 잘 낸 패턴 + 입력한 소구 내용’을 근거로 카피를 만들어요.")
        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 08 — 전체 효율·추이 (send_dashboard 피로도 관점 계승)
    # ══════════════════════════════════════════════════════════════
    elif "전체 효율" in page:
        st.title("전체 효율·추이")
        st.caption("전체 발송을 주차 단위로 집계한 효율 흐름이에요. "
                   "사이드바 필터는 반영되고, '최소 발송수'는 제외돼요.")
        g = dff_all.dropna(subset=["dt"]).copy()
        g = g[g["send"].fillna(0) > 0]
        if len(g) < 3:
            st.info("데이터가 부족해요. 더 많은 주차를 올려 주세요."); st.stop()

        g["주"] = g["dt"].dt.to_period("W").apply(lambda p: p.start_time)
        rows = []
        for wkstart, d in g.groupby("주"):
            s, u, o, a = d["send"].sum(), d["uv"].sum(), d["oc"].sum(), d["amt"].sum()
            rows.append(dict(주=wkstart, 발송=s, 거래액=a, 캠페인수=len(d), 유입UV=u, 주문건수=o,
                             유입전환율=(u / s if s else np.nan),
                             주문전환율=(o / u if u else np.nan),
                             RPS=(a / s if s else np.nan)))
        wk = pd.DataFrame(rows).sort_values("주").reset_index(drop=True)
        # 발송이 미미한 주(오조작·이월 날짜 1~2건으로 생긴 뜬금없는 주)는 추세에서 제외 —
        # 중앙값 주 발송의 5% 미만이면 버림(x축이 통째로 늘어나 차트가 깨지는 것 방지)
        _wk_all = len(wk)
        if len(wk) >= 4:
            _med_send = wk["발송"].median()
            if _med_send and _med_send > 0:
                # 발송 극소(중앙값 5% 미만) 또는 캠페인 3건 미만(오날짜 1~2건) 주 제외
                _wk_f = wk[(wk["발송"] >= _med_send * 0.05)
                           & (wk["캠페인수"] >= 3)].reset_index(drop=True)
                if len(_wk_f) >= 2:          # 필터가 최소 2주는 남길 때만 적용(과필터 방지)
                    wk = _wk_f
        _wk_dropped = _wk_all - len(wk)

        c = st.columns(4)
        c[0].metric("누적 발송", won(g["send"].sum()))
        c[1].metric("누적 거래액", won(g["amt"].sum()))
        c[2].metric("전체 가중 주문전환율",
                    f"{(g['oc'].sum()/g['uv'].sum())*100:.2f}%" if g["uv"].sum() else "–")
        c[3].metric("전체 RPS", won(g["amt"].sum() / g["send"].sum() if g["send"].sum() else np.nan))

        # 주차별 이중축 — 좌(막대)/우(선) 지표 선택해 상관관계 확인
        WKM = {"발송량": ("발송", False), "거래액": ("거래액", False), "캠페인수": ("캠페인수", False),
               "유입UV": ("유입UV", False), "주문건수": ("주문건수", False),
               "CTR(유입전환율)": ("유입전환율", True), "주문전환율": ("주문전환율", True),
               "RPS": ("RPS", False)}
        _keys = list(WKM.keys())
        mc1, mc2 = st.columns(2)
        llab = mc1.selectbox("막대 지표 (위)", _keys, index=_keys.index("발송량"), key="p08_wk_left")
        rlab = mc2.selectbox("선 지표 (아래)", _keys, index=_keys.index("RPS"), key="p08_wk_right")
        lc, lpct = WKM[llab]; rc, rpct = WKM[rlab]
        # 좌(막대) 우(선) 이중축 오버레이 차트로 변경
        fig = overlay_dual(wk["주"], wk[lc] * (100 if lpct else 1), llab,
                           wk[rc] * (100 if rpct else 1), rlab,
                           PALETTE["slate"], PALETTE["green"], h=430,
                           bar_suffix=("%" if lpct else ""), line_suffix=("%" if rpct else ""),
                           title=f"{llab}(좌·막대) ↔ {rlab}(우·선)")
        st.plotly_chart(fig, width="stretch")
        if _wk_dropped:
            st.caption(f"발송이 미미한 주 {_wk_dropped}개는 추세에서 제외했어요 "
                       "(오조작·이월 날짜로 생긴 뜬금없는 주 — x축 왜곡 방지).")

        # 선택한 두 지표의 상관관계
        x = wk[lc].values.astype(float); y = wk[rc].values.astype(float)
        m = ~np.isnan(x) & ~np.isnan(y)
        if m.sum() >= 5 and np.std(x[m]) > 0 and np.std(y[m]) > 0:
            r = float(np.corrcoef(x[m], y[m])[0, 1])
            _, _, _, p, _ = stats.linregress(x[m], y[m])
            arrow = "같이 움직임(양의 상관)" if r > 0 else "반대로 움직임(음의 상관)"
            st.markdown(
                f'<div class="appendix"><b>상관관계</b> — 주차 {llab} ↔ {rlab}: 상관 r={r:.2f} ({arrow}), '
                f'{sig_label(p)}. r이 +1/−1에 가까울수록 강하고 0이면 관계가 없어요. '
                f'상관은 인과가 아니며 시즌·구성 변화가 섞일 수 있어요 '
                f'(예: 발송량↑인데 RPS·전환율↓이면 발송 피로도 신호).</div>', unsafe_allow_html=True)
        elif m.sum() < 5:
            st.caption("상관을 계산하려면 5주차 이상 필요해요.")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 주차별 실적")
        show = wk.copy()
        show["주"] = show["주"].dt.strftime("%Y-%m-%d")
        st.dataframe(show.style.format({
            "발송": "{:,.0f}", "거래액": "{:,.0f}", "캠페인수": "{:,.0f}",
            "유입전환율": "{:.2%}", "주문전환율": "{:.2%}", "RPS": "{:,.0f}"}),
            hide_index=True, width="stretch", height=360)
        st.markdown("<div class=\"appendix\">‘인당 발송 건수’ 기반 피로도(고객 중복 제거)는 이 데이터만으론 계산되지 않습니다 "
                    "— 전사 MTD 발송상세가 필요해요. 여기서는 캠페인 합산 기준 전체 효율을 봐요.</div>",
                    unsafe_allow_html=True)

        # ── 주차별 드릴다운: 해당 주의 캠페인 목록 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📋 주차별 캠페인 보기")
        wk_labels = [f"{r['주'].strftime('%Y-%m-%d')} (캠페인 {r['캠페인수']:.0f}건)" for _, r in wk.iterrows()]
        if wk_labels:
            guard_select("p08_wk", wk_labels)
            sel_wk = st.selectbox("주차 선택", wk_labels, key="p08_wk")
            wk_idx = wk_labels.index(sel_wk)
            wk_start = wk.iloc[wk_idx]["주"]
            wk_end = wk_start + pd.Timedelta(days=7)
            sub8 = g[(g["dt"] >= wk_start) & (g["dt"] < wk_end)]
            st.caption(f"해당 주 캠페인 {len(sub8)}건 — 거래액 높은 순")
            render_messages(sub8, "amt", "p08_drill")

        # ── 퍼널 분해 (발송→UV→VISIT→고객→주문) ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 🪜 퍼널 분석 — 어디서 빠질까")
        st.caption("발송 → UV → 주문고객 → 주문 단계별 전환율이에요. "
                   "전환율이 급락하는 단계가 개선 포인트예요.")
        steps = [("발송", "send"), ("UV(유입)", "uv"),
                 ("주문고객수", "cust"), ("주문", "oc")]
        avail = [(lab, c) for lab, c in steps if c in g.columns and g[c].fillna(0).sum() > 0]
        if len(avail) >= 2:
            vals = [float(g[c].fillna(0).sum()) for _, c in avail]
            labs = [lab for lab, _ in avail]
            base_v = vals[0] if vals[0] else 1
            # 로그 폭 퍼널은 '좁아지는 정도'가 실제 전환율과 무관해 모양 자체가 왜곡 —
            # 막대 길이 = '직전 단계 대비 전환율'인 수평 막대로 교체 (라벨 = 실제 건수)
            rates = [100.0] + [(vals[i] / vals[i - 1] * 100 if vals[i - 1] else 0.0)
                               for i in range(1, len(vals))]
            figf = go.Figure(go.Bar(
                x=rates, y=labs, orientation="h",
                marker_color=[PALETTE["slate"], PALETTE["blue"], PALETTE["teal"],
                              PALETTE["green"], PALETTE["purple"]][:len(labs)],
                text=[(f"{v:,.0f}건" + ("" if i == 0 else f" · 직전 대비 {rates[i]:.2f}%"))
                      for i, v in enumerate(vals)],
                textposition="outside",
                hovertemplate="%{y}<br>직전 단계 대비: %{x:.2f}%<extra></extra>"))
            layf2 = base_layout(h=340, title="단계별 전환 — 막대 길이 = 직전 단계 대비 전환율(%)")
            layf2["xaxis"]["range"] = [0, 130]
            layf2["xaxis"]["ticksuffix"] = "%"
            figf.update_layout(**layf2)
            figf.update_yaxes(autorange="reversed")
            st.plotly_chart(figf, width="stretch")
            # 단계별 전환율 표
            frows = []
            for i in range(1, len(avail)):
                prev_l, prev_c = avail[i - 1]; cur_l, cur_c = avail[i]
                pv = float(g[prev_c].fillna(0).sum()); cv = float(g[cur_c].fillna(0).sum())
                frows.append(dict(단계=f"{prev_l} → {cur_l}", 직전대비=f"{(cv/pv if pv else 0)*100:.2f}%",
                                  발송대비=f"{(cv/base_v)*100:.2f}%"))
            st.dataframe(pd.DataFrame(frows), hide_index=True, width="stretch")
            st.markdown("<div class=\"appendix\">’직전대비’가 가장 낮은 단계가 병목이에요. UV→주문이 낮으면 "
                        "랜딩·오퍼·상품 매력도, 발송→UV가 낮으면 제목·발송시점·타겟이 개선 포인트예요.</div>",
                        unsafe_allow_html=True)
        else:
            st.info("퍼널 분석에 필요한 데이터(UV·방문·고객·주문)가 부족해요.")

        # ── 기간 비교 (직전 대비) ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📈 기간 비교 — 최근 vs 직전")
        st.caption("기간을 둘로 나눠서 효율이 어떻게 바뀌었는지 봐요.")
        gdt = g.dropna(subset=["dt"]).sort_values("dt")
        if gdt["dt"].nunique() < 2:
            st.info("발송일이 2일 이상 있어야 비교할 수 있어요.")
        else:
            uniq_days = sorted(gdt["dt"].dt.normalize().unique())
            mid = uniq_days[len(uniq_days) // 2]
            prev = gdt[gdt["dt"] < mid]; recent = gdt[gdt["dt"] >= mid]
            if len(prev) == 0 or len(recent) == 0:
                st.info("기간을 나누기에 데이터가 부족해요.")
            else:
                def _agg(d):
                    s, u, o, a = d["send"].sum(), d["uv"].sum(), d["oc"].sum(), d["amt"].sum()
                    return dict(캠페인수=len(d), 발송=s, 거래액=a,
                                유입전환율=(u / s if s else np.nan), 주문전환율=(o / u if u else np.nan),
                                RPS=(a / s if s else np.nan))
                pa, ra = _agg(prev), _agg(recent)
                p_lab = f"직전 (~{pd.Timestamp(mid).strftime('%m/%d')} 전)"
                r_lab = f"최근 ({pd.Timestamp(mid).strftime('%m/%d')}~)"
                cmp_rows = []
                for k in ["캠페인수", "발송", "거래액", "유입전환율", "주문전환율", "RPS"]:
                    pv, rv = pa[k], ra[k]
                    if k in ("유입전환율", "주문전환율"):
                        pvs = f"{pv*100:.2f}%" if pd.notna(pv) else "–"; rvs = f"{rv*100:.2f}%" if pd.notna(rv) else "–"
                        chg = f"{(rv-pv)*100:+.2f}%p" if (pd.notna(pv) and pd.notna(rv)) else "–"
                    elif k in ("발송", "거래액", "RPS"):
                        pvs, rvs = won(pv), won(rv)
                        chg = f"{((rv/pv-1)*100):+.1f}%" if (pv and pd.notna(pv) and pd.notna(rv)) else "–"
                    else:
                        pvs, rvs = f"{pv:,.0f}", f"{rv:,.0f}"
                        chg = f"{((rv/pv-1)*100):+.1f}%" if pv else "–"
                    cmp_rows.append({"지표": k, p_lab: pvs, r_lab: rvs, "변화": chg})
                st.dataframe(pd.DataFrame(cmp_rows), hide_index=True, width="stretch")
                # 카테고리별 주문전환율 변화 Top
                if "cat" in gdt.columns:
                    def _catcr(d):
                        gg = d.groupby("cat").apply(
                            lambda x: (x["oc"].sum() / x["uv"].sum()) if x["uv"].sum() else np.nan)
                        return gg
                    pc, rc = _catcr(prev), _catcr(recent)
                    cats_common = [c for c in rc.index if c in pc.index and pd.notna(pc[c]) and pd.notna(rc[c])]
                    if cats_common:
                        ch = pd.DataFrame({"카테고리": cats_common,
                                           "직전_주문CR": [pc[c] for c in cats_common],
                                           "최근_주문CR": [rc[c] for c in cats_common]})
                        ch["변화%p"] = (ch["최근_주문CR"] - ch["직전_주문CR"]) * 100
                        ch = ch.sort_values("변화%p", ascending=False)
                        ch["직전_주문CR"] = ch["직전_주문CR"].map(lambda v: f"{v*100:.2f}%")
                        ch["최근_주문CR"] = ch["최근_주문CR"].map(lambda v: f"{v*100:.2f}%")
                        ch["변화%p"] = ch["변화%p"].map(lambda v: f"{v:+.2f}%p")
                        st.markdown("**카테고리별 주문전환율 변화 (최근 − 직전)**")
                        st.dataframe(ch, hide_index=True, width="stretch", height=260)

        # 인당 발송 기반 피로도 분석은 같은 그룹 하위탭으로 통합됨 (중복 차트 제거)
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.caption("🌡️ 인당 발송 건수 기반 피로도 분석은 상단 하위탭 "
                   "**피로도 시계열 · 발송 빈도·한계수익 · 요일 패턴**에서 볼 수 있어요"
                   + ("" if mtd_data is not None else " (전사 MTD 파일 업로드 필요)") + ".")

        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 09 — BPU·우선순위 효율
    # ══════════════════════════════════════════════════════════════
    elif "BPU" in page:
        st.title("BPU·우선순위 효율")
        st.caption("사업부별, 발송 순번별 효율을 비교해요. 어디서, 몇 번째로 보냈을 때 잘 먹히는지 알 수 있어요. "
                   "전환율·RPS는 합산 기준 가중 평균이에요.")
        mlabel = st.selectbox("지표", list(METRIC_OPTS.keys()))
        mcol, _msuf, mclr = METRIC_OPTS[mlabel]
        is_pct = mcol in ("ord_cr", "infl_cr")
        base = fdf
        if len(base) == 0:
            st.info("조건에 맞는 결과가 없어요. 조건을 넓혀 보세요."); st.stop()

        def agg_eff(d, by):
            out = []
            for key, g in d.groupby(by, dropna=True):
                if str(key).strip() in ("", "nan", "None"):
                    continue
                s, u, o, a = g["send"].sum(), g["uv"].sum(), g["oc"].sum(), g["amt"].sum()
                out.append(dict(_key=key, 캠페인수=len(g), 발송=s, 거래액=a,
                                infl_cr=(u / s if s else np.nan), ord_cr=(o / u if u else np.nan),
                                rps=(a / s if s else np.nan), aov=(a / o if o else np.nan), amt=a))
            return pd.DataFrame(out)

        def eff_table(t, keyname):
            ren = {"_key": keyname, "infl_cr": "CTR", "ord_cr": "주문CR", "rps": "RPS",
                   "aov": "객단가", "amt": "거래액"}
            show = t[["_key", "캠페인수", "발송", "infl_cr", "ord_cr", "rps", "aov", "거래액"]].rename(columns=ren)
            return show.style.format({"캠페인수": "{:,.0f}", "발송": "{:,.0f}", "CTR": "{:.2%}",
                                      "주문CR": "{:.2%}", "RPS": "{:,.0f}", "객단가": "{:,.0f}", "거래액": "{:,.0f}"})

        # ── BPU별 ──
        st.markdown("##### BPU별 효율")
        bp = agg_eff(base, "bpu")
        if len(bp):
            bp = bp.sort_values(mcol, ascending=False)
            y = bp[mcol] * (100 if is_pct else 1)
            fig = go.Figure(go.Bar(x=bp["_key"].astype(str), y=y, marker_color=mclr,
                                   text=[bar_label(v, mcol, is_pct) for v in y], textposition="outside"))
            fig.update_layout(**base_layout(h=340, ysuffix=("%" if is_pct else ""),
                                            title=f"BPU별 (가중) {mlabel}"))
            st.plotly_chart(fig, width="stretch")
            st.dataframe(eff_table(bp.sort_values("발송", ascending=False), "BPU"),
                         hide_index=True, width="stretch")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)

        # ── 우선순위별 ──
        st.markdown("##### 우선순위별 효율")
        base2 = base.copy()
        base2["_prio"] = pd.to_numeric(
            base2["prio"].astype(str).str.replace(r'\.0$', '', regex=True), errors="coerce")
        pr = agg_eff(base2.dropna(subset=["_prio"]), "_prio")
        if len(pr):
            pr["_key"] = pr["_key"].astype(float)
            pr = pr.sort_values("_key")
            pr_chart = pr[pr["캠페인수"] > 1]
            n_excluded = len(pr) - len(pr_chart)
            xlab = pr_chart["_key"].astype(int).astype(str) + "순위"
            y = pr_chart[mcol] * (100 if is_pct else 1)
            fig = go.Figure(go.Bar(x=xlab, y=y, marker_color=mclr,
                                   text=[bar_label(v, mcol, is_pct) for v in y], textposition="outside"))
            fig.update_layout(**base_layout(h=340, ysuffix=("%" if is_pct else ""),
                                            title=f"우선순위별 (가중) {mlabel}"))
            st.plotly_chart(fig, width="stretch")
            if n_excluded:
                excluded_labels = ", ".join(
                    f"{int(k)}순위" for k in pr.loc[pr["캠페인수"] <= 1, "_key"])
                st.caption(f"캠페인 1건뿐인 {excluded_labels}는 표본이 부족해 차트에서 제외했어요. (표에는 포함)")
            tshow = pr.copy(); tshow["_key"] = tshow["_key"].astype(int).astype(str) + "순위"
            st.dataframe(eff_table(tshow, "우선순위"), hide_index=True, width="stretch")
            # 포지션 효과 간단 진단 (차트에 반영된 표본만 사용)
            if len(pr_chart) >= 3 and pr_chart[mcol].notna().sum() >= 3:
                r = float(np.corrcoef(pr_chart["_key"], pr_chart[mcol].fillna(pr_chart[mcol].mean()))[0, 1])
                msg = ("앞 순번일수록 효율이 높습니다 (노출 우위)." if r < -0.3 else
                       "뒤 순번일수록 효율이 높아요." if r > 0.3 else
                       "순번과 효율 사이에 뚜렷한 관계는 약해요.")
                st.markdown(f'<div class="appendix">순번↔{mlabel} 상관 r={r:.2f} → {msg}</div>',
                            unsafe_allow_html=True)

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)

        # ── BPU × 우선순위 히트맵 ──
        st.markdown(f"##### BPU × 우선순위 — 평균 {mlabel}")
        hb = base2.dropna(subset=["_prio"]).copy()
        hb["_prio"] = hb["_prio"].astype(int)
        pv = hb.pivot_table(index="bpu", columns="_prio", values=mcol, aggfunc="mean")
        if not pv.empty:
            _cntb = hb.pivot_table(index="bpu", columns="_prio", values=mcol, aggfunc="count")
            _cntb = _cntb.reindex(index=pv.index, columns=pv.columns)
            pv = pv.where(_cntb >= 3)                  # n<3 셀 마스킹
            z = pv.values * (100 if is_pct else 1)
            txt_arr = np.where(np.isnan(z), "", np.round(z, 2).astype(str))
            fig = go.Figure(go.Heatmap(z=z, x=[f"{c}순위" for c in pv.columns],
                                       y=[str(i) for i in pv.index], colorscale="Blues",
                                       text=txt_arr, texttemplate="%{text}",
                                       textfont=dict(size=10), colorbar=dict(thickness=10),
                                       customdata=_cntb.fillna(0).values,
                                       hovertemplate="%{y} · %{x}<br>" + mlabel + ": %{z:.2f}"
                                                     + ("%" if is_pct else "")
                                                     + " · n=%{customdata:.0f}<extra></extra>"))
            fig.update_layout(**base_layout(h=420, title=f"BPU × 우선순위 평균 {mlabel}"))
            st.plotly_chart(fig, width="stretch")
        st.markdown('<div class="appendix">빈 셀은 캠페인 3건 미만(우연값 방지). hover에 표본수(n)가 함께 떠요.</div>',
                    unsafe_allow_html=True)

        # ── BPU·우선순위 드릴다운 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📋 BPU·우선순위별 실제 메시지 보기")
        bc1, bc2 = st.columns(2)
        bpu_opts = sorted(base["bpu"].dropna().astype(str).unique()) if "bpu" in base else []
        bpu_opts = [b for b in bpu_opts if b.strip() not in ("", "nan", "None")]
        prio_opts = sorted(base2["_prio"].dropna().unique()) if "_prio" in base2 else []
        guard_select("p09_bpu", ["전체"] + bpu_opts)
        guard_select("p09_prio", ["전체"] + [f"{int(p)}순위" for p in prio_opts])
        sel_bpu9 = bc1.selectbox("BPU", ["전체"] + bpu_opts, key="p09_bpu")
        sel_prio9 = bc2.selectbox("우선순위", ["전체"] + [f"{int(p)}순위" for p in prio_opts], key="p09_prio")
        sub9 = base.copy()
        if sel_bpu9 != "전체" and "bpu" in sub9:
            sub9 = sub9[sub9["bpu"].astype(str) == sel_bpu9]
        if sel_prio9 != "전체" and "prio" in sub9:
            pval = sel_prio9.replace("순위", "")
            sub9 = sub9[pd.to_numeric(sub9["prio"].astype(str).str.replace(r'\.0$', '', regex=True),
                                       errors="coerce") == int(pval)]
        st.caption(f"조건 일치 {len(sub9)}건 — {mlabel} 높은 순")
        render_messages(sub9, mcol, "p09_drill")

        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 10 — 키워드·이모지 성과
    # ══════════════════════════════════════════════════════════════
    elif "키워드" in page:
        st.title("키워드·이모지 성과")
        st.caption("발송 문구에 사용된 단어 및 이모지 단위의 개별 성과(전체 평균 대비 상대적 효율)를 분석합니다.")
        mlabel = st.selectbox("성과 지표", list(METRIC_OPTS.keys()), key="p10_metric")
        mcol, _ms, mclr = METRIC_OPTS[mlabel]
        is_pct = mcol in ("ord_cr", "infl_cr")
        base = fdf
        if len(base) < 5:
            st.info("데이터가 부족해요. '최소 발송수'를 낮춰 보세요."); st.stop()

        c1, c2 = st.columns(2)
        kmin = c1.number_input("키워드 최소 표본(캠페인 수)", value=5, min_value=2, step=1, key="p10_kmin")
        ktop = c2.number_input("상위 N개", value=20, min_value=5, step=5, key="p10_ktop")

        def _barfig(d, namecol, title, h):
            up = d.head(int(ktop))
            yv = up["평균"] * (100 if is_pct else 1)
            # 유의(FDR p<0.1) 항목만 본색, 나머지 회색 — 요행 단어가 진하게 강조되는 것 방지
            if "유의성" in up.columns:
                bar_colors = [mclr if s.startswith("✱") else "#cbd5e1" for s in up["유의성"]]
            else:
                bar_colors = mclr
            fig = go.Figure(go.Bar(
                x=yv, y=up[namecol], orientation="h", marker_color=bar_colors, customdata=up["캠페인수"],
                text=[f"{v:.2f}{'%' if is_pct else ''} (n={int(n)})" for v, n in zip(yv, up["캠페인수"])],
                textposition="outside",
                hovertemplate="%{y}<br>평균: %{x:.2f}<br>캠페인수: %{customdata}<extra></extra>"))
            lay = base_layout(h=h, title=title)
            lay["xaxis"]["range"] = bar_xrange(yv, 1.18)
            fig.update_layout(**lay)
            fig.update_yaxes(autorange="reversed")
            return fig

        st.markdown("##### 🔤 키워드별 성과")
        _kw_view = st.radio("보기", ["평균 상위", "차이 ±상하위 (평균을 깎는 단어까지)"],
                            horizontal=True, key="p10_kwview")
        _kdf_full = cached_keyword_perf(base, mcol, int(kmin), 500)
        kdf = _kdf_full.head(int(ktop))
        if len(_kdf_full) == 0:
            st.info("조건에 맞는 키워드가 없어요. 기준을 낮춰 보세요.")
        else:
            if _kw_view.startswith("평균"):
                st.plotly_chart(_barfig(kdf, "단어", f"키워드별 평균 {mlabel} (상위 {int(ktop)})",
                                        max(360, 40 + 24 * min(len(kdf), int(ktop)))),
                                width="stretch")
            else:
                # diverging — 전체평균 대비 +/− 상하위 각 10개 (역효과 단어가 안 보이던 문제)
                _hi = _kdf_full.sort_values("차이", ascending=False).head(10)
                _lo = _kdf_full.sort_values("차이").head(10)
                _dv = pd.concat([_hi, _lo]).drop_duplicates("단어").sort_values("차이", ascending=False)
                kdf = _dv
                _yv = _dv["차이"] * (100 if is_pct else 1)
                _sig_col = _dv["유의성"] if "유의성" in _dv.columns else pd.Series("✱", index=_dv.index)
                _cols_dv = [("#cbd5e1" if not str(s).startswith("✱")
                             else (PALETTE["green"] if v >= 0 else PALETTE["red"]))
                            for v, s in zip(_dv["차이"], _sig_col)]
                figd = go.Figure(go.Bar(
                    x=_yv, y=_dv["단어"], orientation="h", marker_color=_cols_dv,
                    customdata=_dv["캠페인수"],
                    text=[f"{v:+.2f}{'%p' if is_pct else ''} (n={int(n)})"
                          for v, n in zip(_yv, _dv["캠페인수"])],
                    textposition="outside",
                    hovertemplate="%{y}<br>전체평균 대비: %{x:+.2f}<br>캠페인수: %{customdata}<extra></extra>"))
                layd = base_layout(h=max(420, 40 + 24 * len(_dv)),
                                   title=f"전체평균 대비 차이 ±상하위 — {mlabel} (회색=유의하지 않음)")
                figd.update_layout(**layd)
                figd.update_yaxes(autorange="reversed")
                st.plotly_chart(figd, width="stretch")
            sel_kw = st.selectbox("조회 대상 키워드 선택 (실제 발송 메시지 드릴다운)", list(kdf["단어"]), key="p10_kw")
            if sel_kw:
                hay = (base["title"].astype(str) + " " + base["body"].astype(str))
                subk = base[hay.str.contains(re.escape(sel_kw), na=False)]
                st.caption(f"'{sel_kw}' 포함 {len(subk)}건 — {mlabel} 높은 순")
                render_messages(subk, mcol, f"p10kw_{sel_kw}")

        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 😀 이모지별 성과")
        emin = st.number_input("이모지 최소 표본(캠페인 수)", value=3, min_value=2, step=1, key="p10_emin")
        edf = cached_emoji_perf(base, mcol, int(emin), 30)
        if len(edf) == 0:
            st.info("조건에 맞는 이모지가 없어요. 이모지를 쓴 캠페인이 적을 수 있어요.")
        else:
            st.plotly_chart(_barfig(edf, "이모지", f"이모지별 평균 {mlabel}",
                                    max(320, 40 + 28 * len(edf))), width="stretch")
            eshow = edf.copy()
            if is_pct:
                eshow["평균"] = eshow["평균"].map(lambda v: f"{v*100:.2f}%")
                eshow["차이"] = eshow["차이"].map(lambda v: f"{v*100:+.2f}%p")
            elif mcol in ("rps", "aov", "amt"):
                eshow["평균"] = eshow["평균"].map(won); eshow["차이"] = eshow["차이"].map(lambda v: f"{v:+,.0f}")
            else:
                eshow["평균"] = eshow["평균"].map(lambda v: f"{v:,.1f}"); eshow["차이"] = eshow["차이"].map(lambda v: f"{v:+,.1f}")
            eshow = eshow.drop(columns=["_p_adj"], errors="ignore")
            st.dataframe(eshow.style.format({"캠페인수": "{:,.0f}"}), hide_index=True, width="stretch")
        st.markdown('<div class="appendix">단어 및 이모지 성과 분석은 캠페인 단위 평균 데이터입니다. 표본 건수(n)가 적은 항목은 편차가 존재할 수 있으니 유의하시기 바랍니다. (동일 캠페인 내 중복 단어는 1회 집계, 일부 의미 없는 기호/불용어/숫자는 분석에서 제외됨) 차트의 <b>회색 막대</b>는 보유 vs 미보유 차이가 통계적으로 유의하지 않은(FDR 보정 p≥0.1) 항목 — 우연일 수 있으니 참고만 하세요.</div>',
                    unsafe_allow_html=True)
        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 11 — 소구 추세·마모 (시계열)
    # ══════════════════════════════════════════════════════════════
    elif "소구 추세" in page:
        st.title("소구 추세·마모")
        st.caption("동일 소구의 지속 노출에 따른 성과 반응 및 피로도 누적 추세를 시계열로 분석하여 피로 임계 구간을 진단합니다.")
        mlabel = st.selectbox("성과 지표", list(METRIC_OPTS.keys()), key="p11_metric")
        mcol, _ms, mclr = METRIC_OPTS[mlabel]
        is_pct = mcol in ("ord_cr", "infl_cr")
        base = fdf.dropna(subset=["dt"]).copy()
        if len(base) < 8:
            st.info("데이터가 부족해요. 더 많은 주차를 쌓거나 '최소 발송수'를 낮춰 보세요."); st.stop()
        base["주"] = base["dt"].dt.to_period("W").apply(lambda p: p.start_time)

        sel_attrs = st.multiselect("추세를 볼 속성(소구)", [t for t in TAG_BOOLS if t in base.columns],
                                   default=[t for t in ["할인율소구", "마감임박"] if t in base.columns],
                                   key="p11_attrs")
        if not sel_attrs:
            st.info("속성을 하나 이상 골라 주세요."); st.stop()

        fig = go.Figure()
        # 발송이 미미한 주(오조작·이월 날짜 1~2건으로 생긴 뜬금없는 주)는 추세에서 제외
        _wk_send = base.groupby("주")["send"].sum()
        _wk_cnt = base.groupby("주").size()
        _enough = len(_wk_send) >= 4                      # 주가 적을 땐 필터 미적용
        _med_ws = float(_wk_send.median()) if _enough else 0.0
        weeks = sorted(w for w in base["주"].unique()
                       if not _enough or (_wk_send.get(w, 0) >= _med_ws * 0.05
                                          and _wk_cnt.get(w, 0) >= 3))
        _wk_drop = base["주"].nunique() - len(weeks)
        if _wk_drop:
            st.caption(f"데이터가 미미한 주 {_wk_drop}개(캠페인 3건 미만 또는 발송 극소 — "
                       "오조작·이월 날짜)는 추세에서 제외했어요.")
        for t in sel_attrs:
            ys = []
            for w in weeks:
                vv = base[(base["주"] == w) & (base[t])][mcol].dropna()
                ys.append(vv.mean() * (100 if is_pct else 1) if len(vv) else np.nan)
            fig.add_trace(go.Scatter(x=list(weeks), y=ys, mode="lines+markers", name=t,
                                     line=dict(color=tag_color(t), width=2),
                                     connectgaps=True))
        lay = base_layout(h=420, ysuffix=("%" if is_pct else ""),
                          title=f"속성별 주차 추이 — 평균 {mlabel}", hover="x")
        lay["showlegend"] = True
        lay["legend"] = legend_h()
        fig.update_layout(**lay)
        st.plotly_chart(fig, width="stretch")

        # 마모 진단: 각 속성의 주차 추세 회귀(기울기) + 사용 빈도 추이
        st.markdown("##### 📉 반복 소구 노출에 따른 성과 감쇠(마모) 진단")
        diag = []
        for t in sel_attrs:
            pts = []
            for j, w in enumerate(weeks):
                vv = base[(base["주"] == w) & (base[t])][mcol].dropna()
                if len(vv):
                    pts.append((j, float(vv.mean())))
            if len(pts) >= 4:
                xs = np.array([p[0] for p in pts]); ys2 = np.array([p[1] for p in pts])
                sl, ic, rr, pp, _ = stats.linregress(xs, ys2)
                trend = "마모(하락)" if (rr < 0 and pp < 0.1) else ("상승" if (rr > 0 and pp < 0.1) else "변화 약함")
                diag.append(dict(속성=t, 주차수=len(pts), 추세=trend,
                                 상관r=round(float(rr), 2), 유의성=sig_label(pp)))
            else:
                diag.append(dict(속성=t, 주차수=len(pts), 추세="표본부족", 상관r=np.nan, 유의성="–"))
        st.dataframe(pd.DataFrame(diag), hide_index=True, width="stretch")

        # ── 노출 강도 기반 마모 진단 — '시즌 하락'과 '자주 써서 무뎌짐'을 분리 ──
        st.markdown("##### 🔁 노출 강도 기반 마모 진단 — 최근 28일 사용량이 성과를 깎나")
        _wear_rows = []
        for t in sel_attrs:
            vals = base.loc[base[t], [mcol, "dt"]].dropna()
            if len(vals) < 25:
                _wear_rows.append({"속성": t, "n": len(vals), "판정": "표본 부족(25건 미만)"})
                continue
            _days = vals["dt"].values.astype("datetime64[D]").astype(int)
            _order = np.argsort(_days)
            _ds = _days[_order]
            y = vals[mcol].values.astype(float)[_order]
            # 노출 강도 = 각 캠페인 시점 직전 28일간 같은 소구 캠페인 수(같은 날 포함, 자기 제외).
            # side="left"로 하면 같은 날 캠페인을 통째로 빼서 노출을 과소 추정한다 — 실데이터는
            # 같은 소구를 같은 날 여러 세그먼트로 보내는 비율이 99%라 (자기≤, 자기 1개만 제외)로 세야 정확.
            expo = (np.searchsorted(_ds, _ds, side="right") - 1
                    - np.searchsorted(_ds, _ds - 28, side="left")).astype(float)
            wk_idx = (_ds - _ds.min()) / 7.0                # 주차 추세 통제 (시즌 드리프트 분리)
            X = np.column_stack([np.ones(len(y)), expo, wk_idx])
            try:
                beta, *_ = np.linalg.lstsq(X, y, rcond=None)
                dof = len(y) - 3
                s2 = float(((y - X @ beta) ** 2).sum() / max(dof, 1))
                cov = s2 * np.linalg.inv(X.T @ X)
                se1 = float(np.sqrt(cov[1, 1]))
                p1 = float(2 * (1 - stats.t.cdf(abs(beta[1] / se1), dof))) if se1 > 0 else 1.0
            except Exception:
                _wear_rows.append({"속성": t, "n": len(y), "판정": "계산 불가"})
                continue
            _q33, _q67 = np.percentile(expo, 33), np.percentile(expo, 67)
            lo_m = float(y[expo <= _q33].mean()) if (expo <= _q33).any() else np.nan
            hi_m = float(y[expo >= _q67].mean()) if (expo >= _q67).any() else np.nan
            _mul = 100 if is_pct else 1
            if beta[1] < 0 and p1 < 0.1:
                _verdict = "🔻 마모 신호 — 빈도 축소·휴지기 검토"
            elif beta[1] > 0 and p1 < 0.1:
                _verdict = "🔺 빈도 여유 (강도↑에도 성과↑)"
            else:
                _verdict = "신호 없음"
            _wear_rows.append({
                "속성": t, "n": len(y),
                "28일 노출 중앙값": f"{np.median(expo):.0f}회",
                "노출 1회당 효과": (f"{beta[1]*_mul:+.3f}%p" if is_pct else f"{beta[1]:+,.1f}"),
                "p값": f"{p1:.3f}",
                "저노출 평균": (f"{lo_m*_mul:.2f}{'%' if is_pct else ''}" if pd.notna(lo_m) else "–"),
                "고노출 평균": (f"{hi_m*_mul:.2f}{'%' if is_pct else ''}" if pd.notna(hi_m) else "–"),
                "판정": _verdict})
        st.dataframe(pd.DataFrame(_wear_rows), hide_index=True, width="stretch")
        st.markdown('<div class="appendix">캠페인마다 <b>직전 28일간 같은 소구를 몇 번 썼는지(노출 강도)</b>를 '
                    '세고, 주차 추세를 통제한 회귀로 노출 강도가 성과를 깎는지 확인해요 — 위의 달력 추세 '
                    "진단과 달리 '시즌이라 떨어진 것'과 '과사용으로 무뎌진 것'을 분리해 줘요. "
                    "'노출 1회당 효과'가 유의한 음수(🔻)면 그 소구는 쉬어 갈 때예요. "
                    '(카테고리 구성까지 통제하지는 않으니 보조 지표로 참고하세요)</div>',
                    unsafe_allow_html=True)

        # 속성 사용 빈도(발송수) 추이 — 너무 자주 쓰면 마모 위험
        st.markdown("##### 📨 속성별 발송 빈도 추이")
        figf = go.Figure()
        for t in sel_attrs:
            cnts = [int(((base["주"] == w) & (base[t])).sum()) for w in weeks]
            figf.add_trace(go.Scatter(x=list(weeks), y=cnts, mode="lines+markers", name=t,
                                      line=dict(color=tag_color(t), width=2)))
        layf = base_layout(h=320, title="속성별 주차 사용 빈도(캠페인 수)", hover="x")
        layf["showlegend"] = True
        layf["legend"] = legend_h()
        figf.update_layout(**layf)
        st.plotly_chart(figf, width="stretch")
        st.markdown('<div class="appendix">상관계수(r)가 통계적으로 유의한 음수(-)이면, 반복 소구 노출로 인한 성과 저하(마모) 가능성이 높음을 시사합니다. 집행 빈도 대비 효율 감소세가 지속될 경우 해당 메시지의 소구 휴지기 수립 및 대체 오퍼 도입을 권장합니다. (단, 카테고리 구성 비율이나 시즌 이벤트 영향성이 혼재할 수 있으므로 보조 지표로 참고하십시오)</div>', unsafe_allow_html=True)

        glossary()

    # ══════════════════════════════════════════════════════════════
    # 발송피로도 (전사 MTD) — F1~F4
    # ══════════════════════════════════════════════════════════════
    elif page in ("피로도 시계열", "발송 빈도·한계수익", "요일 패턴"):
        MTDOPT = {"CTR": "ctr", "구매전환율(CR)": "purchaseRate", "발송건당거래액(RPS)": "rps",
                  "거래액": "revenue", "인당 발송 건수": "perSend", "객단가": "avgOrderVal",
                  "총거래액": "revenue_sum", "총유입": "totalInflow_sum"}
        MTD_PCT = {"ctr", "purchaseRate"}
        # ctr는 blue — 대시보드 전역에서 빨강=감소/나쁨(△·급락·이탈)이라 CTR 시리즈가
        # 빨강이면 '나쁜 지표'로 오독됨 (METRIC_OPTS의 CTR 색과도 통일)
        MCLR = {"ctr": PALETTE["blue"], "purchaseRate": PALETTE["purple"], "rps": PALETTE["green"],
                "revenue": PALETTE["blue"], "perSend": PALETTE["amber"], "avgOrderVal": PALETTE["teal"],
                "revenue_sum": PALETTE["blue"], "totalInflow_sum": PALETTE["teal"]}

        def mfmt(col, v):
            if v is None or (isinstance(v, float) and np.isnan(v)): return "–"
            if col in MTD_PCT: return f"{v*100:.2f}%"
            if col in ("revenue",): return won(v)
            if col == "perSend": return f"{v:.2f}건"
            return f"{v:,.0f}"

        if mtd_data is None:
            st.title("발송피로도")
            st.info("전사 MTD 파일을 왼쪽 📂 파일 올리기에 올리고 「MTD 저장하기」를 눌러 주세요.")
            st.stop()
        md = mtd_data["df"]
        st.caption(f"전사 MTD · {mtd_data['meta']['start']} ~ {mtd_data['meta']['end']} "
                   f"({mtd_data['meta']['days']:,}일)")

        # ── 피로도 시계열 ──
        if page == "피로도 시계열":
            st.title("피로도 시계열·CTR")
            st.markdown("인당 누적 발송 빈도 대비 성과 지표(CTR, RPS)의 시계열 추이 분석")
            gran = st.radio("집계", ["월별", "분기별"], horizontal=True)
            agg = mtd_data["monthly"] if gran == "월별" else mtd_data["quarterly"]
            xcol = "month" if gran == "월별" else "quarter"
            _f1_opts = ["CTR", "구매전환율(CR)", "발송건당거래액(RPS)"]
            _f1_opts += [o for o in ("총거래액", "총유입") if MTDOPT[o] in agg.columns]
            ylab = st.selectbox("효율 지표(선·아래)", _f1_opts)
            yc = MTDOPT[ylab]
            fig = overlay_dual(agg[xcol], agg["perSend"], "인당 발송 건수",
                               agg[yc] * (100 if yc in MTD_PCT else 1), ylab,
                               PALETTE["amber"], MCLR[yc], h=430,
                               line_suffix=("%" if yc in MTD_PCT else ""),
                               title=f"인당 발송 건수(좌·막대) ↔ {ylab}(우·선)")
            st.plotly_chart(fig, width="stretch")

            st.markdown("##### 추세 분석")
            rows = []
            for k in ["perSend", "ctr", "purchaseRate", "rps", "revenue"]:
                r = mtd_data["reg"][k]
                unit = "%p/일" if k in MTD_PCT else ("원/일" if k in ("rps", "revenue") else "/일")
                sl = r["slope"] * (100 if k in MTD_PCT else 1)
                rows.append(dict(지표=MTD_LABELS[k], **{"일변화": f"{sl:+.4g}{unit}"},
                                 R2=f"{r['r2']:.3f}", 유의성=sig_label(r["p"])))
            st.dataframe(pd.DataFrame(rows), hide_index=True, width="stretch")
            st.markdown('<div class="appendix">인당 발송량은 증가하는 반면 CTR·주문CR·RPS가 동반 하락할 경우 피로도 임계점에 도달했다는 위험 신호일 수 있습니다. '
                        '<br>· <b>R²(결정계수, 0~1)</b>: 추세선의 설명력을 의미하며, 1에 가까울수록 경향성이 뚜렷하고 0에 가까울수록 불규칙한 변동을 보입니다. '
                        '· <b>유의성</b>: 산출된 추세의 통계적 유효성(p-value)을 나타내며, ‘유의함’일 경우 우연이 아닌 일관된 흐름으로 판단할 수 있습니다.</div>',
                        unsafe_allow_html=True)

        # ── 발송 빈도·한계수익 (구 F2+F3 통합) ──
        elif page == "발송 빈도·한계수익":
            st.title("발송 빈도 효율·한계수익")
            st.markdown("고객 인당 발송 빈도에 따른 효율 변화와, 발송을 한 구간 더 늘렸을 때의 "
                        "한계 효율을 함께 분석합니다.")
            lab = st.selectbox("지표", ["CTR", "구매전환율(CR)", "발송건당거래액(RPS)", "거래액", "객단가"])
            mc = MTDOPT[lab]
            b = mtd_data["buckets"]
            if len(b):
                y = b[mc] * (100 if mc in MTD_PCT else 1)
                fig = go.Figure(go.Bar(x=b["bucket"].astype(str), y=y, marker_color=MCLR[mc],
                                       text=[bar_label(v, mc, mc in MTD_PCT) for v in y], textposition="outside"))
                fig.update_layout(**base_layout(h=340, ysuffix=("%" if mc in MTD_PCT else ""),
                                                title=f"인당 발송 구간별 평균 {lab} (표본 30일+)"))
                st.plotly_chart(fig, width="stretch")

            # 한계수익 — 같은 구간에서 한 단계 더 늘렸을 때의 증분 (구 F3 페이지 통합)
            st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
            st.markdown("##### 한계수익 — 발송 구간을 한 단계 올리면 얼마나 변할까")
            b3 = mtd_data["buckets"].reset_index(drop=True)
            if len(b3) >= 2:
                vals = (b3[mc] * (100 if mc in MTD_PCT else 1)).values
                diff = np.diff(vals)
                labels = [f"{b3['bucket'].astype(str).iloc[i]}→{b3['bucket'].astype(str).iloc[i+1]}"
                          for i in range(len(b3) - 1)]
                fig = go.Figure(go.Bar(x=labels, y=diff,
                                       marker_color=[PALETTE["green"] if v >= 0 else PALETTE["red"] for v in diff],
                                       text=[f"{v:+.2f}" for v in diff], textposition="outside"))
                fig.update_layout(**base_layout(h=360, title=f"인당 발송 구간 상승 시 {lab} 한계 변화"))
                st.plotly_chart(fig, width="stretch")
                st.markdown('<div class="appendix">한계효율이 음수(-)인 구간은 추가 발송 시 고객 반응 및 '
                            '효율이 감소하는 감쇠 국면임을 뜻합니다.</div>', unsafe_allow_html=True)
            else:
                st.info("구간별 데이터가 부족해요.")

            st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
            st.markdown("##### 발송량 5분위 효율")
            q = mtd_data["quintile"]
            if len(q):
                cc = st.columns(2)
                with cc[0]:
                    l1 = st.selectbox("막대 지표 (위)", ["발송건당거래액(RPS)", "CTR", "거래액"], key="q_l")
                with cc[1]:
                    l2 = st.selectbox("선 지표 (아래)", ["CTR", "구매전환율(CR)", "객단가"], key="q_r")
                m1, m2 = MTDOPT[l1], MTDOPT[l2]
                fig = stacked_panels(q["label"], q[m1] * (100 if m1 in MTD_PCT else 1), l1,
                                     q[m2] * (100 if m2 in MTD_PCT else 1), l2,
                                     MCLR[m1], MCLR[m2], h=400,
                                     bar_suffix=("%" if m1 in MTD_PCT else ""),
                                     line_suffix=("%" if m2 in MTD_PCT else ""),
                                     title="발송량 5분위(Q1 소량→Q5 대량)")
                st.plotly_chart(fig, width="stretch")

        # ── 요일 패턴 ──
        else:
            st.title("요일 패턴")
            lab = st.selectbox("지표", ["CTR", "구매전환율(CR)", "발송건당거래액(RPS)", "거래액", "인당 발송 건수"])
            mc = MTDOPT[lab]
            dm = mtd_data["dow_mean"]
            order = ["월", "화", "수", "목", "금", "토", "일"]
            dm = dm.set_index("요일").reindex([o for o in order if o in dm["요일"].values]).reset_index()
            y = dm[mc] * (100 if mc in MTD_PCT else 1)
            fig = go.Figure(go.Bar(x=dm["요일"], y=y, marker_color=MCLR[mc],
                                   text=[bar_label(v, mc, mc in MTD_PCT) for v in y], textposition="outside"))
            fig.update_layout(**base_layout(h=320, ysuffix=("%" if mc in MTD_PCT else ""),
                                            title=f"요일별 평균 {lab}"))
            st.plotly_chart(fig, width="stretch")
            dc = pd.DataFrame(mtd_data["dow_comp"])
            if len(dc):
                st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
                st.markdown("##### 같은 요일에서 적게 보낸 날 vs 많이 보낸 날")
                show = pd.DataFrame({
                    "요일": dc["요일"],
                    "저발송일 CTR": (dc["lowCtr"] * 100).map("{:.2f}%".format),
                    "고발송일 CTR": (dc["highCtr"] * 100).map("{:.2f}%".format),
                    "저발송일 RPS": dc["lowRps"].map("{:,.0f}".format),
                    "고발송일 RPS": dc["highRps"].map("{:,.0f}".format),
                })
                st.dataframe(show, hide_index=True, width="stretch")
                st.markdown('<div class="appendix">동일 요일 내 발송 모수가 적을 때 효율(CTR, RPS)이 유의미하게 높다면, 무리한 모수 확장보다 타겟팅 세분화를 통해 반응률을 개선하는 것이 유효함을 의미합니다.</div>',
                            unsafe_allow_html=True)
        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 14 — 전환·AOV 진단
    # ══════════════════════════════════════════════════════════════
    elif "전환·AOV 진단" in page:
        st.title("전환·AOV 진단")
        st.caption("유입 효율(CTR)과 구매 전환 효율(주문CR)을 분리 진단하여, 거래액 증대 요인(객단가 중심 vs 모수 유입 중심)을 정량 분석합니다.")
        base = fdf
        if len(base) < 6:
            st.info("데이터가 부족해요. '최소 발송수'를 낮춰 보세요."); st.stop()

        # ① 전환 2단 분해 (사분면)
        st.markdown("##### ① CTR vs 주문전환율")
        d = base.dropna(subset=["infl_cr", "ord_cr"]).copy()
        if "uv" in d:
            d = d[d["uv"].fillna(0) >= 50]
        if len(d) >= 6:
            mx = float(d["infl_cr"].median()); my = float(d["ord_cr"].median())

            def _quad(r):
                hi_c = r["infl_cr"] >= mx; hi_o = r["ord_cr"] >= my
                if hi_c and hi_o: return "🟢 둘다 좋음"
                if hi_c and not hi_o: return "🟡 유입O 주문X(오퍼/랜딩 점검)"
                if not hi_c and hi_o: return "🔵 유입X 주문O(타겟/제목 점검)"
                return "🔴 둘다 약함"
            d["사분면"] = d.apply(_quad, axis=1)
            cmap = {"🟢 둘다 좋음": PALETTE["green"], "🟡 유입O 주문X(오퍼/랜딩 점검)": PALETTE["amber"],
                    "🔵 유입X 주문O(타겟/제목 점검)": PALETTE["blue"], "🔴 둘다 약함": PALETTE["red"]}
            # send 전부 NaN이면 `nan or 1`이 nan(truthy)을 반환 → 마커 크기가 NaN이 됨. 명시 가드.
            _sm = d["send"].max() if "send" in d else np.nan
            _smax = float(_sm) if (pd.notna(_sm) and _sm > 0) else 1.0
            fig = go.Figure()
            for q, sub in d.groupby("사분면"):
                _sz = 8 + (sub["send"].fillna(0) / _smax * 22) if "send" in sub else 10
                fig.add_trace(go.Scatter(
                    x=sub["infl_cr"] * 100, y=sub["ord_cr"] * 100, mode="markers", name=q,
                    marker=dict(color=cmap.get(q, PALETTE["slate"]), size=_sz),
                    text=sub.get("title", ""),
                    hovertemplate="%{text}<br>CTR:%{x:.2f}%<br>주문CR:%{y:.2f}%<extra></extra>"))
            fig.add_vline(x=mx * 100, line_dash="dash", line_color="#cbd5e1")
            fig.add_hline(y=my * 100, line_dash="dash", line_color="#cbd5e1")
            lay = base_layout(h=420, title="캠페인 전환 사분면 (점 크기=발송량)")
            lay["showlegend"] = True
            lay["xaxis"]["ticksuffix"] = "%"; lay["yaxis"]["ticksuffix"] = "%"
            fig.update_layout(**lay)
            st.plotly_chart(fig, width="stretch")
            qsum = d["사분면"].value_counts().reset_index()
            qsum.columns = ["사분면", "캠페인수"]
            st.dataframe(qsum.style.format({"캠페인수": "{:,.0f}"}), hide_index=True, width="stretch")
            st.caption(f"기준선: CTR 중앙값 {mx*100:.2f}% · 주문CR 중앙값 {my*100:.2f}%. "
                       "🟡 유입O 주문CRX 영역은 오퍼 구조 및 랜딩 페이지 정비를 권장하며, 🔵 유입X 주문CRO 영역은 발송 타겟 확장 및 제목 메시지 보완을 권장합니다.")
            guard_select("p14_quad", list(qsum["사분면"]))
            qpick = st.selectbox("사분면 선택 → 캠페인 보기", list(qsum["사분면"]), key="p14_quad")
            render_messages(d[d["사분면"] == qpick], "ord_cr", f"p14_{qpick}")
        else:
            st.info("UV 50 이상인 캠페인이 6건 이상 있어야 해요.")

        # ② AOV vs 전환 기여
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### ② 거래액 드라이버 분석: 판매 수량(Q) vs 객단가(P)")
        _dim = {"카테고리": "cat", "세그먼트": "target", "브랜드": "brand"}
        _dim = {k: v for k, v in _dim.items() if v in base.columns and base[v].fillna("").ne("").any()}
        if not _dim:
            _dim = {"카테고리": "cat"}
        dimname = st.radio("기준 차원", list(_dim.keys()), horizontal=True, key="p14_dim")
        dcol = _dim[dimname]
        a = base.copy(); a[dcol] = a[dcol].fillna("(미지정)").replace("", "(미지정)")
        gg = a.groupby(dcol).agg(주문CR=("ord_cr", "mean"), AOV=("aov", "mean"),
                                 거래액=("amt", "sum"), 캠페인수=(dcol, "size")).reset_index()
        gg = gg[gg["캠페인수"] >= 3]
        if len(gg) >= 2:
            mcr = float(gg["주문CR"].median()); mav = float(gg["AOV"].median())
            _amax = float(gg["거래액"].max() or 1)
            fig = go.Figure(go.Scatter(
                x=gg["주문CR"] * 100, y=gg["AOV"], mode="markers+text",
                text=gg[dcol], textposition="top center",
                marker=dict(size=10 + gg["거래액"] / _amax * 26, color=PALETTE["teal"]),
                hovertemplate="%{text}<br>주문CR:%{x:.2f}%<br>AOV:%{y:,.0f}<extra></extra>"))
            fig.add_vline(x=mcr * 100, line_dash="dash", line_color="#cbd5e1")
            fig.add_hline(y=mav, line_dash="dash", line_color="#cbd5e1")
            lay = base_layout(h=420, title="주문CR(많이) × AOV(비싸게) — 점 크기=거래액")
            lay["xaxis"]["ticksuffix"] = "%"
            fig.update_layout(**lay)
            st.plotly_chart(fig, width="stretch")
            gs = gg.sort_values("거래액", ascending=False).copy()
            gs["전략"] = [("박리다매(전환↑·객단↓)" if cr >= mcr and av < mav else
                          "고가저빈도(전환↓·객단↑)" if cr < mcr and av >= mav else
                          "올라운더(둘다↑)" if cr >= mcr and av >= mav else "약세(둘다↓)")
                         for cr, av in zip(gs["주문CR"], gs["AOV"])]
            gs["주문CR"] = gs["주문CR"].map(lambda v: f"{v*100:.2f}%")
            gs["AOV"] = gs["AOV"].map(won); gs["거래액"] = gs["거래액"].map(won)
            st.dataframe(gs.rename(columns={dcol: dimname})[[dimname, "캠페인수", "주문CR", "AOV", "거래액", "전략"]]
                         .style.format({"캠페인수": "{:,.0f}"}),
                         hide_index=True, width="stretch")
            st.caption("고효율 영역(올라운더)은 마케팅 리소스를 집중하고, 다빈도저단가(박리다매) 영역은 업셀링을 통한 객단가 제고, 고단가저빈도 영역은 구매 전환율 보완 전략 수립을 권장합니다.")
        else:
            st.caption("차원별 3캠페인 이상 있어야 해요.")
        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 15 — 발송유형·브랜드 랭킹 (미사용 차원)
    # ══════════════════════════════════════════════════════════════
    elif "발송유형·브랜드" in page:
        st.title("발송유형·브랜드 랭킹")
        st.caption("발송유형별, 브랜드별 성과를 비교해요.")
        base = fdf
        if len(base) < 6:
            st.info("데이터가 부족해요. '최소 발송수'를 낮춰 보세요."); st.stop()
        mlabel = st.selectbox("성과 지표", list(METRIC_OPTS.keys()), key="p15_metric")
        mcol, msuf, mclr = METRIC_OPTS[mlabel]
        is_pct = mcol in ("ord_cr", "infl_cr")

        def _rank(dimcol, title, minn, topn):
            if dimcol not in base or base[dimcol].fillna("").eq("").all():
                st.caption(f"{title}: 데이터 없음"); return
            a = base.copy(); a[dimcol] = a[dimcol].fillna("(미지정)").replace("", "(미지정)")
            g = a.groupby(dimcol).agg(캠페인수=(dimcol, "size"), 발송=("send", "sum"),
                                      지표=(mcol, "mean"), 거래액=("amt", "sum")).reset_index()
            g = g[g["캠페인수"] >= minn].sort_values("지표", ascending=False)
            if len(g) == 0:
                st.caption(f"{title}: 최소 표본 미달"); return
            g = g.head(topn)
            yv = g["지표"] * (100 if is_pct else 1)
            overall = base[mcol].mean() * (100 if is_pct else 1)
            fig = go.Figure(go.Bar(
                x=yv, y=g[dimcol], orientation="h",
                marker_color=[PALETTE["green"] if v >= overall else PALETTE["slate"] for v in yv],
                text=[f"{v:.2f}{'%' if is_pct else ''} (n={int(n)})" for v, n in zip(yv, g["캠페인수"])],
                textposition="outside"))
            fig.add_vline(x=overall, line_dash="dash", line_color=PALETTE["red"])
            lay = base_layout(h=max(280, 40 + 28 * len(g)), title=f"{title} — 평균 {mlabel} (빨간선=전체평균)")
            lay["xaxis"]["range"] = bar_xrange(yv, 1.22)
            fig.update_layout(**lay); fig.update_yaxes(autorange="reversed")
            st.plotly_chart(fig, width="stretch")
            gs = g.copy()
            gs["지표"] = gs["지표"].map(lambda v: f"{v*100:.2f}%" if is_pct else (won(v) if mcol in ("rps", "aov", "amt") else f"{v:,.1f}"))
            gs["거래액"] = gs["거래액"].map(won)
            st.dataframe(gs.rename(columns={dimcol: title})[[title, "캠페인수", "발송", "지표", "거래액"]]
                         .style.format({"캠페인수": "{:,.0f}", "발송": "{:,.0f}"}),
                         hide_index=True, width="stretch")

        st.markdown("##### ① 발송유형별 성과")
        _rank("stype", "발송유형", minn=3, topn=15)
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            bminn = st.number_input("브랜드 최소 캠페인수", value=4, min_value=2, step=1, key="p15_bminn")
        with c2:
            btopn = st.number_input("브랜드 표시 상위", value=20, min_value=5, step=5, key="p15_btopn")
        st.markdown("##### ② 브랜드별 성과 (상위)")
        _rank("brand", "브랜드", minn=int(bminn), topn=int(btopn))
        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 16 — 다음주 발송 플레이북 (실행)
    # ══════════════════════════════════════════════════════════════
    elif "다음주 발송 플레이북" in page or "플레이북" in page:
        st.title("다음 주 발송 플레이북")
        st.caption("과거 성과 분석 데이터에 기반하여 차주 발송을 위한 세그먼트별 요일·시간·소구 매칭 테이블을 제공합니다.")
        base = fdf
        if len(base) < 8:
            st.info("데이터가 부족해요. '최소 발송수'를 낮춰 보세요."); st.stop()
        goal = st.selectbox("최적화 목표 지표", list(METRIC_OPTS.keys()), key="p16_goal")
        gcol = METRIC_OPTS[goal][0]; is_pct = gcol in ("ord_cr", "infl_cr")

        def _fmtv(v):
            return f"{v*100:.2f}%" if is_pct else (won(v) if gcol in ("rps", "aov", "amt") else f"{v:,.1f}")

        st.markdown("##### 📅 추천 발송 슬롯 (요일 × 시간)")
        if "dow_k" in base and "hour" in base:
            slot = base.dropna(subset=["hour"]).groupby(["dow_k", "hour"]).agg(
                캠페인수=("hour", "size"), 지표=(gcol, "mean"), 발송=("send", "sum"),
                _uv=("uv", "sum"), _oc=("oc", "sum"), _amt=("amt", "sum")).reset_index()
            _sf = slot.rename(columns={"_uv": "uv", "_oc": "oc", "발송": "send", "_amt": "amt"})
            slot["_rk"] = rank_adjusted(_sf, gcol, ascending=False).values
            slot = slot[slot["캠페인수"] >= 2].sort_values("_rk", ascending=False).head(8)
            if len(slot):
                sshow = slot.copy()
                sshow["시간"] = sshow["hour"].map(fmt_hhmm)
                sshow["지표"] = sshow["지표"].map(_fmtv)
                st.dataframe(sshow.rename(columns={"dow_k": "요일"})[["요일", "시간", "캠페인수", "지표", "발송"]]
                             .style.format({"캠페인수": "{:,.0f}", "발송": "{:,.0f}"}),
                             hide_index=True, width="stretch")
            else:
                st.caption("슬롯 데이터가 부족해요.")

        st.markdown("##### 💡 추천 소구 및 리프트 성과")
        trows = []
        for tag in TAG_BOOLS:
            if tag not in base:
                continue
            yes = base.loc[base[tag] == True, gcol].dropna()
            no = base.loc[base[tag] != True, gcol].dropna()
            if len(yes) < 3 or len(no) < 3:
                continue
            _l = yes.mean() - no.mean()
            trows.append({"소구": tag, "보유평균": _fmtv(yes.mean()),
                          "리프트": (f"{_l*100:+.2f}%p" if is_pct else f"{_l:+,.0f}"),
                          "_l": _l, "보유n": len(yes), "유의성": sig_label(welch(yes.values, no.values))})
        top_tags = []
        if trows:
            tdf = pd.DataFrame(trows).sort_values("_l", ascending=False)
            st.dataframe(tdf[["소구", "보유평균", "리프트", "보유n", "유의성"]].head(6)
                         .style.format({"보유n": "{:,.0f}"}), hide_index=True, width="stretch")
            top_tags = tdf.head(3)["소구"].tolist()
        else:
            st.caption("소구 리프트 표본 부족.")

        st.markdown("##### 📋 세그먼트별 실행 플레이북 (추천 조합)")
        play = []
        if "target" in base and base["target"].fillna("").ne("").any():
            bp = base.copy(); bp["target"] = bp["target"].fillna("(미지정)").replace("", "(미지정)")
            for sname, sg in bp.groupby("target"):
                if len(sg) < 3:
                    continue
                row = {"세그먼트": sname}
                if "cat" in sg and sg["cat"].fillna("").ne("").any():
                    # 카테고리 추천도 합산+표본 보정 점수로 — n=2 raw 평균 argmax는 노이즈 추첨
                    cg = sg.groupby("cat").agg(n=("af", "size"), oc=("oc", "sum"), uv=("uv", "sum"),
                                               send=("send", "sum"), amt=("amt", "sum"))
                    cg = cg[cg["n"] >= 2]
                    if len(cg):
                        row["추천 카테고리"] = (cg.assign(_rk=rank_adjusted(cg, gcol))
                                            .sort_values("_rk", na_position="first").index[-1])
                best = None
                for tag in TAG_BOOLS:
                    if tag not in sg:
                        continue
                    yes = sg.loc[sg[tag] == True, gcol].dropna()
                    no = sg.loc[sg[tag] != True, gcol].dropna()
                    if len(yes) < 3 or len(no) < 3:          # n=2 argmax는 우연 — 최소 3건
                        continue
                    lift = yes.mean() - no.mean()
                    if best is None or lift > best[1]:
                        best = (tag, lift)
                if best:
                    row["추천 소구"] = best[0]
                if "dow_k" in sg and "hour" in sg and sg["hour"].notna().any():
                    sl = sg.dropna(subset=["hour"]).groupby(["dow_k", "hour"]).agg(
                        oc=("oc", "sum"), uv=("uv", "sum"), send=("send", "sum"), amt=("amt", "sum"))
                    if len(sl):
                        bd, bh = (sl.assign(_rk=rank_adjusted(sl, gcol))
                                  .sort_values("_rk", na_position="first").index[-1])
                        row["추천 슬롯"] = f"{bd} {fmt_hhmm(bh)}"
                row["기대 " + goal] = _fmtv(sg[gcol].mean())
                play.append(row)
        if play:
            st.dataframe(pd.DataFrame(play), hide_index=True, width="stretch")
            st.caption("과거 실적 기준 최적 반응 조합 매칭 테이블입니다. 차주 발송 캘린더 기획 시 우선 배치를 권장합니다.")
        else:
            st.caption("세그먼트(target) 데이터가 없어 조합 추천을 건너뜁니다.")
        if top_tags:
            _rec = " · ".join(top_tags)
            st.markdown(f'<div class="appendix">추천 소구: {_rec} → 「AI 처방·카피」 페이지에서 해당 소구를 기반으로 신규 카피 초안 생성이 가능합니다.</div>', unsafe_allow_html=True)
        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 10 — 기획전 비교분석 (발송 promo × 기획전 매출)
    # ══════════════════════════════════════════════════════════════
    elif "기획전 비교분석" in page:
        st.title("기획전 비교분석")
        st.caption("발송 이력 내 기획전 번호(promo)와 기획전 전체 성과를 매핑하여, 발송 기여 수준 및 상대적 효율을 종합 정량 분석합니다. (발송 데이터는 현재 지정된 사이드바 필터가 적용됨)")
        if promo_df is None or len(promo_df) == 0:
            st.info("사이드바 **📂 파일 업로드**에 **기획전 성과시트(xlsx)**를 올려 주세요(자동 인식). "
                    "업로드 후 **「💾 기획전 저장」**을 누르면 누적돼요.")
            st.stop()

        # 발송측: promo 단위 집계 (현재 필터 적용분 · 매칭 여부 무관)
        src = dff_all.copy()
        src["promo"] = src["promo"].map(norm_promo) if "promo" in src else ""
        sent = src[src["promo"] != ""].copy()
        if len(sent) == 0:
            st.warning("발송 데이터에 기획전번호(promo)가 채워진 건이 없어요. "
                       "실적시트의 '기획전' 컬럼을 확인해 보세요.")
            st.stop()
        # 발송 일시(날짜+시간대) — 같은 기획전에 여러 캠페인이면 가장 이른 발송 기준
        _min = sent["hour"].map(hhmm_to_minutes) if "hour" in sent else 0   # 시간대 HHMM→분(정렬·표기)
        sent["_dth"] = (sent["dt"] + pd.to_timedelta(_min, unit="m")) if "dt" in sent else pd.NaT
        g = sent.groupby("promo").agg(
            n_camp=("af", "size"), send=("send", "sum"), s_amt=("amt", "sum"),
            s_oc=("oc", "sum"), s_uv=("uv", "sum"), s_visit=("visit", "sum"),
            first_dth=("_dth", "min"),
        ).reset_index()
        g["s_rps"] = np.where(g["send"] > 0, g["s_amt"] / g["send"], 0.0)
        # 발송성과 비율 — 대시보드 정의와 동일(가중: 합산÷합산)
        g["s_ctr"] = np.where(g["send"] > 0, g["s_uv"] / g["send"], np.nan)   # 유입전환율 = UV÷발송
        g["s_cr"] = np.where(g["s_uv"] > 0, g["s_oc"] / g["s_uv"], np.nan)    # 주문전환율 = 주문÷UV

        def _fmt_date(r):
            t = r.get("first_dth")
            if t is None or pd.isna(t):
                return "–"
            s = f"{t.year}년 {t.month}월 {t.day}일"
            return s + (f" 외 {int(r['n_camp'])-1}건" if r.get("n_camp", 1) > 1 else "")

        def _fmt_time(r):
            t = r.get("first_dth")
            if t is None or pd.isna(t):
                return "–"
            return f"{t.hour:02d}시" + (f" {t.minute:02d}분" if t.minute else "")
        g["발송일자"] = g.apply(_fmt_date, axis=1)
        g["발송시간"] = g.apply(_fmt_time, axis=1)

        # 기획전 매출 조인
        P = promo_df.copy()
        m = g.merge(P, on="promo", how="left")
        matched = m[m["pname"].notna() & (m["pname"].astype(str).str.strip() != "")].copy()
        n_unmatched = len(m) - len(matched)
        sent_ids = set(g["promo"])

        c = st.columns(4)
        c[0].metric("발송된 기획전 수", f"{g['promo'].nunique():,}")
        c[1].metric("기획전시트 매칭", f"{len(matched):,}",
                    delta=(f"미매칭 {n_unmatched}" if n_unmatched else "전건 매칭"),
                    delta_color="off")
        c[2].metric("발송 추적 거래액(필터)", won(g["s_amt"].sum()))
        c[3].metric("매칭 기획전 유입거래액", won(matched["inf_amt"].sum()))
        if n_unmatched:
            st.caption(f"⚠️ 성과시트에서 번호를 못 찾은 기획전 {n_unmatched}건은 "
                       "①·②·④ 분석에서 빠져요. 성과시트를 업데이트해 주세요.")

        tabA, tabB, tabC, tabD = st.tabs(
            ["① 발송 기여율", "② 발송 효율 순위", "③ 발송 유무별 매출", "④ 매출 추세"])

        # ── ① 발송 기여율 (분모 = 유입 거래액) ──
        with tabA:
            st.markdown("##### 발송 기여율 = 발송 추적 거래액 ÷ 기획전 **유입** 거래액")
            st.caption("’유입 거래액’은 해당 기획전을 경유해 발생한 총 매출액이며, 발송 기여율은 이 중 발송 추적을 통해 발생한 매출의 비중을 의미합니다. (어트리뷰션 기준 차이로 인해 100%를 초과할 수 있음)")
            a = matched[matched["inf_amt"].fillna(0) > 0].copy()
            if len(a) == 0:
                st.info("유입 거래액이 있는 매칭 기획전이 없어요.")
            else:
                a["기여율"] = a["s_amt"] / a["inf_amt"]
                tot_contrib = a["s_amt"].sum() / a["inf_amt"].sum() if a["inf_amt"].sum() > 0 else np.nan
                cc = st.columns(3)
                cc[0].metric("전체 발송 기여율(합산)", f"{tot_contrib*100:.1f}%")
                cc[1].metric("기획전 중앙값 기여율", f"{a['기여율'].median()*100:.1f}%")
                cc[2].metric("대상 기획전 수", f"{len(a):,}")
                clip = (a["기여율"].clip(upper=2.0) * 100)
                fig = go.Figure(go.Histogram(x=clip, nbinsx=30, marker_color=PALETTE["blue"]))
                lay = base_layout(h=300, title="기여율 분포 (200% 초과는 200%로 표시)")
                fig.update_layout(**lay)
                fig.update_xaxes(ticksuffix="%")
                st.plotly_chart(fig, width="stretch")
                show = a.sort_values("기여율", ascending=False)
                st.markdown("**기여율 우수 기획전 — 발송 주도형 매출**")
                st.dataframe(promo_perf_table(show.head(15)),
                             hide_index=True, width="stretch")
                lowbase = show[show["inf_amt"] >= show["inf_amt"].median()]
                st.markdown("**기여율 미흡 기획전 (유입 거래액 중앙값 이상) — 자연 유입 의존도 우세, 마케팅 강화 권장**")
                st.dataframe(promo_perf_table(lowbase.sort_values("기여율").head(15)),
                             hide_index=True, width="stretch")

        # ── ② 발송 효율 순위 ──
        with tabB:
            st.markdown("##### 기획전별 발송 효율 — 발송 대비 실매출")
            b = matched.copy()
            b["기여율"] = np.where(b["inf_amt"].fillna(0) > 0, b["s_amt"] / b["inf_amt"], np.nan)
            order = st.radio("정렬 기준", ["발송 RPS", "기여율", "유입거래액", "발송수"],
                             horizontal=True, key="promoB_sort")
            sortmap = {"발송 RPS": "s_rps", "기여율": "기여율", "유입거래액": "inf_amt",
                       "발송수": "send"}
            b = b.sort_values(sortmap[order], ascending=False, na_position="last")
            st.dataframe(promo_perf_table(b.head(50)),
                         hide_index=True, width="stretch", height=520)
            st.markdown('<div class="appendix">'
                        '본 분석의 테이블 열은 <b>발송 실적</b>(발송 로그 추적 데이터)과 <b>기획전 성과</b>(기획전 성과 마스터 시트) 데이터로 구분됩니다. '
                        '유입 고객수(UV) 역시 데이터 소스에 따라 <b>발송 실적 UV</b>(발송 링크 유입 고객수)와 <b>기획전 성과 UV</b>(해당 기획전의 전체 유입 고객수)로 구분되므로 주의하시기 바랍니다.<br>'
                        '· CTR = 발송 유입 UV ÷ 발송 수량  · 주문CR = 주문건수 ÷ 발송 유입 UV  · RPS = 발송 추적 거래액 ÷ 발송 수량  · 기여율 = 발송 추적 거래액 ÷ 유입 거래액<br>'
                        '· 단일 기획전 번호 내 다수의 캠페인이 중복 집행된 경우, 거래액/발송/UV/주문 등의 정량 지표는 합산 처리하며 비율 지표(CTR, CR, RPS)는 합계 기준의 가중 평균으로 계산합니다. 발송 일시는 가장 조기에 집행된 일자 기준입니다.</div>',
                        unsafe_allow_html=True)

        # ── ③ 발송 유무별 매출 ──
        with tabC:
            st.markdown("##### 발송한 기획전 vs 발송 안 한 기획전 — 매출 비교")
            st.caption("기획전 성과 마스터의 전체 모수를 발송 집행 여부로 구분하여 평균 및 중앙값을 대조합니다. (단, 대형 기획전 위주의 편중 집행이 존재할 수 있어 단순 성과 비교 시 선택 편향이 있을 수 있습니다)")
            base_lbl = "유입 거래액"
            col = "inf_amt"
            allp = P.copy()
            allp["발송"] = allp["promo"].isin(sent_ids)
            valid = allp[allp[col].notna()].copy()
            sset = valid[valid["발송"]][col]
            uset = valid[~valid["발송"]][col]
            summary = pd.DataFrame({
                "구분": ["발송함", "발송안함"],
                "기획전수": [int(len(sset)), int(len(uset))],
                "평균매출": [sset.mean(), uset.mean()],
                "중앙값매출": [sset.median(), uset.median()],
                "합계매출": [sset.sum(), uset.sum()],
            })
            st.dataframe(summary.style.format({"기획전수": "{:,.0f}", "평균매출": "{:,.0f}",
                         "중앙값매출": "{:,.0f}", "합계매출": "{:,.0f}"}),
                         hide_index=True, width="stretch")
            fig = go.Figure(go.Bar(x=["발송함", "발송안함"], y=[sset.mean(), uset.mean()],
                                   marker_color=[PALETTE["green"], PALETTE["slate"]]))
            fig.update_layout(**base_layout(h=320, title=f"발송 유무별 평균 {base_lbl}"))
            st.plotly_chart(fig, width="stretch")
            p = welch(sset.dropna().values, uset.dropna().values)
            st.caption(f"평균 차이 통계 유의성: {sig_label(p)} · 발송 {len(sset):,}건 vs 미발송 {len(uset):,}건")
            st.markdown('<div class="appendix">평균값은 일부 초대형 기획전(아웃라이어)에 왜곡되기 쉬우므로 중앙값을 병행하여 검토하십시오. 집행 집단의 매출 규모가 월등히 높은 경우, 발송 액션의 실질 성과인지 혹은 대형 기획전 대상의 선택 편향인지에 대해서는 통계적 추가 검증이 필요합니다.</div>',
                        unsafe_allow_html=True)

        # ── ④ 매출 추세 ──
        with tabD:
            st.markdown("##### 기획전 매출 추세 (기획전 시작월 기준)")
            P2 = P.copy()
            P2["dt"] = pd.to_datetime(P2["pstart"], errors="coerce")
            P2 = P2.dropna(subset=["dt"])
            if len(P2) == 0:
                st.info("기획전 시작일 정보가 없어서 추세를 그릴 수 없어요.")
            else:
                base_lbl = "유입 거래액"
                col = "inf_amt"
                P2["발송"] = P2["promo"].isin(sent_ids)
                P2["월"] = P2["dt"].dt.to_period("M").apply(lambda pp: pp.start_time)
                gm = P2.groupby(["월", "발송"])[col].sum().reset_index()
                fig = go.Figure()
                for flag, name, clr in [(True, "발송 기획전", PALETTE["green"]),
                                        (False, "미발송 기획전", PALETTE["slate"])]:
                    sub = gm[gm["발송"] == flag].sort_values("월")
                    fig.add_trace(go.Scatter(x=sub["월"], y=sub[col], mode="lines+markers",
                                             name=name, line=dict(color=clr, width=2)))
                lay = base_layout(h=420, title=f"월별 {base_lbl} 합계 추이 (발송/미발송 기획전)", hover="x")
                lay["showlegend"] = True
                lay["legend"] = legend_h()
                fig.update_layout(**lay)
                st.plotly_chart(fig, width="stretch")
                cnt = P2.groupby(["월", "발송"]).size().reset_index(name="기획전수")
                figc = go.Figure()
                for flag, name, clr in [(True, "발송 기획전", PALETTE["green"]),
                                        (False, "미발송 기획전", PALETTE["slate"])]:
                    sub = cnt[cnt["발송"] == flag].sort_values("월")
                    figc.add_trace(go.Scatter(x=sub["월"], y=sub["기획전수"], mode="lines+markers",
                                              name=name, line=dict(color=clr, width=2)))
                layc = base_layout(h=300, title="월별 기획전 수 (발송/미발송)", hover="x")
                layc["showlegend"] = True
                layc["legend"] = legend_h()
                figc.update_layout(**layc)
                st.plotly_chart(figc, width="stretch")
                st.caption("기획전 런칭 시작월 기준 집계 데이터입니다. 마케팅(발송) 집행이 활성화된 전후 시점의 발송 기획전(녹색 선) 매출 규모 추이를 검토하십시오.")

        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 07 — 데이터·다운로드
    # ══════════════════════════════════════════════════════════════
    elif "데이터·다운로드" in page or "데이터" in page:
        st.title("데이터 · 다운로드")
        st.markdown(f"**머지 결과** — 전체 {len(raw)}건 · 문구 매칭 {raw['matched'].sum()}건 "
                    f"({raw['matched'].mean()*100:.0f}%)")
        st.dataframe(df.drop(columns=["dt"], errors="ignore"), hide_index=True,
                     width="stretch", height=420)
        # ── 머지 전체 데이터 다운로드 (기획 문구 + 실적 성과) ──
        st.markdown("##### 📊 머지 전체 데이터 다운로드")
        st.caption("문구와 성과를 합친 전체 데이터예요. 자동분류 속성 컬럼도 포함돼요.")
        dl_scope = st.radio("범위", ["전체 (필터 무관 · 매칭+미매칭 포함)", "현재 필터 적용분"],
                            horizontal=True, key="full_dl_scope")
        dl_df = raw if dl_scope.startswith("전체") else df
        dl_df = dl_df.drop(columns=["dt"], errors="ignore")
        d1, d2 = st.columns(2)
        d1.download_button(
            f"📥 CSV 다운로드 ({len(dl_df):,}건)",
            dl_df.to_csv(index=False).encode("utf-8-sig"),
            file_name=f"발송성과_머지전체_{datetime.date.today():%Y%m%d}.csv", mime="text/csv", width="stretch")
        if d2.button(f"📊 엑셀 생성 ({len(dl_df):,}건)", key="gen_full_xlsx", width="stretch"):
            try:
                with st.spinner("엑셀 생성 중…"):
                    st.session_state["full_xlsx"] = df_to_xlsx_bytes(dl_df)
                    st.session_state["full_xlsx_n"] = len(dl_df)
                st.success(f"엑셀을 만들었어요 — {len(dl_df):,}건")
            except Exception as e:
                st.error(f"엑셀 생성에 실패했어요: {e}")
        if st.session_state.get("full_xlsx"):
            st.download_button(
                f"📥 머지 전체 데이터 엑셀(xlsx) 다운로드 ({st.session_state.get('full_xlsx_n', 0):,}건)",
                st.session_state["full_xlsx"], file_name=f"발송성과_머지전체_{datetime.date.today():%Y%m%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # ── 종합 리포트(엑셀) 내보내기 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📑 종합 리포트 내보내기")
        st.caption("분석 결과를 여러 시트로 담은 엑셀이에요. 현재 필터가 적용된 표본 기준이에요.")
        if st.button("📑 리포트 생성", key="gen_report"):
            try:
                with st.spinner("리포트 생성 중…"):
                    st.session_state["report_xlsx"] = build_report_excel(fdf)
                st.success(f"리포트를 만들었어요 — {len(fdf)}건 기준")
            except Exception as e:
                st.error(f"리포트 생성에 실패했어요: {e}")
        if st.session_state.get("report_xlsx"):
            st.download_button(
                "📥 종합 리포트(xlsx) 다운로드", st.session_state["report_xlsx"],
                file_name=f"발송성과_리포트_{datetime.date.today():%Y%m%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # ── 매칭 품질 상시 감시 — 특정 주차부터 매칭이 무너지면(시트 형식 변경 등) 즉시 보이게 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 🧩 매칭 품질 — 주차별 매칭률 추이")
        _mq = raw.dropna(subset=["dt"]).copy()
        if len(_mq):
            _mq["주"] = _mq["dt"].dt.to_period("W").dt.start_time
            wk_m = (_mq.groupby("주").agg(건수=("af", "size"), 매칭률=("matched", "mean"))
                    .reset_index().sort_values("주"))
            _mcolors = [PALETTE["red"] if v < 0.8 else PALETTE["green"] for v in wk_m["매칭률"]]
            figm = go.Figure(go.Bar(
                x=wk_m["주"], y=wk_m["매칭률"] * 100, marker_color=_mcolors,
                customdata=wk_m["건수"],
                hovertemplate="%{x|%m/%d} 주<br>매칭률 %{y:.0f}% · 캠페인 %{customdata}건<extra></extra>"))
            laym = base_layout(h=260, title="주차별 문구 매칭률(%) — 80% 미만 주는 빨강")
            laym["yaxis"]["range"] = [0, 108]
            figm.update_layout(**laym)
            st.plotly_chart(figm, width="stretch")
            _low = wk_m[wk_m["매칭률"] < 0.8]
            if len(_low):
                st.caption("⚠️ 매칭률 80% 미만 주: "
                           + ", ".join(pd.Timestamp(w).strftime("%m/%d") for w in _low["주"])
                           + " — 해당 주차 기획 시트의 날짜·AF코드 형식을 확인해 보세요.")

        st.markdown("##### ⚠️ 매칭 진단 — 실적은 있지만 문구를 못 찾은 건")
        miss = raw[~raw["matched"]][["date", "af", "cat", "brand", "send", "amt"]].copy()
        if len(miss):
            # 원인 4분류 — 기획 lookup이 세션에 있을 때만 (업로드 1회성 로그를 상시 진단으로)
            _lk = st.session_state.get("plan_lookup_gs")
            if _lk:
                _pdates = {d for (d, a) in _lk}
                _pafs = {a for (d, a) in _lk}

                def _cause(r):
                    d_ok = str(r["date"]) in _pdates
                    a_ok = str(r["af"]) in _pafs
                    if d_ok and a_ok:
                        return "날짜·AF 각각은 있음 (조합 불일치)"
                    if d_ok:
                        return "AF코드가 기획에 없음"
                    if a_ok:
                        return "날짜가 기획에 없음"
                    return "날짜·AF 모두 없음 (기획 미적재 주차)"
                miss["원인 추정"] = [_cause(r) for _, r in miss.iterrows()]
                _cnt = miss["원인 추정"].value_counts()
                st.caption("원인 분류 — " + " · ".join(f"**{k}** {v}건" for k, v in _cnt.items()))
            st.dataframe(miss.rename(columns={"date": "날짜", "af": "AF코드", "cat": "카테고리",
                                              "brand": "브랜드", "send": "발송", "amt": "거래액"}),
                         hide_index=True, width="stretch")
            st.caption("AF코드 오타·미등록·날짜 불일치 가능성이 있어요. 기획 파일을 확인해 보세요. "
                       "(원인 분류는 「기획 문구 가져오기」로 적재된 기획 기준)")
        else:
            st.success("모든 실적 캠페인에 문구가 매칭됐어요.")
        glossary()

    # ══════════════════════════════════════════════════════════════
    # PAGE 📱 — 앱푸시 동의 현황
    # ══════════════════════════════════════════════════════════════
    elif "앱푸시 동의 현황" in page:
        import plotly.express as px
        st.title("📱 앱푸시 수신동의 현황")

        if push_consent_df is None or push_consent_df.empty:
            st.info("👈 사이드바에서 **앱푸시 동의 현황 xlsx** 파일을 올려주세요.")
            st.stop()
        
        # date 컬럼을 안전하게 datetime 형식으로 통일
        push_consent_df = push_consent_df.copy()
        push_consent_df["date"] = pd.to_datetime(push_consent_df["date"])
        # 이상치 재계산(런타임) — 신규추가·기존이탈뿐 아니라 '순증감(diff) 급락'도 배치/이관
        # 이벤트로 보고 제외한다. (예전에 저장된 데이터도 재업로드 없이 −20k 스파이크가 걸러짐)
        for _c in ("added", "removed", "diff"):
            if _c in push_consent_df:
                push_consent_df[_c] = pd.to_numeric(push_consent_df[_c], errors="coerce")
        _a = push_consent_df.get("added", pd.Series(0, index=push_consent_df.index)).fillna(0)
        _r = push_consent_df.get("removed", pd.Series(0, index=push_consent_df.index)).fillna(0)
        _f = push_consent_df.get("diff", pd.Series(0, index=push_consent_df.index)).fillna(0)
        push_consent_df["is_outlier"] = (
            (_a.abs() > PUSH_OUTLIER_THRESHOLD) | (_r.abs() > PUSH_OUTLIER_THRESHOLD)
            | (_f.abs() > PUSH_OUTLIER_THRESHOLD))

        # ── 필터: 그룹 & 날짜 범위 ──
        c_grp, c_date = st.columns([1, 2])
        with c_grp:
            sel_group = st.selectbox("그룹", ["Total", "기존", "신규"], index=0, key="pc_group")
        with c_date:
            _pc_dates = push_consent_df["date"].dropna()
            _pc_min, _pc_max = _pc_dates.min().date(), _pc_dates.max().date()
            pc_date_range = st.date_input(
                "기간", value=(_pc_min, _pc_max),
                min_value=_pc_min, max_value=_pc_max, key="pc_dates")

        # 날짜 필터 적용
        try:
            _d0, _d1 = pd.Timestamp(pc_date_range[0]), pd.Timestamp(pc_date_range[1])
        except Exception:
            _d0, _d1 = pd.Timestamp(_pc_min), pd.Timestamp(_pc_max)

        pc_all = push_consent_df[push_consent_df["group"] == sel_group].copy()
        pc_all = pc_all[(pc_all["date"] >= _d0) & (pc_all["date"] <= _d1)]
        pc_clean = pc_all[~pc_all["is_outlier"]].copy()
        n_outlier = int(pc_all["is_outlier"].sum())

        if pc_clean.empty:
            st.warning("선택한 기간에 유효 데이터가 없어요.")
            st.stop()

        # ── KPI 카드 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📊 주요 지표 (이상치 제외 기간 기준)")
        _pc_sorted = pc_clean.sort_values("date")
        _latest = _pc_sorted.iloc[-1]
        # '최근 7일' 비교는 행 위치(iloc[-8])가 아니라 실제 날짜 기준 — 중간에 결측일이
        # 있으면 7일 전이 아닌 날과 비교되던 문제 방지
        _prev_cand = _pc_sorted[_pc_sorted["date"] <= _latest["date"] - pd.Timedelta(days=7)]
        _prev = _prev_cand.iloc[-1] if len(_prev_cand) else _pc_sorted.iloc[0]
        _consent_now  = int(_latest["consent"])  if pd.notna(_latest["consent"])  else 0
        _consent_prev = int(_prev["consent"])    if pd.notna(_prev["consent"])    else 0
        _consent_delta = _consent_now - _consent_prev

        _avg_added   = pc_clean["added"].mean()
        _avg_removed = pc_clean["removed"].mean()
        _avg_diff    = pc_clean["diff"].mean()
        _net_ratio   = (_avg_diff / _avg_added * 100) if (_avg_added and _avg_added > 0) else 0

        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("총 수신동의 수", f"{_consent_now:,}명",
                  delta=f"{_consent_delta:+,}명 (최근 7일)", delta_color="normal")
        k2.metric("일평균 신규추가",  f"{_avg_added:,.0f}명")
        k3.metric("일평균 기존이탈",    f"{_avg_removed:,.0f}명")
        k4.metric("일평균 순증감",    f"{_avg_diff:+,.0f}명",
                  delta_color="normal")
        k5.metric("신규추가 대비 순증 비율", f"{_net_ratio:.1f}%",
                  help="순증감 ÷ 신규추가 × 100. 100%에 가까울수록 탈퇴가 적음")

        if n_outlier > 0:
            st.caption(
                f"⚠️ 배치/이관 이벤트로 추정되는 이상치 **{n_outlier}일** 은 KPI·차트에서 제외됐어요 "
                f"(신규추가 또는 기존이탈 > {PUSH_OUTLIER_THRESHOLD:,}건 기준).")

        # ── 주간 지표 테이블 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📅 주간 지표 (이상치 제외)")
        wk_df = push_weekly(push_consent_df[(push_consent_df["date"] >= _d0) & (push_consent_df["date"] <= _d1)],
                            group=sel_group)
        if not wk_df.empty:
            wk_show = wk_df[["주시작", "주차", "일수", "기말동의수", "순증감합", "신규추가합", "탈퇴합"]].copy()
            wk_show = wk_show.sort_values("주시작", ascending=False).reset_index(drop=True)
            wk_show = wk_show.drop(columns=["주시작"])
            wk_show["기말동의수"] = wk_show["기말동의수"].apply(lambda v: f"{v:,.0f}" if pd.notna(v) else "—")
            wk_show["순증감합"]  = wk_show["순증감합"].apply(lambda v: f"{v:+,.0f}" if pd.notna(v) else "—")
            wk_show["신규추가합"] = wk_show["신규추가합"].apply(lambda v: f"{v:,.0f}" if pd.notna(v) else "—")
            wk_show["탈퇴합"]   = wk_show["탈퇴합"].apply(lambda v: f"{v:,.0f}" if pd.notna(v) else "—")
            wk_show.rename(columns={
                "주차": "주차", "일수": "집계일수", "기말동의수": "기말 동의수",
                "순증감합": "순증감(합)", "신규추가합": "신규추가(합)", "탈퇴합": "기존이탈(합)"}, inplace=True)
            st.dataframe(wk_show, hide_index=True, width="stretch", height=360)
            st.caption("**기말 동의수**는 그 주 마지막 날 기준 누적 동의자 수(스냅샷)이고, "
                       "**순증감·신규추가·기존이탈**은 그 주 전체를 더한 값(합계)이에요.")

            # 이탈률 급등 조기 경보 — 최근 주 이탈률이 과거 추세 대비 z>2면 배너
            if len(wk_df) >= 9:
                _wr_al = wk_df.sort_values("주시작").copy()
                _wr_al["이탈률"] = _wr_al["탈퇴합"] / _wr_al["기말동의수"].replace(0, np.nan)
                _hist_r = _wr_al["이탈률"].iloc[:-1].tail(26).dropna()
                _now_r = _wr_al["이탈률"].iloc[-1]
                if len(_hist_r) >= 8 and _hist_r.std(ddof=0) > 0 and pd.notna(_now_r):
                    _z_r = (float(_now_r) - float(_hist_r.mean())) / float(_hist_r.std(ddof=0))
                    if _z_r > 2:
                        st.warning(f"🚨 최근 주 이탈률 {_now_r*100:.2f}%가 과거 26주 평균"
                                   f"({_hist_r.mean()*100:.2f}%) 대비 +{_z_r:.1f}σ — 급등 신호예요. "
                                   "최근 발송 강도·캠페인 내용을 점검해 보세요.")

            # 주간 CSV 다운로드 — 위에서 계산한 wk_df 재사용 (동일 인자 재계산 제거)
            _wk_raw = wk_df
            if not _wk_raw.empty:
                st.download_button(
                    "📥 주간 데이터 CSV",
                    _wk_raw.to_csv(index=False).encode("utf-8-sig"),
                    file_name=f"앱푸시동의_주간_{sel_group}_{pd.Timestamp.today():%Y%m%d}.csv",
                    mime="text/csv")

        # ── 차트 1: 수신동의 수 추이 (일별·이상치 제외) ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📈 수신동의 수 추이 (일별)")
        _pc_plot = pc_clean.sort_values("date")
        fig_consent = go.Figure()
        # 단일 날짜면 lines만으론 아무것도 안 그려짐(라인은 2점 필요) → 마커 포함
        _cons_mode = "lines" if len(_pc_plot) >= 2 else "markers"
        fig_consent.add_trace(go.Scatter(
            x=_pc_plot["date"], y=_pc_plot["consent"],
            mode=_cons_mode, name="수신동의 수",
            line=dict(color=PALETTE["blue"], width=2),
            marker=dict(color=PALETTE["blue"], size=7),
            hovertemplate="%{x|%Y-%m-%d}<br>수신동의: %{y:,}명<extra></extra>"))
        # 7일 이동평균
        if len(_pc_plot) >= 7:
            _ma7 = _pc_plot["consent"].rolling(7, center=True).mean()
            fig_consent.add_trace(go.Scatter(
                x=_pc_plot["date"], y=_ma7,
                mode="lines", name="7일 이동평균",
                line=dict(color=PALETTE["amber"], width=2, dash="dot"),
                hovertemplate="%{x|%Y-%m-%d}<br>7일 평균: %{y:,.0f}명<extra></extra>"))
        
        lay_consent = base_layout(h=360, title="일별 수신동의 수 추이")
        lay_consent["showlegend"] = True
        lay_consent["yaxis"]["gridcolor"] = "#f1f5f9"
        lay_consent["yaxis"]["title"] = "수신동의 수"
        fig_consent.update_layout(**lay_consent)
        st.plotly_chart(fig_consent, width="stretch")

        # ── 차트 2: 신규추가 / 기존이탈 추이 (일별) ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📊 일별 신규추가 · 기존이탈 추이")
        fig_addrem = go.Figure()
        fig_addrem.add_trace(go.Bar(
            x=_pc_plot["date"], y=_pc_plot["added"],
            name="신규추가", marker_color=PALETTE["green"],
            hovertemplate="%{x|%Y-%m-%d}<br>신규추가: %{y:,}명<extra></extra>"))
        fig_addrem.add_trace(go.Bar(
            x=_pc_plot["date"], y=-_pc_plot["removed"],
            name="기존이탈(아래 방향)", marker_color=PALETTE["red"],
            customdata=_pc_plot["removed"],           # hover는 양수로 — "-1,234명" 오독 방지
            hovertemplate="%{x|%Y-%m-%d}<br>기존이탈: %{customdata:,}명<extra></extra>"))
        # 순증감 라인
        fig_addrem.add_trace(go.Scatter(
            x=_pc_plot["date"], y=_pc_plot["diff"],
            mode="lines", name="순증감",
            line=dict(color=PALETTE["amber"], width=2),
            hovertemplate="%{x|%Y-%m-%d}<br>순증감: %{y:+,}명<extra></extra>"))
        
        lay_addrem = base_layout(h=360, title="일별 신규추가 및 기존이탈 추이")
        lay_addrem["showlegend"] = True
        lay_addrem["barmode"] = "overlay"
        lay_addrem["yaxis"]["gridcolor"] = "#f1f5f9"
        lay_addrem["yaxis"]["zeroline"] = True
        lay_addrem["yaxis"]["zerolinecolor"] = "#cbd5e1"
        lay_addrem["yaxis"]["title"] = "명"
        fig_addrem.update_layout(**lay_addrem)
        st.plotly_chart(fig_addrem, width="stretch")

        # ── 차트 3: 주간 순증감 바차트 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📊 주간 순증감 추이")
        _wk_chart = wk_df.copy()                      # 동일 인자 재계산 제거 — 상단 주간 집계 재사용
        if not _wk_chart.empty:
            _wk_chart = _wk_chart.sort_values("주시작")
            _colors_wk = [PALETTE["green"] if v >= 0 else PALETTE["red"] for v in _wk_chart["순증감합"]]
            fig_wk = go.Figure()
            # customdata로 주차 문자열 전달
            _wk_custom = _wk_chart["주차"].values
            fig_wk.add_trace(go.Bar(
                x=_wk_chart["주시작"], y=_wk_chart["순증감합"],
                name="주간 순증감", marker_color=_colors_wk,
                customdata=_wk_custom,
                hovertemplate="%{customdata}<br>순증감: %{y:+,.0f}명<extra></extra>"))
            fig_wk.add_trace(go.Scatter(
                x=_wk_chart["주시작"], y=_wk_chart["신규추가합"],
                mode="lines+markers", name="주간 신규추가",
                line=dict(color=PALETTE["blue"], width=2),
                customdata=_wk_custom,
                hovertemplate="%{customdata}<br>신규추가: %{y:,.0f}명<extra></extra>"))
            fig_wk.add_trace(go.Scatter(
                x=_wk_chart["주시작"], y=_wk_chart["탈퇴합"],
                mode="lines+markers", name="주간 기존이탈",
                line=dict(color=PALETTE["amber"], width=2),
                customdata=_wk_custom,
                hovertemplate="%{customdata}<br>기존이탈: %{y:,.0f}명<extra></extra>"))
            
            lay_wk = base_layout(h=400, title="주간 신규추가/탈퇴 및 순증감 추이")
            lay_wk["showlegend"] = True
            lay_wk["yaxis"]["gridcolor"] = "#f1f5f9"
            lay_wk["yaxis"]["zeroline"] = True
            lay_wk["yaxis"]["zerolinecolor"] = "#cbd5e1"
            lay_wk["yaxis"]["title"] = "명"
            lay_wk["xaxis"]["type"] = "date"  # 명시적으로 날짜 축 설정
            lay_wk["xaxis"]["tickformat"] = "%Y-%m"  # 연도-월 형태로 눈금 깔끔하게 포맷
            lay_wk["xaxis"]["dtick"] = "M3"  # 3개월 간격으로 격자선 및 눈금 표시
            fig_wk.update_layout(**lay_wk)
            st.plotly_chart(fig_wk, width="stretch")

        # ── 차트 4: 기존 vs 신규 동의수 비교 (스택 영역) ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📊 기존 · 신규 수신동의 구성 추이")
        _old_g = push_consent_df[
            (push_consent_df["group"] == "기존") &
            (~push_consent_df["is_outlier"]) &
            (push_consent_df["date"] >= _d0) &
            (push_consent_df["date"] <= _d1)].sort_values("date")
        _new_g = push_consent_df[
            (push_consent_df["group"] == "신규") &
            (~push_consent_df["is_outlier"]) &
            (push_consent_df["date"] >= _d0) &
            (push_consent_df["date"] <= _d1)].sort_values("date")
        if not _old_g.empty and not _new_g.empty:
            fig_stack = go.Figure()
            fig_stack.add_trace(go.Scatter(
                x=_old_g["date"], y=_old_g["consent"],
                mode="lines", name="기존 동의",
                stackgroup="one", line=dict(color=PALETTE["blue"]),
                hovertemplate="%{x|%Y-%m-%d}<br>기존: %{y:,}명<extra></extra>"))
            fig_stack.add_trace(go.Scatter(
                x=_new_g["date"], y=_new_g["consent"],
                mode="lines", name="신규 동의",
                stackgroup="one", line=dict(color=PALETTE["green"]),
                hovertemplate="%{x|%Y-%m-%d}<br>신규: %{y:,}명<extra></extra>"))
            
            lay_stack = base_layout(h=360, title="기존/신규 수신동의 구성 추이 (절대값 누적)")
            lay_stack["showlegend"] = True
            lay_stack["yaxis"]["gridcolor"] = "#f1f5f9"
            lay_stack["yaxis"]["title"] = "수신동의 수"
            fig_stack.update_layout(**lay_stack)
            st.plotly_chart(fig_stack, width="stretch")

        # ── 크로스 분석: 발송 강도(MTD) ↔ 수신동의 증감 ──
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 🔀 발송 강도 ↔ 수신동의 증감 (전사 MTD 연동)")
        if mtd_data is None:
            st.info("전사 **MTD 발송상세** 파일을 함께 올리면, 발송량이 많은 주/월에 "
                    "수신동의 순증감이 정체·역전되는지 크로스로 분석해 드려요.")
        else:
            st.caption("일별은 노이즈가 커서 주간·월간으로 묶어서 봐요. "
                       "발송이 많이 나간 기간에 순증감(신규−탈퇴)이 줄거나 마이너스면 "
                       "‘발송 피로 → 동의 이탈’ 신호예요. (선택한 그룹·기간·이상치 제외 기준)")
            @st.fragment
            def _cross_analysis_block():
                # fragment — 집계 단위·지표 변경 시 페이지 전체가 아니라 이 블록만 다시 그린다
                xc1, xc2 = st.columns(2)
                _gran = xc1.radio("집계 단위", ["주간", "월간"], horizontal=True, key="pc_x_gran")
                _sendlab = xc2.selectbox("발송 강도 지표", ["총 발송 건수", "인당 발송 건수"], key="pc_x_send")
                _sendcol = "totalSend" if _sendlab == "총 발송 건수" else "perSend"
                _sendagg = "sum" if _sendcol == "totalSend" else "mean"

                def _pk(dts):
                    per = "M" if _gran == "월간" else "W-SUN"
                    return dts.dt.to_period(per).apply(lambda p: p.start_time)

                # 동의(선택 그룹·이상치 제외·기간 필터 = pc_clean)를 기간 단위로 집계
                cc = pc_clean.copy()
                cc["_pk"] = _pk(cc["date"])
                cons_g = cc.groupby("_pk").agg(
                    순증감=("diff", "sum"), 신규추가=("added", "sum"),
                    탈퇴=("removed", "sum"), 기말동의=("consent", "last"),
                    동의일수=("date", "count")).reset_index()
                # MTD 발송(같은 기간 필터)을 기간 단위로 집계
                _mdf = mtd_data["df"]
                mm = _mdf[(_mdf["date"] >= _d0) & (_mdf["date"] <= _d1)].copy()
                mm["_pk"] = _pk(mm["date"])
                send_g = mm.groupby("_pk").agg(
                    발송지표=(_sendcol, _sendagg), 총발송=("totalSend", "sum"),
                    CTR=("ctr", "mean"), 발송일수=("date", "count")).reset_index()
                mrg = cons_g.merge(send_g, on="_pk", how="inner").sort_values("_pk")
                # 주간은 양쪽 4일 이상 있는 기간만(부분 주 왜곡 방지)
                if _gran == "주간":
                    mrg = mrg[(mrg["동의일수"] >= 4) & (mrg["발송일수"] >= 4)]

                if len(mrg) < 3:
                    st.warning("두 데이터가 겹치는 기간이 3개 미만이라 크로스 분석을 못 해요. "
                               "발송(MTD)과 동의 데이터의 기간이 겹치는지 확인해 주세요.")
                else:
                    _sfx = "" if _sendcol == "totalSend" else "건"
                    fig_x = overlay_dual(
                        mrg["_pk"], mrg["발송지표"], _sendlab,
                        mrg["순증감"], "동의 순증감(명)",
                        PALETTE["slate"], PALETTE["purple"], h=440,
                        line_suffix="명",
                        title=f"{_gran} {_sendlab}(좌·막대) ↔ 수신동의 순증감(우·선)")
                    st.plotly_chart(fig_x, width="stretch")

                    # 상관 + 산점도(추세선) — 발송 강도 vs 순증감
                    x = mrg["발송지표"].values.astype(float)
                    y = mrg["순증감"].values.astype(float)
                    msk = ~np.isnan(x) & ~np.isnan(y)
                    if msk.sum() >= 3 and np.std(x[msk]) > 0 and np.std(y[msk]) > 0:
                        r = float(np.corrcoef(x[msk], y[msk])[0, 1])
                        sl, ic, _, p, _ = stats.linregress(x[msk], y[msk])
                        xs = np.linspace(x[msk].min(), x[msk].max(), 50)
                        _pk_labels = [pd.Timestamp(d).strftime("%Y-%m-%d")
                                      for d in mrg["_pk"].to_numpy()[msk]]
                        fig_sc = go.Figure()
                        fig_sc.add_trace(go.Scatter(
                            x=x[msk], y=y[msk], mode="markers",
                            marker=dict(color=PALETTE["blue"], size=10),
                            customdata=_pk_labels,
                            hovertemplate="%{customdata}<br>" + _sendlab + ": %{x:,.0f}"
                                          + _sfx + "<br>순증감: %{y:+,.0f}명<extra></extra>",
                            name="기간"))
                        fig_sc.add_trace(go.Scatter(
                            x=xs, y=sl * xs + ic, mode="lines", name="추세",
                            line=dict(color=PALETTE["red"], width=2, dash="dot")))
                        lay_sc = base_layout(h=340, title=f"{_sendlab} vs 순증감 — 기간 단위 산점도")
                        lay_sc["showlegend"] = False
                        lay_sc["yaxis"]["gridcolor"] = "#f1f5f9"
                        lay_sc["yaxis"]["zeroline"] = True
                        lay_sc["yaxis"]["zerolinecolor"] = "#cbd5e1"
                        fig_sc.update_layout(**lay_sc)
                        st.plotly_chart(fig_sc, width="stretch")

                        if r <= -0.3:
                            _msg = ("발송이 많은 기간일수록 순증감이 낮아져요 — <b>발송 피로로 동의가 "
                                    "정체·이탈</b>하는 신호예요. 발송 강도 상한을 검토해 보세요.")
                        elif r >= 0.3:
                            _msg = ("발송이 많은 기간에 순증감도 높아요 — 아직 피로 구간은 아니고 "
                                    "발송이 동의 확보와 함께 가는 편이에요.")
                        else:
                            _msg = "발송 강도와 순증감 사이에 뚜렷한 관계는 약해요."
                        st.markdown(
                            f'<div class="appendix"><b>상관</b> — {_sendlab} ↔ 순증감: r={r:.2f}, '
                            f'{sig_label(p)}. {_msg}<br>상관은 인과가 아니며 시즌·이벤트가 섞일 수 있어요. '
                            '순증감이 마이너스인 기간의 발송량을 특히 함께 보세요.</div>',
                            unsafe_allow_html=True)

                    # 기간별 표 (발송 강도 · 신규추가 · 탈퇴 · 순증감)
                    _xt = mrg.copy()
                    _xt["기간"] = _xt["_pk"].dt.strftime("%Y-%m-%d")
                    _xt["발송지표"] = _xt["발송지표"].map(
                        lambda v: f"{v:,.0f}" + ("" if _sendcol == "totalSend" else "건"))
                    for _c in ("신규추가", "탈퇴", "총발송"):
                        _xt[_c] = _xt[_c].map(lambda v: f"{v:,.0f}")
                    _xt["순증감"] = _xt["순증감"].map(lambda v: f"{v:+,.0f}")
                    _xt["CTR"] = _xt["CTR"].map(lambda v: f"{v*100:.2f}%" if pd.notna(v) else "–")
                    st.dataframe(
                        _xt[["기간", "발송지표", "총발송", "CTR", "신규추가", "탈퇴", "순증감"]]
                        .rename(columns={"발송지표": _sendlab, "총발송": "총 발송(합)"}),
                        hide_index=True, width="stretch", height=300)

                # ── 시차(lead-lag) 탐색 — '이번 주 발송 폭증 → k주 후 이탈'의 선행 신호 ──
                _cw = pc_clean.copy()
                _cw["_wk"] = _cw["date"].dt.to_period("W-SUN").dt.start_time
                _rem_w = _cw.groupby("_wk")["removed"].sum()
                _mmw = _mdf[(_mdf["date"] >= _d0) & (_mdf["date"] <= _d1)].copy()
                _mmw["_wk"] = _mmw["date"].dt.to_period("W-SUN").dt.start_time
                _snd_w = _mmw.groupby("_wk")["totalSend"].sum()
                _joined = pd.concat([_snd_w.rename("send"), _rem_w.rename("removed")],
                                    axis=1).dropna().sort_index()
                if len(_joined) >= 8:
                    _ll_rows = []
                    for _k in range(0, 5):
                        _x = _joined["send"].shift(_k)
                        _y = _joined["removed"]
                        _mk = _x.notna() & _y.notna()
                        if _mk.sum() >= 6 and _x[_mk].std() > 0 and _y[_mk].std() > 0:
                            _r_k = float(np.corrcoef(_x[_mk], _y[_mk])[0, 1])
                            _p_k = float(stats.linregress(_x[_mk], _y[_mk]).pvalue)
                            _ll_rows.append((_k, _r_k, _p_k, int(_mk.sum())))
                    if _ll_rows:
                        st.markdown("##### ⏱ 발송량 → 이탈 시차 탐색 (주간)")
                        _lt = pd.DataFrame(_ll_rows, columns=["시차(주)", "상관 r", "p값", "표본(주)"])
                        _lt["상관 r"] = _lt["상관 r"].map(lambda v: f"{v:+.2f}")
                        _lt["p값"] = _lt["p값"].map(lambda v: f"{v:.3f}")
                        st.dataframe(_lt, hide_index=True, width="stretch")
                        _best_ll = max(_ll_rows, key=lambda t: abs(t[1]))
                        if _best_ll[0] > 0 and _best_ll[1] > 0.3 and _best_ll[2] < 0.1:
                            st.markdown(f'<div class="appendix">발송량 증가가 <b>{_best_ll[0]}주 후</b> '
                                        f'이탈 증가와 가장 강하게 연결돼요 (r={_best_ll[1]:+.2f}, '
                                        f'p={_best_ll[2]:.3f}). 발송 폭증 주 이후 {_best_ll[0]}주간은 '
                                        '이탈 모니터링과 발송 강도 조절을 권장해요.</div>',
                                        unsafe_allow_html=True)
                        else:
                            st.caption("현재 데이터에선 '발송 증가 → 이후 주 이탈 증가'의 뚜렷한 "
                                       "시차 신호가 없어요 (시차 0~4주 상관 모두 약함).")
            _cross_analysis_block()

        # ── 이상치 제외 목록 ──
        with st.expander(f"⚠️ 제외된 이상치 날짜 ({n_outlier}일)"):
            _outlier_rows = pc_all[pc_all["is_outlier"]][["date", "consent", "diff", "added", "removed"]]
            if not _outlier_rows.empty:
                st.dataframe(
                    _outlier_rows.rename(columns={
                        "date": "날짜", "consent": "수신동의수", "diff": "증감",
                        "added": "신규추가", "removed": "기존이탈"}),
                    hide_index=True, width="stretch")
                st.caption(f"신규추가 또는 기존이탈이 {PUSH_OUTLIER_THRESHOLD:,}명 초과한 날짜예요. 배치 이관·대규모 재동의 등 특이 이벤트로 판단해 제외했어요.")
            else:
                st.success("이상치 없음")

        glossary()                                     # 유일하게 빠져 있던 페이지 (개발 규칙)

    # ── 페이지 리포트 다운로드 (HTML → 브라우저 인쇄로 PDF) ──
    if _REPORT:
        st.markdown('<div class="sdiv"></div>', unsafe_allow_html=True)
        st.markdown("##### 📄 이 페이지 리포트 다운로드")
        try:
            _rep_html = build_report_html(str(page), list(_REPORT))
            _safe = re.sub(r'[^0-9A-Za-z가-힣]+', '_', str(page)).strip('_') or "report"
            st.download_button(
                "📄 리포트 다운로드 (HTML)", _rep_html.encode("utf-8"),
                file_name=f"리포트_{_safe}.html", mime="text/html")
            st.caption("받은 HTML 파일을 열고 **Ctrl+P → PDF로 저장**하면 돼요.")
        except Exception as e:
            st.caption(f"리포트 생성 오류: {str(e)[:80]}")




def build_facts(df, with_attr=False, metric_col="ord_cr"):
    """AI에 넘길 사실 요약 텍스트 생성 (Streamlit 비의존)."""
    if df is None or len(df) == 0:
        return "데이터 없음"
    lines = []
    _ds = sorted(x for x in df["date"].dropna().astype(str) if re.match(r'^\d{8}$', x))
    _rng = f"{_ds[0]} ~ {_ds[-1]}" if _ds else "–"
    lines.append(f"[기간] {_rng} · 캠페인 {len(df)}건")
    lines.append(f"[평균] 유입전환율 {df['infl_cr'].mean()*100:.2f}% · "
                 f"주문전환율 {df['ord_cr'].mean()*100:.2f}% · "
                 f"RPS {df['rps'].mean():,.0f}원 · 객단가 {df['aov'].mean():,.0f}원")
    def _bodyprev(r):
        b = " ".join(_s(r.get("body", "")).split())
        return f" │ 내용: {b[:50]}" if b else ""
    # 상·하위 선정: UV 100 이상(충분할 때) + 표본 보정 순위 — AI가 소표본 요행 캠페인을
    # '성공/실패 공식'으로 학습해 지어낸 패턴을 서술하는 것을 막는다
    rank_base = df
    if "uv" in df:
        _big = df[df["uv"].fillna(0) >= 100]
        if len(_big) >= 12:
            rank_base = _big
            lines.append("[순위 기준] UV 100 이상 캠페인만 · 표본 크기 보정 순위 (소표본 요행 제외)")
    _rk_top = rank_adjusted(rank_base, metric_col, ascending=False)
    _rk_bot = rank_adjusted(rank_base, metric_col, ascending=True)
    win = rank_base.assign(_rk=_rk_top).sort_values("_rk", ascending=False).head(6)
    los = rank_base.assign(_rk=_rk_bot).sort_values("_rk").head(6)

    def _send_n(r):
        v = r.get("send")
        return 0 if (v is None or pd.isna(v)) else int(v)   # 최소 발송수 0이면 NaN 발송 행이 섞임
    lines.append("\n[주문전환율 상위 문구] (제목 │ 내용)")
    for _, r in win.iterrows():
        lines.append(f" - CR {r['ord_cr']*100:.2f}% / 발송 {_send_n(r):,} / {r['cat']} / “{r['title']}”{_bodyprev(r)}")
    lines.append("\n[주문전환율 하위 문구] (제목 │ 내용)")
    for _, r in los.iterrows():
        lines.append(f" - CR {r['ord_cr']*100:.2f}% / 발송 {_send_n(r):,} / {r['cat']} / “{r['title']}”{_bodyprev(r)}")
    if with_attr:
        lines.append("\n[문구 속성별 평균 주문전환율 (보유 vs 미보유)] — '유의'가 붙지 않은 차이는 "
                     "우연일 수 있으니 단정적으로 서술하지 말 것")

        def _welch_p(a, b):
            try:
                from scipy import stats as _ss
                a = np.asarray(a, float); a = a[~np.isnan(a)]
                b = np.asarray(b, float); b = b[~np.isnan(b)]
                if len(a) < 3 or len(b) < 3:
                    return np.nan
                return float(_ss.ttest_ind(a, b, equal_var=False).pvalue)
            except Exception:
                return np.nan
        _rows = []
        for tag in TAG_BOOLS:
            if tag not in df: continue
            yes = df[df[tag]]["ord_cr"]; no = df[~df[tag]]["ord_cr"]
            if len(yes) and len(no):
                _rows.append((tag, yes, no, _welch_p(yes.values, no.values)))
        _adj = fdr_bh([r[3] for r in _rows]) if _rows else []
        for (tag, yes, no, _p), pa in zip(_rows, _adj):
            mark = "[유의함]" if (not np.isnan(pa) and pa < 0.05) else "(유의하지 않음·참고만)"
            lines.append(f" - {tag}: 보유 {yes.mean()*100:.2f}%(n={len(yes)}) vs "
                         f"미보유 {no.mean()*100:.2f}%(n={len(no)}) "
                         f"→ {(yes.mean()-no.mean())*100:+.2f}%p {mark}")
    return "\n".join(lines)


def df_to_xlsx_bytes(d, sheet_name="머지데이터"):
    """DataFrame → 단일 시트 엑셀(bytes). 발송기획+실적 머지 전체 데이터 다운로드용."""
    import io as _io
    buf = _io.BytesIO()
    out = d.drop(columns=["dt"], errors="ignore")
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        out.to_excel(xw, sheet_name=sheet_name, index=False)
    return buf.getvalue()


def build_report_excel(fdf):
    """분석 결과를 다중 시트 엑셀(bytes)로 — 머지데이터/속성별/조합/키워드/이모지/카테고리별."""
    import io as _io
    import itertools as _it
    buf = _io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        fdf.drop(columns=["dt"], errors="ignore").to_excel(xw, sheet_name="머지데이터", index=False)
        # 속성별 성과
        rows = []
        for t in TAG_BOOLS:
            if t not in fdf:
                continue
            y = fdf[fdf[t]]; n = fdf[~fdf[t]]
            if not len(y) or not len(n):
                continue
            rows.append(dict(속성=t, 보유n=len(y), 보유_CTR=y["infl_cr"].mean(),
                             보유_주문CR=y["ord_cr"].mean(), 보유_RPS=y["rps"].mean(),
                             미보유n=len(n), 미보유_주문CR=n["ord_cr"].mean(),
                             주문CR차이=y["ord_cr"].mean() - n["ord_cr"].mean()))
        if rows:
            pd.DataFrame(rows).sort_values("주문CR차이", ascending=False).to_excel(
                xw, sheet_name="속성별성과", index=False)
        # 속성 2개 조합
        bm = fdf["ord_cr"].mean() if "ord_cr" in fdf else np.nan
        crows = []
        for a, b in _it.combinations([t for t in TAG_BOOLS if t in fdf], 2):
            mask = fdf[a] & fdf[b]
            s = fdf[mask]["ord_cr"].dropna()
            if len(s) < 5:
                continue
            crows.append(dict(조합=f"{a}+{b}", 캠페인수=len(s), 주문CR=s.mean(), 차이=s.mean() - bm))
        if crows:
            pd.DataFrame(crows).sort_values("주문CR", ascending=False).to_excel(
                xw, sheet_name="속성조합", index=False)
        # 키워드 / 이모지
        kp = keyword_perf(fdf, "ord_cr", min_n=5, top=50)
        if len(kp):
            kp.to_excel(xw, sheet_name="키워드성과", index=False)
        ep = emoji_perf(fdf, "ord_cr", min_n=3, top=50)
        if len(ep):
            ep.to_excel(xw, sheet_name="이모지성과", index=False)
        # 카테고리별
        if "cat" in fdf and len(fdf):
            cg = fdf.groupby("cat").agg(캠페인수=("af", "size"), 발송=("send", "sum"),
                                        CTR=("infl_cr", "mean"), 주문CR=("ord_cr", "mean"),
                                        RPS=("rps", "mean"), 거래액=("amt", "sum")).reset_index()
            cg.rename(columns={"cat": "카테고리"}).to_excel(xw, sheet_name="카테고리별", index=False)
    return buf.getvalue()


if __name__ == "__main__":
    main()
