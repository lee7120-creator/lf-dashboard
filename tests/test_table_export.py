"""표 엑셀 내보내기 테스트 — 화면 서식이 엑셀 숫자 서식으로 넘어가는지.

값을 문자열로 넣으면 받는 쪽에서 정렬·합계가 안 된다. 그래서 **숫자 + 표시형식**으로
써야 하는데, pandas Styler는 서식을 엑셀로 안 옮겨 준다(전부 General). 화면 표시
문자열에서 형식을 역으로 추론하는 부분이라 조용히 깨지기 쉽다.

로컬 실행:
    python tests/test_table_export.py
"""
import io
import pathlib
import sys

import pandas as pd

ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

from table_export import _css_color, _xl_numfmt, xlsx_bytes  # noqa: E402

CASES = []


def case(fn):
    CASES.append(fn)
    return fn


def _ws(b, name=None):
    wb = openpyxl.load_workbook(io.BytesIO(b))
    return wb[name] if name else wb.active


@case
def t_numfmt_inference():
    """화면 문자열 → 엑셀 표시형식."""
    assert _xl_numfmt(["14,843,856", "461,799"]) == "#,##0"
    assert _xl_numfmt(["3.11%", "0.59%"]) == "0.00%"
    assert _xl_numfmt(["+5.3%", "-0.9%"]) == "+0.0%;-0.0%"
    assert _xl_numfmt(["+1,234", "-99"]) == "+#,##0;-#,##0"
    assert _xl_numfmt(["2.90억", "3.92억"]) is None       # 숫자로 못 읽으면 문자로
    assert _xl_numfmt(["–", ""]) is None


@case
def t_styler_format_becomes_excel_format():
    """Styler의 서식이 엑셀 숫자 서식으로 옮겨지고, 값은 숫자로 들어간다."""
    df = pd.DataFrame({"지표": ["발송", "CTR"], "값": [14843856, 461799],
                       "전주비": [0.053, -0.0009]})
    b = xlsx_bytes(df.style.format({"값": "{:,.0f}", "전주비": "{:+.1%}"}), "t")
    ws = _ws(b)
    assert ws.cell(2, 2).value == 14843856, ws.cell(2, 2).value
    assert ws.cell(2, 2).number_format == "#,##0"
    assert abs(ws.cell(2, 3).value - 0.053) < 1e-9
    assert ws.cell(2, 3).number_format == "+0.0%;-0.0%"


@case
def t_dark_header_style():
    """헤더는 검은 배경 + 흰 굵은 글씨."""
    ws = _ws(xlsx_bytes(pd.DataFrame({"a": [1, 2]}), "t"))
    h = ws.cell(1, 1)
    assert h.fill.fgColor.rgb.endswith("111827"), h.fill.fgColor.rgb
    assert h.font.color.rgb.endswith("FFFFFF")
    assert h.font.bold


@case
def t_title_row_and_freeze():
    """제목을 주면 위에 붙고, 헤더 아래로 틀이 고정된다."""
    ws = _ws(xlsx_bytes(pd.DataFrame({"a": [1]}), "t", title="주요 지표 현황"))
    assert ws.cell(1, 1).value == "주요 지표 현황"
    assert ws.cell(3, 1).value == "a"                     # 제목 1행 + 빈 줄 → 헤더 3행
    assert ws.freeze_panes == "A4"


@case
def t_index_export():
    """축을 인덱스에 세운 표는 인덱스도 같이 나가야 뜻이 통한다."""
    df = pd.DataFrame({"거래액": [120000000]},
                      index=pd.Index(["1월 1주차"], name="주차"))
    ws = _ws(xlsx_bytes(df, "t", index=True))
    assert [ws.cell(1, c).value for c in (1, 2)] == ["주차", "거래액"]
    assert ws.cell(2, 1).value == "1월 1주차"


@case
def t_text_column_stays_text():
    """숫자로 못 읽는 칸은 화면 문자열 그대로 — 엉뚱한 숫자로 바꾸지 않는다."""
    df = pd.DataFrame({"거래액": ["2.90억", "3.92억"], "유의성": ["유의함", "–"]})
    ws = _ws(xlsx_bytes(df, "t"))
    assert ws.cell(2, 1).value == "2.90억"
    assert ws.cell(3, 2).value == "–"


@case
def t_nan_and_empty_safe():
    """NaN·빈 프레임에서 죽지 않는다."""
    import numpy as np
    df = pd.DataFrame({"a": [1.0, np.nan], "b": ["x", None]})
    ws = _ws(xlsx_bytes(df, "t"))
    assert ws.cell(3, 1).value in (None, "")
    assert xlsx_bytes(pd.DataFrame(), "t")                # 빈 표도 파일은 나온다


@case
def t_css_color_parsing():
    """CSS 색 표기 → openpyxl RGB."""
    assert _css_color("color:#dc2626".split(":")[1]) == "DC2626"
    assert _css_color("#16a34a") == "16A34A"
    assert _css_color("#f00") == "FF0000"
    assert _css_color("rgb(220, 38, 38)") == "DC2626"
    assert _css_color("red") == "FF0000"
    assert _css_color("") is None and _css_color("chartreuse-ish") is None


@case
def t_delta_font_color_carries_to_excel():
    """화면의 증감 색(빨강 △ / 초록 +)이 엑셀 글자색으로 그대로 넘어간다.

    Styler._translate()는 끝나면서 ctx를 비운다. 표시값을 먼저 읽고 색을 나중에
    읽으면 전부 기본색으로 나간다 — 조용히 깨지는 회귀라 여기서 못 박아 둔다.
    """
    df = pd.DataFrame({"지표": ["캠페인수", "발송", "CTR"],
                       "기준주": [93, 14843856, 0.0311],
                       "전주비": ["△21.8%", "+5.3%", "–"]})

    def _clr(v):
        s = str(v)
        if s.startswith("△"):
            return "color:#dc2626;font-weight:600"
        if s.startswith("+"):
            return "color:#16a34a;font-weight:600"
        return ""

    ws = _ws(xlsx_bytes(df.style.map(_clr, subset=["전주비"]).format({"기준주": "{:,.0f}"}), "t"))
    assert ws.cell(2, 3).font.color.rgb.endswith("DC2626"), ws.cell(2, 3).font.color.rgb
    assert ws.cell(2, 3).font.bold
    assert ws.cell(3, 3).font.color.rgb.endswith("16A34A")
    assert not ws.cell(4, 3).font.bold                    # 색 없는 칸은 기본 글자
    assert ws.cell(2, 2).number_format == "#,##0"         # 색을 입혀도 숫자 서식은 유지


@case
def t_cell_background_beats_zebra():
    """화면에서 칠한 셀 배경은 줄무늬보다 우선한다."""
    df = pd.DataFrame({"a": [1, 2]})
    ws = _ws(xlsx_bytes(df.style.map(lambda v: "background-color:#fff3cd" if v == 2 else ""), "t"))
    assert ws.cell(3, 1).fill.fgColor.rgb.endswith("FFF3CD"), ws.cell(3, 1).fill.fgColor.rgb


@case
def t_color_column_shift_with_index():
    """인덱스를 앞에 끼워도 색이 원래 칸에 붙어 있어야 한다."""
    df = pd.DataFrame({"전주비": ["△9.0%"]}, index=pd.Index(["발송"], name="지표"))
    styled = df.style.map(lambda v: "color:#dc2626")
    ws = _ws(xlsx_bytes(styled, "t", index=True))
    assert ws.cell(2, 1).value == "발송"
    assert ws.cell(2, 2).font.color.rgb.endswith("DC2626"), ws.cell(2, 2).font.color.rgb


@case
def t_sheet_name_sanitized():
    """엑셀이 금지하는 문자가 시트명에 들어가도 저장된다."""
    ws = _ws(xlsx_bytes(pd.DataFrame({"a": [1]}), "① 거래액 드라이버: Q vs P [비교]/2"))
    assert "[" not in ws.title and ":" not in ws.title and "/" not in ws.title
    assert len(ws.title) <= 31


def main():
    fails = []
    for fn in CASES:
        try:
            fn()
            print(f"  OK   {fn.__name__}")
        except Exception as e:                            # noqa: BLE001
            print(f"  FAIL {fn.__name__}: {e}")
            fails.append(fn.__name__)
    print()
    if fails:
        print(f"실패 {len(fails)}건: {fails}")
        return 1
    print(f"엑셀 내보내기 테스트 {len(CASES)}건 통과 ✅")
    return 0


if __name__ == "__main__":
    sys.exit(main())
